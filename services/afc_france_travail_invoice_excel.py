from __future__ import annotations

import io
import re
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

MODULE_ORDER = ["FT", "SP", "RAN", "PAF", "FESTE"]
MODULE_LABELS = {
    "FT": "Formation professionnelle ou technique",
    "SP": "Appui pédagogique ou Soutien pédagogique",
    "RAN": "RAN",
    "PAF": "Préparation à l’après-formation",
    "FESTE": "FESTE",
}
INVOICE_ROWS = range(33, 38)
ERROR_MARKERS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A")
MONEY_FORMAT = '# ##0,00 [$€-fr-FR]'


def dec(value: Any) -> Decimal:
    try:
        return Decimal(str(value or 0).replace(",", ".")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(f"Valeur numérique invalide : {value!r}") from exc


def load_invoice_template(app_root: str | Path):
    path = Path(app_root) / "static" / "upload" / "facture.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"Modèle de facture France Travail introuvable : {path}")
    return load_workbook(path)


def select_invoice_template_sheet(wb):
    for ws in wb.worksheets:
        if str(ws["A2"].value or "").strip().lower() == "organisme dispensateur" and "facture" in ws.title.lower():
            return ws
    return wb[wb.sheetnames[0]]


def clear_invoice_template_data(ws):
    for cell in ["H2", "H4", "D6", "H6", "D22", "G22", "D24", "B26", "B28", "D28", "B30", "L38", "M38"]:
        ws[cell].value = None
    # dynamic lines and totals; keep labels/borders/styles
    for row in INVOICE_ROWS:
        for col in range(1, 10):
            c = ws.cell(row, col)
            if isinstance(c, MergedCell):
                continue
            c.value = None
            if col >= 3:
                c.number_format = "General"
    for cell in ["D38", "E38", "F38", "G38", "H38", "I38", "C42", "F42"]:
        ws[cell].value = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and any(marker in cell.value for marker in ERROR_MARKERS):
                cell.value = None


def fmt_date(value: Any) -> str:
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")


def write_hour(cell, value: Any):
    number = dec(value)
    if number == number.to_integral_value():
        cell.value = int(number)
        cell.number_format = "0"
    else:
        cell.value = float(number)
        cell.number_format = "0.##"


def write_money(cell, value: Any):
    cell.value = float(dec(value))
    cell.number_format = MONEY_FORMAT


def _module_total(dsf_snapshot: dict[str, Any], module: str) -> Decimal:
    totals = dsf_snapshot.get("moduleTotals") or {}
    if module in totals:
        return dec(totals.get(module))
    return sum((dec((st.get("modules") or {}).get(module)) for st in (dsf_snapshot.get("students") or [])), Decimal("0")).quantize(Decimal("0.01"))


def build_invoice_snapshot(session_data: dict[str, Any], dsf: dict[str, Any], invoice_data: dict[str, Any], created_by: str = "admin") -> dict[str, Any]:
    dsf_snapshot = dsf.get("franceTravailExcelSnapshot") or {}
    if not dsf_snapshot:
        raise ValueError("Snapshot DSF France Travail absent : impossible de générer la facture.")
    rate = dec(dsf_snapshot.get("hourlyRate") or invoice_data.get("hourlyRate"))
    modules = []
    for code in MODULE_ORDER:
        hours = _module_total(dsf_snapshot, code)
        distance = sum((dec(st.get("distanceHours")) if code != "FESTE" else Decimal("0") for st in dsf_snapshot.get("students") or []), Decimal("0")) if code in (dsf.get("modules") or []) else Decimal("0")
        non_billable = Decimal("0")
        absences = Decimal("0")
        total = hours + absences
        amount = (total * rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if hours or distance or non_billable or absences:
            modules.append({"code": code, "label": MODULE_LABELS[code], "billable_hours": str(hours), "distance_hours": str(distance), "non_billable_hours": str(non_billable), "justified_absence_hours": str(absences), "total_hours": str(total), "unit_price": str(rate), "amount": str(amount)})
    amount_total = sum((dec(m["amount"]) for m in modules), Decimal("0")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    dsf_amount = dec(dsf.get("amountTotal") or dsf_snapshot.get("amountTotal") or amount_total)
    snapshot = {
        "invoice_id": f"invoice:{dsf.get('id')}", "dsf_id": dsf.get("id"), "session_id": session_data.get("id"),
        "invoice_number": str(invoice_data.get("invoice_number") or "").strip(), "dsf_number": str(dsf.get("number") or dsf_snapshot.get("number") or ""),
        "dsf_reference": str(invoice_data.get("kairos_engagement_reference") or "").strip(), "kairos_engagement_reference": str(invoice_data.get("kairos_engagement_reference") or "").strip(),
        "market_number": str(((dsf_snapshot.get("session") or {}).get("france_travail") or {}).get("marche_afc") or ""),
        "invoice_date": str(invoice_data.get("invoice_date") or datetime.now().date().isoformat()), "invoice_place": str(invoice_data.get("invoice_place") or "").strip(),
        "invoice_type": invoice_data.get("invoice_type") or "intermediate", "period_start": dsf_snapshot.get("periodStart"), "period_end": dsf_snapshot.get("periodEnd"),
        "session_start": (dsf_snapshot.get("session") or {}).get("date_debut"), "session_end": (dsf_snapshot.get("session") or {}).get("date_fin"),
        "title": (dsf_snapshot.get("session") or {}).get("intitule") or (dsf_snapshot.get("session") or {}).get("name") or "",
        "execution_place": (dsf_snapshot.get("session") or {}).get("lieu") or "", "hourly_rate": str(rate), "modules": modules,
        "student_count": dsf_snapshot.get("studentCount") or len(dsf_snapshot.get("students") or []), "total_hours": str(sum((dec(m["total_hours"]) for m in modules), Decimal("0"))),
        "amount_total": str(amount_total), "amount_total_words": amount_to_french_words(amount_total),
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "created_by": created_by,
        "organization": (dsf_snapshot.get("session") or {}), "dsf_snapshot": dsf_snapshot,
    }
    validate_invoice_against_dsf(snapshot, dsf)
    return snapshot


def validate_invoice_against_dsf(invoice_snapshot: dict[str, Any], dsf: dict[str, Any]):
    if dec(invoice_snapshot.get("amount_total")) != dec(dsf.get("amountTotal")):
        raise ValueError(f"Montant facture ({invoice_snapshot.get('amount_total')} €) différent du montant DSF ({dsf.get('amountTotal')} €).")


def populate_invoice_references(ws, snap):
    ws["H2"] = snap.get("kairos_engagement_reference") or snap.get("dsf_reference") or ""; ws["H2"].number_format = "@"
    ws["H4"] = snap.get("invoice_number") or ""; ws["H4"].number_format = "@"
    ws["D6"] = f"A {snap.get('invoice_place') or ''}".strip(); ws["H6"] = fmt_date(snap.get("invoice_date"))
    ws["H13"] = snap.get("market_number") or ws["H13"].value
    ws["D24"] = snap.get("dsf_reference") or f"{snap.get('kairos_engagement_reference','')}_N{snap.get('dsf_number','')}".strip("_")


def populate_invoice_period(ws, snap):
    ws["D22"] = fmt_date(snap.get("period_start")); ws["G22"] = fmt_date(snap.get("period_end"))


def populate_invoice_session(ws, snap):
    ws["B26"] = snap.get("title") or ""
    ws["B28"] = fmt_date(snap.get("session_start")); ws["D28"] = fmt_date(snap.get("session_end"))
    ws["B30"] = snap.get("execution_place") or ""


def populate_invoice_lines(ws, snap):
    for row, module in zip(INVOICE_ROWS, snap.get("modules") or []):
        ws.cell(row, 1).value = module["label"]
        write_hour(ws.cell(row, 3), module["billable_hours"]); write_hour(ws.cell(row, 4), module["distance_hours"])
        write_hour(ws.cell(row, 5), module["non_billable_hours"]); write_hour(ws.cell(row, 6), module["justified_absence_hours"])
        write_hour(ws.cell(row, 7), module["total_hours"]); write_money(ws.cell(row, 8), module["unit_price"]); write_money(ws.cell(row, 9), module["amount"])


def populate_invoice_totals(ws, snap):
    mods = snap.get("modules") or []
    totals = {k: sum((dec(m[k]) for m in mods), Decimal("0")) for k in ["distance_hours", "non_billable_hours", "justified_absence_hours", "total_hours"]}
    write_hour(ws["D38"], totals["distance_hours"]); write_hour(ws["E38"], totals["non_billable_hours"]); write_hour(ws["F38"], totals["justified_absence_hours"]); write_hour(ws["G38"], totals["total_hours"])
    write_money(ws["H38"], snap.get("hourly_rate")); write_money(ws["I38"], snap.get("amount_total")); write_money(ws["C42"], snap.get("amount_total")); ws["F42"] = snap.get("amount_total_words")


def mark_invoice_type(ws, snap):
    ws["A22"] = ("☒ " if snap.get("invoice_type") == "intermediate" else "☐ ") + "Facture intermédiaire"
    ws["I22"] = ("☒ " if snap.get("invoice_type") == "final" else "☐ ") + "Facture de solde"


def configure_invoice_print_settings(ws):
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "portrait"; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.print_area = "A1:J51"


def generate_invoice_excel_from_snapshot(snapshot: dict[str, Any], app_root: str | Path):
    wb = load_invoice_template(app_root); ws = select_invoice_template_sheet(wb)
    for other in list(wb.worksheets):
        if other is not ws: wb.remove(other)
    ws.title = (f"FACTURE DSF{snapshot.get('dsf_number') or ''}")[:31]
    clear_invoice_template_data(ws); populate_invoice_references(ws, snapshot); populate_invoice_period(ws, snapshot); populate_invoice_session(ws, snapshot)
    mark_invoice_type(ws, snapshot); populate_invoice_lines(ws, snapshot); populate_invoice_totals(ws, snapshot); configure_invoice_print_settings(ws)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0); return bio


def generate_invoice_excel(snapshot: dict[str, Any], app_root: str | Path):
    return generate_invoice_excel_from_snapshot(snapshot, app_root)


def amount_to_french_words(amount: Any) -> str:
    amount = dec(amount); euros = int(amount); cents = int((amount - Decimal(euros)) * 100)
    words = _int_to_fr(euros) + (" EURO" if euros == 1 else " EUROS")
    if cents:
        words += " ET " + _int_to_fr(cents) + (" CENTIME" if cents == 1 else " CENTIMES")
    return words.upper()


def _int_to_fr(n: int) -> str:
    units = ["zéro","un","deux","trois","quatre","cinq","six","sept","huit","neuf","dix","onze","douze","treize","quatorze","quinze","seize"]
    tens = {20:"vingt",30:"trente",40:"quarante",50:"cinquante",60:"soixante"}
    if n < 17: return units[n]
    if n < 20: return "dix-" + units[n-10]
    if n < 70:
        t,u = divmod(n,10); base=tens[t*10]
        return base if u==0 else base + (" et " if u==1 else "-") + units[u]
    if n < 80: return "soixante-" + _int_to_fr(n-60)
    if n < 100:
        return "quatre-vingts" if n==80 else "quatre-vingt-" + _int_to_fr(n-80)
    if n < 1000:
        h,r=divmod(n,100); head="cent" if h==1 else units[h]+" cent"+("s" if r==0 else "")
        return head if r==0 else head+" "+_int_to_fr(r)
    if n < 1000000:
        th,r=divmod(n,1000); head="mille" if th==1 else _int_to_fr(th)+" mille"
        return head if r==0 else head+" "+_int_to_fr(r)
    m,r=divmod(n,1000000); head=_int_to_fr(m)+(" million" if m==1 else " millions")
    return head if r==0 else head+" "+_int_to_fr(r)


def safe_filename_part(value: str) -> str:
    value = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-").lower() or "facture"


def invoice_excel_filename(snapshot: dict[str, Any]) -> str:
    return f"facture_france_travail_{safe_filename_part(snapshot.get('invoice_number'))}_dsf_{safe_filename_part(snapshot.get('dsf_number'))}.xlsx"
