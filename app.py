import os
import json
import uuid
import base64
import time
import tempfile
import zipfile
import hashlib
import hmac
import importlib.util
import smtplib
import urllib.parse
import urllib.request
import urllib.error
import re
import shutil
import subprocess
import secrets
from io import BytesIO
from datetime import datetime, timedelta, date, time as dt_time
from functools import wraps
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import logging
import math
import threading

from flask import (
    Flask, render_template, request, redirect, url_for,
    abort, flash, send_file, send_from_directory, session, Response, jsonify
)
from werkzeug.utils import secure_filename

from yousign_service import YousignClient, YousignError, detect_yousign_environment, get_yousign_config, is_yousign_configured, mask_phone_number, normalizeFrenchPhoneNumber, sanitize_yousign_external_id, test_yousign_connection, yousign_config_diagnostics, yousign_service_access_message

from prospecting import prospecting_bp
from a3p_program import A3P_TOTAL_HOURS, A3P_MODULES, A3P_FORBIDDEN_TERMS, generateA3pSchedule, validate_a3p_planning, is_a3p_non_working_day
from desp_program import DESP_LABEL, DESP_TOTAL_HOURS, DESP_ELEARNING_HOURS, DESP_PRESENTIEL_HOURS, generate_desp_planning, desp_summary_from_planning



# --- 🔧 Forcer le fuseau horaire français ---
os.environ['TZ'] = 'Europe/Paris'
import time
time.tzset()

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s %(message)s")

app = Flask(__name__)
app.register_blueprint(prospecting_bp)
app.secret_key = os.environ.get("SECRET_KEY", "change-me")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get("DATA_DIR", "/mnt/data")
os.makedirs(DATA_DIR, exist_ok=True)
SHORTCUTS_DATA_DIR = os.path.join(DATA_DIR, "shortcuts")
SHORTCUTS_FILE = os.path.join(SHORTCUTS_DATA_DIR, "shortcuts.json")
SHORTCUT_UPLOAD_DIR = os.path.join(SHORTCUTS_DATA_DIR, "images")
LEGACY_SHORTCUTS_FILE = os.path.join(BASE_DIR, "data", "shortcuts.json")
LEGACY_SHORTCUT_UPLOAD_DIR = os.path.join(BASE_DIR, "static", "uploads", "shortcuts")
ALLOWED_SHORTCUT_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}

logger = logging.getLogger("jury-notify")

yousign_startup_diagnostic = yousign_config_diagnostics()
logging.getLogger("yousign").info(
    "Yousign configuration detected environment=%s base_url=%s api_key=%s workspace_id_present=%s",
    yousign_startup_diagnostic["environment"],
    yousign_startup_diagnostic["base_url"],
    yousign_startup_diagnostic["api_key"],
    yousign_startup_diagnostic["workspace_id_present"],
)

from datetime import timedelta

IS_RENDER = os.environ.get("RENDER", "").lower() == "true"

# ✅ Cookies de session uniquement : l’utilisateur doit se reconnecter
# à chaque nouvelle ouverture du navigateur / nouvelle session.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=IS_RENDER,   # ✅ Secure seulement sur Render
)



ADMIN_USER = os.environ.get("ADMIN_USER")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")
ADMIN_SESSION_VERSION = os.environ.get("ADMIN_SESSION_VERSION", "2026-07-10-force-relogin-v2")

def shortcut_image_url(filename):
    return url_for("shortcut_image", filename=filename)


def normalize_shortcut_image(shortcut):
    image = shortcut.get("image") or ""
    filename = os.path.basename(urllib.parse.urlparse(image).path)
    if filename:
        shortcut["image"] = shortcut_image_url(filename)
    return shortcut


def migrate_legacy_shortcuts_storage():
    migrated = False

    if os.path.exists(LEGACY_SHORTCUT_UPLOAD_DIR):
        for entry in os.scandir(LEGACY_SHORTCUT_UPLOAD_DIR):
            if not entry.is_file():
                continue

            destination = os.path.join(SHORTCUT_UPLOAD_DIR, entry.name)
            if os.path.exists(destination):
                continue

            with open(entry.path, "rb") as source_file, open(destination, "wb") as destination_file:
                destination_file.write(source_file.read())
            migrated = True

    if os.path.exists(LEGACY_SHORTCUTS_FILE):
        try:
            with open(LEGACY_SHORTCUTS_FILE, "r", encoding="utf-8") as f:
                legacy_shortcuts = json.load(f)
        except (json.JSONDecodeError, FileNotFoundError):
            legacy_shortcuts = []

        if isinstance(legacy_shortcuts, list):
            normalized_shortcuts = []
            for shortcut in legacy_shortcuts:
                if not isinstance(shortcut, dict):
                    continue
                normalized_shortcuts.append(normalize_shortcut_image(shortcut))

            if normalized_shortcuts:
                with open(SHORTCUTS_FILE, "w", encoding="utf-8") as f:
                    json.dump(normalized_shortcuts, f, ensure_ascii=False, indent=2)
                migrated = True

    return migrated


def ensure_shortcuts_storage():
    os.makedirs(SHORTCUTS_DATA_DIR, exist_ok=True)
    os.makedirs(SHORTCUT_UPLOAD_DIR, exist_ok=True)
    if not os.path.exists(SHORTCUTS_FILE):
        if not migrate_legacy_shortcuts_storage():
            with open(SHORTCUTS_FILE, "w", encoding="utf-8") as f:
                json.dump([], f, ensure_ascii=False, indent=2)


def load_shortcuts():
    ensure_shortcuts_storage()
    try:
        with open(SHORTCUTS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            shortcuts = data if isinstance(data, list) else []
    except (FileNotFoundError, json.JSONDecodeError):
        return []

    updated = False
    for shortcut in shortcuts:
        if not isinstance(shortcut, dict):
            continue
        if not shortcut.get("id"):
            shortcut["id"] = uuid.uuid4().hex
            updated = True
        existing_image = shortcut.get("image")
        normalize_shortcut_image(shortcut)
        if shortcut.get("image") != existing_image:
            updated = True

    if updated:
        save_shortcuts(shortcuts)

    return shortcuts


def save_shortcuts(shortcuts):
    ensure_shortcuts_storage()
    with open(SHORTCUTS_FILE, "w", encoding="utf-8") as f:
        json.dump(shortcuts, f, ensure_ascii=False, indent=2)


def allowed_shortcut_image(filename):
    if "." not in filename:
        return False
    return filename.rsplit(".", 1)[1].lower() in ALLOWED_SHORTCUT_IMAGE_EXTENSIONS


STAGIAIRES_DOCS_TO_CONTROL_URL = os.environ.get(
    "STAGIAIRES_DOCS_TO_CONTROL_URL",
    "https://gestionstagiaires-r5no.onrender.com/docs_to_control.json",
)
STAGIAIRES_DOCS_RETRY_SECONDS = 60
_stagiaires_docs_cache = {"payload": None, "retry_after": 0.0}
_stagiaires_docs_cache_lock = threading.Lock()


def stagiaires_docs_request_headers():
    headers = {
        "Accept": "application/json",
        "User-Agent": "plateformegestion/1.0 (+https://plateformegestion.onrender.com)",
    }
    token = (os.environ.get("STAGIAIRES_DOCS_TO_CONTROL_TOKEN") or "").strip()
    if token:
        headers["Authorization"] = f"Bearer {token}"
        headers["X-API-Key"] = token
    return headers


def fetch_json_url(url, timeout=10, headers=None):
    request_obj = urllib.request.Request(
        url,
        headers=headers or {
            "Accept": "application/json",
            "User-Agent": "plateformegestion/1.0 (+https://plateformegestion.onrender.com)",
        },
    )
    with urllib.request.urlopen(request_obj, timeout=timeout) as response:
        charset = response.headers.get_content_charset() or "utf-8"
        body = response.read().decode(charset)
        return json.loads(body)


def count_pending_stagiaires_documents(payload):
    if not isinstance(payload, dict):
        return 0

    for key in ("pending_count", "docs_to_control", "documents_to_control", "total", "count"):
        value = payload.get(key)
        if isinstance(value, (int, float)):
            return max(int(value), 0)
        if isinstance(value, str) and value.strip().isdigit():
            return int(value.strip())

    items = payload.get("items")
    if not isinstance(items, list):
        return 0

    total = 0
    for item in items:
        if not isinstance(item, dict):
            continue
        pending_count = item.get("pending_count", 0)
        try:
            total += max(int(pending_count), 0)
        except (TypeError, ValueError):
            continue
    return total


def stagiaires_docs_response(payload, stale=False):
    return {
        "ok": True,
        "stale": stale,
        "pending_count": count_pending_stagiaires_documents(payload),
        "items": payload.get("items", []),
    }


# ------------------------------------------------------------
# 🔐 AUTHENTIFICATION ADMIN
# ------------------------------------------------------------

def is_admin_session_valid():
    return (
        session.get("admin_logged")
        and session.get("admin_session_version") == ADMIN_SESSION_VERSION
    )


def require_fresh_admin_session(next_path):
    if is_admin_session_valid():
        return None

    session.clear()
    return redirect(url_for("login", next=next_path))


def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):

        # ✅ Autoriser le lien formateur public avec token
        if request.path.startswith("/formateurs/") and "/upload" in request.path:
            return f(*args, **kwargs)

        # 🔐 Vérification session admin
        redirect_response = require_fresh_admin_session(request.path)
        if redirect_response:
            return redirect_response

        return f(*args, **kwargs)
    return wrapped


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()

        if email == ADMIN_USER and password == ADMIN_PASSWORD:
            session.clear()
            session.permanent = False       # ✅ session navigateur : pas de connexion persistante
            session["admin_logged"] = True
            session["admin_email"] = email
            session["admin_session_version"] = ADMIN_SESSION_VERSION
            return redirect(request.args.get("next") or url_for("index"))


        flash("Identifiant ou mot de passe incorrect.", "error")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.before_request
def protect_all_routes():
    path = request.path

    # ✅ autoriser page login / logout
    if path.startswith("/login") or path.startswith("/logout"):
        return None

    # ✅ autoriser les fichiers statiques (css/js/images)
    if path.startswith("/static/"):
        return None

    # ✅ autoriser les webhooks (Salesforce, etc.)
    if path.startswith("/webhooks/"):
        return None

    # ✅ autoriser lien public formateur (upload avec token)
    if path.startswith("/formateurs/") and "/upload" in path:
        return None

    # ✅ autoriser réponses jury (lien email)
    if path.startswith("/jury-response/"):
        return None

    # ✅ autoriser lien public A3P formateur
    if path.startswith("/public/a3p-planning/") or path.startswith("/api/public/a3p-planning/"):
        return None

    # ✅ autoriser accès préfecture (auth basic gérée dans la route)
    if path.startswith("/prefecture/"):
        return None

    # ✅ autoriser les routes cron (Render Cron)
    if path.startswith("/cron-"):
        return None

    # ✅ autoriser routes publiques utiles (dashboard / tests)
    if path in ("/healthz", "/data.json", "/dotations_data.json", "/formateurs_data.json", "/tz-test"):
        return None

    # 🔐 tout le reste nécessite une session admin fraîche.
    # Les sessions créées avant ADMIN_SESSION_VERSION sont ainsi déconnectées.
    return require_fresh_admin_session(path)


@app.get("/api/yousign/health")
def yousign_health():
    diagnostic = test_yousign_connection()
    status = int(diagnostic.get("status") or (200 if diagnostic.get("ok") else 502))
    http_status = 200 if status == 200 else status if status in {401, 403} else 502
    return jsonify(diagnostic), http_status


# --- Filtres Jinja ---
def format_date(value):
    try:
        dt = datetime.strptime(value, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return value
app.jinja_env.filters['datefr'] = format_date

def format_datetime_fr(value):
    if not value:
        return value
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(str(value), fmt)
            return dt.strftime("%d/%m/%Y %H:%M") if "H" in fmt else dt.strftime("%d/%m/%Y")
        except Exception:
            continue
    return value
app.jinja_env.filters['datetimefr'] = format_datetime_fr

def to_datetime(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except Exception:
        try:
            return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        except Exception:
            return datetime.now()
app.jinja_env.filters['datetime'] = to_datetime

# --- Helper utilisable dans Jinja ---
def get_status_label(step_index, session):
    """Renvoie un dict {status, deadline} lisible dans Jinja"""
    status, dl = status_for_step(step_index, session)
    return {"status": status, "deadline": dl}

app.jinja_env.globals['get_status_label'] = get_status_label


# --- Persistance ---
SESSIONS_FILE = os.path.join(DATA_DIR, "sessions.json")
PRICE_ADAPTATOR_FILE = os.path.join(DATA_DIR, "price_adaptator.json")

PRICE_ADAPTATOR_DEFAULT_DISCOUNT = 30
PRICE_ADAPTATOR_FOLLOWUP_DAYS = 21
PRICE_ADAPTATOR_FORMATION_PRICES = {
    "APS": 1650,
    "A3P": 4200,
    "Dirigeant": 4300,
}
PRICE_ADAPTATOR_FORMATION_LABELS = {
    "APS": "Agent de prévention et de sécurité (APS)",
    "A3P": "Agent de protection physique des personnes (A3P)",
    "Dirigeant": "Dirigeant d'entreprise de sécurité privée (DESP)",
}
PRICE_ADAPTATOR_ALLOWED_FORMATIONS = {
    "APS": "APS",
    "A3P": "A3P",
    "DIRIGEANT": "Dirigeant",
}

# -----------------------
# 📅 Planning PDF (par session)
# -----------------------
PLANNING_DIR = os.path.join(DATA_DIR, "plannings")
os.makedirs(PLANNING_DIR, exist_ok=True)
CONVOCATION_DIR = os.path.join(DATA_DIR, "convocations")
os.makedirs(CONVOCATION_DIR, exist_ok=True)
APS_CONTRACT_DIR = os.path.join(DATA_DIR, "aps_trainer_contracts")
os.makedirs(APS_CONTRACT_DIR, exist_ok=True)
APS_CONTRACT_SIGNED_DIR = os.path.join(APS_CONTRACT_DIR, "_signed_yousign")
os.makedirs(APS_CONTRACT_SIGNED_DIR, exist_ok=True)
YOUSIGN_TRAINER_SIGNATURE_TAG = "{{s1|signature|160|60}}"
APS_ATTENDANCE_DIR = os.path.join(DATA_DIR, "aps_attendance_sheets")
os.makedirs(APS_ATTENDANCE_DIR, exist_ok=True)
A3P_DOC_DIR = os.path.join(DATA_DIR, "a3p_documents")
os.makedirs(A3P_DOC_DIR, exist_ok=True)
APS_CONVOCATION_TEMPLATE = os.path.join(BASE_DIR, "gestionstagiaires", "templates_word", "convocationaps.docx")

APS_TOTAL_HOURS = 175
APS_TOTAL_MINUTES = APS_TOTAL_HOURS * 60
APS_ELEARNING_HOURS = 62
APS_ELEARNING_MINUTES = APS_ELEARNING_HOURS * 60
APS_PRESENTIEL_HOURS = APS_TOTAL_HOURS - APS_ELEARNING_HOURS
APS_PRESENTIEL_MINUTES = APS_PRESENTIEL_HOURS * 60
APS_MAX_DAILY_MINUTES = 7 * 60
APS_EXTENDED_DAILY_MINUTES = 8 * 60

SSIAP1_TOTAL_HOURS = 67
SSIAP1_TOTAL_MINUTES = SSIAP1_TOTAL_HOURS * 60
SSIAP1_SEQUENCE_TOTALS = {
    "P1-S1": 4, "P1-S2": 2,
    "P2-S1": 2.5, "P2-S2": 2, "P2-S3": 2, "P2-S4": 2, "P2-S5": 1.5, "P2-S6": 4, "P2-S7": 2, "P2-S8": 1,
    "P3-S1": 1, "P3-S2": 2, "P3-S3": 2, "P3-S4": 1, "P3-S5": 3,
    "P4-S1": 1, "P4-S2": 1, "P4-S3": 2.5, "P4-S4": 4, "P4-S5": 4, "P4-S6": 4, "P4-S7": 1.5,
    "P5-S1": 10, "P5-S2": 7,
}
SSIAP1_PART_TOTALS = {"1re partie": 6, "2e partie": 17, "3e partie": 9, "4e partie": 18, "5e partie": 17}
SSIAP1_SEQUENCE_LABELS = {
    "P1-S1": "LE FEU", "P1-S2": "COMPORTEMENT AU FEU",
    "P2-S1": "PRINCIPES DE CLASSEMENT DES ÉTABLISSEMENTS", "P2-S2": "FONDAMENTAUX ET PRINCIPES GÉNÉRAUX DE SÉCURITÉ INCENDIE", "P2-S3": "DESSERTE DES BÂTIMENTS", "P2-S4": "CLOISONNEMENT D’ISOLATION DES RISQUES", "P2-S5": "ÉVACUATION DU PUBLIC ET DES OCCUPANTS", "P2-S6": "DÉSENFUMAGE", "P2-S7": "ÉCLAIRAGE DE SÉCURITÉ", "P2-S8": "PRÉSENTATION DES DIFFÉRENTS MOYENS DE SECOURS",
    "P3-S1": "INSTALLATIONS ÉLECTRIQUES", "P3-S2": "ASCENSEURS ET NACELLES", "P3-S3": "INSTALLATIONS FIXES D’EXTINCTION AUTOMATIQUE", "P3-S4": "COLONNES SÈCHES ET HUMIDES", "P3-S5": "SYSTÈME DE SÉCURITÉ INCENDIE",
    "P4-S1": "LE SERVICE DE SÉCURITÉ", "P4-S2": "PRÉSENTATION DES CONSIGNES DE SÉCURITÉ ET MAIN COURANTE", "P4-S3": "POSTE DE SÉCURITÉ", "P4-S4": "RONDES DE SÉCURITÉ ET SURVEILLANCE DES TRAVAUX", "P4-S5": "MISE EN ŒUVRE DES MOYENS D’EXTINCTION", "P4-S6": "APPEL ET RÉCEPTION DES SERVICES PUBLICS DE SECOURS", "P4-S7": "SENSIBILISATION DES OCCUPANTS",
    "P5-S1": "VISITES APPLICATIVES", "P5-S2": "MISES EN SITUATION D’INTERVENTION",
}
SSIAP1_PART_LABELS = {"P1": "1re partie — LE FEU ET SES CONSÉQUENCES", "P2": "2e partie — SÉCURITÉ INCENDIE", "P3": "3e partie — INSTALLATIONS TECHNIQUES", "P4": "4e partie — RÔLE ET MISSIONS DES AGENTS DE SÉCURITÉ INCENDIE", "P5": "5e partie — CONCRÉTISATION DES ACQUIS"}
SSIAP1_CONFIG = {"code": "SSIAP1", "durationHours": SSIAP1_TOTAL_HOURS, "sequenceTotals": SSIAP1_SEQUENCE_TOTALS, "partTotals": SSIAP1_PART_TOTALS, "examRequired": True, "examInPlanning": True, "examAttendance": True, "contract": True, "yousign": True}

SSIAP1_AGREMENT_LINE = "Agrément SSIAP n°8323 délivré par la Préfecture du Var en date du 29/05/2026"

SSIAP1_SST_TOTAL_HOURS = 14
SSIAP1_REVISION_HOURS = 3
SSIAP1_PRESENCE_TOTAL_HOURS = SSIAP1_SST_TOTAL_HOURS + SSIAP1_TOTAL_HOURS + SSIAP1_REVISION_HOURS
SSIAP1_SST_PART = "FORMATION SAUVETEUR SECOURISTE DU TRAVAIL — SST — 14 h"
SSIAP1_REVISION_PART = "RÉVISIONS GÉNÉRALES ET PRÉPARATION À L’EXAMEN SSIAP 1 — 3 h"
SSIAP1_SST_DAYS = [
    {"code": "SST-J1", "title": "SST — JOURNÉE 1", "items": ["Présentation de la formation et du rôle du SST", "Prévention des risques professionnels", "Protection adaptée face à une situation d’accident", "Examen de la victime", "Transmission de l’alerte", "Premiers apprentissages des gestes de secours"]},
    {"code": "SST-J2", "title": "SST — JOURNÉE 2", "items": ["Secours adaptés aux différentes situations", "Mises en situation pratiques", "Cas concrets", "Prévention dans l’entreprise", "Évaluations certificatives SST"]},
]
SSIAP1_REVISION_ITEMS = ["Révision des principales notions réglementaires", "Révision du rôle et des missions de l’agent SSIAP 1", "Exploitation du SSI", "Révision des conduites à tenir", "Préparation aux épreuves écrites et pratiques", "Questions-réponses avant l’examen"]

def _ssiap1_detail(code, part_number, part_title, sequence_number, sequence_title, total_minutes, content_minutes, content_items, application_minutes=0, application_items=None):
    return {
        "code": code,
        "part_number": part_number,
        "part_title": part_title,
        "sequence_number": sequence_number,
        "sequence_title": sequence_title,
        "total_duration_minutes": total_minutes,
        "content_duration_minutes": content_minutes,
        "content_items": content_items,
        "application_duration_minutes": application_minutes,
        "application_items": application_items or [],
    }

SSIAP1_SEQUENCE_DETAILS = [
    _ssiap1_detail("P1-S1", 1, "LE FEU ET SES CONSÉQUENCES", 1, "LE FEU", 240, 120, ["Théorie du feu : triangle du feu, classes de feux et causes", "La fumée et ses dangers", "Propagation du feu : conduction, convection, rayonnement et projection", "Conduite à tenir face à un local enfumé sans mise en danger pour l’intervenant"], 120, ["Exercice de sortie d’un local enfumé par des fumées odorantes, froides et non toxiques"]),
    _ssiap1_detail("P1-S2", 1, "LE FEU ET SES CONSÉQUENCES", 2, "COMPORTEMENT AU FEU", 120, 120, ["Principe de la résistance au feu des éléments de construction", "Principe de la réaction au feu des matériaux d’aménagement", "Critères de classement de ces comportements"]),
    _ssiap1_detail("P2-S1", 2, "SÉCURITÉ INCENDIE", 1, "PRINCIPES DE CLASSEMENT DES ÉTABLISSEMENTS", 150, 120, ["Définition d’un ERP", "Définition du public", "Différents types d’établissements", "Catégories selon l’effectif", "Définition du seuil de la 5e catégorie", "Méthode de détermination de l’effectif", "Définition et classification d’un IGH"], 30, ["Exercices simples de classement d’établissements"]),
    _ssiap1_detail("P2-S2", 2, "SÉCURITÉ INCENDIE", 2, "FONDAMENTAUX ET PRINCIPES GÉNÉRAUX DE SÉCURITÉ INCENDIE", 120, 120, ["Fondamentaux de sécurité : évacuation des occupants, accessibilité et mise en service des moyens de secours", "Principes généraux : implantation, desserte, isolement, matériaux, cloisonnement, aménagement, dégagement, désenfumage, éclairage de sécurité et moyens de secours"]),
    _ssiap1_detail("P2-S3", 2, "SÉCURITÉ INCENDIE", 3, "DESSERTE DES BÂTIMENTS", 120, 90, ["Desserte et voiries", "Voies engins", "Voies échelles", "Espaces libres", "Baies accessibles"], 30, ["Repérage de dessertes et baies accessibles sur plans"]),
    _ssiap1_detail("P2-S4", 2, "SÉCURITÉ INCENDIE", 4, "CLOISONNEMENT D’ISOLATION DES RISQUES", 120, 120, ["Cloisonnement traditionnel", "Secteurs", "Compartiments", "Locaux à risques", "Recoupement des vides"]),
    _ssiap1_detail("P2-S5", 2, "SÉCURITÉ INCENDIE", 5, "ÉVACUATION DU PUBLIC ET DES OCCUPANTS", 90, 90, ["Définition du dégagement", "Notion d’unité de passage", "Balisage des dégagements", "Manœuvre et déverrouillage des issues", "Espace d’attente sécurisé et évacuation différée"]),
    _ssiap1_detail("P2-S6", 2, "SÉCURITÉ INCENDIE", 6, "DÉSENFUMAGE", 240, 120, ["Objectifs du désenfumage", "Désenfumage des dégagements", "Désenfumage des locaux", "Déclenchement manuel", "Entretien et vérification", "Remise en position d’attente des dispositifs"], 120, ["Réarmement d’un volet, d’un clapet ou d’un exutoire"]),
    _ssiap1_detail("P2-S7", 2, "SÉCURITÉ INCENDIE", 7, "ÉCLAIRAGE DE SÉCURITÉ", 120, 90, ["Définition de l’éclairage de sécurité", "Éclairage d’évacuation", "Éclairage d’ambiance ou anti-panique", "Blocs autonomes et sources centrales", "Maintenance et vérifications"], 30, ["Identification des différents blocs d’éclairage et vérification simple"]),
    _ssiap1_detail("P2-S8", 2, "SÉCURITÉ INCENDIE", 8, "PRÉSENTATION DES DIFFÉRENTS MOYENS DE SECOURS", 60, 60, ["Moyens d’extinction", "Dispositions visant à faciliter l’action des sapeurs-pompiers", "Service de sécurité incendie", "Système de sécurité incendie", "Système d’alerte"]),
    _ssiap1_detail("P3-S1", 3, "INSTALLATIONS TECHNIQUES", 1, "INSTALLATIONS ÉLECTRIQUES", 60, 60, ["Impact des installations électriques sur la sécurité", "Maintien de l’alimentation des installations de sécurité", "Éclairage normal, de remplacement et de sécurité", "Notions de sources de sécurité"]),
    _ssiap1_detail("P3-S2", 3, "INSTALLATIONS TECHNIQUES", 2, "ASCENSEURS ET NACELLES", 120, 90, ["Ascenseurs et monte-charge", "Dispositifs de sécurité", "Conduite à tenir en cas de personne bloquée", "Nacelles et équipements assimilés"], 30, ["Procédure de dégagement et consignes d’exploitation"]),
    _ssiap1_detail("P3-S3", 3, "INSTALLATIONS TECHNIQUES", 3, "INSTALLATIONS FIXES D’EXTINCTION AUTOMATIQUE", 120, 90, ["Principes de fonctionnement", "Différents types d’installations fixes", "Poste de contrôle", "Alarme et report d’information", "Entretien et vérifications"], 30, ["Lecture d’un poste de contrôle et identification des organes principaux"]),
    _ssiap1_detail("P3-S4", 3, "INSTALLATIONS TECHNIQUES", 4, "COLONNES SÈCHES ET HUMIDES", 60, 60, ["Colonnes sèches", "Colonnes en charge ou humides", "Emplacements et signalisation", "Alimentation et utilisation par les services de secours", "Vérifications"]),
    _ssiap1_detail("P3-S5", 3, "INSTALLATIONS TECHNIQUES", 5, "SYSTÈME DE SÉCURITÉ INCENDIE", 180, 120, ["Définition du SSI", "Catégories de SSI", "Système de détection incendie", "Centralisateur de mise en sécurité incendie", "Dispositifs actionnés de sécurité", "Unité de gestion d’alarme"], 60, ["Exploitation d’un SSI de catégorie A et lecture des informations au poste de sécurité"]),
    _ssiap1_detail("P4-S1", 4, "RÔLE ET MISSIONS DES AGENTS DE SÉCURITÉ INCENDIE", 1, "LE SERVICE DE SÉCURITÉ", 60, 60, ["Rôle et missions du service de sécurité incendie", "Composition du service", "Qualification des personnels", "Missions générales et particulières"]),
    _ssiap1_detail("P4-S2", 4, "RÔLE ET MISSIONS DES AGENTS DE SÉCURITÉ INCENDIE", 2, "PRÉSENTATION DES CONSIGNES DE SÉCURITÉ ET MAIN COURANTE", 60, 60, ["Consignes générales", "Consignes particulières", "Main courante et enregistrements", "Transmission des informations", "Rédaction d’une main courante à partir d’un événement simple"]),
    _ssiap1_detail("P4-S3", 4, "RÔLE ET MISSIONS DES AGENTS DE SÉCURITÉ INCENDIE", 3, "POSTE DE SÉCURITÉ", 150, 90, ["Organisation du poste de sécurité", "Documents présents au poste", "Réception et traitement des alarmes", "Moyens de communication", "Gestion des clés et moyens d’accès"], 60, ["Exploitation du poste de sécurité et traitement d’une alarme"]),
    _ssiap1_detail("P4-S4", 4, "RÔLE ET MISSIONS DES AGENTS DE SÉCURITÉ INCENDIE", 4, "RONDES DE SÉCURITÉ ET SURVEILLANCE DES TRAVAUX", 240, 120, ["Objectifs de la ronde", "Itinéraires et points de contrôle", "Anomalies et mesures conservatoires", "Permis de feu", "Surveillance des travaux par points chauds"], 120, ["Ronde avec résolution d’anomalies et contrôle d’un permis de feu"]),
    _ssiap1_detail("P4-S5", 4, "RÔLE ET MISSIONS DES AGENTS DE SÉCURITÉ INCENDIE", 5, "MISE EN ŒUVRE DES MOYENS D’EXTINCTION", 240, 60, ["Différents moyens d’extinction", "Agents extincteurs", "Règles de sécurité lors de l’attaque d’un feu"], 180, ["Extinction de feux réels ou simulés au moyen d’extincteurs adaptés", "Mise en œuvre d’un robinet d’incendie armé"]),
    _ssiap1_detail("P4-S6", 4, "RÔLE ET MISSIONS DES AGENTS DE SÉCURITÉ INCENDIE", 6, "APPEL ET RÉCEPTION DES SERVICES PUBLICS DE SECOURS", 240, 90, ["Différents moyens d’alerte", "Message d’alerte", "Accueil et guidage des secours", "Mise à disposition des informations utiles"], 150, ["Exercices d’appel des secours", "Mise en situation d’accueil et de guidage des secours"]),
    _ssiap1_detail("P4-S7", 4, "RÔLE ET MISSIONS DES AGENTS DE SÉCURITÉ INCENDIE", 7, "SENSIBILISATION DES OCCUPANTS", 90, 60, ["Information des occupants", "Actions de prévention", "Conduite à tenir en cas d’alarme", "Adaptation du message au public"], 30, ["Présentation orale d’une consigne de sécurité à des occupants"]),
    _ssiap1_detail("P5-S1", 5, "CONCRÉTISATION DES ACQUIS", 1, "VISITES APPLICATIVES", 600, 0, [], 600, ["Visite applicative d’au moins deux établissements recevant du public", "Identification des risques et des dispositions de sécurité", "Repérage des dégagements, désenfumage, SSI et moyens de secours", "Lecture de plans et documents de sécurité"]),
    _ssiap1_detail("P5-S2", 5, "CONCRÉTISATION DES ACQUIS", 2, "MISES EN SITUATION D’INTERVENTION", 420, 0, [], 420, ["Mises en situation d’intervention sur départ de feu", "Application des consignes et levée de doute", "Gestion d’une évacuation", "Alerte, accueil et guidage des secours", "Exploitation du SSI et de la main courante"]),
]
SSIAP1_SEQUENCE_DETAIL_BY_CODE = {item["code"]: item for item in SSIAP1_SEQUENCE_DETAILS}

APS_EXPECTED_UV_TOTALS = {
    "UV1": 14,
    "UV2": 22,
    "UV3": 14,
    "UV4": 7,
    "UV5": 7,
    "UV6": 7,
    "UV7": 13,
    "UV8": 45,
    "UV9": 7,
    "UV10": 7,
    "UV11": 11,
    "UV12": 7,
    "UV13": 7,
    "UV14": 7,
}

APS_UV_LABELS = {
    "UV1": "Secouriste Sauveteur du Travail (SST)",
    "UV2": "Environnement juridique de la sécurité privée",
    "UV3": "Gestion des conflits",
    "UV4": "Stratégique",
    "UV5": "Prévention des risques incendie",
    "UV6": "Appréhension au cours de l’exercice",
    "UV7": "Risques terroristes",
    "UV8": "Professionnel",
    "UV9": "Palpation de sécurité et inspection visuelle des bagages",
    "UV10": "Surveillance par moyens électroniques",
    "UV11": "Gestion des risques",
    "UV12": "Événementiel spécifique",
    "UV13": "Gestion des situations conflictuelles dégradées",
    "UV14": "Industriel spécifique",
}

APS_LEGAL_LINES = [
    "ORGANISME DE FORMATION CERTIFIÉ QUALIOPI",
    "La certification qualité a été délivrée au titre de la ou des catégories d’actions suivantes : actions de formation, actions de formation par apprentissage.",
    "Autorisation d'exercice CNAPS n°FOR-083-2027-02-08-20200755135",
    "Agrément ADEF APS : 8320032701 - Agrément ADEF A3P : 8320111201",
]

APS_MODULES = [
    ("UV2 ENVIRONNEMENT JURIDIQUE DE LA SÉCURITÉ PRIVÉE", 22),
    ("UV8 PROFESSIONNEL", 6),
    ("UV14 INDUSTRIEL SPÉCIFIQUE", 7),
    ("UV1 SECOURISTE SAUVETEUR DU TRAVAIL (SST)", 14),
    ("UV7 RISQUES TERRORISTES", 13),
    ("UV8 PROFESSIONNEL", 1),
    ("UV9 PALPATION DE SÉCURITÉ ET INSPECTION VISUELLE DES BAGAGES", 7),
    ("UV3 GESTION DES CONFLITS", 14),
    ("UV4 STRATÉGIQUE", 7),
    ("UV6 APPRÉHENSION AU COURS DE L’EXERCICE", 7),
    ("UV5 PRÉVENTION DES RISQUES INCENDIE", 7),
    ("UV10 SURVEILLANCE PAR MOYENS ÉLECTRONIQUES", 7),
    ("UV12 ÉVÉNEMENTIEL SPÉCIFIQUE", 7),
    ("UV11 GESTION DES RISQUES", 11),
    ("UV8 PROFESSIONNEL", 31),
    ("UV13 GESTION DES SITUATIONS CONFLICTUELLES DÉGRADÉES", 7),
    ("UV8 PROFESSIONNEL", 7),
]


APS_ELEARNING_PRESENTIEL_MODULES = [
    {"part": "PÉRIODE 1 — E-LEARNING / DISTANCIEL — 62h", "modality": "elearning", "uv": "UV2", "title": "Environnement juridique de la sécurité privée", "durationMinutes": 17 * 60},
    {"part": "PÉRIODE 1 — E-LEARNING / DISTANCIEL — 62h", "modality": "elearning", "uv": "UV3", "title": "Gestion des risques et situations conflictuelles", "durationMinutes": 5 * 60},
    {"part": "PÉRIODE 1 — E-LEARNING / DISTANCIEL — 62h", "modality": "elearning", "uv": "UV4", "title": "Transmission des consignes et informations", "durationMinutes": 5 * 60},
    {"part": "PÉRIODE 1 — E-LEARNING / DISTANCIEL — 62h", "modality": "elearning", "uv": "UV2", "title": "Environnement juridique de la sécurité privée", "durationMinutes": 3 * 60},
    {"part": "PÉRIODE 1 — E-LEARNING / DISTANCIEL — 62h", "modality": "elearning", "uv": "UV11", "title": "Gestion des risques / connaissances des vecteurs d’incendie", "durationMinutes": 9 * 60},
    {"part": "PÉRIODE 1 — E-LEARNING / DISTANCIEL — 62h", "modality": "elearning", "uv": "UV7", "title": "Prévention des risques terroristes", "durationMinutes": 6 * 60},
    {"part": "PÉRIODE 1 — E-LEARNING / DISTANCIEL — 62h", "modality": "elearning", "uv": "UV1", "title": "Secourir", "durationMinutes": 1 * 60},
    {"part": "PÉRIODE 1 — E-LEARNING / DISTANCIEL — 62h", "modality": "elearning", "uv": "UV10", "title": "Connaissance de l’outil informatique / transmission", "durationMinutes": 2 * 60},
    {"part": "PÉRIODE 1 — E-LEARNING / DISTANCIEL — 62h", "modality": "elearning", "uv": "UV8", "title": "Surveillance et gardiennage", "durationMinutes": 7 * 60},
    {"part": "PÉRIODE 1 — E-LEARNING / DISTANCIEL — 62h", "modality": "elearning", "uv": "UV12", "title": "Événementiel", "durationMinutes": 7 * 60},
    {"part": "PÉRIODE 2 — PRÉSENTIEL AU CENTRE — 113h", "modality": "presentiel", "uv": "UV1", "title": "Gestion des premiers secours", "durationMinutes": 14 * 60},
    {"part": "PÉRIODE 2 — PRÉSENTIEL AU CENTRE — 113h", "modality": "presentiel", "uv": "UV2", "title": "Environnement juridique de la sécurité privée", "durationMinutes": 2 * 60},
    {"part": "PÉRIODE 2 — PRÉSENTIEL AU CENTRE — 113h", "modality": "presentiel", "uv": "UV5", "title": "Gestion des risques / connaissances des vecteurs d’incendie", "durationMinutes": 16 * 60},
    {"part": "PÉRIODE 2 — PRÉSENTIEL AU CENTRE — 113h", "modality": "presentiel", "uv": "UV7", "title": "Prévention des risques terroristes", "durationMinutes": 270},
    {"part": "PÉRIODE 2 — PRÉSENTIEL AU CENTRE — 113h", "modality": "presentiel", "uv": "UV1", "title": "Secourir", "durationMinutes": 90},
    {"part": "PÉRIODE 2 — PRÉSENTIEL AU CENTRE — 113h", "modality": "presentiel", "uv": "UV3", "title": "Gestion des risques et des situations conflictuelles", "durationMinutes": 9 * 60},
    {"part": "PÉRIODE 2 — PRÉSENTIEL AU CENTRE — 113h", "modality": "presentiel", "uv": "UV13", "title": "Gestion des risques de situations conflictuelles dégradées", "durationMinutes": 7 * 60},
    {"part": "PÉRIODE 2 — PRÉSENTIEL AU CENTRE — 113h", "modality": "presentiel", "uv": "UV8", "title": "Surveillance et gardiennage", "durationMinutes": 45 * 60},
    {"part": "PÉRIODE 2 — PRÉSENTIEL AU CENTRE — 113h", "modality": "presentiel", "uv": "UV12", "title": "Événementiel", "durationMinutes": 14 * 60},
]

APS_RECAP_ROWS = [
    ("UV1", "Secouriste Sauveteur du Travail (SST)"),
    ("UV2", "Environnement juridique de la sécurité privée"),
    ("UV3", "Gestion des conflits"),
    ("UV4", "Stratégique"),
    ("UV5", "Prévention des risques incendie"),
    ("UV6", "Appréhension au cours de l’exercice"),
    ("UV7", "Risques terroristes"),
    ("UV8", "Professionnel"),
    ("UV9", "Palpation de sécurité et inspection visuelle des bagages"),
    ("UV10", "Surveillance par moyens électroniques"),
    ("UV11", "Gestion des risques"),
    ("UV12", "Événementiel spécifique"),
    ("UV13", "Gestion des situations conflictuelles dégradées"),
    ("UV14", "Industriel spécifique"),
]

def easter_date(year):
    """Retourne la date du dimanche de Pâques (algorithme de Meeus/Jones/Butcher)."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)

def french_holidays(year):
    easter = easter_date(year)
    return {
        easter + timedelta(days=1),
        date(year, 5, 1),
        date(year, 5, 8),
        easter + timedelta(days=39),
        easter + timedelta(days=50),
        date(year, 7, 14),
        date(year, 8, 15),
        date(year, 11, 1),
        date(year, 11, 11),
        date(year, 12, 25),
    }

def is_french_working_day(day):
    return day.weekday() < 5 and day not in french_holidays(day.year)

def aps_local_date_iso(value):
    """Return a YYYY-MM-DD string without timezone conversion."""
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    parsed = parse_date(text)
    return parsed.strftime("%Y-%m-%d") if parsed else text[:10]

def is_aps_training_day(day, exam_iso=""):
    return is_french_working_day(day) and day.isoformat() != exam_iso

def next_aps_training_day(day, exam_iso=""):
    day += timedelta(days=1)
    while not is_aps_training_day(day, exam_iso):
        day += timedelta(days=1)
    return day

def add_hours_to_time(start_time, hours):
    base = datetime.combine(date.today(), start_time)
    return (base + timedelta(hours=hours)).time()

def add_minutes_to_time(start_time, minutes):
    base = datetime.combine(date.today(), start_time)
    return (base + timedelta(minutes=minutes)).time()

def next_french_working_day(day):
    day += timedelta(days=1)
    while not is_french_working_day(day):
        day += timedelta(days=1)
    return day

def format_duration_from_minutes(minutes):
    hours = minutes // 60
    rest = minutes % 60
    return f"{hours:g}h" if rest == 0 else f"{hours:g}h{rest:02d}"

def is_ssiap1_exam_part(part):
    return (part or "").strip().upper() == "EXAMEN SSIAP 1"

def aps_working_days_between(start_date, end_date, exam_iso=""):
    if not start_date or not end_date or start_date > end_date:
        return []
    days = []
    current = start_date
    while current <= end_date:
        if is_aps_training_day(current, exam_iso):
            days.append(current)
        current += timedelta(days=1)
    return days

def aps_impossible_period_message(start_date, end_date, available_minutes, required_minutes, extended_minutes=None):
    if extended_minutes is not None:
        return (
            "Impossible de générer le planning : "
            f"{available_minutes / 60:g} heures disponibles à 7h/jour "
            f"({extended_minutes / 60:g} heures maximum à 8h/jour) entre le {format_date(start_date)} "
            f"et le {format_date(end_date)}, mais {required_minutes / 60:g} heures nécessaires."
        )
    return (
        "Impossible de générer le planning : "
        f"{available_minutes / 60:g} heures disponibles entre le {format_date(start_date)} "
        f"et le {format_date(end_date)}, mais {required_minutes / 60:g} heures nécessaires."
    )

def log_aps_generation_diagnostics(session_id=None, planning_mode="full_presentiel", start_date=None, end_date=None, exam_iso="", available_days=0, available_minutes=0, elearning_minutes=0, presentiel_minutes=0, total_minutes=0, extended_minutes=None, elongated_days=0, day_distribution=None, level="error"):
    logger = app.logger.error if level == "error" else app.logger.info
    logger(
        "Diagnostic planning APS session_id=%s planning_mode=%s start_date=%s end_date=%s exam_iso=%s heures_necessaires=%s jours_disponibles=%s capacite_7h=%s capacite_8h=%s heures_elearning=%s heures_presentiel=%s total_heures_attendu=%s journees_allongees=%s repartition_heures_par_jour=%s",
        session_id,
        planning_mode,
        start_date.isoformat() if hasattr(start_date, "isoformat") else start_date,
        end_date.isoformat() if hasattr(end_date, "isoformat") else end_date,
        exam_iso,
        presentiel_minutes / 60 if planning_mode == "elearning_presentiel" else total_minutes / 60,
        available_days,
        available_minutes / 60,
        (extended_minutes if extended_minutes is not None else available_minutes) / 60,
        elearning_minutes / 60,
        presentiel_minutes / 60,
        total_minutes / 60,
        elongated_days,
        day_distribution or [],
    )

def build_aps_planning(start_date, end_date=None, exam_iso=""):
    modules = [{"name": name, "hours": float(hours), "remaining": float(hours)} for name, hours in APS_MODULES]
    module_idx = 0
    days = []
    totals = {}
    total_hours = 0.0
    current_day = start_date

    while round(total_hours, 2) < APS_TOTAL_HOURS:
        if end_date and current_day > end_date:
            raise ValueError("La période disponible avant l’examen ne permet pas de placer toutes les heures de formation APS. Merci d’avancer la date de début ou de reculer la date d’examen.")
        if not is_aps_training_day(current_day, exam_iso):
            current_day += timedelta(days=1)
            continue

        day_blocks = []
        for slot_start, slot_hours in ((dt_time(8, 30), 4.0), (dt_time(13, 30), 3.0)):
            cursor = slot_start
            remaining_slot = slot_hours
            while remaining_slot > 0 and module_idx < len(modules):
                module = modules[module_idx]
                duration = min(remaining_slot, module["remaining"])
                end_time = add_hours_to_time(cursor, duration)
                day_blocks.append({
                    "uv": module["name"],
                    "start": cursor,
                    "end": end_time,
                    "hours": duration,
                })
                module["remaining"] = round(module["remaining"] - duration, 2)
                remaining_slot = round(remaining_slot - duration, 2)
                total_hours = round(total_hours + duration, 2)
                totals[module["name"]] = round(totals.get(module["name"], 0) + duration, 2)
                cursor = end_time
                if module["remaining"] == 0:
                    module_idx += 1
                if round(total_hours, 2) == APS_TOTAL_HOURS:
                    break
            if round(total_hours, 2) == APS_TOTAL_HOURS:
                break
        if day_blocks:
            days.append({"date": current_day, "blocks": day_blocks})
        current_day += timedelta(days=1)

    return days, totals, total_hours


def normalize_formation_code(value):
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


UNSUPPORTED_TRAINING_TYPE_MESSAGE = "Impossible de générer le document : le type de formation de cette session n’est pas reconnu."
TRAINING_CODE_ALIASES = {
    "AFCAPSSSIAP": "AFC_APS_SSIAP",
    "AFCAPSSSIAP1": "AFC_APS_SSIAP",
    "AFCCFRANCETRAVAILAPSSSIAP": "AFC_APS_SSIAP",
    "AFC_APS_SSIAP": "AFC_APS_SSIAP",
    "APS": "APS",
    "A3P": "A3P",
    "SSIAP": "SSIAP1",
    "SSIAP1": "SSIAP1",
    "DIRIGEANT": "DESP",
    "DESP": "DESP",
}


def normalize_training_code(session_data):
    """Return the stable business training code, never inferred from display title."""
    session_data = session_data or {}
    for key in ("training_code", "formation_type", "program_code", "formation"):
        raw = session_data.get(key)
        if raw in (None, ""):
            continue
        code = normalize_formation_code(raw)
        if code in TRAINING_CODE_ALIASES:
            return TRAINING_CODE_ALIASES[code]
        raise ValueError(UNSUPPORTED_TRAINING_TYPE_MESSAGE)
    raise ValueError(UNSUPPORTED_TRAINING_TYPE_MESSAGE)


def is_ssiap1_session(session_data):
    try:
        return normalize_training_code(session_data) == "SSIAP1"
    except ValueError:
        return False


def select_training_builder(session_data, builders):
    code = normalize_training_code(session_data)
    builder = builders.get(code)
    if builder is None:
        raise ValueError(UNSUPPORTED_TRAINING_TYPE_MESSAGE)
    return code, builder



def build_aps_session_planning(session_data, payload=None):
    payload = payload or {}
    mode = (payload.get("planningMode") or session_data.get("apsPlanningMode") or "").strip()
    if mode not in {"full_presentiel", "elearning_presentiel"}:
        raise ValueError("Le type de planning APS est obligatoire.")
    start_dt = parse_date(session_data.get("date_debut"))
    end_dt = parse_date(session_data.get("date_fin"))
    if not start_dt:
        raise ValueError("La date de début de session est obligatoire.")
    formateur = (payload.get("trainer") or payload.get("formateur") or "").strip()
    room = (payload.get("room") or session_data.get("salle") or "Intégrale Academy – 54 chemin du Carreou – 83480 PUGET-SUR-ARGENS").strip()
    return build_aps_planning_data(start_dt.date(), formateur, room, mode, end_date=end_dt.date() if end_dt else None, exam_iso=aps_local_date_iso(session_data.get("date_exam")), session_id=session_data.get("id"))[0]


def build_ssiap_planning(session_data, payload=None):
    payload = payload or {}
    formateur = (payload.get("trainer") or payload.get("formateur") or "").strip()
    room = (payload.get("room") or session_data.get("salle") or "Intégrale Academy – 54 chemin du Carreou – 83480 PUGET-SUR-ARGENS").strip()
    exam = ssiap1_exam_payload(session_data, payload)
    end_dt = parse_date(session_data.get("date_fin"))
    return build_ssiap1_planning_data(parse_date(session_data.get("date_debut")).date(), formateur, room, end_date=end_dt.date() if end_dt else None, exam_iso=exam["date"], exam_payload=exam, excluded_dates=ssiap1_excluded_dates_from_payload(session_data, payload))[0]


def build_desp_planning(session_data, payload=None):
    payload = payload or {}
    elearning_start = parse_date(payload.get("despElearningStart") or session_data.get("despElearningStart") or session_data.get("date_debut"))
    elearning_end = parse_date(payload.get("despElearningEnd") or session_data.get("despElearningEnd") or session_data.get("date_distanciel_fin") or session_data.get("date_elearning_fin"))
    presentiel_start = parse_date(payload.get("despPresentielStart") or session_data.get("despPresentielStart") or session_data.get("date_presentiel_debut"))
    presentiel_end = parse_date(payload.get("despPresentielEnd") or session_data.get("despPresentielEnd") or session_data.get("date_fin"))
    if not all([elearning_start, elearning_end, presentiel_start, presentiel_end]):
        raise ValueError("Les dates de début/fin distanciel et début/fin présentiel DESP sont obligatoires.")
    return generate_desp_planning(elearning_start.date(), elearning_end.date(), presentiel_start.date(), presentiel_end.date(), payload.get("trainer") or payload.get("formateur") or "", payload.get("room") or session_data.get("salle") or "", exam_iso=aps_local_date_iso(session_data.get("date_exam")), allow_saturday=bool(payload.get("allowSaturday") or session_data.get("despAllowSaturday")))


def build_a3p_planning(session_data, payload=None):
    return session_data.get("a3pPlanningData") or generateA3pSchedule(session_data.get("a3pPlanningConfig") or {})


def build_aps_presence_days(session_data):
    return _aps_presentiel_days(session_data.get("apsPlanningData"), session_data.get("apsPlanningMode") or "full_presentiel")


def build_a3p_presence_days(session_data):
    _, converted = _a3p_session_for_shared_docs(session_data)
    return _aps_presentiel_days(converted, "full_presentiel")


def build_ssiap_presence_days(session_data):
    days = []
    for day in session_data.get("apsPlanningData") or []:
        slots = [slot for slot in day.get("slots", []) if is_in_person_slot(slot) or _normalized_slot_value(slot, "modality", "delivery_mode", "period_type") in {"sst", "revision"}]
        if slots:
            copied = dict(day); copied["slots"] = slots; days.append(copied)
    return days


def build_desp_presence_days(session_data):
    return _aps_presentiel_days(session_data.get("apsPlanningData"), "desp")


PLANNING_BUILDERS = {"APS": build_aps_session_planning, "A3P": build_a3p_planning, "SSIAP1": build_ssiap_planning, "DESP": build_desp_planning}
PRESENCE_BUILDERS = {"APS": build_aps_presence_days, "A3P": build_a3p_presence_days, "SSIAP1": build_ssiap_presence_days, "DESP": build_desp_presence_days}

def ssiap1_exam_payload(session_data, payload=None):
    payload = payload or {}
    exam_date = aps_local_date_iso(payload.get("examDate") or session_data.get("exam_date") or session_data.get("date_exam"))
    exam_start = (payload.get("examStartTime") or session_data.get("exam_start_time") or session_data.get("ssiapExamStartTime") or "08:30").strip()
    exam_end = (payload.get("examEndTime") or session_data.get("exam_end_time") or session_data.get("ssiapExamEndTime") or "16:30").strip()
    exam_room = (payload.get("examRoom") or session_data.get("exam_room") or session_data.get("ssiapExamRoom") or session_data.get("salle") or "Intégrale Academy – 54 chemin du Carreou – 83480 PUGET-SUR-ARGENS").strip()
    if not exam_date:
        raise ValueError("La date d'examen SSIAP 1 est obligatoire.")
    try:
        start_dt = datetime.strptime(f"{exam_date} {exam_start}", "%Y-%m-%d %H:%M")
        end_dt = datetime.strptime(f"{exam_date} {exam_end}", "%Y-%m-%d %H:%M")
    except Exception as exc:
        raise ValueError("Les horaires d'examen SSIAP 1 sont obligatoires au format HH:MM.") from exc
    if end_dt <= start_dt:
        raise ValueError("L'heure de fin d'examen SSIAP 1 doit être après l'heure de début.")
    return {"date": exam_date, "start": exam_start, "end": exam_end, "room": exam_room, "durationMinutes": int((end_dt - start_dt).total_seconds() // 60), "sstTrainer": (payload.get("sstTrainer") or session_data.get("ssiapSstTrainer") or "").strip(), "ssiapTrainer": (payload.get("ssiapTrainer") or payload.get("trainer") or payload.get("formateur") or session_data.get("ssiapTrainer") or "").strip(), "revisionTrainer": (payload.get("revisionTrainer") or session_data.get("ssiapRevisionTrainer") or "").strip(), "examTrainer": (payload.get("examTrainer") or session_data.get("ssiapExamTrainer") or "").strip()}

def _ssiap1_required_training_days():
    full_days, remainder = divmod(SSIAP1_TOTAL_MINUTES, APS_MAX_DAILY_MINUTES)
    return full_days + (1 if remainder else 0)

def _ssiap1_role_trainer(exam_payload, role, fallback=""):
    return (exam_payload or {}).get(role) or fallback or ""

def _ssiap1_next_training_days(start_date, count, exam_iso=""):
    days = []
    current = start_date
    while len(days) < count:
        if is_aps_training_day(current, exam_iso):
            days.append(current)
        current += timedelta(days=1)
    return days

def _ssiap1_day_slot(date_value, start, minutes, *, uv, part, title, items, trainer, room, modality, content=None):
    end = add_minutes_to_time(start, minutes)
    return {
        "start": start.strftime("%H:%M"), "end": end.strftime("%H:%M"),
        "duration": round(minutes / 60, 2), "durationMinutes": minutes,
        "uv": uv, "part": part, "sequence": uv, "title": title, "content": content or title,
        "subpartItems": list(items or []), "subpartDisplayItems": list(items or []),
        "room": room, "trainer": trainer, "modality": modality,
    }

def ssiap1_excluded_dates_from_payload(session_data, payload=None):
    payload = payload or {}
    raw_values = []
    for source in (session_data or {}, payload):
        for key in (
            "ssiapExcludedDates",
            "ssiap_excluded_dates",
            "ssiapNonTrainingDays",
            "nonTrainingDays",
            "excludedDates",
            "excluded_dates",
        ):
            value = source.get(key) if isinstance(source, dict) else None
            if value:
                raw_values.append(value)
    excluded = set()
    for value in raw_values:
        items = value if isinstance(value, (list, tuple, set)) else re.split(r"[,;\n\s]+", str(value))
        for item in items:
            iso = aps_local_date_iso(item)
            if iso:
                excluded.add(iso)
    return excluded

def build_ssiap1_planning_data(start_date, formateur, salle, end_date=None, exam_iso="", exam_payload=None, excluded_dates=None):
    exam_payload = exam_payload or {}
    excluded_dates = {aps_local_date_iso(value) for value in (excluded_dates or []) if aps_local_date_iso(value)}
    if end_date is None:
        raise ValueError("La date de fin de formation SSIAP 1 est obligatoire.")
    sst_trainer = _ssiap1_role_trainer(exam_payload, "sstTrainer", formateur)
    ssiap_trainer = _ssiap1_role_trainer(exam_payload, "ssiapTrainer", formateur)
    revision_trainer = _ssiap1_role_trainer(exam_payload, "revisionTrainer", ssiap_trainer)
    exam_trainer = _ssiap1_role_trainer(exam_payload, "examTrainer", ssiap_trainer)

    all_days = [day for day in aps_working_days_between(start_date, end_date, exam_iso) if day.isoformat() not in excluded_dates]
    required_total_days = 2 + _ssiap1_required_training_days()
    if len(all_days) < required_total_days:
        raise ValueError(aps_impossible_period_message(start_date, end_date, len(all_days) * APS_MAX_DAILY_MINUTES, (SSIAP1_PRESENCE_TOTAL_HOURS * 60)))
    session_days = all_days[:required_total_days]
    if session_days[0] != start_date or session_days[1] != start_date + timedelta(days=1):
        # Keep the rule generic but explicit: SST occupies the first two working days.
        pass
    sst_days = session_days[:2]
    training_days = session_days[2:]
    if training_days[0] != start_date + timedelta(days=2):
        raise ValueError("La formation SSIAP 1 doit commencer le troisième jour ouvré de la session, après les 14h SST.")
    if training_days[-1] != end_date:
        raise ValueError(f"La dernière journée SSIAP 1 doit être le {end_date.strftime('%d/%m/%Y')} afin d'y placer les 4 dernières heures et les révisions.")

    planning, totals, occurrence_counts, total_occurrences = [], {}, {}, {}
    for idx, day in enumerate(sst_days):
        detail = SSIAP1_SST_DAYS[idx]
        slots = [
            _ssiap1_day_slot(day, dt_time(8, 30), 240, uv=detail["code"], part=SSIAP1_SST_PART, title=detail["title"], items=detail["items"][:3], trainer=sst_trainer, room=salle, modality="sst"),
            _ssiap1_day_slot(day, dt_time(13, 30), 180, uv=detail["code"], part=SSIAP1_SST_PART, title=detail["title"], items=detail["items"][3:], trainer=sst_trainer, room=salle, modality="sst"),
        ]
        planning.append({"date": day.isoformat(), "dayLabel": aps_day_label(day), "category": "sst", "slots": slots})

    modules = []
    for detail in SSIAP1_SEQUENCE_DETAILS:
        part_label = f"{detail['part_number']}{'re' if detail['part_number'] == 1 else 'e'} partie — {detail['part_title']}"
        for kind, duration_key, items_key, label in (("content", "content_duration_minutes", "content_items", "CONTENU"), ("application", "application_duration_minutes", "application_items", "APPLICATION")):
            duration = int(detail.get(duration_key) or 0)
            if duration <= 0:
                continue
            modules.append({**detail, "part": part_label, "subpart_type": kind, "subpart_label": label, "subpart_duration_minutes": duration, "subpart_items": list(detail.get(items_key) or []), "remainingMinutes": duration, "offsetMinutes": 0})

    module_idx = 0; total_minutes = 0
    for current_day in training_days:
        slots = []
        day_capacity = min(APS_MAX_DAILY_MINUTES, SSIAP1_TOTAL_MINUTES - total_minutes)
        slot_definitions = [(dt_time(8, 30), min(240, day_capacity))]
        if day_capacity > 240:
            slot_definitions.append((dt_time(13, 30), day_capacity - 240))
        for slot_start, slot_minutes in slot_definitions:
            if slot_minutes <= 0:
                continue
            cursor = slot_start; remaining_slot = slot_minutes
            while remaining_slot > 0 and module_idx < len(modules):
                module = modules[module_idx]
                duration_minutes = min(remaining_slot, module["remainingMinutes"])
                end_time = add_minutes_to_time(cursor, duration_minutes)
                code = module["code"]
                occurrence_counts[(code, module["subpart_type"])] = occurrence_counts.get((code, module["subpart_type"]), 0) + 1
                total_occurrences[(code, module["subpart_type"])] = total_occurrences.get((code, module["subpart_type"]), 0) + 1
                slots.append({
                    "start": cursor.strftime("%H:%M"), "end": end_time.strftime("%H:%M"),
                    "duration": round(duration_minutes/60, 2), "durationMinutes": duration_minutes,
                    "uv": code, "part": module["part"], "sequence": code,
                    "partNumber": module["part_number"], "partTitle": module["part_title"],
                    "sequenceNumber": module["sequence_number"], "sequenceTitle": module["sequence_title"],
                    "title": module["sequence_title"], "content": module["sequence_title"],
                    "totalSequenceDurationMinutes": module["total_duration_minutes"],
                    "subpartType": module["subpart_type"], "subpartLabel": module["subpart_label"],
                    "subpartDurationMinutes": module["subpart_duration_minutes"],
                    "subpartItems": module["subpart_items"], "subpartOffsetMinutes": module["offsetMinutes"],
                    "splitIndex": occurrence_counts[(code, module["subpart_type"])],
                    "room": salle, "trainer": ssiap_trainer, "modality": "presentiel"})
                module["remainingMinutes"] -= duration_minutes; module["offsetMinutes"] += duration_minutes
                remaining_slot -= duration_minutes; total_minutes += duration_minutes; totals[code] = totals.get(code, 0) + duration_minutes; cursor = end_time
                if module["remainingMinutes"] == 0: module_idx += 1
                if total_minutes == SSIAP1_TOTAL_MINUTES: break
            if total_minutes == SSIAP1_TOTAL_MINUTES: break
        if current_day == training_days[-1]:
            slots.append(_ssiap1_day_slot(current_day, dt_time(13, 30), SSIAP1_REVISION_HOURS * 60, uv="REV-EXAM", part=SSIAP1_REVISION_PART, title="RÉVISIONS GÉNÉRALES ET PRÉPARATION À L’EXAMEN SSIAP 1", items=SSIAP1_REVISION_ITEMS, trainer=revision_trainer, room=salle, modality="revision"))
        if slots:
            planning.append({"date": current_day.isoformat(), "dayLabel": aps_day_label(current_day), "category": "ssiap1", "slots": slots})

    counts = {}
    for day in planning:
        for slot in day.get("slots", []):
            if not slot.get("partNumber"):
                continue
            key = (slot.get("sequence"), slot.get("subpartType")); counts[key] = counts.get(key, 0) + 1
            total_for_key = total_occurrences.get(key, 1)
            slot["splitLabel"] = "" if total_for_key == 1 else ("début" if counts[key] == 1 else "suite")
            items = list(slot.get("subpartItems") or [])
            if total_for_key > 1:
                total_minutes_for_subpart = max(1, int(slot.get("subpartDurationMinutes") or 0))
                start_offset = int(slot.get("subpartOffsetMinutes") or 0); end_offset = start_offset + int(slot.get("durationMinutes") or 0)
                selected = []
                for idx, item in enumerate(items):
                    item_midpoint = ((idx + 0.5) * total_minutes_for_subpart) / max(1, len(items))
                    is_last_chunk = counts[key] == total_for_key
                    if item_midpoint >= start_offset and (item_midpoint < end_offset or is_last_chunk): selected.append(item)
                if not selected and counts[key] > 1: selected = ["Poursuite et approfondissement du contenu commencé lors du créneau précédent"]
                slot["subpartDisplayItems"] = selected
                slot["subpartProgressLabel"] = f"{slot.get('subpartLabel', 'Sous-partie').capitalize()} : créneau {counts[key]} sur {total_for_key} — total {format_duration_from_minutes(total_minutes_for_subpart)}"
            else:
                slot["subpartDisplayItems"] = items

    exam_date = exam_payload.get("date") or exam_iso
    last_training = datetime.strptime(planning[-1]["date"], "%Y-%m-%d").date() if planning else None
    exam_date_obj = datetime.strptime(exam_date, "%Y-%m-%d").date() if exam_date else None
    if not exam_date_obj or (last_training and exam_date_obj <= last_training): raise ValueError("L'examen SSIAP 1 doit être placé après la fin des 67 heures de formation.")
    planning.append({"date": exam_date, "dayLabel": aps_day_label(exam_date_obj), "exam": True, "category": "exam", "slots": [{"start": exam_payload.get("start", "08:30"), "end": exam_payload.get("end", "16:30"), "duration": round((exam_payload.get("durationMinutes") or 0)/60, 2), "durationMinutes": exam_payload.get("durationMinutes") or 0, "uv": "EXAMEN", "part": "EXAMEN SSIAP 1", "sequence": "EXAMEN", "title": "EXAMEN SSIAP 1", "content": "Épreuves d'examen SSIAP 1", "room": exam_payload.get("room") or salle, "trainer": exam_trainer, "modality": "exam"}]})
    return planning, {k: round(v/60, 2) for k, v in totals.items()}, round(total_minutes/60, 2)

def ssiap1_summary_from_data(planning_data):
    totals = {k: 0 for k in SSIAP1_SEQUENCE_TOTALS}
    part_totals = {k: 0 for k in SSIAP1_PART_TOTALS}
    errors = []
    order = list(SSIAP1_SEQUENCE_TOTALS)
    seen_order = []
    total = 0
    sst_minutes = 0
    revision_minutes = 0
    exam = None
    exam_dates = []
    previous = None
    ssiap_day_minutes = {}
    presence_day_minutes = {}
    sst_day_minutes = {}
    revision_day_minutes = {}
    for day in planning_data or []:
        day_date = day.get("date")
        for slot in day.get("slots", []):
            minutes = int(round(float(slot.get("durationMinutes") or (float(slot.get("duration") or 0) * 60))))
            modality = (slot.get("modality") or "").strip().lower()
            if modality in {"exam", "examen"} or slot.get("uv") == "EXAMEN":
                exam = {"date": day_date, "start": slot.get("start"), "end": slot.get("end"), "room": slot.get("room"), "durationMinutes": minutes}
                exam_dates.append(day_date)
                continue
            try:
                start_dt = datetime.strptime(f"{day_date} {slot.get('start')}", "%Y-%m-%d %H:%M")
                end_dt = datetime.strptime(f"{day_date} {slot.get('end')}", "%Y-%m-%d %H:%M")
                if int((end_dt-start_dt).total_seconds()//60) != minutes: errors.append(f"Durée incohérente le {day_date} {slot.get('start')}-{slot.get('end')}.")
                if start_dt.minute not in {0,30} or end_dt.minute not in {0,30}: errors.append(f"Horaire irrégulier le {day_date} {slot.get('start')}-{slot.get('end')}.")
                if previous and start_dt < previous: errors.append(f"Ordre chronologique incohérent le {day_date} {slot.get('start')}.")
                previous = end_dt
            except Exception: errors.append(f"Horaire invalide le {day_date}: {slot.get('start')}-{slot.get('end')}.")
            presence_day_minutes[day_date] = presence_day_minutes.get(day_date, 0) + minutes
            if modality == "sst":
                sst_minutes += minutes
                sst_day_minutes[day_date] = sst_day_minutes.get(day_date, 0) + minutes
                continue
            if modality == "revision":
                revision_minutes += minutes
                revision_day_minutes[day_date] = revision_day_minutes.get(day_date, 0) + minutes
                continue
            code = slot.get("sequence") or slot.get("uv")
            if code not in totals:
                errors.append(f"Séquence SSIAP 1 inconnue: {code}"); continue
            ssiap_day_minutes[day_date] = ssiap_day_minutes.get(day_date, 0) + minutes
            totals[code] += round(minutes/60, 2); total += round(minutes/60, 2)
            part_totals[SSIAP1_PART_LABELS[code.split('-')[0]].split(' — ')[0]] += round(minutes/60, 2)
            if not seen_order or seen_order[-1] != code: seen_order.append(code)
    if sst_minutes != SSIAP1_SST_TOTAL_HOURS * 60: errors.append(f"Le total SST doit être exactement de 14h (actuel: {sst_minutes/60:g}h).")
    if sorted(sst_day_minutes.values()) != [APS_MAX_DAILY_MINUTES, APS_MAX_DAILY_MINUTES]: errors.append("Le SST doit comporter exactement deux journées de 7h.")
    if round(total, 2) != SSIAP1_TOTAL_HOURS: errors.append(f"Le total formation SSIAP 1 doit être exactement de 67h (actuel: {total:g}h).")
    day_values = list(ssiap_day_minutes.values())
    if len(day_values) != _ssiap1_required_training_days(): errors.append(f"La formation SSIAP 1 doit comporter exactement {_ssiap1_required_training_days()} journées de formation (actuel: {len(day_values)}).")
    full_days = [m for m in day_values if m == APS_MAX_DAILY_MINUTES]
    partial_days = [m for m in day_values if m != APS_MAX_DAILY_MINUTES]
    if len(full_days) != 9: errors.append(f"La formation SSIAP 1 doit comporter exactement 9 journées complètes de 7h (actuel: {len(full_days)}).")
    if partial_days != [240]: errors.append("La dernière journée SSIAP 1 doit être la seule journée partielle réglementaire et durer exactement 4h.")
    if revision_minutes != SSIAP1_REVISION_HOURS * 60: errors.append(f"Les révisions complémentaires doivent totaliser exactement 3h (actuel: {revision_minutes/60:g}h).")
    if revision_day_minutes and list(revision_day_minutes.values()) != [SSIAP1_REVISION_HOURS * 60]: errors.append("Les révisions SSIAP 1 doivent être placées en un bloc de 3h.")
    if presence_day_minutes.get(max(presence_day_minutes or {"":0})) != APS_MAX_DAILY_MINUTES: errors.append("La dernière journée de présence avant examen doit totaliser 7h.")
    for d in presence_day_minutes:
        parsed = parse_date(d)
        if not parsed or not is_french_working_day(parsed.date()): errors.append(f"Aucun créneau SSIAP 1/SST/révision ne doit être placé le week-end ou un jour férié: {d}.")
    if any(d in ssiap_day_minutes for d in sst_day_minutes): errors.append("Aucune séquence SSIAP 1 ne doit être placée pendant les journées SST.")
    if exam and exam.get("date") in presence_day_minutes: errors.append("La date d’examen SSIAP 1 ne doit pas contenir de créneau de formation.")
    for part, expected in SSIAP1_PART_TOTALS.items():
        if round(part_totals.get(part, 0), 2) != expected: errors.append(f"{part} doit totaliser {expected}h (actuel: {part_totals.get(part, 0):g}h).")
    for code, expected in SSIAP1_SEQUENCE_TOTALS.items():
        if round(totals.get(code, 0), 2) != expected: errors.append(f"{code} — {SSIAP1_SEQUENCE_LABELS[code]} doit totaliser {expected:g}h (actuel: {totals.get(code, 0):g}h).")
    if seen_order != order: errors.append("Les 24 séquences SSIAP 1 doivent être présentes et dans l'ordre réglementaire.")
    if not exam: errors.append("L'examen SSIAP 1 doit être présent dans le planning.")
    return {"total_hours": round(total, 2), "sst_hours": round(sst_minutes/60, 2), "revision_hours": round(revision_minutes/60, 2), "presence_total_hours": round((sst_minutes + int(total*60) + revision_minutes)/60, 2), "uv_totals": totals, "part_totals": part_totals, "uv_rows": [{"uv": c, "label": SSIAP1_SEQUENCE_LABELS[c], "title": SSIAP1_SEQUENCE_LABELS[c], "hours": totals[c], "expected": SSIAP1_SEQUENCE_TOTALS[c], "modality": "presentiel"} for c in order], "modality_totals": {"sst": round(sst_minutes/60, 2), "presentiel": round(total, 2), "revision": round(revision_minutes/60, 2)}, "daily_totals": {d: round(m/60, 2) for d, m in presence_day_minutes.items()}, "ssiap_daily_totals": {d: round(m/60, 2) for d, m in ssiap_day_minutes.items()}, "exam": exam, "errors": errors}

def find_center_image(*keywords):
    normalized_keywords = tuple((keyword or "").lower() for keyword in keywords)
    explicit_assets = (
        os.path.join(BASE_DIR, "templates", "signature"),
        os.path.join(BASE_DIR, "templates", "signature.png"),
        os.path.join(BASE_DIR, "templates", "Tampon.png"),
        os.path.join(BASE_DIR, "templates", "tampon.png"),
    )
    for asset_path in explicit_assets:
        name = os.path.basename(asset_path).lower()
        if os.path.isfile(asset_path) and any(keyword in name for keyword in normalized_keywords):
            return asset_path

    image_dir = os.path.join(BASE_DIR, "static", "img")
    if not os.path.isdir(image_dir):
        return None
    for entry in os.scandir(image_dir):
        if not entry.is_file():
            continue
        name = entry.name.lower()
        extension = name.rsplit(".", 1)[-1] if "." in name else ""
        if any(keyword in name for keyword in normalized_keywords) and extension in {"png", "jpg", "jpeg"}:
            return entry.path
    return None

def wrap_text_lines(text, max_width, font="Helvetica", size=9):
    from reportlab.pdfbase.pdfmetrics import stringWidth
    words = str(text or "").split()
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if stringWidth(candidate, font, size) <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines

def draw_wrapped_text(canvas, text, x, y, max_width, font="Helvetica", size=9, leading=11):
    from reportlab.pdfbase.pdfmetrics import stringWidth
    words = text.split()
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if stringWidth(candidate, font, size) <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    canvas.setFont(font, size)
    for line in lines:
        canvas.drawString(x, y, line)
        y -= leading
    return y


def split_uv_title(module_name):
    parts = (module_name or "").split(" ", 1)
    uv = parts[0].strip() if parts else ""
    title = parts[1].strip() if len(parts) > 1 else APS_UV_LABELS.get(uv, module_name)
    return uv, title

def aps_day_label(day_date):
    weekdays = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    months = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    return f"{weekdays[day_date.weekday()]} {day_date.day} {months[day_date.month - 1]} {day_date.year}"

def aps_blocks_to_planning_data(days, formateur, salle, planning_mode="full_presentiel"):
    planning = []
    elearning_remaining = 62.0 if planning_mode == "elearning_presentiel" else 0.0
    for day in days:
        day_date = day["date"]
        slots = []
        for block in day.get("blocks", []):
            uv, title = split_uv_title(block.get("uv"))
            duration = float(block.get("hours", 0))
            modality = "presentiel"
            if elearning_remaining > 0:
                modality = "elearning"
                elearning_remaining = round(elearning_remaining - duration, 2)
            slots.append({
                "start": block["start"].strftime("%H:%M"),
                "end": block["end"].strftime("%H:%M"),
                "duration": duration,
                "uv": uv,
                "title": title,
                "room": "" if modality == "elearning" else (salle or "Intégrale Academy – 54 chemin du Carreou – 83480 PUGET-SUR-ARGENS"),
                "trainer": "" if modality == "elearning" else (formateur or ""),
                "modality": modality,
            })
        planning.append({"date": day_date.isoformat(), "dayLabel": aps_day_label(day_date), "slots": slots})
    return planning

def generateApsFullPresentielPlanning(start_date, formateur, salle, end_date=None, exam_iso=""):
    days, totals, total_hours = build_aps_planning(start_date, end_date=end_date, exam_iso=exam_iso)
    return aps_blocks_to_planning_data(days, formateur, salle, "full_presentiel"), totals, total_hours

def generateApsElearningPresentielPlanning(start_date, formateur, salle, end_date=None, exam_iso="", session_id=None):
    sequence = [dict(item, remainingMinutes=int(item["durationMinutes"])) for item in APS_ELEARNING_PRESENTIEL_MODULES]
    expected_elearning = sum(item["durationMinutes"] for item in sequence if item["modality"] == "elearning")
    expected_presentiel = sum(item["durationMinutes"] for item in sequence if item["modality"] == "presentiel")
    if expected_elearning != APS_ELEARNING_MINUTES or expected_presentiel != APS_PRESENTIEL_MINUTES:
        raise ValueError("Configuration APS incohérente : la répartition e-learning / présentiel ne correspond pas au total attendu.")

    idx = 0
    current_day = start_date
    planning = []
    totals = {}

    # 1) Les 62h e-learning sont placées au début de session, sans salle ni formateur.
    while idx < len(sequence) and sequence[idx]["modality"] == "elearning":
        if end_date and current_day > end_date:
            available_days = aps_working_days_between(start_date, end_date, exam_iso)
            log_aps_generation_diagnostics(session_id, "elearning_presentiel", start_date, end_date, exam_iso, len(available_days), len(available_days) * APS_MAX_DAILY_MINUTES, APS_ELEARNING_MINUTES, APS_PRESENTIEL_MINUTES, APS_TOTAL_MINUTES)
            raise ValueError(aps_impossible_period_message(start_date, end_date, len(available_days) * APS_MAX_DAILY_MINUTES, APS_TOTAL_MINUTES))
        if not is_aps_training_day(current_day, exam_iso):
            current_day += timedelta(days=1)
            continue
        slots = []
        for slot_start, slot_minutes in ((dt_time(8, 30), 240), (dt_time(13, 30), 180)):
            cursor = slot_start
            remaining_slot = slot_minutes
            while remaining_slot > 0 and idx < len(sequence) and sequence[idx]["modality"] == "elearning":
                module = sequence[idx]
                duration_minutes = min(remaining_slot, module["remainingMinutes"])
                end_time = add_minutes_to_time(cursor, duration_minutes)
                slots.append({
                    "start": cursor.strftime("%H:%M"), "end": end_time.strftime("%H:%M"),
                    "duration": round(duration_minutes / 60, 2), "durationMinutes": duration_minutes,
                    "uv": module["uv"], "title": module["title"], "part": module["part"],
                    "room": "", "trainer": "", "modality": "elearning",
                })
                module["remainingMinutes"] -= duration_minutes
                remaining_slot -= duration_minutes
                cursor = end_time
                totals[module["title"]] = totals.get(module["title"], 0) + duration_minutes
                if module["remainingMinutes"] == 0:
                    idx += 1
            if idx >= len(sequence) or sequence[idx]["modality"] != "elearning":
                break
        if slots:
            planning.append({"date": current_day.isoformat(), "dayLabel": aps_day_label(current_day), "slots": slots})
        current_day += timedelta(days=1)

    # 2) Le présentiel démarre après l'e-learning et doit tenir jusqu'à la fin réelle de formation.
    presentiel_start = current_day
    while not is_aps_training_day(presentiel_start, exam_iso):
        presentiel_start += timedelta(days=1)
    presentiel_end = end_date
    if not presentiel_end:
        presentiel_end = presentiel_start + timedelta(days=60)
    presentiel_days = aps_working_days_between(presentiel_start, presentiel_end, exam_iso)
    standard_presentiel_minutes = len(presentiel_days) * APS_MAX_DAILY_MINUTES
    extended_presentiel_minutes = len(presentiel_days) * APS_EXTENDED_DAILY_MINUTES
    if APS_PRESENTIEL_MINUTES > extended_presentiel_minutes:
        log_aps_generation_diagnostics(
            session_id, "elearning_presentiel", start_date, presentiel_end, exam_iso,
            len(presentiel_days), standard_presentiel_minutes, APS_ELEARNING_MINUTES,
            APS_PRESENTIEL_MINUTES, APS_TOTAL_MINUTES, extended_presentiel_minutes,
        )
        raise ValueError(aps_impossible_period_message(presentiel_start, presentiel_end, standard_presentiel_minutes, APS_PRESENTIEL_MINUTES, extended_presentiel_minutes))

    # Capacité journalière réelle : 7h par défaut, puis jusqu'à 8h seulement si la
    # période présentielle est trop courte à 7h/jour. Le dépassement est posé en
    # priorité sur les derniers jours afin de garder un maximum de journées à 7h.
    day_capacities = {day: APS_MAX_DAILY_MINUTES for day in presentiel_days}
    missing_minutes = max(0, APS_PRESENTIEL_MINUTES - standard_presentiel_minutes)
    for day in reversed(presentiel_days):
        if missing_minutes <= 0:
            break
        extra = min(APS_EXTENDED_DAILY_MINUTES - APS_MAX_DAILY_MINUTES, missing_minutes)
        day_capacities[day] += extra
        missing_minutes -= extra
    elongated_days = sum(1 for minutes in day_capacities.values() if minutes > APS_MAX_DAILY_MINUTES)
    log_aps_generation_diagnostics(
        session_id, "elearning_presentiel", presentiel_start, presentiel_end, exam_iso,
        len(presentiel_days), standard_presentiel_minutes, APS_ELEARNING_MINUTES,
        APS_PRESENTIEL_MINUTES, APS_TOTAL_MINUTES, extended_presentiel_minutes, elongated_days,
        [(day.isoformat(), day_capacities[day] / 60) for day in presentiel_days], level="info",
    )

    # Si la plage présentielle est plus large que nécessaire, on conserve des journées de formation
    # au début et à la fin pour terminer explicitement à date_fin_session.
    final_day_distribution = []
    for current_day in presentiel_days:
        if idx >= len(sequence):
            break
        days_after = [d for d in presentiel_days if d > current_day]
        remaining_presentiel = sum(item["remainingMinutes"] for item in sequence[idx:] if item["modality"] == "presentiel")
        future_capacity = sum(day_capacities[d] for d in days_after)
        min_today = max(0, remaining_presentiel - future_capacity)
        is_last_training_day = current_day == presentiel_days[-1]
        if is_last_training_day:
            daily_limit = min(remaining_presentiel, day_capacities[current_day])
        else:
            # Répartition souple : le présentiel commence dès le prochain jour ouvré,
            # tout en gardant assez d'heures à placer pour finir sur date_fin_session.
            average_today = ((remaining_presentiel + len(days_after)) // (len(days_after) + 1) + 59) // 60 * 60
            daily_limit = min(day_capacities[current_day], max(min_today, min(average_today, remaining_presentiel)))
        if daily_limit <= 0:
            continue
        slots = []
        for slot_start, slot_minutes in ((dt_time(8, 30), 240), (dt_time(13, 30), 180), (dt_time(16, 30), 60)):
            cursor = slot_start
            remaining_slot = min(slot_minutes, daily_limit - sum(s["durationMinutes"] for s in slots))
            while remaining_slot > 0 and idx < len(sequence) and sequence[idx]["modality"] == "presentiel":
                module = sequence[idx]
                duration_minutes = min(remaining_slot, module["remainingMinutes"])
                end_time = add_minutes_to_time(cursor, duration_minutes)
                slots.append({
                    "start": cursor.strftime("%H:%M"), "end": end_time.strftime("%H:%M"),
                    "duration": round(duration_minutes / 60, 2), "durationMinutes": duration_minutes,
                    "uv": module["uv"], "title": module["title"], "part": module["part"],
                    "room": salle or "Intégrale Academy – 54 chemin du Carreou – 83480 PUGET-SUR-ARGENS", "trainer": formateur or "", "modality": "presentiel",
                })
                module["remainingMinutes"] -= duration_minutes
                remaining_slot -= duration_minutes
                cursor = end_time
                totals[module["title"]] = totals.get(module["title"], 0) + duration_minutes
                if module["remainingMinutes"] == 0:
                    idx += 1
            if idx >= len(sequence) or sequence[idx]["modality"] != "presentiel":
                break
        if slots:
            day_minutes = sum(slot["durationMinutes"] for slot in slots)
            final_day_distribution.append((current_day.isoformat(), day_minutes / 60))
            planning.append({"date": current_day.isoformat(), "dayLabel": aps_day_label(current_day), "slots": slots})

    app.logger.info(
        "Planning APS e-learning + présentiel généré session_id=%s heures_necessaires=%s jours_disponibles=%s capacite_7h=%s capacite_8h=%s journees_allongees=%s repartition_finale_heures_par_jour=%s",
        session_id, APS_PRESENTIEL_MINUTES / 60, len(presentiel_days), standard_presentiel_minutes / 60,
        extended_presentiel_minutes / 60, sum(1 for _, hours in final_day_distribution if hours > 7), final_day_distribution,
    )
    total_hours = sum(slot["durationMinutes"] for day in planning for slot in day["slots"]) / 60
    return planning, {k: round(v / 60, 2) for k, v in totals.items()}, total_hours


def _a3p_contract_days(planning_data):
    return len([d for d in planning_data or [] if d.get("slots")])

def _assert_a3p_pdf_text_safe(*parts):
    text = "\n".join(str(p or "") for p in parts)
    forbidden = [term for term in A3P_FORBIDDEN_TERMS if term in text]
    if forbidden:
        raise ValueError("Document A3P invalide: mentions interdites détectées (" + ", ".join(forbidden) + ").")

def _a3p_slot_to_aps_slot(slot):
    minutes = int(slot.get("durationMinutes") or (_minutes_from_hhmm(slot.get("end")) - _minutes_from_hhmm(slot.get("start"))))
    return {"start": slot.get("start") or "", "end": slot.get("end") or "", "duration": round(minutes / 60, 2), "durationMinutes": minutes, "uv": slot.get("code") or slot.get("uv") or "", "title": slot.get("title") or "", "trainer": (slot.get("trainer") or "").strip(), "room": (slot.get("room") or "").strip(), "modality": "presentiel"}

def _a3p_full_day_label(iso_date):
    parsed = parse_date(iso_date)
    if not parsed:
        return iso_date or "—"
    day_date = parsed.date()
    weekday = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"][day_date.weekday()]
    return f"{weekday} {day_date.strftime('%d/%m/%Y')}"


def _a3p_planning_as_aps_data(planning):
    converted = []
    for day in planning or []:
        converted.append({"date": day.get("date"), "dayLabel": _a3p_full_day_label(day.get("date")), "slots": [_a3p_slot_to_aps_slot(slot) for slot in day.get("slots", [])]})
    return converted

def _a3p_document_profile(summary=None, planning=None):
    summary = summary or {}
    module_totals = summary.get("moduleTotals") or {}
    rows = [{"uv": m["code"], "label": m["title"], "hours": module_totals.get(m["code"], m["hours"]), "expected": m["hours"]} for m in A3P_MODULES]
    total = summary.get("totalHours", A3P_TOTAL_HOURS)
    return {"validate": "a3p", "source_planning": planning or [], "short_label": "A3P", "planning_title": "PLANNING DE FORMATION A3P", "subtitle": f"Agent de protection physique des personnes — {A3P_TOTAL_HOURS} heures hors examen", "modality_line": f"Modalité : 100% présentiel • Présentiel : {A3P_TOTAL_HOURS}h • Examen séparé", "summary": {"total_hours": total, "uv_totals": module_totals, "uv_rows": rows, "modality_totals": {"presentiel": total}, "days_count": len(planning or []), "slots_count": sum(len(d.get("slots", [])) for d in planning or []), "errors": []}}

def _a3p_session_for_shared_docs(session_data):
    planning = session_data.get("a3pPlanningData") or []
    converted = _a3p_planning_as_aps_data(planning)
    fallback_trainer = (session_data.get("a3pTrainerName") or "").strip()
    fallback_room = (session_data.get("a3pRoom") or session_data.get("salle") or session_data.get("room") or "").strip()
    for day in converted:
        for slot in day.get("slots", []):
            if fallback_trainer and not slot.get("trainer"):
                slot["trainer"] = fallback_trainer
            if fallback_room and not slot.get("room"):
                slot["room"] = fallback_room
    copied = dict(session_data)
    copied.update({"formation": "A3P", "apsPlanningMode": "full_presentiel", "apsPlanningData": converted, "apsAttendanceStudents": session_data.get("a3pAttendanceStudents") or session_data.get("apsAttendanceStudents") or [], "salle": session_data.get("a3pRoom") or session_data.get("salle") or session_data.get("room") or ""})
    return copied, converted

def generate_a3p_planning_pdf(session_data, output_path):
    planning = session_data.get("a3pPlanningData") or []
    errors, summary = validate_a3p_planning(planning, session_data.get("date_exam"))
    if errors:
        raise ValueError(" ".join(errors))
    _assert_a3p_pdf_text_safe(session_data.get("a3pTrainerName"), session_data.get("a3pRoom"))
    shared_session, converted = _a3p_session_for_shared_docs(session_data)
    return generate_aps_planning_pdf(shared_session, session_data.get("a3pTrainerName") or "", output_path, planning_data=converted, planning_mode="full_presentiel", document_profile=_a3p_document_profile(summary, planning))

def generate_a3p_attendance_pdf(session_data, output_path):
    planning = session_data.get("a3pPlanningData") or []
    errors, _summary = validate_a3p_planning(planning, session_data.get("date_exam"))
    if errors:
        raise ValueError(" ".join(errors))
    shared_session, _converted = _a3p_session_for_shared_docs(session_data)
    return generate_attendance_pdf_common(shared_session, output_path, training_type="A3P", subtitle="TFP Agent de Protection Physique des Personnes (A3P)")

def _a3p_trainer_contract_data(session_data, contract):
    shared_session, converted = _a3p_session_for_shared_docs(session_data)
    trainer_name = session_data.get("a3pTrainerName") or contract.get("trainerName") or ""
    contract = merge_formateur_contract_defaults(contract, find_formateur_by_identity(name=trainer_name, email=contract.get("trainerEmail") or contract.get("email") or contract.get("trainerEmail")))
    interventions = aps_trainer_interventions(converted, trainer_name)
    daily = float(contract.get("dailyRate") or 0)
    billed_days = float(contract.get("billedDays") or interventions["calendarDays"] or _a3p_contract_days(session_data.get("a3pPlanningData") or []))
    total_ht = round(daily * billed_days, 2)
    vat_enabled = bool(contract.get("vatEnabled"))
    vat_rate = float(contract.get("vatRate") or 20)
    vat_amount = round(total_ht * vat_rate / 100, 2) if vat_enabled else 0
    payload = merge_formateur_contract_defaults(contract, find_formateur_by_identity(name=trainer_name, email=contract.get("trainerEmail") or contract.get("email") or contract.get("trainerEmail")))
    payload.update({"trainerName": trainer_name, "interventions": interventions["interventions"], "calculatedHours": interventions["totalHours"], "calculatedDays": interventions["calculatedDays"], "billedDays": billed_days, "dailyRate": daily, "totalHT": total_ht, "vatEnabled": vat_enabled, "vatRate": vat_rate, "vatAmount": vat_amount, "totalTTC": round(total_ht + vat_amount, 2)})
    return shared_session, payload

def generate_a3p_trainer_contract_pdf(session_data, contract, output_path):
    errors, _summary = validate_a3p_planning(session_data.get("a3pPlanningData") or [], session_data.get("date_exam"))
    if errors:
        raise ValueError(" ".join(errors))
    shared_session, payload = _a3p_trainer_contract_data(session_data, contract or {})
    return generate_aps_trainer_contract_pdf(shared_session, payload, output_path)

def generate_a3p_simple_pdf(session_data, output_path, kind="planning", contract=None):
    if kind == "planning":
        return generate_a3p_planning_pdf(session_data, output_path)
    if kind == "attendance":
        return generate_a3p_attendance_pdf(session_data, output_path)
    if kind == "contract":
        return generate_a3p_trainer_contract_pdf(session_data, contract or {}, output_path)
    raise ValueError("Type de document A3P invalide.")

def build_aps_planning_data(start_date, formateur, salle, planning_mode="full_presentiel", end_date=None, exam_iso="", session_id=None):
    if planning_mode == "elearning_presentiel":
        return generateApsElearningPresentielPlanning(start_date, formateur, salle, end_date=end_date, exam_iso=exam_iso, session_id=session_id)
    return generateApsFullPresentielPlanning(start_date, formateur, salle, end_date=end_date, exam_iso=exam_iso)


AFC_APS_SSIAP_LABEL = "AFC France Travail APS + SSIAP"
AFC_APS_SSIAP_TOTAL_HOURS = 393
AFC_APS_SSIAP_TECHNICAL_HOURS = 273
AFC_APS_SSIAP_EXPECTED_MINUTES = {
    "RAN": 55 * 60,
    "ACCUEIL": 210,
    "APS": 175 * 60,
    "EXAM_APS": 7 * 60,
    "H0B0": 7 * 60,
    "SSIAP1": 70 * 60,
    "EXAM_SSIAP1": 7 * 60,
    "BILAN": 210,
    "SP": 45 * 60,
    "PAF": 20 * 60,
}
AFC_APS_SSIAP_SUMMARY_ORDER = ["RAN", "ACCUEIL", "APS", "EXAM_APS", "H0B0", "SSIAP1", "EXAM_SSIAP1", "SP", "PAF", "BILAN"]
AFC_APS_SSIAP_LABELS = {
    "RAN": "Remise à niveau (RAN)",
    "ACCUEIL": "Accueil",
    "APS": "Formation Agent de prévention et de sécurité (APS)",
    "SP": "Soutien personnalisé (SP)",
    "EXAM_APS": "Examen Agent de prévention et de sécurité (APS)",
    "H0B0": "Habilitation électrique H0B0",
    "SSIAP1": "Formation SSIAP 1",
    "EXAM_SSIAP1": "Examen SSIAP 1",
    "BILAN": "Bilan de formation",
    "PAF": "Préparation à l’après-formation (PAF)",
}
AFC_DAY_SEGMENTS = ((8 * 60 + 30, 12 * 60 + 30), (13 * 60 + 30, 16 * 60 + 30))
AFC_TECHNICAL_CODES = {"ACCUEIL", "APS", "EXAM_APS", "H0B0", "SSIAP1", "EXAM_SSIAP1", "BILAN"}
AFC_ACCOMPANIMENT_CODES = {"SP", "PAF"}
AFC_CATEGORY_COLORS = {
    "RAN": "#bfdbfe",
    "ACCUEIL": "#5eead4",
    "APS": "#1d4ed8",
    "EXAM_APS": "#581c87",
    "H0B0": "#fde047",
    "SSIAP1": "#fb923c",
    "EXAM_SSIAP1": "#dc2626",
    "SP": "#d8b4fe",
    "PAF": "#86efac",
    "BILAN": "#374151",
}


DSF_DIR = os.path.join(DATA_DIR, "afc_dsf")
os.makedirs(DSF_DIR, exist_ok=True)
AFC_DSF_STATUS_FINALIZED = "Finalisée"
AFC_DSF_STATUS_CANCELLED = "Annulée"
AFC_DSF_MODULES = {
    "FT": {"label": "Formation technique (FT)", "theoreticalHours": 273, "colors": ["#1e3a8a", "#dbeafe"]},
    "RAN": {"label": "Remise à niveau (RAN)", "theoreticalHours": 55, "colors": ["#38bdf8", "#e0f2fe"]},
    "SP": {"label": "Soutien personnalisé (SP)", "theoreticalHours": 45, "colors": ["#a855f7", "#f3e8ff"]},
    "PAF": {"label": "Préparation à l’après-formation (PAF)", "theoreticalHours": 20, "colors": ["#22c55e", "#dcfce7"]},
}
AFC_DSF_FT_CATEGORIES = {"ACCUEIL", "APS", "EXAM_APS", "H0B0", "SSIAP1", "EXAM_SSIAP1", "BILAN"}

def is_afc_aps_ssiap_session(session_data):
    try:
        return normalize_training_code(session_data) == "AFC_APS_SSIAP"
    except ValueError:
        return False

def afc_dsf_students(session_data):
    students = session_data.get("apsAttendanceStudents") or []
    result = []
    for idx, st in enumerate(students):
        last = (st.get("lastName") or st.get("nom") or "").strip()
        first = (st.get("firstName") or st.get("prenom") or "").strip()
        if not (last or first):
            continue
        result.append({"id": st.get("id") or st.get("studentId") or f"student-{idx+1}", "lastName": last, "firstName": first, "displayName": f"{last} {first}".strip(), "entryDate": st.get("entryDate") or st.get("dateEntree") or st.get("startDate") or ""})
    return result

def afc_dsf_module_for_slot(slot):
    category = slot.get("afcCategory") or slot.get("category") or slot.get("uv")
    kind = slot.get("afcKind")
    if kind == "FT" or category in AFC_DSF_FT_CATEGORIES:
        return "FT"
    if category in {"RAN", "SP", "PAF"}:
        return category
    return None

def afc_dsf_slot_key(session_id, student_id, day, slot, module):
    return "|".join([str(session_id), str(student_id), day.get("date", ""), slot.get("start", ""), slot.get("end", ""), module])

def afc_dsf_planned_slots(session_data, start_iso=None, end_iso=None, modules=None):
    modules = set(modules or AFC_DSF_MODULES)
    slots = []
    for day in session_data.get("apsPlanningData") or []:
        d = day.get("date")
        if start_iso and d < start_iso: continue
        if end_iso and d > end_iso: continue
        for slot in day.get("slots") or []:
            module = afc_dsf_module_for_slot(slot)
            if module not in modules: continue
            minutes = int(slot.get("durationMinutes") or round(float(slot.get("duration") or 0) * 60))
            if minutes <= 0: continue
            slots.append({"date": d, "start": slot.get("start"), "end": slot.get("end"), "module": module, "minutes": minutes, "title": slot.get("title") or slot.get("content") or slot.get("uv"), "afcCategory": slot.get("afcCategory")})
    return slots

def afc_dsf_finalized(dsfs):
    return [d for d in dsfs or [] if d.get("status") == AFC_DSF_STATUS_FINALIZED]

def afc_dsf_billed_keys(dsfs):
    keys = set()
    for dsf in afc_dsf_finalized(dsfs):
        for key in dsf.get("billedSlotKeys") or []:
            keys.add(key)
    return keys

def afc_dsf_next_number(session_data):
    return max([int(d.get("number") or 0) for d in session_data.get("afcDsfs") or []] or [0]) + 1

def afc_dsf_effective_start(period_start, student):
    entry = (student.get("entryDate") or "").strip()
    if entry and re.match(r"^\d{4}-\d{2}-\d{2}$", entry):
        return max(period_start, entry)
    return period_start

def afc_dsf_compute(session_data, start_iso, end_iso, modules):
    if not is_afc_aps_ssiap_session(session_data): raise ValueError("Session AFC France Travail APS + SSIAP requise.")
    if not modules or len(modules) > 2 or len(set(modules)) != len(modules): raise ValueError("Sélectionnez un ou deux modules maximum.")
    if any(m not in AFC_DSF_MODULES for m in modules): raise ValueError("Module DSF invalide.")
    if start_iso > end_iso: raise ValueError("La date de début ne peut pas être postérieure à la date de fin.")
    if session_data.get("date_debut") and start_iso < session_data.get("date_debut"): raise ValueError("La période doit être comprise dans la session.")
    if session_data.get("date_fin") and end_iso > session_data.get("date_fin"): raise ValueError("La période doit être comprise dans la session.")
    students = afc_dsf_students(session_data)
    if not students: raise ValueError("Au moins un stagiaire est requis.")
    requested = afc_dsf_planned_slots(session_data, start_iso, end_iso, modules)
    if not requested:
        labels = ", ".join(AFC_DSF_MODULES[m]["label"] for m in modules)
        raise ValueError(f"Aucune heure planifiée pour {labels} sur cette période.")
    billed_keys = afc_dsf_billed_keys(session_data.get("afcDsfs") or [])
    student_rows, billed_slot_keys, billed_slots = [], [], []
    per_student_hours = {m: 0 for m in modules}; already = {m: 0 for m in modules}; requested_hours = {m: 0 for m in modules}
    for slot in requested:
        requested_hours[slot["module"]] += slot["minutes"] / 60
    for st in students:
        row = {"id": st["id"], "lastName": st["lastName"], "firstName": st["firstName"], "displayName": st["displayName"], "entryDate": st.get("entryDate") or "", "modules": {m: 0 for m in modules}, "totalHours": 0}
        effective_start = afc_dsf_effective_start(start_iso, st)
        for slot in requested:
            if slot["date"] < effective_start:
                continue
            key = afc_dsf_slot_key(session_data.get("id"), st["id"], {"date": slot["date"]}, slot, slot["module"])
            if key in billed_keys:
                already[slot["module"]] += slot["minutes"] / 60
                continue
            hours = slot["minutes"] / 60
            row["modules"][slot["module"]] += hours; row["totalHours"] += hours
            billed_slot_keys.append(key); billed_slots.append({**slot, "studentId": st["id"], "studentName": st["displayName"]})
        student_rows.append(row)
    for m in modules:
        per_student_hours[m] = round(sum(r["modules"].get(m, 0) for r in student_rows) / len(students), 2)
    totals = {m: round(sum(r["modules"].get(m, 0) for r in student_rows), 2) for m in modules}
    total = round(sum(totals.values()), 2)
    if total <= 0: raise ValueError("Aucune heure restante à facturer pour les modules et la période sélectionnés.")
    return {"periodStart": start_iso, "periodEnd": end_iso, "modules": modules, "students": student_rows, "studentCount": len(students), "hoursPerStudent": per_student_hours, "moduleTotals": totals, "totalHours": total, "requestedHours": {m: round(requested_hours[m] * len(students), 2) for m in modules}, "alreadyBilledHours": {m: round(already[m], 2) for m in modules}, "billedSlotKeys": billed_slot_keys, "billedSlots": billed_slots, "hasAlreadyBilled": any(v > 0 for v in already.values())}

def afc_dsf_summary(session_data):
    students = afc_dsf_students(session_data)
    all_slots = afc_dsf_planned_slots(session_data)
    planned_by_student = {st["id"]: {m: 0 for m in AFC_DSF_MODULES} for st in students}
    for st in students:
        effective_start = afc_dsf_effective_start(session_data.get("date_debut") or "", st)
        for slot in all_slots:
            if slot["date"] < effective_start:
                continue
            planned_by_student[st["id"]][slot["module"]] += slot["minutes"] / 60
    billed_by_student = {st["id"]: {m: 0 for m in AFC_DSF_MODULES} for st in students}
    for dsf in afc_dsf_finalized(session_data.get("afcDsfs") or []):
        for row in dsf.get("students") or []:
            sid = row.get("id")
            if sid in billed_by_student:
                for m, h in (row.get("modules") or {}).items():
                    if m in billed_by_student[sid]: billed_by_student[sid][m] += float(h or 0)
    cards = []
    for m, meta in AFC_DSF_MODULES.items():
        planned_values = [planned_by_student.get(st["id"], {}).get(m, 0) for st in students]
        planned_total = sum(planned_values)
        billed_total = sum(v[m] for v in billed_by_student.values())
        remaining_total = planned_total - billed_total
        overbilled_total = max(0, -remaining_total)
        if overbilled_total > 0.0001:
            app.logger.warning(
                "DSF AFC overbilled summary session=%s module=%s planned=%.2f billed=%.2f overbilled=%.2f",
                session_data.get("id"), m, planned_total, billed_total, overbilled_total,
            )
        remaining_total = max(0, remaining_total)
        cards.append({"code": m, **meta, "plannedPerStudent": round((planned_total / len(students)) if students else 0,2), "plannedTotal": round(planned_total,2), "billedTotal": round(billed_total,2), "overbilledTotal": round(overbilled_total,2), "remainingPerStudent": round((remaining_total / len(students)) if students else 0,2), "remainingTotal": round(remaining_total,2)})
    detail=[]
    for st in students:
        row={"student":st,"modules":{},"plannedTotal":0,"billedTotal":0,"remainingTotal":0}
        for m in AFC_DSF_MODULES:
            p=planned_by_student.get(st["id"],{}).get(m,0); b=billed_by_student.get(st["id"],{}).get(m,0); r=max(0,p-b); row["modules"][m]={"planned":round(p,2),"billed":round(b,2),"remaining":round(r,2)}; row["plannedTotal"]+=p; row["billedTotal"]+=b; row["remainingTotal"]+=r
        detail.append(row)
    return {"cards": cards, "detail": detail, "total": {"theoreticalHours":393,"billedTotal":round(sum(c["billedTotal"] for c in cards),2),"remainingTotal":round(sum(c["remainingTotal"] for c in cards),2)}}

def parse_interruption_ranges(value):
    ranges = []
    if not value:
        return ranges
    if isinstance(value, (list, tuple)):
        items = []
        for item in value:
            if isinstance(item, dict):
                items.append(f"{item.get('start') or item.get('startDate') or ''} au {item.get('end') or item.get('endDate') or ''}")
            else:
                items.append(item)
    else:
        items = re.split(r"[,;\n]+", str(value))
    for item in items:
        parts = re.split(r"\s+(?:au|à|a|-|→)\s+", str(item).strip(), maxsplit=1, flags=re.I)
        if len(parts) == 1 and ".." in parts[0]:
            parts = parts[0].split("..", 1)
        start = parse_date(parts[0].strip()) if parts and parts[0].strip() else None
        end = parse_date(parts[1].strip()) if len(parts) > 1 and parts[1].strip() else start
        if start and end:
            if end < start:
                start, end = end, start
            ranges.append((start.date(), end.date()))
    return ranges

def is_interrupted_day(day, interruptions):
    return any(start <= day <= end for start, end in (interruptions or []))

def is_afc_working_day(day, interruptions=None):
    return is_french_working_day(day) and not is_interrupted_day(day, interruptions or [])

def afc_minutes_to_hhmm(minute):
    return f"{minute//60:02d}:{minute%60:02d}"

def afc_build_main_sequence():
    seq = [
        {"code":"RAN", "category":"RAN", "uv":"RAN", "title":"Remise à niveau (RAN)", "durationMinutes":55*60, "afcKind":"RAN"},
        {"code":"ACCUEIL", "category":"ACCUEIL", "uv":"ACCUEIL", "title":"Accueil", "durationMinutes":210, "afcKind":"FT"},
    ]
    for name, hours in APS_MODULES:
        uv, title = split_uv_title(name)
        seq.append({"code":"APS", "category":"APS", "uv":uv, "title":title, "content":title, "durationMinutes":int(hours*60), "afcKind":"FT"})
    seq.extend([
        {"code":"EXAM_APS", "category":"EXAM_APS", "uv":"EXAM_APS", "title":"Examen APS", "durationMinutes":420, "afcKind":"FT"},
        {"code":"H0B0", "category":"H0B0", "uv":"H0B0", "title":"Habilitation électrique H0B0", "durationMinutes":420, "afcKind":"FT"},
    ])
    for detail in SSIAP1_SEQUENCE_DETAILS:
        seq.append({"code":"SSIAP1", "category":"SSIAP1", "uv":detail["code"], "sequence":detail["code"], "part":detail["part_title"], "title":detail["sequence_title"], "content":detail["sequence_title"], "durationMinutes":detail["total_duration_minutes"], "afcKind":"FT", "partNumber":detail["part_number"], "sequenceNumber":detail["sequence_number"], "sequenceTitle":detail["sequence_title"], "totalSequenceDurationMinutes":detail["total_duration_minutes"], "subpartItems":detail.get("content_items") or detail.get("application_items") or []})
    seq.append({"code":"SSIAP1", "category":"SSIAP1", "uv":"REV-SSIAP1", "title":"Révisions générales SSIAP 1", "content":"Révisions générales et préparation pédagogique SSIAP 1", "durationMinutes":180, "afcKind":"FT"})
    seq.extend([
        {"code":"EXAM_SSIAP1", "category":"EXAM_SSIAP1", "uv":"EXAM_SSIAP1", "title":"Examen SSIAP 1", "durationMinutes":420, "afcKind":"FT"},
        {"code":"BILAN", "category":"BILAN", "uv":"BILAN", "title":"Bilan de formation", "durationMinutes":210, "afcKind":"FT"},
    ])
    return seq

def afc_slot_from_module(day, start, end, module, trainer, room):
    minutes = end - start
    slot = {"start": afc_minutes_to_hhmm(start), "end": afc_minutes_to_hhmm(end), "duration": round(minutes/60,2), "durationMinutes": minutes, "uv": module.get("uv") or module["code"], "sequence": module.get("sequence") or module.get("uv") or module["code"], "part": module.get("part") or AFC_APS_SSIAP_LABELS.get(module.get("category") or module["code"], module.get("title")), "title": module.get("title"), "content": module.get("content") or module.get("title"), "room": room, "trainer": trainer, "modality": "presentiel", "afcKind": module.get("afcKind"), "afcCategory": module.get("category") or module["code"]}
    for key in ("partNumber", "sequenceNumber", "sequenceTitle", "totalSequenceDurationMinutes", "subpartItems"):
        if module.get(key) is not None: slot[key] = module[key]
    return slot

def afc_active_weeks(start_date, interruptions, count):
    weeks, seen, day = [], set(), start_date
    while len(weeks) < count:
        if is_afc_working_day(day, interruptions):
            key = day.isocalendar()[:2]
            if key not in seen:
                seen.add(key); weeks.append(key)
        day += timedelta(days=1)
    return weeks

def afc_nth_working_day(start_date, interruptions=None, count=57):
    if not start_date:
        return None
    remaining, day = count, start_date
    while remaining > 0:
        if is_afc_working_day(day, interruptions or []):
            remaining -= 1
            if remaining == 0:
                return day
        day += timedelta(days=1)
    return None

def build_afc_aps_ssiap_planning_data(start_date, trainer="", room="", interruptions=None, contractual_end_date=None):
    """Génère le parcours AFC APS + SSIAP France Travail sur les 57 jours contractuels."""
    interruptions = interruptions or [(date(2026, 12, 23), date(2027, 1, 4))]
    end_date = contractual_end_date or date(2027, 2, 15)
    if isinstance(end_date, str):
        parsed = parse_date(end_date); end_date = parsed.date() if parsed else None
    if not end_date:
        raise ValueError("La date de fin contractuelle AFC est invalide.")

    eligible_dates = []
    day = start_date
    while day <= end_date:
        if is_afc_working_day(day, interruptions):
            eligible_dates.append(day)
        day += timedelta(days=1)
    if len(eligible_dates) != 57:
        raise ValueError("La période sélectionnée ne contient pas les 57 jours prévus par la commande France Travail.")
    if eligible_dates[-1] != end_date:
        raise ValueError("La date de fin AFC doit être un jour admissible et contenir l’examen SSIAP 1.")

    six_hour_days = {date(2026, 11, 25), date(2026, 12, 1), date(2026, 12, 8), date(2027, 1, 6), date(2027, 2, 3), date(2027, 2, 11)}
    if not six_hour_days.issubset(set(eligible_dates)):
        raise ValueError("Les journées AFC de 6h ne correspondent pas aux jours admissibles.")
    daily_targets = {d: (6 * 60 if d in six_hour_days else 7 * 60) for d in eligible_dates}

    raw = afc_build_main_sequence()
    ran = dict(next(m for m in raw if m["category"] == "RAN"), remainingMinutes=55 * 60)
    accueil = dict(next(m for m in raw if m["category"] == "ACCUEIL"), remainingMinutes=210)
    exam = dict(next(m for m in raw if m["category"] == "EXAM_SSIAP1"), remainingMinutes=420)
    sequence = [dict(m, remainingMinutes=m["durationMinutes"]) for m in raw if m["category"] not in {"RAN", "ACCUEIL", "EXAM_SSIAP1"}]

    planning_map = {}
    def item(d):
        return planning_map.setdefault(d.isoformat(), {"date": d.isoformat(), "dayLabel": aps_day_label(d), "category": "afc_aps_ssiap", "slots": []})
    def add(d, start, minutes, mod):
        item(d)["slots"].append(afc_slot_from_module(d, start, start + minutes, mod, trainer, room))
    def module_for(code):
        return {"code": code, "category": code, "uv": code, "title": AFC_APS_SSIAP_LABELS[code], "content": "Accompagnement transversal", "afcKind": code}
    # RAN strictement en premier : 7h du 16 au 24/11, puis 6h le 25/11.
    for d in eligible_dates[:8]:
        minutes = daily_targets[d]
        if minutes > ran["remainingMinutes"]:
            minutes = ran["remainingMinutes"]
        add(d, 8 * 60 + 30, min(240, minutes), ran)
        if minutes > 240:
            add(d, 13 * 60 + 30, minutes - 240, ran)
        ran["remainingMinutes"] -= minutes
    if ran["remainingMinutes"] != 0:
        raise ValueError("Les 55h de RAN n’ont pas pu être placées avant toute autre activité.")

    # Accueil sur une matinée puis démarrage APS dans la même journée.
    accueil_day = eligible_dates[8]
    add(accueil_day, 8 * 60 + 30, 210, accueil)
    free_intervals = {d: [(8 * 60 + 30, 12 * 60 + 30), (13 * 60 + 30, 13 * 60 + 30 + max(0, daily_targets[d] - 240))] for d in eligible_dates}
    free_intervals[accueil_day] = [(12 * 60, 12 * 60 + 30), (13 * 60 + 30, 16 * 60 + 30)]
    free_intervals[end_date] = []

    # Accompagnements en fin de semaine, avant le dernier jour et jamais en début de parcours.
    # Quand une semaine contient SP et PAF, le SP précède immédiatement le PAF et le PAF
    # occupe les cinq dernières heures de la semaine.
    sp_weeks = {date(2026,11,30), date(2026,12,7), date(2026,12,14), date(2026,12,21), date(2027,1,4), date(2027,1,11), date(2027,1,18), date(2027,1,25), date(2027,2,1)}
    paf_weeks = {date(2027,1,11), date(2027,1,18), date(2027,1,25), date(2027,2,8)}
    for monday in sorted(sp_weeks | paf_weeks):
        week_days = [d for d in eligible_dates if (d - timedelta(days=d.weekday())) == monday and d < end_date]
        if not week_days:
            continue
        last = week_days[-1]
        has_sp = monday in sp_weeks
        has_paf = monday in paf_weeks
        if has_sp and has_paf:
            prev = week_days[-2] if len(week_days) >= 2 else None
            if not prev:
                raise ValueError("Impossible de placer SP puis PAF sur la semaine AFC.")
            add(prev, 13*60+30, 180, module_for("SP"))
            add(last, 8*60+30, 120, module_for("SP"))
            add(last, 10*60+30, 120, module_for("PAF"))
            add(last, 13*60+30, 180, module_for("PAF"))
            free_intervals[prev] = [(8*60+30, 12*60+30)]
            free_intervals[last] = []
        elif has_sp:
            add(last, 10*60+30, 120, module_for("SP")); add(last, 13*60+30, 180, module_for("SP"))
            free_intervals[last] = [(8*60+30, 10*60+30)]
        elif has_paf:
            add(last, 10*60+30, 120, module_for("PAF")); add(last, 13*60+30, 180, module_for("PAF"))
            free_intervals[last] = [(8*60+30, 10*60+30)]

    idx = 0
    def add_technical(d, start, minutes):
        nonlocal idx
        cur = start; left = minutes
        while left > 0 and idx < len(sequence):
            m = sequence[idx]
            take = min(left, m["remainingMinutes"])
            add(d, cur, take, m)
            m["remainingMinutes"] -= take; cur += take; left -= take
            if m["remainingMinutes"] == 0:
                idx += 1
        return left

    for d in eligible_dates[8:-1]:
        intervals = list(free_intervals.get(d, []))
        if not intervals:
            continue
        for start, end in intervals:
            if idx >= len(sequence):
                break
            add_technical(d, start, end - start)
    if idx < len(sequence) or any(m.get("remainingMinutes", 0) for m in sequence[idx:]):
        raise ValueError(f"Impossible de planifier les 393 heures entre le {start_date.strftime('%d/%m/%Y')} et le {end_date.strftime('%d/%m/%Y')} avec les règles AFC France Travail.")
    add(end_date, 8*60+30, 240, exam); add(end_date, 13*60+30, 180, exam)

    for it in planning_map.values():
        it["slots"].sort(key=lambda x: x["start"])
    planning = [planning_map[d.isoformat()] for d in eligible_dates if planning_map.get(d.isoformat(), {}).get("slots")]
    summary = afc_aps_ssiap_summary_from_data(planning, interruptions=interruptions, contractual_end_date=end_date)
    if summary["errors"]:
        raise ValueError(" ".join(summary["errors"]))
    return planning

def calculate_actual_afc_hours(events):
    """Additionne les durées réelles AFC par catégorie depuis les créneaux planifiés."""
    totals = {k: 0 for k in AFC_APS_SSIAP_EXPECTED_MINUTES}
    for day in events or []:
        for slot in day.get("slots", []) or []:
            cat = slot.get("afcCategory") or slot.get("category") or slot.get("uv") or ""
            minutes = int(round(float(slot.get("durationMinutes") or float(slot.get("duration") or 0) * 60)))
            totals[cat] = totals.get(cat, 0) + minutes
    return {k: round(v / 60, 2) for k, v in totals.items()}

def validate_actual_afc_hours(events):
    actual_hours = calculate_actual_afc_hours(events)
    expected_hours = {k: round(v / 60, 2) for k, v in AFC_APS_SSIAP_EXPECTED_MINUTES.items()}
    errors = []
    for code in AFC_APS_SSIAP_SUMMARY_ORDER:
        actual = actual_hours.get(code, 0)
        expected = expected_hours.get(code, 0)
        if actual != expected:
            errors.append(f"- {AFC_APS_SSIAP_LABELS.get(code, code)} : {actual:g} h planifiées au lieu de {expected:g} h")
    total = round(sum(actual_hours.values()), 2)
    if total != AFC_APS_SSIAP_TOTAL_HOURS:
        errors.append(f"- Total AFC : {total:g} h planifiées au lieu de {AFC_APS_SSIAP_TOTAL_HOURS:g} h")
    if errors:
        return ["Planning AFC invalide :\n" + "\n".join(errors)]
    return []

def afc_aps_ssiap_summary_from_data(planning_data, interruptions=None, contractual_end_date=None):
    actual_hours = calculate_actual_afc_hours(planning_data)
    totals = {k: int(round(actual_hours.get(k, 0) * 60)) for k in AFC_APS_SSIAP_EXPECTED_MINUTES}
    errors = []; week_buckets = {}; total = 0
    for day in planning_data or []:
        d = parse_date(day.get("date")); day_minutes = 0; intervals = []
        if not d or not is_french_working_day(d.date()): errors.append(f"Jour non ouvré dans le planning AFC: {day.get('date')}.")
        if d and is_interrupted_day(d.date(), interruptions or []): errors.append(f"Jour d'interruption dans le planning AFC: {day.get('date')}.")
        for slot in sorted(day.get("slots", []), key=lambda s: s.get("start") or ""):
            minutes = int(round(float(slot.get("durationMinutes") or float(slot.get("duration") or 0)*60)))
            cat = slot.get("afcCategory") or slot.get("uv") or ""; total += minutes; day_minutes += minutes
            try:
                sm = int(slot["start"][:2])*60 + int(slot["start"][3:5]); em = int(slot["end"][:2])*60 + int(slot["end"][3:5])
                if em <= sm or em-sm != minutes: errors.append(f"Créneau invalide le {day.get('date')} {slot.get('start')}-{slot.get('end')}.")
                if not any(sm >= a and em <= b for a,b in AFC_DAY_SEGMENTS): errors.append(f"Créneau traversant une demi-journée le {day.get('date')} {slot.get('start')}-{slot.get('end')}.")
                if any(sm < old_end and em > old_start for old_start, old_end in intervals): errors.append(f"Chevauchement le {day.get('date')} {slot.get('start')}-{slot.get('end')}.")
                intervals.append((sm, em))
            except Exception: errors.append(f"Horaire invalide le {day.get('date')}: {slot.get('start')}-{slot.get('end')}.")
            if d:
                b = week_buckets.setdefault(d.date().isocalendar()[:2], {"total":0,"technical":0,"RAN":0,"SP":0,"PAF":0})
                b["total"] += minutes
                if cat in AFC_TECHNICAL_CODES: b["technical"] += minutes
                if cat in {"RAN", "SP", "PAF"}: b[cat] += minutes
        # Contrôle facturable : FT/RAN/SP/PAF en heures entières par demi-journée.
        halfday_totals = {}
        for slot in day.get("slots", []):
            minutes = int(round(float(slot.get("durationMinutes") or float(slot.get("duration") or 0)*60)))
            sm = int(slot["start"][:2])*60 + int(slot["start"][3:5])
            half = "AM" if sm < 13*60 else "PM"
            cat = slot.get("afcCategory") or slot.get("uv") or ""
            billing = "FT" if cat in AFC_TECHNICAL_CODES else cat
            if billing in {"FT","RAN","SP","PAF"}:
                halfday_totals[(half,billing)] = halfday_totals.get((half,billing), 0) + minutes
        for (half,billing), minutes in halfday_totals.items():
            if minutes % 60 != 0:
                errors.append(f"La demi-journée {half} du {day.get('date')} contient {format_duration_from_minutes(minutes)} en {billing}, non facturable en heures entières.")
        if day_minutes > APS_MAX_DAILY_MINUTES: errors.append(f"La journée {day.get('date')} dépasse 7h.")
    for week, b in week_buckets.items():
        support_minutes = b["SP"] + b["PAF"]
        if b["total"] > 35*60: errors.append(f"La semaine {week} dépasse 35h.")
        if b["technical"] > 30*60: errors.append(f"La semaine {week} dépasse 30h de formation technique.")
        if support_minutes not in {0, 5*60, 10*60}: errors.append(f"La semaine {week} doit contenir exactement 0h, 5h ou 10h de SP/PAF (actuel: {format_duration_from_minutes(support_minutes)}).")
        if b["SP"] not in {0, 5*60}: errors.append(f"La semaine {week} doit contenir exactement 0h ou 5h de SP (actuel: {format_duration_from_minutes(b['SP'])}).")
        if b["PAF"] not in {0, 5*60}: errors.append(f"La semaine {week} doit contenir exactement 0h ou 5h de PAF (actuel: {format_duration_from_minutes(b['PAF'])}).")
    errors.extend(validate_actual_afc_hours(planning_data))
    eligible_dates = []
    if planning_data:
        first = parse_date(planning_data[0].get("date")); last = parse_date((contractual_end_date.isoformat() if hasattr(contractual_end_date, "isoformat") else contractual_end_date) or planning_data[-1].get("date"))
        if first and last:
            cur = first.date(); end = last.date()
            while cur <= end:
                if is_afc_working_day(cur, interruptions or []): eligible_dates.append(cur.isoformat())
                cur += timedelta(days=1)
            programmed_dates = sorted({d.get("date") for d in planning_data or [] if d.get("slots")})
            if len(eligible_dates) != 57: errors.append(f"La période AFC doit contenir exactement 57 dates admissibles (actuel: {len(eligible_dates)}).")
            if len(programmed_dates) != 57: errors.append(f"Le planning AFC doit contenir exactement 57 dates programmées (actuel: {len(programmed_dates)}).")
            if set(programmed_dates) != set(eligible_dates): errors.append("Chaque date admissible AFC doit contenir de la formation, sans date supplémentaire.")
    day_values = []
    for day in planning_data or []:
        dm = sum(int(round(float(slot.get("durationMinutes") or float(slot.get("duration") or 0)*60))) for slot in day.get("slots", []))
        day_values.append(dm)
        if dm < 6*60: errors.append(f"La journée {day.get('date')} contient moins de 6h.")
        if dm > 7*60: errors.append(f"La journée {day.get('date')} dépasse 7h.")
        if day.get("date", "") >= "2027-03-01": errors.append("Aucune activité AFC ne doit être planifiée en mars 2027.")
    if day_values and day_values.count(7*60) != 51: errors.append(f"Le planning AFC doit contenir 51 journées de 7h (actuel: {day_values.count(7*60)}).")
    if day_values and day_values.count(6*60) != 6: errors.append(f"Le planning AFC doit contenir 6 journées de 6h (actuel: {day_values.count(6*60)}).")
    ordered_slots = []
    for day in sorted(planning_data or [], key=lambda d: d.get("date") or ""):
        for slot in sorted(day.get("slots", []), key=lambda s: s.get("start") or ""):
            cat = slot.get("afcCategory") or slot.get("uv") or ""
            ordered_slots.append((day.get("date"), slot.get("start"), slot.get("end"), cat, slot))
    weekly_ordered = {}
    for d, st, e, cat, slot in ordered_slots:
        parsed = parse_date(d)
        if parsed:
            weekly_ordered.setdefault(parsed.date().isocalendar()[:2], []).append((d, st, e, cat, slot))
    for week, slots in weekly_ordered.items():
        support_slots = [slot for slot in slots if slot[3] in AFC_ACCOMPANIMENT_CODES]
        if support_slots:
            if slots[-len(support_slots):] != support_slots:
                errors.append(f"La semaine {week} ne place pas le bloc SP/PAF dans les 5 dernières heures.")
    def first_time(cat):
        return next(((d, st) for d, st, _e, ccat, _s in ordered_slots if ccat == cat), None)
    def last_end(cat):
        matches = [(d, e) for d, _st, e, ccat, _s in ordered_slots if ccat == cat]
        return matches[-1] if matches else None
    strict_order = ["RAN", "ACCUEIL", "APS", "EXAM_APS", "H0B0", "SSIAP1", "EXAM_SSIAP1"]
    for before, after in zip(strict_order, strict_order[1:]):
        if first_time(after) and last_end(before) and first_time(after) < last_end(before):
            errors.append(f"Ordre AFC invalide: {AFC_APS_SSIAP_LABELS[after]} commence avant la fin de {AFC_APS_SSIAP_LABELS[before]}.")
    ran_end = last_end("RAN")
    accueil_end = last_end("ACCUEIL")
    if ran_end:
        for d, st, _e, cat, _slot in ordered_slots:
            if cat != "RAN" and (d, st) < ran_end:
                errors.append("Une activité non RAN est planifiée avant la fin complète des 55h de RAN.")
                break
    if accueil_end:
        for d, st, _e, cat, _slot in ordered_slots:
            if cat in {"APS", "EXAM_APS", "H0B0", "SSIAP1", "EXAM_SSIAP1", "SP", "PAF", "BILAN"} and (d, st) < accueil_end:
                errors.append("Une activité technique/SP/PAF est planifiée avant la fin complète de l’Accueil.")
                break
    first_sp = first_time("SP")
    first_aps = first_time("APS")
    if first_sp:
        aps_minutes_before_sp = 0
        aps_days_before_sp = set()
        accueil_day = accueil_end[0] if accueil_end else None
        for d, st, e, cat, slot in ordered_slots:
            if (d, st) >= first_sp:
                break
            if cat == "APS":
                aps_minutes_before_sp += int(round(float(slot.get("durationMinutes") or float(slot.get("duration") or 0)*60)))
                aps_days_before_sp.add(d)
        if aps_minutes_before_sp < 7 * 60:
            errors.append("Le premier SP doit être strictement postérieur à au moins 7h d’APS déjà réalisées (aps_hours_completed_before_first_sp >= 7).")
        if accueil_day and first_sp[0] == accueil_day:
            errors.append("Aucun SP ne peut être placé le jour de l’Accueil.")
        if first_aps and first_sp[0] == first_aps[0]:
            errors.append("Le premier SP ne peut pas être placé le jour du premier démarrage APS.")
        if not any(sum(int(round(float(s.get("durationMinutes") or float(s.get("duration") or 0)*60))) for s in day.get("slots", []) if (s.get("afcCategory") or s.get("uv")) == "APS") >= 7*60 for day in planning_data or [] if day.get("date") in aps_days_before_sp):
            errors.append("Le premier SP doit intervenir après une journée complète de 7h d’APS.")
    if ordered_slots and ordered_slots[-1][3] != "EXAM_SSIAP1":
        errors.append("L’examen SSIAP 1 doit être la dernière activité du parcours AFC.")
    if contractual_end_date:
        end_iso = contractual_end_date.isoformat() if hasattr(contractual_end_date, "isoformat") else str(contractual_end_date)
        if ordered_slots and ordered_slots[-1][0] != end_iso:
            errors.append(f"La dernière activité AFC doit être placée le {format_date(end_iso)}.")
        if any(d > end_iso for d, _st, _e, _cat, _slot in ordered_slots):
            errors.append(f"Aucune activité AFC ne doit être postérieure au {format_date(end_iso)}.")
        exam_slots = [x for x in ordered_slots if x[0] == end_iso]
        if len(exam_slots) != 2 or any(x[3] != "EXAM_SSIAP1" for x in exam_slots):
            errors.append("La date de fin contractuelle doit contenir uniquement les deux créneaux d’examen SSIAP 1.")
    for _d, _st, _e, _cat, slot in ordered_slots:
        visible = " ".join(str(slot.get(k) or "") for k in ("uv", "sequence", "part", "title", "content"))
        if any(bad in visible for bad in ("None", "null", "undefined")) or not (slot.get("title") or "").strip():
            errors.append("Un créneau AFC contient un libellé vide ou invalide.")
            break
    rows = [{"uv": code, "label": AFC_APS_SSIAP_LABELS[code], "title": AFC_APS_SSIAP_LABELS[code], "hours": round(totals.get(code,0)/60,2), "expected": round(AFC_APS_SSIAP_EXPECTED_MINUTES[code]/60,2), "modality":"presentiel"} for code in AFC_APS_SSIAP_SUMMARY_ORDER]
    return {"total_hours": round(total/60,2), "uv_totals": {k: round(v/60,2) for k,v in totals.items()}, "uv_rows": rows, "modality_totals": {"presentiel": round(total/60,2)}, "weekly": {str(k): {kk: round(vv/60,2) for kk,vv in b.items()} for k,b in week_buckets.items()}, "days_count": len(planning_data or []), "slots_count": sum(len(d.get("slots", [])) for d in planning_data or []), "errors": errors}

def aps_summary_from_data(planning_data):
    uv_totals = {uv: 0.0 for uv in APS_EXPECTED_UV_TOTALS}
    total = 0.0
    modality_totals = {"elearning": 0.0, "presentiel": 0.0}
    slot_count = 0
    errors = []
    previous = None
    for day in planning_data or []:
        day_date = day.get("date")
        try:
            datetime.strptime(day_date, "%Y-%m-%d")
        except Exception:
            errors.append(f"Date invalide: {day_date}")
        for slot in day.get("slots", []):
            slot_count += 1
            uv = (slot.get("uv") or "").strip().upper()
            duration_minutes = int(round(float(slot.get("durationMinutes") or (float(slot.get("duration") or 0) * 60))))
            duration = round(duration_minutes / 60, 2)
            start = slot.get("start") or ""
            end = slot.get("end") or ""
            try:
                start_dt = datetime.strptime(f"{day_date} {start}", "%Y-%m-%d %H:%M")
                end_dt = datetime.strptime(f"{day_date} {end}", "%Y-%m-%d %H:%M")
                real_duration = round((end_dt - start_dt).total_seconds() / 3600, 2)
                if real_duration != round(duration, 2):
                    errors.append(f"Durée incohérente le {day_date} {start}-{end}.")
                if previous and start_dt < previous:
                    errors.append(f"Ordre chronologique incohérent le {day_date} {start}.")
                previous = end_dt
            except Exception:
                errors.append(f"Horaire invalide le {day_date}: {start}-{end}.")
            modality = (slot.get("modality") or "presentiel").strip()
            if modality in modality_totals:
                modality_totals[modality] = round(modality_totals[modality] + duration, 2)
            if uv not in uv_totals:
                errors.append(f"UV inconnue: {uv}")
            else:
                uv_totals[uv] = round(uv_totals[uv] + duration, 2)
            total = round(total + duration, 2)
    rows = [{"uv": uv, "label": APS_UV_LABELS[uv], "hours": uv_totals.get(uv, 0), "expected": expected} for uv, expected in APS_EXPECTED_UV_TOTALS.items()]
    return {"total_hours": total, "uv_totals": uv_totals, "uv_rows": rows, "modality_totals": modality_totals, "days_count": len(planning_data or []), "slots_count": slot_count, "errors": errors}

def validate_aps_planning_data(planning_data, planning_mode="full_presentiel"):
    summary = aps_summary_from_data(planning_data)
    errors = list(summary["errors"])
    if round(summary["total_hours"], 2) != APS_TOTAL_HOURS:
        errors.append(f"Le total doit être exactement de {APS_TOTAL_HOURS}h (actuel: {summary['total_hours']}h).")
    if planning_mode == "elearning_presentiel":
        errors.extend(validate_aps_elearning_presentiel_rules(planning_data, summary))
    else:
        for uv, expected in APS_EXPECTED_UV_TOTALS.items():
            actual = round(summary["uv_totals"].get(uv, 0), 2)
            if actual != expected:
                errors.append(f"{uv} doit totaliser {expected}h (actuel: {actual}h).")
    return errors, summary

def validate_aps_elearning_presentiel_rules(planning_data, summary=None):
    summary = summary or aps_summary_from_data(planning_data)
    errors = []
    totals = summary.get("modality_totals", {})
    if int(round(totals.get("elearning", 0) * 60)) != APS_ELEARNING_MINUTES:
        errors.append(f"Le total e-learning doit être exactement de 62h (actuel: {totals.get('elearning', 0):g}h).")
    if int(round(totals.get("presentiel", 0) * 60)) != APS_PRESENTIEL_MINUTES:
        errors.append(f"Le total présentiel doit être exactement de 113h (actuel: {totals.get('presentiel', 0):g}h).")
    seen_presentiel = False
    last_elearning_day = None
    first_presentiel = None
    for day in planning_data or []:
        day_modalities = {(slot.get("modality") or "presentiel") for slot in day.get("slots", [])}
        if len(day_modalities) > 1:
            errors.append(f"La journée {day.get('date')} mélange e-learning et présentiel.")
        for slot in day.get("slots", []):
            modality = slot.get("modality") or "presentiel"
            if modality == "presentiel":
                seen_presentiel = True
                if first_presentiel is None:
                    first_presentiel = (day.get("date"), slot.get("start"))
            elif seen_presentiel:
                errors.append("Tous les blocs e-learning doivent être avant les blocs présentiels.")
            if modality == "elearning":
                last_elearning_day = day.get("date")
    if first_presentiel:
        first_date, first_start = first_presentiel
        if first_start != "08:30":
            errors.append("Le présentiel doit commencer sur un nouveau jour ouvré à 08h30.")
        try:
            expected = next_aps_training_day(datetime.strptime(last_elearning_day, "%Y-%m-%d").date()).isoformat()
            if first_date != expected:
                errors.append(f"Le présentiel doit commencer le prochain jour ouvré complet ({expected}) après la période e-learning.")
        except Exception:
            errors.append("Impossible de vérifier le prochain jour ouvré de démarrage présentiel.")
    return errors

def aps_pdf_logo_path():
    public_logo = os.path.join(BASE_DIR, "public", "logo-integrale-academy.png")
    if os.path.exists(public_logo):
        return public_logo
    static_logo = os.path.join(BASE_DIR, "static", "img", "logo-integrale.png")
    return static_logo if os.path.exists(static_logo) else None

def append_planning_history(session_data, label):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    session_data.setdefault("planning_history", []).append({"label": label, "at": now})
    if len(session_data["planning_history"]) > 20:
        session_data["planning_history"] = session_data["planning_history"][-20:]
    return now

def aps_modality_ranges(planning_data):
    ranges = {}
    for day in planning_data or []:
        day_date = day.get("date")
        if not day_date:
            continue
        for slot in day.get("slots", []):
            modality = (slot.get("modality") or "presentiel").strip()
            if modality not in {"elearning", "presentiel"}:
                continue
            ranges.setdefault(modality, {"start": day_date, "end": day_date})
            ranges[modality]["start"] = min(ranges[modality]["start"], day_date)
            ranges[modality]["end"] = max(ranges[modality]["end"], day_date)
    return ranges

def aps_format_range(range_data):
    if not range_data:
        return ""
    return f"du {format_date(range_data.get('start'))} au {format_date(range_data.get('end'))}"

def planning_pdf_profile(document_profile, planning_mode, summary):
    profile = dict(document_profile or {})
    is_desp = profile.get("validate") == "desp"
    is_ssiap1 = profile.get("validate") == "ssiap1"
    is_afc = profile.get("validate") == "afc_aps_ssiap"
    totals = summary.get("modality_totals", {}) if summary else {}
    if is_ssiap1:
        profile.setdefault("planning_title", "PLANNING DE FORMATION SSIAP 1")
        profile.setdefault("subtitle", "Parcours comprenant 14 h de SST, 67 h de formation SSIAP 1 et une préparation à l’examen")
        profile.setdefault("modality_line", "SST : 14h • Formation réglementaire SSIAP 1 : 67h • Révisions : 3h • Total présence avant examen : 84h")
        profile.setdefault("legend_elearning", "Examen SSIAP 1")
        profile.setdefault("legend_presentiel", "Formation SSIAP 1 — 67h")
        profile.setdefault("short_label", "SSIAP 1")
    elif is_afc:
        profile.setdefault("planning_title", "PLANNING AFC FRANCE TRAVAIL APS + SSIAP")
        profile.setdefault("subtitle", "Parcours complet — 393 heures")
        profile.setdefault("modality_line", "Modalité : présentiel • Parcours complet : 393h")
        profile.setdefault("legend_elearning", "Accompagnements SP / PAF")
        profile.setdefault("legend_presentiel", "Formation AFC APS + SSIAP")
        profile.setdefault("short_label", "AFC APS + SSIAP")
    elif is_desp:
        profile.setdefault("planning_title", "PLANNING DE FORMATION DESP")
        profile.setdefault("subtitle", f"{DESP_LABEL} — {DESP_TOTAL_HOURS} heures")
        profile.setdefault("modality_line", f"Modalité : E-learning + présentiel • E-learning : {DESP_ELEARNING_HOURS}h • Présentiel : {DESP_PRESENTIEL_HOURS}h • Total : {DESP_TOTAL_HOURS}h")
        profile.setdefault("legend_elearning", f"E-learning / distanciel — {DESP_ELEARNING_HOURS}h")
        profile.setdefault("legend_presentiel", f"Présentiel au centre — {DESP_PRESENTIEL_HOURS}h")
        profile.setdefault("short_label", "DESP")
    elif planning_mode == "elearning_presentiel":
        profile.setdefault("planning_title", "PLANNING DE FORMATION APS")
        profile.setdefault("subtitle", "Agent de Prévention et de Sécurité — 175 heures")
        profile.setdefault("modality_line", f"Modalité : E-learning + présentiel • E-learning : {totals.get('elearning', APS_ELEARNING_HOURS):g}h • Présentiel : {totals.get('presentiel', APS_PRESENTIEL_HOURS):g}h • Total : {summary.get('total_hours', APS_TOTAL_HOURS):g}h")
        profile.setdefault("legend_elearning", f"E-learning / distanciel — {totals.get('elearning', APS_ELEARNING_HOURS):g}h")
        profile.setdefault("legend_presentiel", f"Présentiel au centre — {totals.get('presentiel', APS_PRESENTIEL_HOURS):g}h")
    else:
        profile.setdefault("planning_title", "PLANNING DE FORMATION APS")
        profile.setdefault("subtitle", "Agent de Prévention et de Sécurité — 175 heures")
        profile.setdefault("modality_line", "Modalité : 100% présentiel • Présentiel : 175h")
    return profile


def planning_day_title(day, document_profile=None):
    if (document_profile or {}).get("validate") == "a3p":
        return _a3p_full_day_label(day.get("date"))
    parsed = parse_date(day.get("date"))
    return aps_day_label(parsed.date()) if parsed else (day.get("dayLabel") or day.get("date"))


def planning_slot_title(slot):
    if slot.get("partNumber") and slot.get("sequenceNumber"):
        split = f" — {slot.get('splitLabel')}" if slot.get("splitLabel") else ""
        return f"Séquence {slot.get('sequenceNumber')} — {slot.get('sequenceTitle') or slot.get('title')}{split}"
    return f"{slot.get('uv')} — {slot.get('title')}"


def planning_card_height(slot, printable_width):
    if slot.get("partNumber") and slot.get("sequenceNumber"):
        text_w = printable_width - 32
        title_lines = max(1, len(wrap_text_lines(planning_slot_title(slot), text_w, "Helvetica-Bold", 9.4)))
        items_h = 0
        for item in slot.get("subpartDisplayItems") or slot.get("subpartItems") or []:
            items_h += max(1, len(wrap_text_lines(f"• {item}", text_w - 14, "Helvetica", 8.7))) * 10
        progress_h = 10 if slot.get("subpartProgressLabel") else 0
        meta_lines = max(1, len(wrap_text_lines(f"Formateur : {slot.get('trainer') or '—'} • Salle : {slot.get('room') or '—'}", text_w, "Helvetica", 7.8)))
        return max(90, 16 + title_lines * 11 + 13 + 16 + progress_h + 8 + items_h + 8 + meta_lines * 9 + 8)
    title_w = printable_width - 225
    title_lines = max(1, len(wrap_text_lines(planning_slot_title(slot), title_w, "Helvetica-Bold", 8.2)))
    title_h = title_lines * 9
    meta_h = 14 if (slot.get("modality") or "presentiel") != "elearning" else 0
    return max(44, 12 + title_h + 8 + meta_h + 14)


def planning_day_height(day, current_part, planning_mode, printable_width):
    needed = 30 + 2
    part_cursor = current_part
    first_part = None
    for slot in day.get("slots", []):
        slot_part = slot.get("part")
        if planning_mode in {"elearning_presentiel", "desp", "ssiap1"} and slot_part and slot_part != part_cursor:
            needed += 34
            part_cursor = slot_part
            first_part = first_part or slot_part
        needed += planning_card_height(slot, printable_width) + 5
    return needed, first_part


def generate_aps_planning_pdf(session_data, formateur, output_path, planning_data=None, planning_mode="full_presentiel", document_profile=None):
    document_profile = document_profile or {}
    if planning_mode not in {"full_presentiel", "elearning_presentiel", "desp", "ssiap1"}:
        raise ValueError("Le type de planning est obligatoire.")
    if document_profile.get("validate") == "ssiap1":
        planning_mode = "ssiap1"
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise RuntimeError("La dépendance reportlab est requise pour générer le PDF APS.") from exc

    start_dt = parse_date(session_data.get("date_debut"))
    exam_dt = parse_date(session_data.get("date_exam"))
    is_afc_document = document_profile.get("validate") == "afc_aps_ssiap"
    if not start_dt:
        raise ValueError("La date de début de session est obligatoire.")
    if not exam_dt and not is_afc_document:
        raise ValueError("La date d'examen est obligatoire pour générer le planning APS.")

    salle = session_data.get("salle") or session_data.get("room") or "Intégrale Academy – 54 chemin du Carreou – 83480 PUGET-SUR-ARGENS"
    exam_iso = aps_local_date_iso(session_data.get("date_exam"))
    session_end = parse_date(session_data.get("date_fin"))
    latest_training_date = session_end.date() if session_end else ((exam_dt.date() - timedelta(days=1)) if exam_dt else None)
    if exam_dt and latest_training_date and latest_training_date >= exam_dt.date() and document_profile.get("validate") not in {"ssiap1", "afc_aps_ssiap"}:
        raise ValueError("Impossible de générer le planning : la date de fin de formation doit être antérieure à la date d’examen.")
    if planning_data is None:
        planning_data, _, _ = (build_ssiap1_planning_data(start_dt.date(), formateur, salle, end_date=latest_training_date, exam_iso=exam_iso, exam_payload=ssiap1_exam_payload(session_data), excluded_dates=ssiap1_excluded_dates_from_payload(session_data)) if document_profile.get("validate") == "ssiap1" else build_aps_planning_data(start_dt.date(), formateur, salle, planning_mode, end_date=latest_training_date, exam_iso=exam_iso, session_id=session_data.get("id")))
    if document_profile.get("validate") == "desp":
        summary = document_profile.get("summary") or desp_summary_from_planning(planning_data)
        errors = list(summary.get("errors") or [])
    elif document_profile.get("validate") == "a3p":
        errors, a3p_summary = validate_a3p_planning(document_profile.get("source_planning") or [], exam_iso)
        summary = document_profile.get("summary") or {"total_hours": a3p_summary.get("totalHours", 0), "uv_totals": a3p_summary.get("moduleTotals", {}), "uv_rows": document_profile.get("summary_rows", []), "modality_totals": {"presentiel": a3p_summary.get("totalHours", 0)}}
    elif document_profile.get("validate") == "ssiap1":
        summary = document_profile.get("summary") or ssiap1_summary_from_data(planning_data)
        errors = list(summary.get("errors") or [])
    elif document_profile.get("validate") == "afc_aps_ssiap":
        summary = document_profile.get("summary") or afc_aps_ssiap_summary_from_data(planning_data)
        errors = list(summary.get("errors") or [])
    else:
        errors, summary = validate_aps_planning_data(planning_data, planning_mode)
    if document_profile.get("validate") not in {"ssiap1", "afc_aps_ssiap"} and any(day.get("date") == exam_iso for day in planning_data or []):
        errors.append(f"Sécurité planning {document_profile.get('short_label', 'APS')}: la date d'examen ({format_date(exam_iso)}) est exclue des journées de formation. Aucun créneau ne peut être généré ce jour-là.")
    if errors:
        raise ValueError(" ".join(errors))

    document_profile = planning_pdf_profile(document_profile, planning_mode, summary)
    c = canvas.Canvas(output_path, pagesize=A4)
    width, height = A4
    margin = 36
    printable_width = width - 2 * margin
    logo_path = aps_pdf_logo_path()
    title = document_profile.get("planning_title") or "PLANNING DE FORMATION APS"
    edited = datetime.now().strftime("%d/%m/%Y à %H:%M")
    computed_end = (planning_data[-1].get("date") if planning_data else session_data.get("date_fin"))
    period = f"Du {format_date(session_data.get('date_debut'))} au {format_date(session_data.get('date_fin') or computed_end) if (session_data.get('date_fin') or computed_end) else '—'}"
    trainers = sorted({(slot.get("trainer") or "").strip() for day in planning_data for slot in day.get("slots", []) if (slot.get("modality") or "presentiel") == "presentiel" and (slot.get("trainer") or "").strip()})
    trainer_label = ", ".join(trainers[:4]) + ("…" if len(trainers) > 4 else "") if trainers else "—"
    modality_ranges = aps_modality_ranges(planning_data)
    if document_profile.get("validate") == "ssiap1":
        for key, modality in (("sst", "sst"), ("presentiel", "presentiel"), ("revision", "revision")):
            dates = [day.get("date") for day in planning_data or [] for slot in day.get("slots", []) if (slot.get("modality") or "") == modality and day.get("date")]
            if dates:
                modality_ranges[key] = {"start": min(dates), "end": max(dates)}
        if modality_ranges.get("presentiel"):
            modality_ranges["presentiel"]["end"] = aps_local_date_iso(session_data.get("date_fin")) or modality_ranges["presentiel"].get("end")
    modality_totals = summary.get("modality_totals", {})
    logical_day_totals = {}
    for source_day in planning_data or []:
        if document_profile.get("validate") == "ssiap1" and source_day.get("exam"):
            continue
        logical_day_totals[source_day.get("date")] = sum(int(round(float(slot.get("durationMinutes") or (float(slot.get("duration") or 0) * 60)))) for slot in source_day.get("slots", []))

    def period_title(part):
        if document_profile.get("validate") == "ssiap1":
            if is_ssiap1_exam_part(part):
                exam = summary.get("exam") or {}
                return f"EXAMEN SSIAP 1 — {format_date(exam.get('date'))} — {exam.get('start') or '—'} / {exam.get('end') or '—'}"
            return part
        modality = "elearning" if "E-LEARNING" in (part or "") else "presentiel"
        base = "PÉRIODE 1 — E-LEARNING / DISTANCIEL" if modality == "elearning" else "PÉRIODE 2 — PRÉSENTIEL AU CENTRE"
        hours = modality_totals.get(modality, 0)
        date_range = aps_format_range(modality_ranges.get(modality))
        return f"{base} — {date_range} — {hours:g}h" if date_range else f"{base} — {hours:g}h"

    first_content_y = height - (146 if planning_mode in {"elearning_presentiel", "desp", "ssiap1"} else 122)
    def build_pages():
        built_pages, current_page, current_part = [], [], None
        planning_day_height_helper = planning_day_height
        y = first_content_y
        bottom_limit = 84
        source_days = [day for day in (planning_data or []) if not (document_profile.get("validate") == "ssiap1" and day.get("exam"))]
        for day in source_days:
            day_header_needed = 32
            day_started_on_page = any(existing.get("date") == day.get("date") for existing in current_page)
            current_fragment = None
            for slot in day.get("slots", []):
                _ = planning_day_height_helper
                slot_part = slot.get("part")
                band_needed = 34 if planning_mode in {"elearning_presentiel", "desp", "ssiap1"} and slot_part and slot_part != current_part else 0
                header_needed = 0 if day_started_on_page and current_fragment is not None else day_header_needed
                needed = header_needed + band_needed + planning_card_height(slot, printable_width) + 5
                if current_page and y - needed < bottom_limit:
                    built_pages.append(current_page)
                    current_page, current_part = [], None
                    y = height - 122
                    current_fragment = None
                    day_started_on_page = False
                    band_needed = 34 if planning_mode in {"elearning_presentiel", "desp", "ssiap1"} and slot_part else 0
                    header_needed = day_header_needed
                    needed = header_needed + band_needed + planning_card_height(slot, printable_width) + 5
                if current_fragment is None:
                    current_fragment = {**day, "slots": []}
                    current_page.append(current_fragment)
                    y -= header_needed
                    day_started_on_page = True
                if band_needed:
                    y -= band_needed
                    current_part = slot_part
                current_fragment["slots"].append(slot)
                y -= planning_card_height(slot, printable_width) + 5
            if current_fragment:
                y -= 2
        if current_page or not built_pages:
            built_pages.append(current_page)
        return built_pages

    pages = build_pages()
    signature_image = find_center_image("signature", "sign")
    stamp_image = find_center_image("tampon", "cachet", "stamp")
    page_no = 1
    total_pages = 1

    def draw_header_footer(page_no, total_pages):
        if logo_path:
            c.drawImage(logo_path, margin, height - 72, width=72, height=49, preserveAspectRatio=True, mask="auto")
        c.setFillColor(colors.HexColor("#111827")); c.setFont("Helvetica-Bold", 16)
        c.drawString(margin + 88, height - 35, title)
        c.setFont("Helvetica", 9); c.setFillColor(colors.HexColor("#4b5563"))
        c.drawString(margin + 88, height - 50, document_profile.get("subtitle"))
        exam_bits = ""
        if document_profile.get("validate") == "afc_aps_ssiap":
            aps_exam = next((d.get("date") for d in planning_data for sl in d.get("slots", []) if sl.get("afcCategory") == "EXAM_APS"), None)
            ssiap_exam = next((d.get("date") for d in planning_data for sl in d.get("slots", []) if sl.get("afcCategory") == "EXAM_SSIAP1"), None)
            exam_bits = f" • Examen APS : {format_date(aps_exam)} • Examen SSIAP 1 : {format_date(ssiap_exam)}"
        else:
            exam_bits = f" • Examen prévu le {format_date(session_data.get('date_exam'))}"
        info_y = draw_wrapped_text(c, f"{period}{exam_bits} • Formateur(s) présentiel : {trainer_label}", margin + 88, height - 64, width - margin - (margin + 88), "Helvetica", 9, 11)
        modality_y = info_y - 1
        c.setFont("Helvetica-Bold", 8); c.setFillColor(colors.HexColor("#111827"))
        c.drawString(margin + 88, modality_y, document_profile.get("modality_line"))
        y_dates = modality_y - 13
        date_lines = []
        if document_profile.get("validate") == "ssiap1":
            date_lines.append(f"Période globale de formation : du {format_date(session_data.get('date_debut'))} au {format_date(session_data.get('date_fin'))}")
            if modality_ranges.get("sst"):
                date_lines.append(f"SST : {aps_format_range(modality_ranges.get('sst'))}")
            if modality_ranges.get("presentiel"):
                date_lines.append(f"SSIAP 1 : {aps_format_range(modality_ranges.get('presentiel'))}")
            date_lines.append(f"Examen SSIAP 1 : le {format_date(session_data.get('date_exam'))}")
        else:
            if document_profile.get("validate") == "afc_aps_ssiap":
                date_lines.append(period)
            if modality_ranges.get("elearning"):
                date_lines.append(f"Période e-learning : {aps_format_range(modality_ranges.get('elearning'))}")
            if modality_ranges.get("presentiel"):
                date_lines.append(f"Période présentiel : {aps_format_range(modality_ranges.get('presentiel'))}")
        draw_wrapped_text(c, " • ".join(date_lines), margin + 88, y_dates, width - margin - (margin + 88), "Helvetica-Bold", 8, 10)
        c.setStrokeColor(colors.HexColor("#e5e7eb")); c.line(margin, height - 112, width - margin, height - 112)
        c.setFont("Helvetica", 6.2); c.setFillColor(colors.HexColor("#6b7280"))
        c.drawString(margin, 40, f"Édité le {edited}, sous réserve de modification.")
        legal_lines = APS_LEGAL_LINES[:1] + APS_LEGAL_LINES[2:]
        if document_profile.get("validate") == "ssiap1":
            legal_lines.append(SSIAP1_AGREMENT_LINE)
        legal = " • ".join(legal_lines)
        draw_wrapped_text(c, legal, margin, 28, width - 2 * margin - 78, "Helvetica", 6.2, 8)
        c.drawRightString(width - margin, 40, f"Page {page_no} / {total_pages}")

    def draw_legend(y):
        c.setFont("Helvetica-Bold", 8); c.setFillColor(colors.HexColor("#111827")); c.drawString(margin, y, "Légende :")
        exam_legend_color = "#b91c1c" if document_profile.get("validate") == "ssiap1" else "#6d28d9"
        c.setFillColor(colors.HexColor(exam_legend_color)); c.roundRect(margin + 54, y - 8, 18, 9, 2, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#111827")); c.drawString(margin + 78, y, document_profile.get("legend_elearning"))
        c.setFillColor(colors.HexColor("#0d9488")); c.roundRect(margin + 238, y - 8, 18, 9, 2, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#111827")); c.drawString(margin + 262, y, document_profile.get("legend_presentiel"))
        return y - 20

    def draw_planning_pages(total_pages):
        nonlocal page_no
        for page_days in pages:
            draw_header_footer(page_no, total_pages)
            y = first_content_y if page_no == 1 else height - 122
            if page_no == 1 and planning_mode in {"elearning_presentiel", "desp", "ssiap1"}:
                y = draw_legend(y)
            current_part = None
            seen_dates_on_previous_pages = {fragment.get("date") for previous_page in pages[:page_no-1] for fragment in previous_page}
            for day in page_days:
                c.setFillColor(colors.HexColor("#f3f4f6")); c.roundRect(margin, y - 18, printable_width, 22, 6, fill=1, stroke=0)
                day_total_minutes = logical_day_totals.get(day.get("date"), sum(int(round(float(slot.get("durationMinutes") or (float(slot.get("duration") or 0) * 60)))) for slot in day.get("slots", [])))
                suite = " — suite" if day.get("date") in seen_dates_on_previous_pages else ""
                day_total_label = format_duration_from_minutes(day_total_minutes).replace("h", " h")
                c.setFillColor(colors.HexColor("#111827")); c.setFont("Helvetica-Bold", 10); c.drawString(margin + 10, y - 12, f"{planning_day_title(day, document_profile)}{suite} — {day_total_label.replace(' ', '')} — Total journée : {day_total_label}")
                y -= 30
                for slot in day.get("slots", []):
                    slot_part = slot.get("part")
                    if planning_mode in {"elearning_presentiel", "desp", "ssiap1"} and slot_part and slot_part != current_part:
                        current_part = slot_part
                        band_color = "#b91c1c" if is_ssiap1_exam_part(slot_part) else ("#6d28d9" if "E-LEARNING" in slot_part else "#0d9488")
                        c.setFillColor(colors.HexColor(band_color)); c.roundRect(margin, y - 20, printable_width, 24, 6, fill=1, stroke=0)
                        c.setFillColor(colors.white); c.setFont("Helvetica-Bold", 10); c.drawString(margin + 10, y - 12, period_title(slot_part))
                        y -= 34
                    h = planning_card_height(slot, printable_width)
                    c.setFillColor(colors.white); c.setStrokeColor(colors.HexColor("#d1d5db")); c.roundRect(margin, y - h + 5, printable_width, h, 6, fill=1, stroke=1)
                    if document_profile.get("validate") == "afc_aps_ssiap":
                        modality_color = AFC_CATEGORY_COLORS.get(slot.get("afcCategory"), "#0d9488")
                    else:
                        modality_color = "#b91c1c" if slot.get("modality") == "exam" else ("#6d28d9" if slot.get("modality") == "elearning" else "#0d9488")
                    c.setFillColor(colors.HexColor(modality_color)); c.roundRect(margin, y - h + 5, 7, h, 2, fill=1, stroke=0)
                    modality_label = "Examen" if slot.get("modality") == "exam" else ("E-learning" if slot.get("modality") == "elearning" else "Présentiel")

                    def draw_modality_badge(badge_x, badge_y):
                        c.setFillColor(colors.HexColor(modality_color))
                        c.roundRect(badge_x, badge_y, 76, 14, 4, fill=1, stroke=0)
                        c.setFillColor(colors.white)
                        c.setFont("Helvetica-Bold", 7)
                        c.drawCentredString(badge_x + 38, badge_y + 4, modality_label)

                    if slot.get("partNumber") and slot.get("sequenceNumber"):
                        text_x = margin + 16; text_w = printable_width - 32
                        cursor_y = draw_wrapped_text(c, planning_slot_title(slot), text_x, y - 8, text_w, "Helvetica-Bold", 9.4, 11)
                        c.setFont("Helvetica", 8); c.setFillColor(colors.HexColor("#374151"))
                        c.drawString(text_x, cursor_y, f"Durée totale de la séquence : {format_duration_from_minutes(int(slot.get('totalSequenceDurationMinutes') or 0))}")
                        badge_y = cursor_y - 16
                        c.setFillColor(colors.HexColor(modality_color)); c.roundRect(text_x, badge_y - 3, 94, 14, 4, fill=1, stroke=0)
                        c.setFillColor(colors.white); c.setFont("Helvetica-Bold", 7.4); c.drawCentredString(text_x + 47, badge_y + 1, f"{slot.get('subpartLabel') or 'Séquence'} — {format_duration_from_minutes(int(slot.get('durationMinutes') or 0))}")
                        c.setFillColor(colors.HexColor("#111827")); c.setFont("Helvetica-Bold", 8.2)
                        c.drawRightString(margin + printable_width - 10, badge_y + 1, f"{slot.get('start')} - {slot.get('end')} ({format_duration_from_minutes(int(slot.get('durationMinutes') or 0))})")
                        item_y = badge_y - 13
                        if slot.get("subpartProgressLabel"):
                            item_y = draw_wrapped_text(c, slot.get("subpartProgressLabel"), text_x + 8, item_y, text_w - 12, "Helvetica-Oblique", 7.6, 9)
                        for item in slot.get("subpartDisplayItems") or slot.get("subpartItems") or []:
                            item_y = draw_wrapped_text(c, f"• {item}", text_x + 8, item_y, text_w - 12, "Helvetica", 8.7, 10)
                        draw_wrapped_text(c, f"Formateur : {slot.get('trainer') or '—'} • Salle : {slot.get('room') or salle}", text_x, y - h + 19, text_w - 112, "Helvetica", 7.8, 9)
                        draw_modality_badge(width - margin - 86, y - h + 14)
                    else:
                        title_w = printable_width - 225
                        draw_wrapped_text(c, planning_slot_title(slot), margin + 16, y - 8, title_w, "Helvetica-Bold", 8.2, 9)
                        c.setFont("Helvetica", 8); c.setFillColor(colors.HexColor("#374151"))
                        c.drawString(width - margin - 168, y - 8, f"{slot.get('start')} - {slot.get('end')} ({format_duration_from_minutes(int(slot.get('durationMinutes') or 0))})")
                        draw_modality_badge(width - margin - 86, y - h + 14)
                        if slot.get("modality") != "elearning":
                            c.setFillColor(colors.HexColor("#374151")); c.setFont("Helvetica", 8)
                            draw_wrapped_text(c, f"Salle : {slot.get('room') or salle} • Formateur : {slot.get('trainer') or '—'}", margin + 14, y - h + 19, printable_width - 112, "Helvetica", 8, 9)
                    y -= h + 5
                y -= 2
            page_no += 1
            c.showPage()

    def summary_table_header(y, continued=False):
        c.setFont("Helvetica-Bold", 10); c.setFillColor(colors.HexColor("#111827")); c.drawString(margin, y, "B. Récapitulatif détaillé" + (" — suite" if continued else ""))
        y -= 14
        c.setFont("Helvetica-Bold", 8); c.drawString(margin, y, "Partie"); c.drawString(margin+92, y, "Module"); c.drawString(width-margin-140, y, "Modalité"); c.drawRightString(width-margin, y, "Heures")
        c.setStrokeColor(colors.HexColor("#e5e7eb")); c.line(margin, y-4, width-margin, y-4)
        return y - 12

    def draw_summary_pages(total_pages):
        nonlocal page_no
        draw_header_footer(page_no, total_pages)
        y = height - 122
        if document_profile.get("validate") == "afc_aps_ssiap":
            c.setFont("Helvetica-Bold", 18); c.setFillColor(colors.HexColor("#0f766e")); c.drawCentredString(width/2, y, "SYNTHÈSE DES HEURES")
            y -= 22; c.setFont("Helvetica-Bold", 12); c.setFillColor(colors.HexColor("#111827")); c.drawCentredString(width/2, y, "AFC FRANCE TRAVAIL APS + SSIAP")
            y -= 34
            col_x=[margin, margin+44, margin+300, width-margin-72]; headers=["Couleur","Intitulé complet","Catégorie de facturation","Durée"]
            c.setFillColor(colors.HexColor("#0f766e")); c.roundRect(margin, y-18, printable_width, 24, 6, fill=1, stroke=0)
            c.setFillColor(colors.white); c.setFont("Helvetica-Bold", 8.5)
            for x,h in zip(col_x,headers): c.drawString(x+4, y-10, h)
            y -= 28
            billing={"RAN":"RAN","SP":"SP","PAF":"PAF","ACCUEIL":"Formation technique (FT)","APS":"Formation technique (FT)","EXAM_APS":"Formation technique (FT)","H0B0":"Formation technique (FT)","SSIAP1":"Formation technique (FT)","EXAM_SSIAP1":"Formation technique (FT)","BILAN":"Formation technique (FT)"}
            for code in AFC_APS_SSIAP_SUMMARY_ORDER:
                c.setFillColor(colors.HexColor("#f8fafc" if AFC_APS_SSIAP_SUMMARY_ORDER.index(code)%2==0 else "#ffffff")); c.rect(margin, y-16, printable_width, 20, fill=1, stroke=0)
                c.setFillColor(colors.HexColor(AFC_CATEGORY_COLORS[code])); c.circle(col_x[0]+14, y-6, 4, fill=1, stroke=0)
                c.setFillColor(colors.HexColor("#111827")); c.setFont("Helvetica", 8.2); c.drawString(col_x[1]+4, y-9, AFC_APS_SSIAP_LABELS[code])
                c.drawString(col_x[2]+4, y-9, billing[code]); c.drawRightString(width-margin-8, y-9, format_duration_from_minutes(AFC_APS_SSIAP_EXPECTED_MINUTES[code]).replace("h", " h"))
                y -= 20
            y -= 10
            recap=[("Formation technique (FT)","273 h"),("Remise à niveau (RAN)","55 h"),("Soutien personnalisé (SP)","45 h"),("Préparation à l’après-formation (PAF)","20 h")]
            c.setFillColor(colors.HexColor("#ecfeff")); c.roundRect(margin, y-86, printable_width, 92, 8, fill=1, stroke=0)
            c.setFillColor(colors.HexColor("#0f172a")); c.setFont("Helvetica-Bold", 10); c.drawString(margin+14, y-12, "Récapitulatif de facturation")
            yy=y-30; c.setFont("Helvetica", 9)
            for label,val in recap:
                c.drawString(margin+18, yy, label); c.drawRightString(width-margin-18, yy, val); yy-=14
            c.setFillColor(colors.HexColor("#0f766e")); c.roundRect(margin+12, yy-18, printable_width-24, 22, 5, fill=1, stroke=0)
            c.setFillColor(colors.white); c.setFont("Helvetica-Bold", 10); c.drawString(margin+22, yy-11, "TOTAL"); c.drawRightString(width-margin-22, yy-11, "393 h")
            y = yy - 46
            box_w=(printable_width-18)/2
            for idx_label,(label,img) in enumerate((("Signature",signature_image),("Tampon",stamp_image))):
                x=margin+idx_label*(box_w+18); c.setFillColor(colors.white); c.setStrokeColor(colors.HexColor("#d1d5db")); c.roundRect(x,y-64,box_w,64,6,fill=1,stroke=1); c.setFillColor(colors.HexColor("#374151")); c.setFont("Helvetica-Bold",9); c.drawString(x+10,y-16,label)
                if img: c.drawImage(img,x+12,y-58,width=box_w-24,height=40,preserveAspectRatio=True,mask="auto")
            page_no += 1
            return
        if document_profile.get("validate") == "ssiap1":
            exam_day = next((day for day in planning_data or [] if day.get("exam")), None)
            exam_slot = (exam_day.get("slots") or [{}])[0] if exam_day else (summary.get("exam") or {})
            c.setFillColor(colors.HexColor("#b91c1c")); c.roundRect(margin, y - 20, printable_width, 24, 6, fill=1, stroke=0)
            c.setFillColor(colors.white); c.setFont("Helvetica-Bold", 10); c.drawString(margin + 10, y - 12, "EXAMEN SSIAP 1")
            y -= 34
            c.setFillColor(colors.white); c.setStrokeColor(colors.HexColor("#fecaca")); c.roundRect(margin, y - 74, printable_width, 78, 6, fill=1, stroke=1)
            c.setFillColor(colors.HexColor("#111827")); c.setFont("Helvetica-Bold", 10)
            c.drawString(margin + 14, y - 16, f"{planning_day_title(exam_day or {'date': (summary.get('exam') or {}).get('date')}, document_profile)} — {exam_slot.get('start') or '—'} - {exam_slot.get('end') or '—'}")
            c.setFont("Helvetica", 8.6); c.setFillColor(colors.HexColor("#374151"))
            c.drawString(margin + 14, y - 32, f"Salle : {exam_slot.get('room') or salle}")
            c.drawString(margin + 14, y - 46, f"Formateur / responsable : {exam_slot.get('trainer') or formateur or '—'}")
            c.drawString(margin + 14, y - 60, exam_slot.get("content") or "Épreuves d'examen SSIAP 1")
            y -= 96
        c.setFont("Helvetica-Bold", 13); c.setFillColor(colors.HexColor("#111827")); c.drawString(margin, y, "Synthèse des heures")
        y -= 18
        c.setFont("Helvetica-Bold", 10); c.drawString(margin, y, "A. AFC France Travail APS + SSIAP" if document_profile.get("validate") == "afc_aps_ssiap" else "A. SST")
        y -= 14
        c.setFont("Helvetica", 9)
        if document_profile.get("validate") == "afc_aps_ssiap":
            for code in AFC_APS_SSIAP_SUMMARY_ORDER:
                c.setFillColor(colors.HexColor(AFC_CATEGORY_COLORS[code])); c.rect(margin, y - 7, 8, 8, fill=1, stroke=0)
                c.setFillColor(colors.HexColor("#111827")); c.drawString(margin + 12, y, f"{AFC_APS_SSIAP_LABELS[code]} : {format_duration_from_minutes(AFC_APS_SSIAP_EXPECTED_MINUTES[code]).replace('h', ' h')}"); y -= 13
            c.setFont("Helvetica-Bold", 9); c.drawString(margin, y, "TOTAL : 393 h"); y -= 20
        elif document_profile.get("validate") == "ssiap1":
            c.drawString(margin, y, f"12/10/2026 : 7 h • 13/10/2026 : 7 h • Total SST : {summary.get('sst_hours', 0):g} h"); y -= 14
            c.setFont("Helvetica-Bold", 10); c.drawString(margin, y, "B. Formation réglementaire SSIAP 1"); y -= 14; c.setFont("Helvetica", 9)
            c.drawString(margin, y, f"du {format_date(modality_ranges.get('presentiel', {}).get('start'))} au {format_date(modality_ranges.get('presentiel', {}).get('end'))} — total : {summary['total_hours']:g} h"); y -= 14
            c.setFont("Helvetica-Bold", 10); c.drawString(margin, y, "C. Révisions et préparation à l’examen"); y -= 14; c.setFont("Helvetica", 9)
            c.drawString(margin, y, f"27/10/2026 de 13h30 à 16h30 — total : {summary.get('revision_hours', 0):g} h — hors total réglementaire des 67 h"); y -= 14
            c.setFont("Helvetica-Bold", 10); c.drawString(margin, y, "D. Examen SSIAP 1"); y -= 14; c.setFont("Helvetica", 9)
            c.drawString(margin, y, f"28/10/2026 — {exam_slot.get('start') or '08:30'} - {exam_slot.get('end') or '16:30'} — journée distincte"); y -= 16
            c.setFont("Helvetica-Bold", 9)
            y = draw_wrapped_text(c, f"Total SST : {summary.get('sst_hours', 0):g} h • Total SSIAP 1 réglementaire : {summary['total_hours']:g} h • Total révisions complémentaires : {summary.get('revision_hours', 0):g} h • Total de présence avant examen : {summary.get('presence_total_hours', 0):g} h • Examen : distinct", margin, y, printable_width, "Helvetica-Bold", 9, 11) - 11
        else:
            if document_profile.get("validate") == "afc_aps_ssiap":
                date_lines.append(period)
            if modality_ranges.get("elearning"):
                c.drawString(margin, y, f"E-learning / distanciel : {modality_totals.get('elearning', 0):g}h — {aps_format_range(modality_ranges.get('elearning'))}"); y -= 13
            if modality_ranges.get("presentiel"):
                c.drawString(margin, y, f"Présentiel : {modality_totals.get('presentiel', 0):g}h — {aps_format_range(modality_ranges.get('presentiel'))}"); y -= 13
            c.drawString(margin, y, f"Total : {summary['total_hours']:g}h")
            y -= 20
        y = summary_table_header(y)
        rows = summary.get("uv_rows") or []
        continued = False
        for row in rows:
            module_text = ((row.get("title") or row.get("label")) if document_profile.get("validate") == "afc_aps_ssiap" else (f"Séquence {str(row.get('uv')).split('-S')[-1]} — {row.get('title') or row.get('label')}" if document_profile.get("validate") == "ssiap1" and str(row.get('uv')).startswith('P') else f"{row.get('uv')} — {row.get('title') or row.get('label')}"))
            module_lines = wrap_text_lines(module_text, width - margin - 300, "Helvetica", 7.2)
            row_h = max(12, len(module_lines) * 8 + 4)
            if y - row_h < 132:
                c.showPage(); page_no += 1; draw_header_footer(page_no, total_pages); y = height - 122; continued = True; y = summary_table_header(y, continued=True)
            c.setFont("Helvetica", 6.4); c.setFillColor(colors.HexColor("#111827"))
            modality = row.get("modality") or ("elearning" if "Distanciel" in str(row.get("label")) else "presentiel")
            part_label = ("Parcours" if document_profile.get("validate") == "afc_aps_ssiap" else (f"Partie {str(row.get('uv')).split('-')[0][1:]}" if document_profile.get("validate") == "ssiap1" and str(row.get('uv')).startswith('P') else ("Période 1" if modality == "elearning" else "Période 2")))
            c.drawString(margin, y, part_label)
            draw_wrapped_text(c, module_text, margin+92, y, width - margin - 300, "Helvetica", 7.2, 8)
            c.drawString(width-margin-140, y, "E-learning" if modality == "elearning" else "Présentiel")
            c.drawRightString(width-margin, y, format_duration_from_minutes(int(round(float(row.get("hours", 0))*60))))
            y -= row_h
        final_block_h = 34 + 24 + 92 + 25 + 70
        if y - final_block_h < 84:
            c.showPage(); page_no += 1; draw_header_footer(page_no, total_pages); y = height - 122
        total_label = f"Total formation SSIAP 1 : {summary['total_hours']:g}h" if document_profile.get("validate") == "ssiap1" else f"TOTAL : {summary['total_hours']:g}h"
        c.setFont("Helvetica-Bold", 10); c.drawString(margin, y - 4, total_label)
        y -= 34
        exam_summary = summary.get("exam") or {}
        if document_profile.get("validate") == "afc_aps_ssiap":
            exam_line = "Examens APS et SSIAP 1 : voir dates calculées en en-tête."
        else:
            exam_line = f"Examen SSIAP 1 : {format_date(exam_summary.get('date'))} de {exam_summary.get('start') or '—'} à {exam_summary.get('end') or '—'}" if document_profile.get("validate") == "ssiap1" else f"Examen le {format_date(session_data.get('date_exam'))}."
        c.setFont("Helvetica-Bold", 10); c.drawString(margin, y, exam_line)
        y -= 24
        box_w = (width - 2 * margin - 18) / 2; signature_box_h = 92; signature_label_h = 18
        for idx, (label, image_path) in enumerate((("Signature", signature_image), ("Tampon", stamp_image))):
            x = margin + idx * (box_w + 18)
            c.setFillColor(colors.white); c.setStrokeColor(colors.HexColor("#d1d5db")); c.roundRect(x, y - signature_box_h, box_w, signature_box_h, 6, fill=1, stroke=1)
            c.setFillColor(colors.HexColor("#374151")); c.setFont("Helvetica-Bold", 9); c.drawString(x + 10, y - 16, label)
            if image_path:
                c.drawImage(image_path, x + 12, y - signature_box_h + 8, width=box_w - 24, height=signature_box_h - signature_label_h - 10, preserveAspectRatio=True, mask="auto")
        y -= signature_box_h + 25
        c.setFont("Helvetica-Bold", 9); c.drawString(margin, y, "Informations légales")
        y -= 14
        for line in APS_LEGAL_LINES + ([SSIAP1_AGREMENT_LINE] if document_profile.get("validate") == "ssiap1" else []):
            y = draw_wrapped_text(c, line, margin, y, printable_width, "Helvetica", 7.5, 10)
        page_no += 1

    def compute_summary_page_count():
        count = 1
        y = height - 122 - (130 if document_profile.get("validate") == "ssiap1" else 0) - 18 - 14
        if modality_ranges.get("elearning"):
            y -= 13
        if modality_ranges.get("presentiel"):
            y -= 13
        y -= 20
        y -= 26
        for row in summary.get("uv_rows") or []:
            module_text = ((row.get("title") or row.get("label")) if document_profile.get("validate") == "afc_aps_ssiap" else (f"Séquence {str(row.get('uv')).split('-S')[-1]} — {row.get('title') or row.get('label')}" if document_profile.get("validate") == "ssiap1" and str(row.get('uv')).startswith('P') else f"{row.get('uv')} — {row.get('title') or row.get('label')}"))
            row_h = max(12, len(wrap_text_lines(module_text, width - margin - 300, "Helvetica", 7.2)) * 8 + 4)
            if y - row_h < 132:
                count += 1
                y = height - 122 - 26
            y -= row_h
        final_block_h = 34 + 24 + 92 + 25 + 70
        if y - final_block_h < 84:
            count += 1
        return count

    calendar_month_count = 0
    if document_profile.get("validate") == "afc_aps_ssiap" and planning_data:
        _sd=parse_date(planning_data[0]["date"]).date(); _ed=parse_date(planning_data[-1]["date"]).date(); _cur=date(_sd.year,_sd.month,1)
        while _cur <= _ed:
            calendar_month_count += 1; _cur = date(_cur.year + (1 if _cur.month==12 else 0), 1 if _cur.month==12 else _cur.month+1, 1)
    extra_calendar_pages = 1 if document_profile.get("validate") == "afc_aps_ssiap" and calendar_month_count else 0
    total_pages = len(pages) + compute_summary_page_count() + extra_calendar_pages
    draw_planning_pages(total_pages)
    draw_summary_pages(total_pages)
    if document_profile.get("validate") == "afc_aps_ssiap":
        c.showPage(); page_no += 1
        from reportlab.lib.pagesizes import landscape, A3
        import calendar
        c.setPageSize(landscape(A3)); lw, lh = landscape(A3)
        logo_path = find_center_image("logo")
        if logo_path:
            c.drawImage(logo_path, 36, lh-58, width=72, height=34, preserveAspectRatio=True, mask="auto")
        c.setFillColor(colors.HexColor("#111827")); c.setFont("Helvetica-Bold", 15)
        c.drawString(118, lh-35, "CALENDRIER RÉCAPITULATIF")
        c.setFont("Helvetica-Bold", 12); c.drawString(118, lh-52, "AFC France Travail APS + SSIAP")
        c.setFont("Helvetica", 9); c.drawString(118, lh-66, f"Du {format_date(planning_data[0]["date"])} au {format_date(planning_data[-1]["date"])}.")
        colors_by_cat = AFC_CATEGORY_COLORS
        by_date={d.get("date"):d for d in planning_data}; start_d=parse_date(planning_data[0]["date"]).date(); end_d=parse_date(planning_data[-1]["date"]).date(); months=[]; cur=date(start_d.year,start_d.month,1)
        while cur <= end_d:
            months.append(cur); cur = date(cur.year + (1 if cur.month==12 else 0), 1 if cur.month==12 else cur.month+1, 1)
        fr_months=["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
        short_labels={"RAN":"RAN","ACCUEIL":"Accueil","APS":"APS","EXAM_APS":"Examen APS","H0B0":"H0B0","SSIAP1":"SSIAP 1","EXAM_SSIAP1":"Examen SSIAP 1","SP":"SP","PAF":"PAF","BILAN":"Bilan"}
        months = months[:4]
        per_page=4
        calendar_page_no = 0
        for page_start in range(0, len(months), per_page):
            if page_start:
                c.showPage(); page_no += 1; c.setPageSize(landscape(A3))
            calendar_page_no += 1
            subset=months[page_start:page_start+per_page]
            cols=2; rows=2
            cell_w_month=(lw-72)/cols; cell_h_month=(lh-176)/rows
            for i,m in enumerate(subset):
                x=36+(i%cols)*cell_w_month; y=lh-78-(i//cols)*cell_h_month
                c.setFillColor(colors.HexColor("#111827")); c.setFont("Helvetica-Bold", 10); c.drawString(x,y, f"{fr_months[m.month]} {m.year}")
                cell_w=(cell_w_month-8)/7; cell_h=(cell_h_month-34)/6; y0=y-16
                c.setFont("Helvetica-Bold",6.3)
                for col,letter in enumerate(["L","M","M","J","V","S","D"]): c.drawCentredString(x+col*cell_w+cell_w/2, y0, letter)
                c.setFont("Helvetica",5.0)
                for r,week in enumerate(calendar.Calendar(firstweekday=0).monthdayscalendar(m.year,m.month)):
                    for col,dd in enumerate(week):
                        cx=x+col*cell_w; cy=y0-5-r*cell_h
                        if not dd: continue
                        curd=date(m.year,m.month,dd); iso=curd.isoformat(); bg="#f3f4f6" if curd.weekday()>=5 else ("#fee2e2" if curd in french_holidays(curd.year) else "#ffffff")
                        interrupted=is_interrupted_day(curd, parse_interruption_ranges(session_data.get("interruptions")))
                        if interrupted: bg="#d1d5db"
                        c.setFillColor(colors.HexColor(bg)); c.rect(cx,cy-cell_h+2,cell_w-1,cell_h-1,fill=1,stroke=1)
                        c.setFillColor(colors.HexColor("#111827")); c.drawString(cx+1,cy-5,str(dd))
                        if iso in by_date:
                            am=[]; pm=[]
                            for sl in by_date[iso].get("slots",[]):
                                sm=int(sl["start"][:2])*60+int(sl["start"][3:5]); (am if sm < 12*60+30 else pm).append(sl.get("afcCategory"))
                            for half,cats in enumerate((am,pm)):
                                if not cats: continue
                                cat=next((c for c in cats if c), cats[0]); yy=cy-9-half*(cell_h/2-2)
                                c.setFillColor(colors.HexColor(colors_by_cat.get(cat,"#e5e7eb"))); c.rect(cx+9, yy-(cell_h/2-6), cell_w-12, cell_h/2-7, fill=1, stroke=0)
                                c.setFillColor(colors.white if cat in {"APS","EXAM_APS","EXAM_SSIAP1","BILAN"} else colors.black)
                                for li,line in enumerate(wrap_text_lines(short_labels.get(cat,cat), cell_w-13, "Helvetica-Bold", 4.6)[:2]):
                                    c.setFont("Helvetica-Bold",4.6); c.drawString(cx+10, yy-4-li*5, line)
                        elif interrupted:
                            c.setFillColor(colors.HexColor("#374151")); c.setFont("Helvetica",4.5); c.drawString(cx+2, cy-14, "INTERRUPTION")
            legend_y=54; legend_cols=5; legend_cell=(lw-72)/legend_cols; c.setFont("Helvetica", 6.1)
            for li, code in enumerate(AFC_APS_SSIAP_SUMMARY_ORDER):
                label=AFC_APS_SSIAP_LABELS[code]; lx=36+(li%legend_cols)*legend_cell; ly=legend_y-(li//legend_cols)*16
                c.setFillColor(colors.HexColor("#f8fafc")); c.roundRect(lx, ly-7, legend_cell-5, 13, 3, fill=1, stroke=0)
                c.setFillColor(colors.HexColor(colors_by_cat[code])); c.circle(lx+6, ly-1, 3, fill=1, stroke=0)
                c.setFillColor(colors.HexColor("#111827")); draw_wrapped_text(c, label, lx+14, ly+1, legend_cell-22, "Helvetica", 6.1, 6.5)
            c.setStrokeColor(colors.HexColor("#e5e7eb")); c.line(36, 28, lw-36, 28)
            c.setFillColor(colors.HexColor("#6b7280")); c.setFont("Helvetica", 6.4)
            c.drawString(36, 18, f"Édité le {edited}")
            c.drawCentredString(lw/2, 18, "Intégrale Academy — calendrier récapitulatif AFC")
            c.drawRightString(lw-36, 18, f"Page {page_no - 1} / {total_pages}")
    c.save()
    return {"planning_data": planning_data, "totals": summary["uv_totals"], "total_hours": summary["total_hours"], "summary": summary}

def _money(value):
    try:
        return f"{float(value):,.2f} €".replace(",", " ").replace(".", ",")
    except Exception:
        return "0,00 €"


def _normalized_slot_value(slot, *keys):
    for key in keys:
        value = (slot or {}).get(key)
        if value is not None and str(value).strip():
            return str(value).strip().lower().replace("é", "e")
    return ""


def is_in_person_slot(slot):
    """Return True only for normalized in-person training slots."""
    return _normalized_slot_value(slot, "modality", "delivery_mode", "period_type") in {"presentiel", "in_person"}


def aps_is_contract_billable_slot(slot):
    normalized = _normalized_slot_value(slot, "modality", "delivery_mode", "period_type")
    if normalized in {"elearning", "e-learning", "distanciel", "distance", "asynchronous", "examen", "exam"}:
        return False
    return True


def aps_detect_trainers(planning_data):
    return sorted({
        (slot.get("trainer") or "").strip()
        for day in planning_data or []
        for slot in day.get("slots", [])
        if aps_is_contract_billable_slot(slot) and (slot.get("trainer") or "").strip()
    })


def aps_trainer_interventions(planning_data, trainer_name):
    interventions = []
    total_hours = 0.0
    dates = set()
    for day in planning_data or []:
        day_date = day.get("date") or ""
        for slot in day.get("slots", []):
            if not aps_is_contract_billable_slot(slot):
                continue
            if (slot.get("trainer") or "").strip() != trainer_name:
                continue
            duration = round(float(slot.get("duration") or 0), 2)
            total_hours = round(total_hours + duration, 2)
            dates.add(day_date)
            interventions.append({
                "date": day_date,
                "dateLabel": format_date(day_date),
                "hours": duration,
                "start": slot.get("start") or "",
                "end": slot.get("end") or "",
                "module": f"{slot.get('uv') or ''} {slot.get('title') or ''}".strip(),
                "modality": "E-learning" if slot.get("modality") == "elearning" else "Présentiel",
                "room": slot.get("room") or "—",
            })
    calculated_days = round(total_hours / 7, 2) if total_hours else 0
    return {"interventions": interventions, "totalHours": total_hours, "calendarDays": len(dates), "calculatedDays": calculated_days}


def generate_aps_trainer_contract_pdf(session_data, contract, output_path):
    """Génère un contrat formateur professionnel en PDF avec ReportLab.

    Le générateur reste autonome (pas de Render/LibreOffice) et s'appuie sur les
    données déjà collectées dans le planning APS et la modale de génération.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            BaseDocTemplate,
            Frame,
            ListFlowable,
            ListItem,
            PageBreak,
            PageTemplate,
            Paragraph,
            Spacer,
            Table,
            TableStyle,
            KeepTogether,
            CondPageBreak,
            NextPageTemplate,
            Image,
        )
    except ImportError as exc:
        raise RuntimeError("La dépendance reportlab est requise pour générer le contrat PDF.") from exc

    width, height = A4
    landscape_width, landscape_height = landscape(A4)
    logo_path = aps_pdf_logo_path()
    generated_dt = datetime.now()
    generated = generated_dt.strftime("%d/%m/%Y")
    generated_full = generated_dt.strftime("%d/%m/%Y à %H:%M")
    formation_name = session_data.get("formation") or "APS"
    session_name = session_data.get("display_name") or session_data.get("name") or formation_name
    start_date = format_date(session_data.get("date_debut"))
    end_date = format_date(session_data.get("date_fin"))
    exam_date = format_date(session_data.get("date_exam"))
    is_mixed = session_data.get("apsPlanningMode") == "elearning_presentiel"
    modality_label = "Mixte : e-learning et présentiel" if is_mixed else "Présentiel"
    interventions = contract.get("interventions") or []
    modules = sorted({(row.get("module") or "Module non renseigné").strip() for row in interventions if (row.get("module") or "").strip()})
    room_label = "Intégrale Academy – 54 chemin du Carreou – 83480 PUGET-SUR-ARGENS"
    has_elearning = is_mixed or any((row.get("modality") or "").lower().find("learning") >= 0 for row in interventions)
    total_ht = float(contract.get("totalHT") or 0)
    tva_label = f"TVA {float(contract.get('vatRate') or 20):g}%" if contract.get("vatEnabled") else (contract.get("vatMention") or "TVA non applicable / franchise de TVA si applicable")

    doc = BaseDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=19 * mm,
        rightMargin=19 * mm,
        topMargin=24 * mm,
        bottomMargin=17 * mm,
        title="Contrat d’intervention formateur",
        author="Intégrale Academy",
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
    cover_frame = Frame(17 * mm, 18 * mm, width - 34 * mm, height - 32 * mm, id="cover")
    landscape_frame = Frame(doc.leftMargin, doc.bottomMargin, landscape_width - doc.leftMargin - doc.rightMargin, landscape_height - doc.topMargin - doc.bottomMargin, id="landscape")

    def page_canvas(canvas, document):
        page_width, page_height = document.pagesize
        canvas.saveState()
        if logo_path:
            canvas.drawImage(logo_path, doc.leftMargin, page_height - 18 * mm, width=24 * mm, height=12 * mm, preserveAspectRatio=True, mask="auto")
        canvas.setFillColor(colors.HexColor("#111827")); canvas.setFont("Helvetica-Bold", 9.5)
        canvas.drawString(doc.leftMargin + 29 * mm, page_height - 10 * mm, "CONTRAT D’INTERVENTION FORMATEUR")
        canvas.setFillColor(colors.HexColor("#6b7280")); canvas.setFont("Helvetica", 7.2)
        canvas.drawString(doc.leftMargin + 29 * mm, page_height - 15 * mm, f"{formation_name} — {session_name} — généré le {generated_full}")
        canvas.setStrokeColor(colors.HexColor("#d1d5db")); canvas.line(doc.leftMargin, page_height - 20 * mm, page_width - doc.rightMargin, page_height - 20 * mm)
        canvas.setFillColor(colors.HexColor("#6b7280")); canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(page_width / 2, 8 * mm, f"Page {document.page}")
        canvas.restoreState()

    doc.addPageTemplates([
        PageTemplate(id="cover", frames=[cover_frame], pagesize=A4),
        PageTemplate(id="contrat", frames=[frame], pagesize=A4, onPage=page_canvas),
        PageTemplate(id="planning_landscape", frames=[landscape_frame], pagesize=landscape(A4), onPage=page_canvas),
    ])
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("CoverTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=17, leading=20, textColor=colors.HexColor("#111827"), alignment=TA_CENTER, spaceAfter=4))
    styles.add(ParagraphStyle("CoverSubtitle", parent=styles["Normal"], fontSize=9.4, leading=11.4, textColor=colors.HexColor("#6b5f4a"), alignment=TA_CENTER, spaceAfter=6))
    styles.add(ParagraphStyle("CardTitle", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=7.6, leading=9.2, textColor=colors.HexColor("#8a5a20"), uppercase=True, spaceAfter=3))
    styles.add(ParagraphStyle("CardText", parent=styles["Normal"], fontSize=7.7, leading=9.1, textColor=colors.HexColor("#1f2937"), wordWrap="CJK"))
    styles.add(ParagraphStyle("Subtle", parent=styles["Normal"], fontSize=8.2, leading=10, textColor=colors.HexColor("#64748b"), alignment=TA_CENTER))
    styles.add(ParagraphStyle("Body", parent=styles["Normal"], fontSize=9.1, leading=11.5, textColor=colors.HexColor("#1f2937"), spaceAfter=4.2))
    styles.add(ParagraphStyle("H", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=11.4, leading=14, textColor=colors.HexColor("#2f2418"), spaceBefore=7, spaceAfter=4, borderPadding=(0, 0, 4, 0), borderColor=colors.HexColor("#d6b26d"), borderWidth=0, borderBottomWidth=0.5))
    styles.add(ParagraphStyle("Small", parent=styles["Normal"], fontSize=8.6, leading=10.5, textColor=colors.HexColor("#334155")))
    styles.add(ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7.2, leading=8.6, textColor=colors.HexColor("#111827"), wordWrap="CJK"))
    styles.add(ParagraphStyle("CellHead", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=7.3, leading=8.8, textColor=colors.white, alignment=TA_CENTER))
    styles.add(ParagraphStyle("SignLabel", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=7.6, leading=9.2, textColor=colors.HexColor("#475569"), alignment=TA_CENTER))
    styles.add(ParagraphStyle(
        "YousignAnchor",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=1,
        leading=1,
        textColor=colors.Color(1, 1, 1, alpha=0),
        alignment=TA_CENTER,
    ))

    def p(txt, style="Body"):
        return Paragraph(str(txt or "—").replace("\n", "<br/>"), styles[style])

    def section(title):
        return KeepTogether([CondPageBreak(22 * mm), Paragraph(title, styles["H"])])

    def compact_module_label(module):
        text = str(module or "Module non renseigné").strip()
        import re
        match = re.search(r"\b(UV\s*\d+)\b\s*[-—:]?\s*(.*)", text, re.IGNORECASE)
        if not match:
            return text
        uv = match.group(1).upper().replace(" ", "")
        label = match.group(2).strip(" -—:")
        label = re.sub(r"^(MODULE|UNIT[ÉE] DE VALEUR)\s*[:\-—]?\s*", "", label, flags=re.IGNORECASE)
        return f"{uv} — {label}" if label else uv

    def bullet(items):
        return ListFlowable([ListItem(p(i, "Body"), leftIndent=3, bulletOffsetY=1) for i in items], bulletType="bullet", start="•", leftIndent=10, bulletIndent=2, bulletFontSize=5.5, spaceBefore=1, spaceAfter=3)

    def kv_table(rows, col_widths=None, shade=True):
        table = Table([[p(k, "Small"), p(v, "Small")] for k, v in rows], colWidths=col_widths or [42 * mm, 52 * mm], hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#d9dee7")),
            ("INNERGRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#e7ebf0")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#faf7f1") if shade else colors.white),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return table

    def card(title, lines):
        body = [p(title, "CardTitle"), p(lines, "CardText")]
        tbl = Table([[body]], colWidths=[(width - 40 * mm) / 2], rowHeights=[40 * mm], hAlign="LEFT")
        tbl.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor("#d7dce3")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fbfaf7")),
            ("LEFTPADDING", (0, 0), (-1, -1), 5.5), ("RIGHTPADDING", (0, 0), (-1, -1), 5.5),
            ("TOPPADDING", (0, 0), (-1, -1), 5.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5.5),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        return tbl

    story = []
    if logo_path:
        story.append(Image(logo_path, width=34 * mm, height=16 * mm, kind="proportional", hAlign="CENTER"))
        story.append(Spacer(1, 3))
    story += [p("Contrat d’intervention formateur", "CoverTitle"), p("Contrat de prestation de services / sous-traitance pédagogique", "CoverSubtitle")]
    cover_cards = [
        [card("Centre de formation", "Intégrale Academy<br/>54 chemin du Carreou<br/>83480 Puget-sur-Argens<br/>SIRET : 840 899 884 00026<br/>Représentant légal : Monsieur Clément VAILLANT"), card("Formateur / prestataire", f"{contract.get('trainerName')}<br/>{contract.get('status') or 'Statut juridique à compléter'}<br/>SIRET : {contract.get('siret') or 'à compléter'}<br/>NDA : {contract.get('activityDeclaration') or 'à compléter'}<br/>{contract.get('address') or 'Adresse à compléter'}<br/>{contract.get('trainerEmail') or 'Email à compléter'} — {contract.get('trainerPhone') or 'Téléphone à compléter'}")],
        [card("Mission", f"{formation_name} — {session_name}<br/>Du {start_date} au {end_date}<br/>Examen : {exam_date}<br/>Modalité : {modality_label}<br/>Volume : {float(contract.get('calculatedHours') or 0):g} h"), card("Rémunération", f"{float(contract.get('billedDays') or 0):g} jour(s) facturé(s)<br/>Tarif journalier : {_money(contract.get('dailyRate'))} HT<br/>Total HT : {_money(total_ht)}<br/>TVA : {tva_label}<br/>Total TTC : {_money(contract.get('totalTTC') or total_ht)}")],
        [card("Lieu d’intervention", room_label), card("Documents contractuels", "Le présent contrat est complété par le planning détaillé des interventions, le récapitulatif financier et l’engagement qualité / traçabilité pédagogique.")],
    ]
    cover_grid = Table(cover_cards, colWidths=[(width - 40 * mm) / 2, (width - 40 * mm) / 2], rowHeights=[46 * mm, 46 * mm, 46 * mm], hAlign="CENTER")
    cover_grid.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 2), ("RIGHTPADDING", (0, 0), (-1, -1), 2), ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
    story += [cover_grid, Spacer(1, 5), p("Document contractuel généré automatiquement à partir des informations de session et du planning validé.", "Subtle"), NextPageTemplate("contrat"), PageBreak()]

    story += [section("1. Nature juridique du contrat et indépendance du prestataire"),
        p("Le présent contrat est un contrat de prestation de services et, le cas échéant, de sous-traitance pédagogique. Le formateur intervient exclusivement en qualité de prestataire indépendant : aucune clause, consigne opérationnelle ou modalité de suivi ne peut être interprétée comme créant un lien de subordination juridique avec Intégrale Academy."),
        p("Le formateur organise librement ses moyens humains, matériels et pédagogiques, sous sa responsabilité professionnelle. Cette autonomie s’exerce toutefois dans le cadre strict des référentiels applicables, du planning validé, des horaires communiqués, du règlement intérieur, des procédures qualité, des consignes de sécurité et des exigences administratives du centre."),
        bullet(["le formateur demeure seul responsable de ses déclarations fiscales et sociales, de ses cotisations, assurances, habilitations, qualifications, autorisations administratives et obligations réglementaires ;", "le présent contrat n’emporte aucune exclusivité au profit d’Intégrale Academy, sauf accord écrit distinct ;", "le formateur s’interdit de se présenter comme salarié, représentant permanent, mandataire social ou agent d’Intégrale Academy auprès des stagiaires, financeurs, autorités ou partenaires."])]
    story += [section("2. Objet de la mission"), kv_table([("Formation", f"{formation_name} / {'Service de Sécurité Incendie et d’Assistance à Personnes — Niveau 1' if is_ssiap1_session(session_data) else ('Agent de protection physique des personnes' if str(formation_name).upper()=='A3P' else 'Agent de Prévention et de Sécurité')}"), ("Session", session_name), ("Dates", f"Du {start_date} au {end_date}"), ("Date d’examen", exam_date), ("Modalité", modality_label), ("Modules / UV confiés", ", ".join(modules) or "Selon planning annexé"), ("Volume horaire formation", f"{float(contract.get('calculatedHours') or 0):g} heures"), ("Intervention examen", "Oui" if contract.get("trainerAttendsExam") else "Non"), ("Heures examen", f"{float(contract.get('examTrainerHours') or 0):g} h"), ("Montant examen HT", _money(contract.get("examTrainerAmount") or 0)), ("Jours facturés", f"{float(contract.get('billedDays') or 0):g}"), ("Lieu d’intervention", room_label)], [43 * mm, 133 * mm]),
        p("La mission confiée porte sur la réalisation de séquences pédagogiques identifiées dans le planning annexé, pour les modules, dates, horaires et volumes horaires générés dynamiquement par la session. En signant le présent contrat, le formateur accepte les modules confiés, les objectifs pédagogiques, le niveau attendu et les contraintes de continuité pédagogique propres à l’action de formation."),
        p("Le formateur adapte son animation au public accueilli, veille à la progression des apprenants et respecte les contenus réglementaires sans modification non validée. Toute modification de planning, d’horaires, de lieu, de module, de modalité ou d’intervenant doit faire l’objet d’une validation écrite préalable d’Intégrale Academy.")]
    story += [section("3. Obligations générales du formateur"),
        p("Le formateur s’engage à préparer sérieusement chaque séance, à assurer une présence effective et ponctuelle pendant toute la durée prévue, et à conduire ses interventions avec rigueur, pédagogie et professionnalisme. Il garantit que son intervention respecte le programme réglementaire applicable, le référentiel applicable, les exigences administratives, les critères Qualiopi, les attentes financeurs et les consignes internes transmises."),
        bullet(["respecter strictement les horaires, dates, durées, lieux, modules et modalités figurant au planning ;", "signer obligatoirement le planning journalier par demi-journée et faire compléter les émargements requis ;", "remettre les feuilles d’émargement, documents de suivi, évaluations, observations et justificatifs demandés avant de quitter le centre ou dans le délai fixé ;", "signaler immédiatement toute absence, retard, incident, difficulté stagiaire, tension de groupe, problème matériel ou risque de non-conformité ;", "ne pas quitter le centre sans avoir remis les documents nécessaires à la traçabilité de la journée ;", "adopter un comportement exemplaire avec les stagiaires, l’équipe administrative, les partenaires, les auditeurs et tout représentant d’une autorité de contrôle ;", "ne pas se faire remplacer ni sous-traiter tout ou partie de la mission sans accord écrit préalable d’Intégrale Academy."])]
    story += [section("4. Obligations d’Intégrale Academy"),
        p("Intégrale Academy met à disposition du formateur, lorsque cela est nécessaire à l’exécution de la mission, les informations et moyens raisonnablement requis : locaux, salle, supports disponibles, listes de stagiaires, feuilles d’émargement, planning, consignes pédagogiques, consignes qualité et documents administratifs utiles."),
        p("Le centre assure la validation des heures réalisées au regard du planning, des émargements, des documents remis et des contrôles internes. Le paiement intervient uniquement après réception d’une facture conforme et des justificatifs attendus. Intégrale Academy peut contrôler la conformité des interventions et demander toute correction documentaire nécessaire avant validation ou paiement.")]
    story += [section("5. Sécurité, locaux et fermeture du centre"),
        p("Lorsque l’intervention se déroule dans les locaux d’Intégrale Academy, le formateur participe activement à la sécurité des personnes, à la protection des biens et à la préservation des équipements. Il utilise les locaux et matériels avec soin et respecte les règles d’accès, de rangement, de fermeture et de confidentialité."),
        p("En fin de journée ou de demi-journée, il vérifie que la salle est propre, rangée et exploitable pour la séance suivante : tables, chaises, supports et matériels doivent être remis en ordre ; aucun document confidentiel, liste stagiaire, feuille d’émargement ou information interne ne doit rester visible ou accessible."),
        bullet(["éteindre vidéoprojecteur, écran, ordinateur, climatisation, chauffage et lumières inutiles lorsque ces équipements ont été utilisés ;", "fermer les fenêtres, vérifier les accès, fermer la salle à clé et restituer les clés ou badges selon les consignes ;", "s’il est le dernier présent dans le centre, s’assurer que les lumières sont éteintes, que les accès sont sécurisés et que les locaux sont laissés en bon état ;", "signaler immédiatement toute anomalie, dégradation, perte de clé, incident matériel, problème de sécurité ou situation susceptible d’engager la responsabilité du centre."])]
    story += [section("6. Traçabilité pédagogique et qualité"),
        p("Les émargements, plannings journaliers, évaluations, observations, justificatifs et preuves de suivi constituent des éléments essentiels de preuve de la réalisation de la formation. Ils conditionnent la conformité Qualiopi, la relation avec les financeurs, la conformité réglementaire et la capacité d’Intégrale Academy à justifier la réalité de l’action."),
        p("Tous les documents doivent être sincères, complets, lisibles, cohérents avec les horaires réellement effectués et remis dans les délais fixés. Le planning journalier doit être signé par demi-journée. Toute anomalie, omission, rature non justifiée ou incohérence doit être signalée et corrigée sans délai."),
        bullet(["l’absence, l’incomplétude ou l’incohérence d’un document peut entraîner la suspension de la validation des heures et du paiement correspondant jusqu’à régularisation ;", "en modalité e-learning ou mixte, le formateur contribue à la conservation des preuves de connexions, accompagnements, évaluations, échanges pédagogiques, relances et suivis individualisés ;", "le formateur coopère à la démarche d’amélioration continue et fournit les éléments nécessaires aux audits, contrôles qualité et demandes des financeurs."])]
    if has_elearning:
        story += [section("Note — Modalités e-learning"), p("Lorsque des séquences e-learning sont associées à la session, le formateur respecte les modalités de suivi prévues par Intégrale Academy et contribue à documenter l’accompagnement pédagogique réalisé à distance."), bullet(["conserver ou transmettre les preuves de suivi, de connexion, d’évaluation et d’échanges pédagogiques ;", "alerter le centre en cas d’inactivité, d’absence de progression ou de difficulté technique d’un stagiaire ;", "ne valider aucun suivi qui ne serait pas réellement effectué ou traçable."])]
    story += [section("7. Rémunération et facturation"), kv_table([("Nombre total d’heures formation", f"{float(contract.get('calculatedHours') or 0):g} h"), ("Heures examen séparées", f"{float(contract.get('examTrainerHours') or 0):g} h" if contract.get("trainerAttendsExam") else "Non incluses"), ("Montant examen HT", _money(contract.get("examTrainerAmount") or 0)), ("Nombre de jours calculés", f"{float(contract.get('calculatedDays') or 0):g}"), ("Nombre de jours facturés retenus", f"{float(contract.get('billedDays') or 0):g}"), ("Tarif journalier HT", f"{_money(contract.get('dailyRate'))} HT"), ("Total HT", f"{_money(total_ht)} HT"), ("TVA", f"{tva_label} — {_money(contract.get('vatAmount') or 0)}"), ("Total TTC", _money(contract.get('totalTTC') or total_ht))], [52 * mm, 124 * mm]),
        p("Seules les heures effectivement réalisées, justifiées par les documents attendus et validées par Intégrale Academy ouvrent droit à rémunération. La facture du formateur doit être conforme aux informations contractuelles, aux règles fiscales applicables et aux heures validées."),
        bullet(["aucun paiement automatique n’est dû en cas d’absence, de retard, d’annulation, de prestation non réalisée ou de document manquant ;", "en cas de réalisation partielle, Intégrale Academy peut proratiser le montant dû selon les heures ou demi-journées réellement effectuées et validées ;", "les frais de déplacement, repas, hébergement, stationnement ou toute indemnité complémentaire ne sont pas inclus sauf accord écrit préalable ;", "le formateur reste seul responsable de ses charges, cotisations, impôts, déclarations et obligations comptables."])]
    story += [section("8. Annulation, report, absence"),
        p("Le formateur informe immédiatement Intégrale Academy par écrit de toute difficulté susceptible d’affecter sa présence, sa ponctualité ou la continuité pédagogique. Il doit respecter un délai de prévenance raisonnable et proposer, lorsque cela est possible, les éléments permettant au centre d’organiser une solution compatible avec les exigences réglementaires et pédagogiques."),
        bullet(["une absence injustifiée, un retard significatif ou une annulation tardive peut constituer un manquement grave si la session est désorganisée ;", "Intégrale Academy peut reporter, modifier ou annuler une intervention pour des raisons pédagogiques, administratives, réglementaires, commerciales, organisationnelles ou liées au nombre de stagiaires ;", "les heures non réalisées ne sont pas dues et aucune indemnité n’est acquise sans validation écrite préalable ;", "en cas de désorganisation de la session ou de risque de non-conformité, le contrat peut être résilié dans les conditions prévues ci-après."])]
    social = ["garantit être régulièrement immatriculé et autorisé à exercer son activité ;", "fournit sur demande son SIRET, assurance RC Pro, attestation de vigilance le cas échéant, NDA si applicable, justificatifs de compétences, diplômes, habilitations et tout document administratif utile ;", "informe immédiatement Intégrale Academy de tout changement de statut, radiation, suspension, interdiction d’exercer, défaut d’assurance ou perte d’habilitation ;", "garantit Intégrale Academy contre tout recours, redressement, sanction ou réclamation lié au travail dissimulé, à un défaut de déclaration, à un défaut d’assurance ou à un manquement réglementaire du formateur."]
    if total_ht >= 5000:
        social.append("Compte tenu du montant de la prestation, le formateur devra fournir une attestation de vigilance URSSAF de moins de six mois.")
    story += [section("9. Conformité administrative, sociale et fiscale"), p("Le formateur garantit que sa situation administrative, sociale, fiscale et professionnelle est régulière pendant toute la durée du contrat."), bullet(social)]
    story += [section("10. Responsabilité et assurance"),
        p("Le formateur est responsable de la qualité pédagogique de ses interventions et des conséquences de ses fautes, négligences, omissions, comportements inadaptés ou non-respects des consignes. Il doit disposer d’une assurance responsabilité civile professionnelle couvrant son activité de formation et en justifier sur demande."),
        bullet(["utiliser correctement les locaux, matériels, supports et équipements mis à disposition ;", "répondre des dégradations volontaires ou résultant d’une négligence caractérisée ;", "signaler immédiatement tout incident, accident, dégradation, réclamation stagiaire ou difficulté susceptible d’engager sa responsabilité ou celle du centre."])]
    story += [section("11. Qualité pédagogique, documents et preuves de réalisation"),
        p("Le formateur prépare ses interventions, adapte ses méthodes au public, respecte les objectifs pédagogiques et utilise des méthodes d’animation compatibles avec le niveau attendu. Il ne modifie pas les contenus réglementaires, volumes horaires, objectifs, évaluations ou modalités de réalisation sans validation préalable du centre."),
        bullet(["identifier et remonter les difficultés stagiaires, besoins d’adaptation, absences, incidents ou risques d’échec ;", "participer aux contrôles qualité, retours d’expérience et actions correctives demandées ;", "remettre évaluations, observations, preuves de réalisation, documents pédagogiques et éléments de suivi selon les délais fixés ;", "contribuer à une animation claire, structurée, professionnelle et conforme aux exigences du référentiel."])]
    story += [section("12. Confidentialité et protection des données"),
        p("Le formateur est tenu à une obligation stricte de confidentialité concernant les données stagiaires, dossiers administratifs, tarifs, supports, procédures internes, informations commerciales, documents pédagogiques et toute information non publique portée à sa connaissance dans le cadre de la mission."),
        bullet(["ne pas copier, transmettre, réutiliser, céder, publier ou diffuser les documents et supports d’Intégrale Academy sans autorisation écrite ;", "protéger les données personnelles et n’y accéder que pour les besoins stricts de la mission ;", "ne pas utiliser les contacts stagiaires à des fins personnelles, commerciales, de prospection ou au profit d’un tiers ;", "maintenir cette obligation pendant toute la durée du contrat et après son terme, quelle qu’en soit la cause."])]
    story += [section("13. Comportement professionnel"),
        p("Le formateur adopte en toutes circonstances une attitude correcte, respectueuse, neutre et professionnelle. Il veille à préserver l’image, la réputation et les intérêts légitimes d’Intégrale Academy auprès des stagiaires, équipes, partenaires, financeurs et autorités."),
        bullet(["porter une tenue adaptée au contexte de formation et respecter le règlement intérieur ;", "s’interdire tout propos discriminatoire, agressif, humiliant, politique, religieux, déplacé ou contraire à la dignité des personnes ;", "gérer les désaccords et conflits avec calme et professionnalisme ;", "prévenir sans délai la direction en cas de difficulté avec un stagiaire, un groupe, un partenaire ou un membre de l’équipe."])]
    story += [section("14. Résiliation"),
        p("Intégrale Academy peut résilier le contrat en cas de manquement contractuel du formateur, notamment en cas d’absence, retard répété, défaut de documents, refus de correction, comportement inadapté, non-respect du référentiel, problème de sécurité, non-respect de la confidentialité ou atteinte à l’image du centre."),
        p("En cas de faute grave, de risque réglementaire, de mise en danger, de désorganisation importante de la session ou de manquement portant atteinte aux intérêts du centre, la résiliation peut être immédiate. Seules les heures effectivement réalisées, justifiées et validées restent dues ; aucune indemnité n’est due pour les heures non réalisées ou annulées.")]
    story += [section("15. Contrôles, audits et conformité"),
        p("Le formateur accepte que ses interventions, documents et preuves de réalisation puissent être vérifiés dans le cadre de contrôles internes, audits Qualiopi, contrôles financeurs, contrôles CNAPS, ADEF ou demandes de toute autorité compétente."),
        bullet(["coopérer loyalement avec Intégrale Academy et fournir rapidement tout document ou explication demandé ;", "corriger sans délai les non-conformités documentaires ou pédagogiques relevant de son intervention ;", "rester disponible pour toute demande d’explication liée à une intervention passée, un audit, un contrôle financeur ou une vérification réglementaire ;", "respecter les plans d’action qualité décidés par Intégrale Academy lorsque ceux-ci concernent la traçabilité ou la conformité de ses prestations."])]

    story += [NextPageTemplate("planning_landscape"), PageBreak(), section("Annexe 1 : Planning détaillé des interventions")]
    planning_rows = [[p("Date", "CellHead"), p("Demi-journée", "CellHead"), p("Horaires", "CellHead"), p("Durée", "CellHead"), p("UV / module", "CellHead"), p("Lieu", "CellHead"), p("Signature", "CellHead")]]
    for r in interventions:
        try:
            hour = int((r.get("start") or "12:00").split(":")[0])
        except Exception:
            hour = 12
        half_day = "Matin" if hour < 12 else "Après-midi"
        modality = r.get("modality") or "Présentiel"
        trace = "Traçabilité e-learning" if "learning" in modality.lower() else ""
        planning_rows.append([p(r.get("dateLabel") or r.get("date"), "Cell"), p(half_day, "Cell"), p(f"{r.get('start')} - {r.get('end')}", "Cell"), p(f"{float(r.get('hours') or 0):g} h", "Cell"), p(compact_module_label(r.get("module")), "Cell"), p(room_label, "Cell"), p(trace, "Cell")])
    table = Table(planning_rows, colWidths=[24*mm, 22*mm, 22*mm, 15*mm, 88*mm, 55*mm, 33*mm], repeatRows=1, splitByRow=1, hAlign="LEFT")
    table.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#3b3026")), ("GRID", (0,0), (-1,-1), 0.22, colors.HexColor("#cbd5e1")), ("VALIGN", (0,0), (-1,-1), "TOP"), ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#fafafa")]), ("LEFTPADDING", (0,0), (-1,-1), 2.5), ("RIGHTPADDING", (0,0), (-1,-1), 2.5), ("TOPPADDING", (0,0), (-1,-1), 3), ("BOTTOMPADDING", (0,0), (-1,-1), 3)]))
    story += [table, NextPageTemplate("contrat"), PageBreak(), section("Annexe 2 : Récapitulatif financier"), kv_table([("Total heures", f"{float(contract.get('calculatedHours') or 0):g} h"), ("Jours facturés", f"{float(contract.get('billedDays') or 0):g}"), ("Tarif journalier HT", f"{_money(contract.get('dailyRate'))} HT"), ("Total HT", f"{_money(total_ht)} HT"), ("TVA", f"{tva_label} — {_money(contract.get('vatAmount') or 0)}"), ("Total TTC", _money(contract.get('totalTTC') or total_ht))], [52*mm, 124*mm])]
    story += [KeepTogether([section("Annexe 3 : Engagement qualité et traçabilité pédagogique"),
        p("Le formateur confirme que la qualité de l’animation et la traçabilité pédagogique constituent des obligations essentielles du contrat. Il s’engage à respecter le référentiel applicable, le programme validé, les objectifs pédagogiques, les modalités d’évaluation et les procédures qualité d’Intégrale Academy."),
        bullet(["faire compléter et contrôler les émargements à chaque demi-journée ;", "signer le planning journalier par demi-journée et signaler immédiatement toute incohérence ;", "remonter sans délai les absences, retards, incidents, difficultés stagiaires, problèmes matériels ou situations de sécurité ;", "remettre en fin de journée, ou dans le délai fixé par le centre, les feuilles d’émargement, observations, évaluations, justificatifs et documents demandés ;", "conserver ou transmettre les preuves e-learning lorsque la session comporte un suivi à distance : connexions, accompagnements, évaluations et échanges pédagogiques ;", "participer à la démarche Qualiopi, aux contrôles financeurs, audits internes, actions correctives et demandes de preuves ;", "garantir une animation professionnelle, adaptée au public, respectueuse des stagiaires et conforme à l’image d’Intégrale Academy ;", "préserver la confidentialité des données stagiaires, supports, documents internes et informations commerciales."])
    ])]

    story += [PageBreak(), p(f"Fait à Puget-sur-Argens, le {generated}.", "Body"), p("Chaque partie reconnaît avoir pris connaissance du présent contrat, de ses annexes éventuelles et en accepter l’ensemble des conditions.", "Body"), Spacer(1, 5)]
    signature_image = find_center_image("signature", "sign")
    stamp_image = find_center_image("tampon", "cachet", "stamp")

    def signature_zone(label, image_path=None, height_mm=31):
        content = [p(label, "SignLabel")]
        if image_path:
            content.append(Image(image_path, width=42 * mm, height=(18 if "Signature" in label else 24) * mm, kind="proportional", hAlign="CENTER"))
        else:
            content.append(p("<br/><br/>", "Body"))
        tbl = Table([[content]], colWidths=[75 * mm], rowHeights=[height_mm * mm])
        tbl.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")), ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ffffff")), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5), ("TOPPADDING", (0, 0), (-1, -1), 5)]))
        return tbl

    def trainer_signature_box(height_mm=31):
        # Signature positionnée exclusivement par champ manuel Yousign.
        # Ne pas réinsérer l’ancre {{s1|signature|160|60}} dans ce PDF.
        content = [p("Signature du formateur", "SignLabel"), Spacer(1, 8), p(" ", "Body")]
        tbl = Table([[content]], colWidths=[75 * mm], rowHeights=[height_mm * mm])
        tbl.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")), ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ffffff")), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5), ("TOPPADDING", (0, 0), (-1, -1), 5)]))
        return tbl

    def party_block(title, name, quality, sig_img=None, stamp_img=None, trainer_signature_anchor=False):
        signature_label = "Signature du formateur" if trainer_signature_anchor else "Signature / cachet du centre"
        inner = [p(f"<b>{title}</b><br/>{name}<br/>{quality}<br/><br/>Mention manuscrite : “Lu et approuvé”", "Body"), trainer_signature_box() if trainer_signature_anchor else signature_zone(signature_label, sig_img), Spacer(1, 3), signature_zone("Tampon / cachet", stamp_img, 28)]
        tbl = Table([[inner]], colWidths=[82 * mm], rowHeights=[108 * mm])
        tbl.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#94a3b8")), ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fbfaf7")), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 7), ("BOTTOMPADDING", (0, 0), (-1, -1), 7)]))
        return tbl

    sign_table = Table([[party_block("Pour Intégrale Academy", "Monsieur Clément VAILLANT", "Directeur général", signature_image, stamp_image), party_block("Pour le formateur / prestataire", contract.get('trainerName') or "Nom à compléter", contract.get('status') or "Qualité à compléter", trainer_signature_anchor=True)]], colWidths=[86 * mm, 86 * mm], hAlign="CENTER")
    sign_table.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP"), ("LEFTPADDING", (0,0), (-1,-1), 2), ("RIGHTPADDING", (0,0), (-1,-1), 2)]))
    story.append(sign_table)
    doc.build(story)

def _aps_presentiel_days(planning_data, planning_mode):
    days = []
    for day in planning_data or []:
        slots = day.get("slots") or []
        if planning_mode in {"elearning_presentiel", "desp"}:
            slots = [slot for slot in slots if is_in_person_slot(slot)]
        if slots:
            copied = dict(day)
            copied["slots"] = slots
            days.append(copied)
    return days


def _minutes_from_hhmm(value):
    try:
        h, m = str(value).split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return 0


def _hhmm_to_fr(value):
    return str(value or "").replace(":", "h")


def _period_label(slots, morning=True):
    selected = []
    for slot in slots:
        midpoint = (_minutes_from_hhmm(slot.get("start")) + _minutes_from_hhmm(slot.get("end"))) / 2
        if (morning and midpoint < 13 * 60) or (not morning and midpoint >= 13 * 60):
            selected.append(slot)
    if not selected:
        return "—"
    return " / ".join(f"{_hhmm_to_fr(s.get('start'))} - {_hhmm_to_fr(s.get('end'))}" for s in selected)


def _aps_normalize_student_name(value, uppercase=False):
    cleaned = re.sub(r"[^A-Za-zÀ-ÿ' ,\-]", "", value or "")
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ,\t;-')(")
    if uppercase:
        return cleaned.upper()
    return cleaned


def _aps_extract_phone_from_line(line):
    phone_pattern = re.compile(r"(?<!\d)((?:\+33|0)\s*[1-9](?:[\s.\-]*\d{2}){4})(?!\d)")
    match = phone_pattern.search(line)
    if not match:
        return "", line
    phone = re.sub(r"\D+", "", match.group(1))
    if phone.startswith("33") and len(phone) == 11:
        phone = "0" + phone[2:]
    if len(phone) == 10:
        phone = " ".join([phone[:2], phone[2:4], phone[4:6], phone[6:8], phone[8:]])
    remaining = (line[:match.start()] + " " + line[match.end():]).strip()
    return phone, remaining


def aps_extract_students_from_text(text):
    students = []
    seen = set()
    email_pattern = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.IGNORECASE)

    for raw_line in (text or "").splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip(" -\t;")
        if not line:
            continue

        numbered_match = re.match(r"^(\d{1,3})\s+(.+)$", line)
        if not numbered_match:
            continue

        candidate = numbered_match.group(2).strip()
        lowered = candidate.lower()
        if any(fragment in lowered for fragment in (
            "numéros trouvés", "numero trouves", "numéro trouvé", "# nom", "nom prénom",
            "email téléphone", "liste stagiaires", "http://", "https://", "page "
        )):
            continue

        email = ""
        email_match = email_pattern.search(candidate)
        if email_match:
            email = email_match.group(0).strip()
            candidate = (candidate[:email_match.start()] + " " + candidate[email_match.end():]).strip()

        phone, candidate = _aps_extract_phone_from_line(candidate)
        name_part = re.sub(r"\s+", " ", candidate).strip(" -\t;")
        parts = name_part.split()
        if len(parts) < 2:
            continue

        last_parts = []
        for part in parts:
            comparable = re.sub(r"[^A-Za-zÀ-ÿ]", "", part)
            if comparable and comparable.upper() == comparable:
                last_parts.append(part)
            else:
                break
        if not last_parts or len(last_parts) >= len(parts):
            continue

        last = _aps_normalize_student_name(" ".join(last_parts), uppercase=True)
        first = _aps_normalize_student_name(" ".join(parts[len(last_parts):]))
        if not last or not first:
            continue

        student = {"lastName": last, "firstName": first}
        if email:
            student["email"] = email
        if phone:
            student["phone"] = phone
        key = (last, first, email, phone)
        if key not in seen:
            seen.add(key)
            students.append(student)
    return students


def aps_extract_students_from_pdf(file_storage):
    suffix = os.path.splitext(file_storage.filename or "")[-1].lower()
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix or ".pdf") as tmp:
        file_storage.save(tmp.name)
        tmp_path = tmp.name
    try:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("La dépendance pypdf est requise pour lire le texte PDF.") from exc
        reader = PdfReader(tmp_path)
        text = "\n".join((page.extract_text() or "") for page in reader.pages)
        return aps_extract_students_from_text(text), bool(text.strip())
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass



def attendance_header_layout(page_width, page_height, margin, string_width, title="FEUILLE DE PRÉSENCE", subtitle=None):
    """Return the shared APS/DESP attendance header geometry."""
    logo_width = 91
    logo_height = 55
    logo_x = margin
    logo_y = page_height - 72
    lines = [
        {"text": title, "font": "Helvetica-Bold", "size": 17, "x": page_width / 2, "y": page_height - 38, "align": "center"},
        {"text": subtitle or "Agent de Prévention et de Sécurité (APS)", "font": "Helvetica", "size": 9, "x": page_width / 2, "y": page_height - 54, "align": "center"},
    ]
    for line in lines:
        line["width"] = string_width(line["text"], line["font"], line["size"])
    return {
        "logo": {"x": logo_x, "y": logo_y, "width": logo_width, "height": logo_height},
        "lines": lines,
        "session_y": page_height - 84,
        "title": lines[0],
        "subtitle": lines[1],
    }


def desp_attendance_header_layout(page_width, page_height, margin, string_width):
    """Compatibility wrapper: DESP now uses the shared APS header layout."""
    return attendance_header_layout(page_width, page_height, margin, string_width, subtitle=DESP_LABEL)

def generate_attendance_pdf_common(session_data, output_path, training_type=None, subtitle=None):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise RuntimeError("La dépendance reportlab est requise pour générer le PDF.") from exc

    training_type = (training_type or session_data.get("formation") or "APS").upper()
    planning_data = session_data.get("apsPlanningData") or []
    planning_mode = session_data.get("apsPlanningMode") or "full_presentiel"
    students = session_data.get("apsAttendanceStudents") or []
    is_afc = training_type == "AFC_APS_SSIAP"
    default_student_start_date = session_data.get("date_debut") or ""

    def students_for_day(day_date):
        if not is_afc:
            return students
        day_key = str(day_date or "")
        return [
            student for student in students
            if not (student.get("startDate") or default_student_start_date)
            or not day_key
            or (student.get("startDate") or default_student_start_date) <= day_key
        ]

    is_ssiap1 = training_type == "SSIAP1" or is_ssiap1_session(session_data)
    if is_ssiap1:
        presentiel_days = []
        exam_days = []
        for day in planning_data:
            training_slots = [slot for slot in (day.get("slots") or []) if is_in_person_slot(slot) or _normalized_slot_value(slot, "modality", "delivery_mode", "period_type") in {"sst", "revision"}]
            exam_slots = [slot for slot in (day.get("slots") or []) if _normalized_slot_value(slot, "modality", "delivery_mode", "period_type") in {"exam", "examen"} or (slot.get("uv") or "").upper() == "EXAMEN"]
            if training_slots:
                copied = dict(day); copied["slots"] = training_slots; presentiel_days.append(copied)
            if exam_slots:
                copied = dict(day); copied["slots"] = exam_slots; exam_days.append(copied)
    else:
        presentiel_days = _aps_presentiel_days(planning_data, planning_mode)
        exam_days = []
    if not presentiel_days:
        raise ValueError(f"Aucun jour présentiel n'est trouvé dans le planning {training_type}.")
    if training_type == "DESP":
        desp_hours = {"presentiel": 0.0, "elearning": 0.0, "exam": 0.0}
        for day in planning_data:
            for slot in day.get("slots", []):
                duration = round(float(slot.get("duration") or 0), 2)
                normalized = _normalized_slot_value(slot, "modality", "delivery_mode", "period_type")
                if is_in_person_slot(slot): desp_hours["presentiel"] = round(desp_hours["presentiel"] + duration, 2)
                elif normalized in {"elearning", "e-learning", "distanciel", "distance", "asynchronous"}: desp_hours["elearning"] = round(desp_hours["elearning"] + duration, 2)
                elif normalized in {"examen", "exam"}: desp_hours["exam"] = round(desp_hours["exam"] + duration, 2)
        if desp_hours["presentiel"] != DESP_PRESENTIEL_HOURS:
            raise ValueError(f"Impossible de générer les feuilles de présence DESP : le planning présentiel contient {desp_hours['presentiel']:g}h alors que {DESP_PRESENTIEL_HOURS}h sont attendues.")

    c = canvas.Canvas(output_path, pagesize=A4)
    width, height = A4
    margin = 10 * mm
    FOOTER_RESERVED_HEIGHT = 16 * mm
    footer_y = 6 * mm
    footer_top_y = footer_y + FOOTER_RESERVED_HEIGHT
    content_bottom_limit = footer_top_y + 6
    logo_path = aps_pdf_logo_path()
    stamp_image = find_center_image("tampon", "cachet", "stamp")
    include_summary_page = training_type != "DESP" and not (is_ssiap1 and len(students) <= 6 and exam_days)
    total_pages = len(presentiel_days) + len(exam_days) + (1 if include_summary_page else 0)
    raw_session_name = (session_data.get("display_name") or session_data.get("name") or "").strip()
    session_id = str(session_data.get("id") or "").strip()
    session_name = raw_session_name if raw_session_name and raw_session_name != f"Session {session_id}" else (session_id or raw_session_name or "—")

    def reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75):
        if hasattr(c, "setFillAlpha"):
            c.setFillAlpha(1)
        if hasattr(c, "setStrokeAlpha"):
            c.setStrokeAlpha(1)
        c.setFillColor(fill)
        c.setStrokeColor(stroke)
        c.setLineWidth(line_width)

    def footer(page_no):
        c.saveState()
        reset_graphics_state(fill=colors.HexColor("#111827"), stroke=colors.HexColor("#4b5563"), line_width=0.75)
        c.line(margin, footer_top_y, width - margin, footer_top_y)
        page_label = f"Page {page_no} / {total_pages}"; c.setFillColor(colors.HexColor("#111827")); c.setFont("Helvetica", 6.4)
        lines = APS_LEGAL_LINES[:]
        if is_ssiap1 and SSIAP1_AGREMENT_LINE not in lines: lines.append(SSIAP1_AGREMENT_LINE)
        y = footer_top_y - 6
        for line in lines[:5]:
            reset_graphics_state(fill=colors.HexColor("#111827"), stroke=colors.HexColor("#4b5563"), line_width=0.75)
            c.setFont("Helvetica", 6.4)
            c.drawString(margin, y, line[:155]); y -= 6.2
        reset_graphics_state(fill=colors.HexColor("#111827"), stroke=colors.HexColor("#4b5563"), line_width=0.75)
        c.setFont("Helvetica", 6.4)
        c.drawRightString(width - margin, footer_y, page_label); c.restoreState()

    def draw_header(title, date_label, slots, page_no, *, exam=False, continuation=False):
        c.saveState()
        reset_graphics_state(fill=colors.white, stroke=colors.white, line_width=0)
        c.rect(0, 0, width, height, fill=1, stroke=0)
        c.restoreState()
        reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75)
        if logo_path:
            c.saveState(); c.drawImage(logo_path, margin, height - 72, width=91, height=55, preserveAspectRatio=True, mask="auto"); c.restoreState()
        reset_graphics_state(fill=colors.HexColor("#111827"), stroke=colors.black, line_width=0.75)
        c.setFont("Helvetica-Bold", 16 if exam else 17)
        c.drawCentredString(width / 2, height - 38, title)
        c.setFont("Helvetica", 9); c.drawCentredString(width / 2, height - 54, subtitle or "Agent de Prévention et de Sécurité (APS)")
        y = height - 84
        if continuation:
            c.setFont("Helvetica-Bold", 9); c.drawString(margin, y, f"Suite de la feuille de présence du {date_label}"); y -= 16
        trainers = sorted({(s.get("trainer") or "").strip() for s in slots if (s.get("trainer") or "").strip()}) or ["—"]
        room = (slots[0].get("room") if slots else "") or session_data.get("exam_room") or session_data.get("salle") or "—"
        am = period_amplitude(slots, True); pm = period_amplitude(slots, False)
        rows = [("Session", session_name, "Date de l’examen" if exam else "Date", date_label), ("Responsable(s) / intervenant(s)" if exam else "Formateur", ", ".join(trainers), "Lieu / salle", room), ("Période de formation", f"du {format_date(session_data.get('date_debut'))} au {format_date(session_data.get('date_fin'))}", "Date d’examen", format_date(session_data.get('date_exam'))), ("Horaires du matin", am if am != "—" else "", "Horaires de l’après-midi", pm if pm != "—" else "")]
        col_w = (width - 2 * margin - 12) / 2; label_w = 112; row_h = 21
        # No filled background box here: DESP must use the same readable APS text flow,
        # without a pale rectangle covering or tinting the session information.
        reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75)
        cy = y - 14
        for l1,v1,l2,v2 in rows:
            for x,l,v in [(margin+6,l1,v1),(margin+6+col_w+12,l2,v2)]:
                reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75)
                c.setFont("Helvetica-Bold", 7.4); c.drawString(x, cy, f"{l} :")
                reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75)
                c.setFont("Helvetica", 7.8)
                tx = x + label_w
                for i,line in enumerate(wrap_text_lines(v, col_w - label_w - 8, "Helvetica", 7.8)[:2]): c.drawString(tx, cy - i*8, line)
            cy -= row_h
        return y - row_h*4 - 18

    def period_amplitude(slots, morning=True):
        start_limit, end_limit = (0, 13 * 60) if morning else (13 * 60, 24 * 60)
        selected = [slot for slot in slots if _minutes_from_hhmm(slot.get("start")) < end_limit and _minutes_from_hhmm(slot.get("end")) > start_limit]
        if not selected:
            return "—"
        if any(_minutes_from_hhmm(slot.get("start")) < 13 * 60 and _minutes_from_hhmm(slot.get("end")) > 13 * 60 for slot in selected):
            return "08h30 - 12h30" if morning else "13h30 - 16h30"
        return f"{_hhmm_to_fr(min(slot.get('start') for slot in selected))} - {_hhmm_to_fr(max(slot.get('end') for slot in selected))}"

    def parts(slots):
        return period_amplitude(slots, True) != "—", period_amplitude(slots, False) != "—"

    def module_rows(slots):
        out=[]
        for slot in slots:
            code=slot.get("uv") or ""; detail=SSIAP1_SEQUENCE_DETAIL_BY_CODE.get(code, {})
            ps = f"Partie {detail.get('part_number')} - Séquence {detail.get('sequence_number')}" if detail else code
            typ = (slot.get("subpartLabel") or slot.get("subpartType") or slot.get("content_type") or slot.get("type") or slot.get("modality") or "Contenu").capitalize()
            if typ.lower() in {"presentiel", "présentiel"}: typ = "Contenu"
            items = slot.get("subpartDisplayItems") or slot.get("subpartItems") or []
            title = " ; ".join(items[:2]) if items else (slot.get("content") or slot.get("title") or detail.get("title") or "")
            row=(f"{_hhmm_to_fr(slot.get('start'))} - {_hhmm_to_fr(slot.get('end'))}", ps, typ, title)
            if not out or out[-1] != row: out.append(row)
        return out

    def draw_modules(y, slots):
        reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75)
        c.setFont("Helvetica-Bold", 8.8); c.drawString(margin, y, "Programme et horaires du jour"); y-=12
        headers=["Horaires","Partie / Séquence","Type","Intitulé"]; ws=[70,105,62,width-2*margin-70-105-62]
        row_data=[headers]+module_rows(slots); x0=margin
        for r,row in enumerate(row_data):
            line_lists=[wrap_text_lines(cell, ws[i]-6, "Helvetica-Bold" if r==0 else "Helvetica", 8.2 if r==0 else 8.5) or [""] for i,cell in enumerate(row)]
            rh=max(16, max(len(ll) for ll in line_lists)*9.5+6)
            c.saveState()
            reset_graphics_state(fill=colors.HexColor("#f3f4f6") if r==0 else colors.white, stroke=colors.black, line_width=0.75)
            c.rect(x0, y-rh, sum(ws), rh, fill=1, stroke=1)
            c.restoreState()
            reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75)
            x=x0
            for i,ll in enumerate(line_lists):
                if i:
                    reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75)
                    c.line(x, y, x, y-rh)
                reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75)
                c.setFont("Helvetica-Bold" if r==0 else "Helvetica", 8.2 if r==0 else 8.5)
                ty=y-11
                for line in ll: c.drawString(x+3, ty, line); ty-=9.5
                x+=ws[i]
            y-=rh
        return y-8

    def attendance_bottom_layout(table_top, people_count, signatures_count):
        signature_h = 50 if len(students)>6 else 62
        observations_h = signature_h
        header_h = 18
        spacing_after_table = 12
        spacing_between_sections = 10
        min_row_h = 23.5
        max_row_h = 45
        available_table_height = (
            content_bottom_limit
            - table_top
            - signature_h
            - observations_h
            - spacing_after_table
            - spacing_between_sections
        )
        # ReportLab coordinates descend as content is drawn; convert the available
        # vertical span to a positive line height while reserving the APS footer.
        available_table_height = abs(available_table_height) - header_h
        row_h = max(min_row_h, min(max_row_h, available_table_height / max(people_count, 1)))
        table_bottom = table_top - header_h - (row_h * max(people_count, 0))
        signatures_y = table_bottom - spacing_after_table
        bottom_boxes_y = signatures_y - signature_h - spacing_between_sections
        bottom_boxes_bottom = bottom_boxes_y - observations_h
        if bottom_boxes_bottom < content_bottom_limit:
            row_h = max(min_row_h, (table_top - header_h - spacing_after_table - signature_h - spacing_between_sections - observations_h - content_bottom_limit) / max(people_count, 1))
            table_bottom = table_top - header_h - (row_h * max(people_count, 0))
            signatures_y = table_bottom - spacing_after_table
            bottom_boxes_y = signatures_y - signature_h - spacing_between_sections
            bottom_boxes_bottom = bottom_boxes_y - observations_h
        if bottom_boxes_bottom < content_bottom_limit - 0.1:
            raise ValueError("La feuille de présence ne peut pas être dessinée sans chevauchement du footer.")
        return {"row_h": row_h, "signature_h": signature_h, "observations_h": observations_h, "signatures_y": signatures_y, "bottom_boxes_y": bottom_boxes_y, "bottom_boxes_bottom": bottom_boxes_bottom}

    def draw_people_table(y, people, has_am, has_pm, row_h):
        reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75)
        headers=["N°","Nom","Prénom"] + (["Signature matin"] if has_am else []) + (["Signature après-midi"] if has_pm else [])
        if has_am and has_pm: ws=[28,115,95,145,width-2*margin-28-115-95-145]
        else: ws=[28,125,105,width-2*margin-28-125-105]
        c.saveState(); reset_graphics_state(fill=colors.HexColor("#f3f4f6"), stroke=colors.black, line_width=0.75); c.rect(margin,y-18,sum(ws),18,fill=1,stroke=1); c.restoreState(); reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75); c.setFont("Helvetica-Bold",8)
        x=margin
        for i,h in enumerate(headers): c.drawString(x+4,y-12,h); x+=ws[i]
        y-=18; reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75); c.setFont("Helvetica",8.5)
        for idx, st in enumerate(people,1):
            reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75)
            c.rect(margin,y-row_h,sum(ws),row_h); x=margin
            for w in ws[:-1]: x+=w; c.line(x,y,x,y-row_h)
            reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75)
            c.setFont("Helvetica",8.5)
            c.drawString(margin+8,y-14,str(idx)); c.drawString(margin+36,y-14,st.get("lastName", "")); c.drawString(margin+36+ws[1],y-14,st.get("firstName", "")); y-=row_h
        return y-12

    def draw_signature_blocks(y, slots, *, exam=False, block_h=None):
        has_am, has_pm = parts(slots); gap=18; bw=(width-2*margin-gap)/2; bh=block_h or (50 if len(students)>6 else 62)
        labels=[]
        if has_am: labels.append(("Signature responsable / intervenant - matin" if exam else ("Signature du formateur - matin" if is_ssiap1 else "Signature formateur matin"), True))
        if has_pm: labels.append(("Signature responsable / intervenant - après-midi" if exam else ("Signature du formateur - après-midi" if is_ssiap1 else "Signature formateur après-midi"), False))
        if len(labels)==1: bw=width-2*margin; gap=0
        for i,(lab,morn) in enumerate(labels):
            x=margin+i*(bw+gap); reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75); c.rect(x,y-bh,bw,bh); c.setFont("Helvetica-Bold",8); c.drawString(x+8,y-14,lab); reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75); c.setFont("Helvetica",8); c.drawString(x+8,y-27,half_day_trainer_label(slots,morn))
        y-=bh+10; bw=(width-2*margin-gap)/2
        for i,lab in enumerate(["Observations éventuelles","Cachet du centre"]):
            x=margin+i*(bw+gap); reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75); c.rect(x,y-bh,bw,bh); c.setFont("Helvetica-Bold",8); c.drawString(x+8,y-14,lab)
            if i==1 and stamp_image: c.drawImage(stamp_image,x+14,y-bh+10,width=bw-28,height=bh-22,preserveAspectRatio=True,anchor="c",mask="auto")
        bottom = y-bh
        if bottom < content_bottom_limit - 0.1:
            raise ValueError("Chevauchement détecté entre les cadres bas et la zone footer réservée.")
        return bottom-8

    def half_day_trainer_label(slots, morning=True):
        names=[]
        for slot in slots:
            in_period = (_minutes_from_hhmm(slot.get("start")) < 13*60 and _minutes_from_hhmm(slot.get("end")) > 0) if morning else (_minutes_from_hhmm(slot.get("end")) > 13*60)
            if in_period and (slot.get("trainer") or "").strip() not in names: names.append((slot.get("trainer") or "").strip())
        return ", ".join([n for n in names if n]) or "—"

    def draw_summary(y):
        c.setFillColor(colors.HexColor("#f9fafb")); c.rect(margin,y-92,width-2*margin,92,fill=1,stroke=1); c.setFillColor(colors.HexColor("#111827")); c.setFont("Helvetica-Bold",10); c.drawString(margin+8,y-14,"Synthèse des feuilles de présence SSIAP 1" if is_ssiap1 else "Synthèse des feuilles de présence")
        total_hours = SSIAP1_TOTAL_HOURS if is_ssiap1 else round(sum(float(slot.get("duration") or 0) for day in presentiel_days for slot in day.get("slots", [])), 2)
        vals=[f"Nombre total de stagiaires : {len(students)}",(f"Nombre de journées de formation : {len(presentiel_days)}" if is_ssiap1 else f"Nombre de journées présentielles : {len(presentiel_days)}"),(f"Total formation : {total_hours:g} h" if is_ssiap1 else f"Nombre total d’heures présentielles : {total_hours:g}h"),f"Période de formation : du {format_date(session_data.get('date_debut'))} au {format_date(session_data.get('date_fin'))}",f"Date d’examen : {format_date(session_data.get('date_exam'))}",("Modalité : présentiel" if is_ssiap1 else "Modalité : Présentiel au centre")]
        c.setFont("Helvetica",8.5); yy=y-30
        for i,v in enumerate(vals): c.drawString(margin+10+(i%2)*250, yy-(i//2)*16, v)
        return y-102

    page_no=1
    for day in presentiel_days:
        slots=day.get("slots") or []; date_label=format_date(day.get("date"));
        if is_ssiap1 and all((slot.get("modality") or "") == "sst" for slot in slots):
            page_title = "FEUILLE DE PRÉSENCE — FORMATION SST"; page_subtitle = "Formation intégrée au parcours SSIAP 1"
        else:
            page_title = "FEUILLE DE PRÉSENCE — FORMATION SSIAP 1" if is_ssiap1 else "FEUILLE DE PRÉSENCE"; page_subtitle = subtitle
        original_subtitle = subtitle
        if page_subtitle != subtitle:
            subtitle = page_subtitle
        y=draw_header(page_title, date_label, slots, page_no)
        subtitle = original_subtitle
        reset_graphics_state(fill=colors.black, stroke=colors.black, line_width=0.75)
        c.setFont("Helvetica-Bold", 8.8); c.drawString(margin, y, "Modules et horaires du jour"); y -= 12
        for slot in slots:
            code = slot.get("code") or slot.get("uv") or ""
            title = slot.get("content") or slot.get("title") or ""
            if is_ssiap1:
                detail = SSIAP1_SEQUENCE_DETAIL_BY_CODE.get(code, {})
                items = slot.get("subpartDisplayItems") or slot.get("subpartItems") or []
                sequence = f"Partie {detail.get('part_number')} - Séquence {detail.get('sequence_number')}" if detail else code
                title = " ; ".join(items[:2]) if items else title or detail.get("title") or ""
                label = f"{code} — {sequence}" if sequence and sequence != code else code
            else:
                label = code
            separator = " — " if label and title else ""
            y = draw_wrapped_text(c, f"{_hhmm_to_fr(slot.get('start'))} - {_hhmm_to_fr(slot.get('end'))} : {label}{separator}{title}", margin + 8, y, width - 2 * margin - 8, "Helvetica", 7.4, 8.5)
        y -= 5
        day_students = students_for_day(day.get("date"))
        has_am,has_pm=parts(slots); layout=attendance_bottom_layout(y, len(day_students), int(has_am)+int(has_pm)); y=draw_people_table(y, day_students, has_am, has_pm, layout["row_h"]); y=draw_signature_blocks(layout["signatures_y"], slots, block_h=layout["signature_h"])
        footer(page_no); c.showPage(); page_no+=1

    for exam_day in exam_days:
        slots=exam_day.get("slots") or []; date_label=format_date(exam_day.get("date")); y=draw_header("FEUILLE DE PRÉSENCE - EXAMEN SSIAP 1", date_label, slots, page_no, exam=True)
        c.setFont("Helvetica-Bold",8.8); c.drawString(margin,y,"Épreuve(s) d’examen"); y-=12
        for slot in slots: y=draw_wrapped_text(c, f"{_hhmm_to_fr(slot.get('start'))} - {_hhmm_to_fr(slot.get('end'))} : {slot.get('title') or 'EXAMEN SSIAP 1'}", margin+8, y, width-2*margin-8, "Helvetica", 8.5, 10)
        day_students = students_for_day(exam_day.get("date"))
        y-=8; layout=attendance_bottom_layout(y, len(day_students), 2); y=draw_people_table(y, day_students, True, True, layout["row_h"]); y=draw_signature_blocks(layout["signatures_y"], slots, exam=True, block_h=layout["signature_h"])
        if is_ssiap1 and len(students)<=6 and y-102>footer_top_y+10: y=draw_summary(y)
        footer(page_no); c.showPage(); page_no+=1

    if include_summary_page:
        y=height-70
        if logo_path: c.drawImage(logo_path, margin, height - 72, width=91, height=55, preserveAspectRatio=True, mask="auto")
        c.setFont("Helvetica-Bold",16)
        title_x = margin + 100 if logo_path else margin
        c.drawString(title_x, y, "Synthèse des feuilles de présence SSIAP 1" if is_ssiap1 else "Synthèse des feuilles de présence"); y-=28
        draw_summary(y); footer(page_no)
    c.save()


def generate_aps_attendance_pdf(session_data, output_path):
    if is_ssiap1_session(session_data):
        training_type = "SSIAP1"
        subtitle = "Service de Sécurité Incendie et d’Assistance à Personnes - Niveau 1"
    else:
        code = normalize_training_code(session_data)
        training_type = "AFC_APS_SSIAP" if code == "AFC_APS_SSIAP" else ("DESP" if (session_data.get("formation") or "").upper() in {"DESP", "DIRIGEANT"} else "APS")
        subtitle = AFC_APS_SSIAP_LABEL if training_type == "AFC_APS_SSIAP" else (DESP_LABEL if training_type == "DESP" else None)
    return generate_attendance_pdf_common(session_data, output_path, training_type=training_type, subtitle=subtitle)


def send_email_with_attachments(to_email, subject, body, attachments):
    smtp_config = get_smtp_config() if 'get_smtp_config' in globals() else {"server": SMTP_SERVER, "port": SMTP_PORT, "username": FROM_EMAIL, "password": EMAIL_PASSWORD, "from_email": FROM_EMAIL}
    if not smtp_config.get("from_email") or not smtp_config.get("username") or not smtp_config.get("password"):
        return False, "SMTP non configuré."
    msg = MIMEMultipart(); msg["From"] = smtp_config["from_email"]; msg["To"] = to_email; msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))
    for path, filename in attachments:
        with open(path, "rb") as f:
            part = MIMEApplication(f.read(), _subtype="pdf")
        part.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(part)
    try:
        with smtplib.SMTP(smtp_config["server"], smtp_config["port"]) as server:
            server.starttls(); server.login(smtp_config["username"], smtp_config["password"]); server.sendmail(smtp_config["from_email"], [to_email], msg.as_string())
        return True, "Email envoyé"
    except Exception as exc:
        return False, f"Erreur email: {exc}"

# -----------------------
# Convocation APS depuis modèle Word officiel
# -----------------------
DOCX_UNRESOLVED_PATTERN = re.compile(r"\[(?:'?[A-Za-zÀ-ÿ0-9_()=]+|:[A-Za-zÀ-ÿ0-9_]+)\]")
WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def _xml_escape(value):
    return ("" if value is None else str(value)).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _format_hour(value, default="08h30"):
    raw = (value or "").strip() if isinstance(value, str) else ""
    if not raw:
        return default
    if re.match(r"^\d{1,2}:\d{2}$", raw):
        h, m = raw.split(":")
        return f"{int(h):02d}h{m}"
    return raw


def _session_label(session_data):
    return session_data.get("display_name") or session_data.get("name") or "TFP APS - Agent de Prévention et de Sécurité"


def _trainee_value(trainee, *keys):
    for key in keys:
        value = trainee.get(key) if isinstance(trainee, dict) else None
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _aps_convocation_context(trainee, session_data):
    start = parse_date(session_data.get("date_debut"))
    exam = parse_date(session_data.get("date_exam"))
    date_convocation = start.strftime("%d/%m/%Y") if start else ""
    date_exam = exam.strftime("%d/%m/%Y") if exam else ""
    civilite = _trainee_value(trainee, "civilite", "NomCivilite", "title")
    prenom = _trainee_value(trainee, "prenom", "Prenom", "first_name")
    nom = _trainee_value(trainee, "nom", "Nom", "last_name")
    ctx = {
        "NomCivilite": civilite,
        "Prenom": prenom,
        "Nom": nom,
        "Ligne1": _trainee_value(trainee, "ligne1", "Ligne1", "adresse", "address1"),
        "Ligne2": _trainee_value(trainee, "ligne2", "Ligne2", "address2"),
        "Ligne3": _trainee_value(trainee, "ligne3", "Ligne3", "address3"),
        "Ligne4": _trainee_value(trainee, "ligne4", "Ligne4", "address4"),
        "CodePostal": _trainee_value(trainee, "code_postal", "CodePostal", "postal_code", "cp"),
        "Ville": _trainee_value(trainee, "ville", "Ville", "city"),
        "NomPedagogique": _session_label(session_data),
        "Libelle": _session_label(session_data),
        "DateConvocation": date_convocation,
        "heureConvocation": _format_hour(session_data.get("heure_convocation") or session_data.get("heureConvocation"), "08h30"),
        "DateExamen": date_exam,
        "heureExamen": _format_hour(session_data.get("heure_exam") or session_data.get("heureExamen"), "08h00"),
        "Duree": "175 heures",
        "LieuFormation": "Intégrale Academy - 54 chemin du Carreou - 83480 Puget-sur-Argens",
        "=TODAY()": datetime.now().strftime("%d/%m/%Y"),
    }
    # Le modèle contient les deux variantes: [Nom] et ['Nom].
    replacements = {}
    for key, value in ctx.items():
        replacements[f"[{key}]"] = value
        replacements[f"['{key}]"] = value
    return ctx, replacements


def _replace_text_preserving_xml_nodes(xml, replacements):
    def repl_paragraph(match):
        block = match.group(0)
        texts = re.findall(r"<w:t[^>]*>(.*?)</w:t>", block, flags=re.S)
        if not texts:
            return block
        plain = "".join(texts)
        original = plain
        for needle, value in replacements.items():
            plain = plain.replace(needle, str(value))
        if plain == original:
            return block
        escaped = _xml_escape(plain)
        first = True
        def replace_t(tmatch):
            nonlocal first
            if first:
                first = False
                return re.sub(r">.*?</w:t>", f">{escaped}</w:t>", tmatch.group(0), flags=re.S)
            return re.sub(r">.*?</w:t>", "></w:t>", tmatch.group(0), flags=re.S)
        return re.sub(r"<w:t[^>]*>.*?</w:t>", replace_t, block, flags=re.S)
    return re.sub(r"<w:p[\s>].*?</w:p>", repl_paragraph, xml, flags=re.S)


def _drop_empty_conditionals(xml, ctx):
    def conditional_repl(match):
        block = match.group(0)
        key_match = re.search(r"\['?([A-Za-zÀ-ÿ0-9_]+)\].*?\[:if\]", block, flags=re.S)
        if key_match and not str(ctx.get(key_match.group(1), "")).strip():
            return ""
        return block.replace("[:if]", "")
    xml = re.sub(r"<w:tr[\s>].*?\[:if\].*?</w:tr>", conditional_repl, xml, flags=re.S)
    xml = re.sub(r"<w:p[\s>].*?\[:if\].*?</w:p>", conditional_repl, xml, flags=re.S)
    return xml


def _expand_afs_block(xml, replacements):
    formation_line = {
        "[Libelle]": replacements.get("[Libelle]", ""),
        "['Libelle]": replacements.get("[Libelle]", ""),
        "[DateConvocation]": replacements.get("[DateConvocation]", ""),
        "[heureConvocation]": replacements.get("[heureConvocation]", ""),
        "[Duree]": "175 heures",
        "[LieuFormation]": "Intégrale Academy - 54 chemin du Carreou - 83480 Puget-sur-Argens",
    }
    def repl(match):
        inner = match.group(1)
        for needle, value in formation_line.items():
            inner = inner.replace(needle, str(value))
        return inner
    return re.sub(r"\[AFs\](.*?)\[:AFs\]", repl, xml, flags=re.S)


def _render_docx_template(template_path, output_docx_path, ctx, replacements):
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Modèle Word APS introuvable: {template_path}")
    unresolved = set()
    with zipfile.ZipFile(template_path, "r") as zin, zipfile.ZipFile(output_docx_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("word/") and item.filename.endswith(".xml"):
                xml = data.decode("utf-8")
                xml = _drop_empty_conditionals(xml, ctx)
                xml = _expand_afs_block(xml, replacements)
                xml = _replace_text_preserving_xml_nodes(xml, replacements)
                xml = xml.replace("[:if]", "")
                found = {m.group(0) for m in DOCX_UNRESOLVED_PATTERN.finditer(xml)}
                unresolved.update(v for v in found if v not in {"[Content_Types]"})
                data = xml.encode("utf-8")
            zout.writestr(item, data)
    return sorted(unresolved)


def _convert_docx_to_pdf(docx_path, output_dir):
    soffice = shutil.which("libreoffice") or shutil.which("soffice")
    if not soffice:
        raise RuntimeError("LibreOffice/soffice est introuvable sur le serveur; impossible de convertir le DOCX en PDF.")
    result = subprocess.run([soffice, "--headless", "--convert-to", "pdf", "--outdir", output_dir, docx_path], capture_output=True, text=True, timeout=90)
    if result.returncode != 0:
        raise RuntimeError(f"Conversion LibreOffice échouée: {result.stderr or result.stdout}")
    pdf_path = os.path.join(output_dir, os.path.splitext(os.path.basename(docx_path))[0] + ".pdf")
    if not os.path.exists(pdf_path):
        raise RuntimeError("Conversion LibreOffice terminée sans fichier PDF généré.")
    return pdf_path


def generateApsConvocationFromDocxTemplate(trainee, session_data):
    ctx, replacements = _aps_convocation_context(trainee or {}, session_data or {})
    sid = session_data.get("id") or uuid.uuid4().hex
    trainee_id = _trainee_value(trainee or {}, "id") or hashlib.sha1((ctx.get("Nom", "") + ctx.get("Prenom", "") + uuid.uuid4().hex).encode()).hexdigest()[:10]
    base_name = secure_filename(f"convocation_aps_session_{sid}_{trainee_id}")
    docx_path = os.path.join(CONVOCATION_DIR, f"{base_name}.docx")
    final_pdf = os.path.join(CONVOCATION_DIR, f"{base_name}.pdf")
    app.logger.info("Convocation APS: template=%s stagiaire=%s %s session=%s date_convocation=%s date_examen=%s", APS_CONVOCATION_TEMPLATE, ctx.get("Prenom"), ctx.get("Nom"), sid, ctx.get("DateConvocation"), ctx.get("DateExamen"))
    unresolved = _render_docx_template(APS_CONVOCATION_TEMPLATE, docx_path, ctx, replacements)
    if unresolved:
        app.logger.error("Convocation APS: variables non remplacées session=%s stagiaire=%s variables=%s", sid, trainee_id, unresolved)
        try:
            os.remove(docx_path)
        except OSError:
            pass
        raise ValueError("Variables non remplacées dans le modèle Word: " + ", ".join(unresolved))
    generated_pdf = _convert_docx_to_pdf(docx_path, CONVOCATION_DIR)
    if generated_pdf != final_pdf:
        os.replace(generated_pdf, final_pdf)
    app.logger.info("Convocation APS générée: docx=%s pdf=%s", docx_path, final_pdf)
    return {"pdf_url": url_for("view_aps_convocation_pdf", filename=os.path.basename(final_pdf)), "docx_url": url_for("download_aps_convocation_docx", filename=os.path.basename(docx_path)), "pdf_path": final_pdf, "docx_path": docx_path}

def get_planning_for_session(sid):
    data = load_sessions()
    s = find_session(data, sid)
    if not s:
        return None
    return s.get("planning_pdf")  # ex: "planning_session_<sid>.pdf"

def set_planning_for_session(sid, filename):
    data = load_sessions()
    s = find_session(data, sid)
    if not s:
        return False
    s["planning_pdf"] = filename
    s["planning_generated_at"] = datetime.now().strftime("%Y-%m-%d")
    save_sessions(data)
    return True


def refresh_aps_planning_pdf_file(session_data, sid):
    if (session_data.get("formation") or "").upper() != "APS" and not is_ssiap1_session(session_data):
        return session_data.get("planning_pdf")
    planning_data = session_data.get("apsPlanningData") or []
    if not planning_data:
        return session_data.get("planning_pdf")

    filename = f"planning_{'ssiap1' if is_ssiap1_session(session_data) else 'aps'}_session_{sid}.pdf"
    output_path = os.path.join(PLANNING_DIR, filename)
    temp_path = f"{output_path}.tmp"
    planning_mode = session_data.get("apsPlanningMode") or (
        "ssiap1" if is_ssiap1_session(session_data) else
        "elearning_presentiel"
        if any(slot.get("modality") == "elearning" for day in planning_data for slot in day.get("slots", []))
        else "full_presentiel"
    )
    try:
        result = generate_aps_planning_pdf(
            session_data,
            "",
            temp_path,
            planning_data=planning_data,
            planning_mode=planning_mode,
            document_profile={"validate":"ssiap1"} if is_ssiap1_session(session_data) else None,
        )
        os.replace(temp_path, output_path)
        session_data["planning_pdf"] = filename
        session_data["apsPlanningSummary"] = result["summary"]
        session_data["apsPlanningMode"] = planning_mode
        session_data["planning_pdf_refreshed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return filename
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass


def send_planning_pdf_file(path, *, as_attachment, download_name=None):
    response = send_file(
        path,
        mimetype="application/pdf",
        as_attachment=as_attachment,
        download_name=download_name,
        conditional=False,
        max_age=0,
    )
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


FROM_EMAIL = os.environ.get("FROM_EMAIL")
EMAIL_PASSWORD = os.environ.get("EMAIL_PASSWORD")
SMTP_SERVER = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))

BREVO_SMTP_LOGIN = os.environ.get("BREVO_SMTP_LOGIN") or os.environ.get("BREVO_SMTP_USER")
BREVO_SMTP_KEY = os.environ.get("BREVO_SMTP_KEY")
BREVO_SMTP_SERVER = os.environ.get("BREVO_SMTP_SERVER", "smtp-relay.brevo.com")
BREVO_SMTP_PORT = int(os.environ.get("BREVO_SMTP_PORT", "587"))
BREVO_FROM_EMAIL = os.environ.get("BREVO_FROM_EMAIL")
BREVO_SENDER_EMAIL = os.environ.get("BREVO_SENDER_EMAIL")
BREVO_SENDER_NAME = os.environ.get("BREVO_SENDER_NAME")
BREVO_API_KEY = os.environ.get("BREVO_API_KEY")
BREVO_SMS_SENDER = os.environ.get("BREVO_SMS_SENDER")

# -----------------------
# Utils persistance
# -----------------------


A3P_TRAINER_MANUAL_CODES = {"UV1", "UV5", "UV6A", "UV9"}
A3P_TRAINER_MODULE_LABELS = {m["code"]: m for m in A3P_MODULES if m["code"] in A3P_TRAINER_MANUAL_CODES}
A3P_TRAINER_STATUS_LABELS = {
    "no_link": "Lien non généré",
    "waiting": "En attente formateur",
    "sent": "Lien envoyé",
    "in_progress": "En cours de complétion",
    "incomplete": "Incomplet",
    "completed": "Modules complétés",
    "validated": "Validé",
    "disabled": "Lien désactivé",
}

def a3p_trainer_public_url(token):
    return url_for("public_a3p_planning_page", token=token, _external=True)

def a3p_trainer_status(session_data):
    token = session_data.get("a3pTrainerPublicToken")
    status = session_data.get("a3pTrainerModulesStatus") or ("waiting" if token else "no_link")
    return {"code": status, "label": A3P_TRAINER_STATUS_LABELS.get(status, status), "url": a3p_trainer_public_url(token) if token and status != "disabled" else ""}

def find_a3p_public_session(data, token):
    for s in data.get("sessions", []):
        if (s.get("formation") or "").upper() == "A3P" and s.get("a3pTrainerPublicToken") == token and s.get("a3pTrainerModulesStatus") != "disabled":
            return s
    return None

def validate_a3p_trainer_manual_data(session_data, modules_data):
    cfg = session_data.get("a3pPlanningDraftJson") or {}
    start = cfg.get("startDate") or session_data.get("date_debut")
    end = cfg.get("endDate") or session_data.get("date_fin")
    exam = cfg.get("examDate") or session_data.get("date_exam")
    errors = []
    by_day = {}
    for code in A3P_TRAINER_MANUAL_CODES:
        expected = int(A3P_TRAINER_MODULE_LABELS[code]["hours"] * 60)
        rows = (modules_data or {}).get(code) or []
        total = 0
        for row in rows:
            d, st, en = row.get("date"), row.get("start"), row.get("end")
            if not d or not st or not en:
                errors.append(f"{A3P_TRAINER_MODULE_LABELS[code]['title']} : date et horaires obligatoires."); continue
            if (start and d < start) or (end and d > end): errors.append(f"{A3P_TRAINER_MODULE_LABELS[code]['title']} : {format_date(d)} hors période de formation.")
            try:
                parsed_day = datetime.strptime(d, "%Y-%m-%d").date()
                if is_a3p_non_working_day(parsed_day): errors.append(f"{A3P_TRAINER_MODULE_LABELS[code]['title']} : {format_date(d)} est un jour non travaillé (week-end ou jour férié français).")
            except Exception:
                errors.append(f"{A3P_TRAINER_MODULE_LABELS[code]['title']} : date invalide.")
            if exam and d == exam: errors.append(f"{A3P_TRAINER_MODULE_LABELS[code]['title']} : impossible le jour de l’examen.")
            try:
                sm = int(st[:2])*60 + int(st[3:5]); em = int(en[:2])*60 + int(en[3:5])
            except Exception:
                errors.append(f"{A3P_TRAINER_MODULE_LABELS[code]['title']} : horaires invalides."); continue
            if em <= sm: errors.append(f"{A3P_TRAINER_MODULE_LABELS[code]['title']} : heure de fin avant début."); continue
            total += em - sm
            by_day.setdefault(d, []).append((sm, em, code))
        if total != expected:
            errors.append(f"{A3P_TRAINER_MODULE_LABELS[code]['title']} : {round(total/60,2)}h saisies / {expected/60:g}h attendues.")
    for d, slots in by_day.items():
        if sum(e-s for s,e,_ in slots) > 480: errors.append(f"{format_date(d)} : plus de 8h de formation.")
        slots = sorted(slots)
        for prev, cur in zip(slots, slots[1:]):
            if cur[0] < prev[1]: errors.append(f"{format_date(d)} : chevauchement entre modules imposés.")
    return errors



def _a3p_manual_modules_from_state(state):
    if not isinstance(state, dict):
        return {}
    cfg = state.get("scheduleConfig") if isinstance(state.get("scheduleConfig"), dict) else state
    return cfg.get("lockedModules") or state.get("lockedModules") or {}

def are_a3p_manual_modules_complete(state):
    modules_data = _a3p_manual_modules_from_state(state)
    for code in A3P_TRAINER_MANUAL_CODES:
        expected = int(A3P_TRAINER_MODULE_LABELS[code]["hours"] * 60)
        total = 0
        for row in modules_data.get(code) or []:
            if row.get("durationMinutes") is not None:
                try:
                    total += int(float(row.get("durationMinutes") or 0))
                    continue
                except (TypeError, ValueError):
                    pass
            try:
                st, en = row.get("start"), row.get("end")
                if not st or not en:
                    continue
                total += (int(en[:2]) * 60 + int(en[3:5])) - (int(st[:2]) * 60 + int(st[3:5]))
            except Exception:
                continue
        if total != expected:
            return False
    return True

def _a3p_planning_total_minutes(planning):
    if not isinstance(planning, list):
        return 0
    return sum(int(float(slot.get("durationMinutes") or 0)) for day in planning for slot in (day.get("slots") or []))

def can_generate_a3p_documents_state(state):
    planning = state.get("planning") or state.get("preview") or state.get("generatedPlanning") or [] if isinstance(state, dict) else []
    total_minutes = _a3p_planning_total_minutes(planning)
    total_hours = total_minutes / 60
    remaining_hours = A3P_TOTAL_HOURS - total_hours
    return (
        are_a3p_manual_modules_complete(state)
        and round(total_minutes) == A3P_TOTAL_HOURS * 60
        and round(remaining_hours * 60) == 0
        and bool(planning)
    )

def mark_a3p_manual_modules_admin_validated(session_data, modules_data=None):
    now = datetime.now().isoformat()
    session_data["manual_modules_source"] = "admin"
    session_data["manual_modules_completed"] = True
    session_data["manual_modules_validated"] = True
    session_data["manual_modules_validated_at"] = now
    session_data["a3pTrainerModulesStatus"] = "validated"
    session_data["a3pTrainerModulesValidatedAt"] = now
    if modules_data is not None:
        session_data["a3pTrainerManualModulesData"] = modules_data

def a3p_public_payload(session_data):
    cfg = session_data.get("a3pPlanningDraftJson") or {}
    return {
        "sessionId": session_data.get("id"), "formation": "A3P",
        "startDate": cfg.get("startDate") or session_data.get("date_debut"),
        "endDate": cfg.get("endDate") or session_data.get("date_fin"),
        "examDate": cfg.get("examDate") or session_data.get("date_exam"),
        "room": cfg.get("room") or session_data.get("a3pRoom") or session_data.get("salle") or "",
        "trainerName": session_data.get("a3pTrainerName") or (cfg.get("trainerFirstName", "") + " " + cfg.get("trainerLastName", "")).strip(),
        "showWeekends": bool(cfg.get("showWeekends")),
        "days": cfg.get("days") or [],
        "modules": [{"code": c, "title": A3P_TRAINER_MODULE_LABELS[c]["title"], "hours": A3P_TRAINER_MODULE_LABELS[c]["hours"]} for c in ("UV1","UV5","UV6A","UV9")],
        "modulesData": session_data.get("a3pTrainerManualModulesData") or {},
        "status": a3p_trainer_status(session_data),
    }

def load_sessions():
    if os.path.exists(SESSIONS_FILE):
        try:
            with open(SESSIONS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    data.setdefault("sessions", [])
                    data.setdefault("jurys", [])
                    return data
        except Exception:
            pass
    return {"sessions": [], "jurys": []}

def save_sessions(data):
    tmp_path = SESSIONS_FILE + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp_path, SESSIONS_FILE)

def load_price_adaptator_data():
    if os.path.exists(PRICE_ADAPTATOR_FILE):
        try:
            with open(PRICE_ADAPTATOR_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    data.setdefault("prospects", [])
                    data.setdefault("dates", {})
                    return data
        except Exception:
            pass
    return {"prospects": [], "dates": {}}

def save_price_adaptator_data(data):
    with open(PRICE_ADAPTATOR_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def normalize_price_adaptator_discount(value):
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return PRICE_ADAPTATOR_DEFAULT_DISCOUNT
    return max(0, min(parsed, 100))

def normalize_price_adaptator_nom(value):
    if value is None:
        return ""
    return str(value).strip().upper()


def normalize_price_adaptator_prenom(value):
    if value is None:
        return ""
    cleaned = " ".join(str(value).strip().split())
    return cleaned.lower().title()


def normalize_price_adaptator_formation(value):
    if value is None:
        return None
    cleaned = str(value).strip().upper()
    return PRICE_ADAPTATOR_ALLOWED_FORMATIONS.get(cleaned)


def normalize_price_adaptator_proposed_price(price_value):
    try:
        return max(float(price_value), 0.0)
    except (TypeError, ValueError):
        return 0.0

def get_price_adaptator_followup_date(dates, formation):
    date_range = (dates or {}).get(formation, {})
    start_value = (date_range or {}).get("start")
    if not start_value:
        return None
    try:
        start_date = datetime.strptime(start_value, "%Y-%m-%d").date()
    except ValueError:
        return None
    return start_date - timedelta(days=PRICE_ADAPTATOR_FOLLOWUP_DAYS)

def format_price_adaptator_date_range(date_range):
    if not date_range:
        return "dates à définir"
    start_value = date_range.get("start")
    end_value = date_range.get("end")
    try:
        start_label = datetime.strptime(start_value, "%Y-%m-%d").strftime("%d/%m/%Y") if start_value else None
    except ValueError:
        start_label = None
    try:
        end_label = datetime.strptime(end_value, "%Y-%m-%d").strftime("%d/%m/%Y") if end_value else None
    except ValueError:
        end_label = None
    if start_label and end_label:
        return f"{start_label} au {end_label}"
    if start_label:
        return start_label
    return "dates à définir"

def build_price_adaptator_message(prospect, dates, price_override=None):
    formation = prospect.get("formation", "")
    formation_full = PRICE_ADAPTATOR_FORMATION_LABELS.get(formation, formation)
    base_price = PRICE_ADAPTATOR_FORMATION_PRICES.get(formation, 0)
    discount_value = normalize_price_adaptator_discount((dates or {}).get(formation, {}).get("discount"))
    discounted_price = round(base_price * (1 - discount_value / 100))
    price_value = price_override if price_override is not None else discounted_price
    if base_price and price_value is not None:
        computed_discount = round((1 - price_value / base_price) * 100)
        discount_value = max(0, min(computed_discount, 100))
    price_label = f"{price_value:,.0f} €".replace(",", " ")
    base_price_label = f"{base_price:,.0f} €".replace(",", " ") if base_price else None
    date_text = format_price_adaptator_date_range((dates or {}).get(formation))
    prenom = normalize_price_adaptator_prenom(prospect.get("prenom"))
    logo_path = os.path.join("static", "img", "logo-integrale.png")
    logo_src = url_for("static", filename="img/logo-integrale.png", _external=True)
    try:
        with open(logo_path, "rb") as logo_file:
            logo_src = "data:image/png;base64," + base64.b64encode(logo_file.read()).decode("utf-8")
    except OSError:
        pass
    html = f"""
    <div style="font-family:'Segoe UI',Arial,Helvetica,sans-serif;background:#f2f4f7;padding:24px;">
      <style>
        @media screen and (max-width: 600px) {{
          .stack-column {{
            display: block !important;
            width: 100% !important;
            box-sizing: border-box !important;
          }}
          .stack-column-right {{
            text-align: left !important;
            padding-top: 0 !important;
          }}
          .email-container {{
            max-width: 100% !important;
            width: 100% !important;
          }}
          .email-body {{
            padding-left: 20px !important;
            padding-right: 20px !important;
          }}
        }}
      </style>
      <table role="presentation" cellspacing="0" cellpadding="0" class="email-container" style="width:100%;max-width:620px;margin:0 auto;background:#ffffff;border-radius:16px;overflow:hidden;border:1px solid #e6e9ef;box-shadow:0 12px 30px rgba(16,24,40,0.08);">
        <tr>
          <td style="background:linear-gradient(135deg,#111827,#1f2937);padding:24px;text-align:center;">
            <img src="{logo_src}" alt="Intégrale Academy" style="max-width:150px;height:auto;">
            <div style="margin-top:12px;color:#e5e7eb;font-size:14px;letter-spacing:0.4px;text-transform:uppercase;">Offre dernière minute</div>
          </td>
        </tr>
        <tr>
          <td class="email-body" style="padding:28px 32px 10px;color:#1f2937;line-height:1.7;">
            <p style="margin:0 0 16px;font-size:16px;">Bonjour {prenom},</p>
            <p style="margin:0 0 16px;font-size:16px;">
              Je me permets de revenir vers vous concernant notre formation <strong>{formation_full}</strong>.
            </p>
            <p style="margin:0 0 16px;font-size:16px;">
              Bonne nouvelle : Suite à des désistements, nous pouvons vous proposer un tarif exceptionnel de dernière
              minute à <strong>{price_label}</strong> au lieu de <strong>{base_price_label or "prix initial"}</strong> (prix initial de
              la formation), soit une remise de <strong>{discount_value:.0f} %</strong> pour notre prochaine session
              qui se déroulera du <strong>{date_text}</strong>.
            </p>
            <table role="presentation" cellspacing="0" cellpadding="0" style="width:100%;background:#f9fafb;border-radius:12px;border:1px solid #eef2f6;margin:16px 0;">
              <tr>
                <td class="stack-column" style="padding:18px 20px;">
                  <div style="font-size:14px;text-transform:uppercase;letter-spacing:0.8px;color:#6b7280;margin-bottom:8px;">Tarif exceptionnel</div>
                  <div style="font-size:28px;font-weight:700;color:#111827;">{price_label}</div>
                  <div style="font-size:14px;color:#6b7280;margin-top:4px;">
                    {f"Au lieu de {base_price_label} • remise de {discount_value:.0f} %" if base_price_label else "Offre dernière minute limitée"}
                  </div>
                </td>
                <td class="stack-column stack-column-right" style="padding:18px 20px;text-align:right;">
                  <div style="font-size:13px;color:#6b7280;text-transform:uppercase;letter-spacing:0.6px;">Prochaine session</div>
                  <div style="font-size:16px;font-weight:600;color:#111827;">{date_text}</div>
                </td>
              </tr>
            </table>
            <p style="margin:0 0 16px;font-size:16px;">
              Pour bénéficier de ce tarif et pour vous inscrire, nous vous remercions de bien vouloir nous contacter au
              <strong>04 22 47 07 68</strong>.
            </p>
            <p style="margin:0 0 20px;font-size:16px;">Cette offre est limitée, profitez-en dès maintenant.</p>
            <p style="margin:0 0 8px;font-size:16px;">
              Je reste à votre disposition pour tous renseignements complémentaires et je vous souhaite une bonne journée,
            </p>
            <p style="margin:0 0 0;font-size:16px;">A très bientôt !</p>
          </td>
        </tr>
        <tr>
          <td style="padding:0 32px 28px;">
            <table role="presentation" cellspacing="0" cellpadding="0" style="width:100%;background:#111827;border-radius:12px;">
              <tr>
                <td style="padding:16px 20px;color:#ffffff;font-size:15px;">
                  <div style="font-weight:600;">Clément VAILLANT</div>
                  <div style="font-size:13px;color:#d1d5db;">Intégrale Academy</div>
                </td>
                <td style="padding:16px 20px;text-align:right;">
                  <span style="display:inline-block;background:#f9fafb;color:#111827;font-weight:600;padding:10px 16px;border-radius:999px;font-size:13px;">
                    04 22 47 07 68
                  </span>
                </td>
              </tr>
            </table>
          </td>
        </tr>
      </table>
    </div>
    """
    subject = f"Proposition tarif dernière minute {formation_full}"
    sms_message = (
        f"Bonjour {prenom}, Tarif exceptionnel dernière minute à {price_label} pour la formation "
        f"{formation_full} (du {date_text}). Offre limitée: contactez-nous au 04 22 47 07 68. "
        "Cordialement, Clément VAILLANT - Intégrale Academy"
    )
    return {
        "subject": subject,
        "html": html,
        "sms": sms_message,
        "price": price_value,
        "base_price": base_price,
        "discount_value": discount_value,
        "date_text": date_text,
    }

def attempt_price_adaptator_send(prospect, dates, price_override=None):
    message = build_price_adaptator_message(prospect, dates, price_override=price_override)
    email_sent = False
    email_error = None
    sms_sent = False
    sms_error = None
    email = (prospect.get("email") or "").strip()
    phone = (prospect.get("telephone") or "").strip()
    if email:
        email_sent, email_error = send_price_adaptator_email(email, message["subject"], message["html"])
    if phone:
        sms_sent, sms_error = send_price_adaptator_sms(phone, message["sms"])
    return {
        "email_sent": email_sent,
        "sms_sent": sms_sent,
        "email_error": email_error,
        "sms_error": sms_error,
        "price": message["price"],
    }

def process_price_adaptator_followups():
    data = load_price_adaptator_data()
    today = datetime.now().date()
    updated = False
    for prospect in data.get("prospects", []):
        if prospect.get("sent") or prospect.get("manual_sent"):
            continue
        followup_date = get_price_adaptator_followup_date(data.get("dates"), prospect.get("formation"))
        if not followup_date or followup_date > today:
            continue
        price_override = prospect.get("proposed_price")
        with app.app_context():
            result = attempt_price_adaptator_send(prospect, data.get("dates"), price_override=price_override)
        prospect["last_attempt_at"] = datetime.now().isoformat()
        prospect["last_error"] = result["email_error"] or result["sms_error"]
        prospect["proposed_price"] = result["price"]
        if result["email_sent"] or result["sms_sent"]:
            prospect["sent"] = True
            prospect["sentAt"] = datetime.now().isoformat()
            prospect["last_sent_price"] = result["price"]
        updated = True
    if updated:
        save_price_adaptator_data(data)

def price_adaptator_scheduler_loop():
    while True:
        try:
            process_price_adaptator_followups()
        except Exception as exc:
            logging.exception("[price-adaptator] Scheduler error: %s", exc)
        time.sleep(60 * 30)

def start_price_adaptator_scheduler():
    if os.environ.get("ENABLE_PRICE_ADAPTATOR_AUTOSEND", "").lower() != "true":
        return
    if app.debug and os.environ.get("WERKZEUG_RUN_MAIN") != "true":
        return
    thread = threading.Thread(target=price_adaptator_scheduler_loop, daemon=True)
    thread.start()

def ensure_jury_defaults(session):
    session.setdefault("jurys", [])
    session.setdefault("jury_notification_status", "to_notify")
    for jury in session["jurys"]:
        jury.setdefault("id", str(uuid.uuid4())[:8])
        jury.setdefault("status", "pending")
        jury.setdefault("token", str(uuid.uuid4()))
        jury.setdefault("notified_at", None)
        jury.setdefault("reminded_at", None)

def ensure_global_jury_defaults(data):
    data.setdefault("jurys", [])
    for jury in data["jurys"]:
        jury.setdefault("id", str(uuid.uuid4())[:8])
        jury.setdefault("nom", "")
        jury.setdefault("prenom", "")
        jury.setdefault("email", "")
        jury.setdefault("telephone", "")

def find_global_jury_by_email(data, email):
    if not email:
        return None
    normalized = email.strip().lower()
    return next((j for j in data.get("jurys", []) if j.get("email", "").strip().lower() == normalized), None)

def find_global_jury_by_id(data, jury_id):
    return next((j for j in data.get("jurys", []) if j.get("id") == jury_id), None)

def sync_global_jurys(data):
    ensure_global_jury_defaults(data)
    for session in data.get("sessions", []):
        for jury in session.get("jurys", []):
            if not jury.get("email") and not jury.get("id"):
                continue
            existing = find_global_jury_by_email(data, jury.get("email"))
            if existing:
                existing.update({
                    "nom": jury.get("nom", existing.get("nom", "")),
                    "prenom": jury.get("prenom", existing.get("prenom", "")),
                    "email": jury.get("email", existing.get("email", "")),
                    "telephone": jury.get("telephone", existing.get("telephone", "")),
                })
                if not existing.get("id") and jury.get("id"):
                    existing["id"] = jury.get("id")
            elif find_global_jury_by_id(data, jury.get("id")) is None:
                data["jurys"].append({
                    "id": jury.get("id") or str(uuid.uuid4())[:8],
                    "nom": jury.get("nom", ""),
                    "prenom": jury.get("prenom", ""),
                    "email": jury.get("email", ""),
                    "telephone": jury.get("telephone", ""),
                })

def find_session(data, sid):
    for s in data["sessions"]:
        if s["id"] == sid:
            return s
    return None

def steps_rules_for_formation(formation):
    if formation in ("APS", "A3P", "DIRIGEANT", "AFC_APS_SSIAP"):
        rules = APS_A3P_STEPS
    elif formation == "SSIAP":
        rules = SSIAP_STEPS
    elif formation == "GENERAL":
        rules = GENERAL_STEPS
    else:
        return []

    return [rule for rule in rules if "formations" not in rule or formation in rule["formations"]]


def sync_steps(session):
    """Reconstruit les étapes selon le modèle officiel (ordre + ajout + évite doublons),
    tout en conservant done/done_at/custom_date des étapes existantes.
    """
    formation = session.get("formation")

    rules = steps_rules_for_formation(formation)
    if not rules:
        return

    # sécurité si steps absent
    session.setdefault("steps", [])

    # index existant par nom
    existing_by_name = {s.get("name"): s for s in session["steps"] if s.get("name")}

    new_steps = []
    for rule in rules:
        name = rule["name"]
        old = existing_by_name.get(name)

        # ✅ on conserve l'état existant si présent
        new_steps.append({
            "name": name,
            "done": bool(old.get("done")) if old else False,
            "done_at": old.get("done_at") if old else None,
            "custom_date": old.get("custom_date") if old else None
        })

    session["steps"] = new_steps


    # Récupère la liste actuelle des règles depuis le code
    rules = steps_rules_for_formation(formation)
    if not rules:
        return
    existing_names = [s["name"] for s in session.get("steps", [])]

    # Pour chaque étape officielle, si elle n’existe pas encore dans la session → on l’ajoute
    for rule in rules:
        if rule["name"] not in existing_names:
            session["steps"].append({
                "name": rule["name"],
                "done": False,
                "done_at": None
            })


# -----------------------
# Modèles d'étapes
# -----------------------
APS_A3P_STEPS = [
    {"name":"Création session CNAPS", "relative_to":"start", "offset_type":"before", "days":20},
    {"name":"Création session ADEF", "relative_to":"start", "offset_type":"before", "days":15},
    {"name":"Envoyer test de français", "relative_to":"start", "offset_type":"before", "days":10},
    {"name":"Nomination jury examen", "relative_to":"start", "offset_type":"before", "days":10},
    {"name":"Planification YPAREO", "relative_to":"start", "offset_type":"before", "days":10},
    {"name":"Envoyer lien à compléter stagiaires", "relative_to":"start", "offset_type":"before", "days":10},
    {"name":"Ajout des stagiaires sur DRACAR", "relative_to":"start", "offset_type":"before", "days":7},
    {"name":"Ajout des formateurs sur DRACAR", "relative_to":"start", "offset_type":"before", "days":7},
    {"name":"Contrat envoyé au formateur", "relative_to":"start", "offset_type":"before", "days":5},
    {"name":"Contrat formateur imprimé", "relative_to":"start", "offset_type":"before", "days":5},
    {"name":"Saisie des candidats ADEF", "relative_to":"start", "offset_type":"before", "days":5},
    {"name":"Impression des fiches CNIL", "relative_to":"start", "offset_type":"before", "days":5},
    {"name":"Fabrication badge formateur", "relative_to":"start", "offset_type":"before", "days":5},
    {"name":"Vérification dossier formateur", "relative_to":"start", "offset_type":"before", "days":5},
    {"name":"Corriger et imprimer test de français", "relative_to":"start", "offset_type":"before", "days":5},
    {"name":"Validation session ADEF", "relative_to":"start", "offset_type":"before", "days":2},
    # AVANT EXAM
    {"name":"Saisie des SST", "relative_to":"exam", "offset_type":"before", "days":7},
    {"name":"Impression des SST", "relative_to":"exam", "offset_type":"before", "days":5},
    {"name":"Impression dossier fin de formation", "relative_to":"exam", "offset_type":"before", "days":5},
    {"name":"Impression évaluation de fin de formation", "relative_to":"exam", "offset_type":"before", "days":5},
    # JOUR EXAM
    {"name":"Session examen clôturée", "relative_to":"exam", "offset_type":"after", "days":0},
    {"name":"Frais ADEF réglés", "relative_to":"exam", "offset_type":"after", "days":0},
    {"name":"Documents examen envoyés à l’ADEF", "relative_to":"exam", "offset_type":"after", "days":0},
    # APRÈS EXAM
    {"name":"Envoyer mail stagiaires attestations de formation", "relative_to":"exam", "offset_type":"after", "days":2},
    {"name":"Message avis Google", "relative_to":"exam", "offset_type":"after", "days":2},
    {"name":"Diplômes reçus", "relative_to":"exam", "offset_type":"after", "days":7},
    {"name":"Diplômes envoyés aux stagiaires", "relative_to":"exam", "offset_type":"after", "days":10},
    {"name": "Saisie entrée en formation EDOF", "relative_to": "start", "offset_type": "after", "days": 0},
    {"name":"Imprimer feuilles de présence et planning", "relative_to":"start", "offset_type":"before", "days":2},
    {"name":"Documents examens imprimés", "relative_to":"exam", "offset_type":"before", "days":1},
    {"name": "Signature fiches CNIL", "relative_to": "start", "offset_type": "after", "days": 0},
    {"name":"Fin de formation EDOF", "relative_to":"exam", "offset_type":"after", "days":1},
    {"name": "Signature registre entretien SST", "relative_to": "start", "offset_type": "after", "days": 15},
    {"name": "Distribution des t-shirts", "relative_to": "start", "offset_type": "after", "days": 1, "formations": ["A3P"]},
    {"name": "Récupérer paiement logement", "relative_to": "start", "offset_type": "after", "days": 0, "formations": ["A3P"]},
    {"name":"Préparation planning de ménage", "relative_to":"start", "offset_type":"before", "days":2, "formations": ["A3P"]},
    {"name":"Créer groupe Whatsapp", "relative_to":"start", "offset_type":"before", "days":7},
]

SSIAP_STEPS = [

    # ============================
    # 📌 SESSION — Article 4
    # ============================
    {"name": "Le formateur a été nommé (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},
    {"name": "Le contrat d'intervention a été envoyé au formateur (7 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 7},
    {"name": "Le contrat d'intervention formateur a été signé et imprimé (5 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 5},
    {"name": "Le nombre de candidats est de 12 maximum (2 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 2},
    {"name": "La préfecture a été avisée de l'ouverture de la session 2 mois avant le démarrage (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},
    {"name": "La préfecture a été avisée de la date d'examen 2 mois avant le démarrage (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},
    {"name": "Les convocations en formation ont été envoyées aux candidats (15 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 15},
    {"name": "Le test de français a été envoyé à tous les candidats (7 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 7},

    # =======================================
    # 📌 DOSSIER CANDIDAT (formation)
    # =======================================
    {"name": "Le dossier comporte la pièce d'identité de chaque candidat (1er jour de formation)", "relative_to": "start", "offset_type": "after", "days": 0},
    {"name": "Le dossier comporte l'attestation de formation au secourisme de chaque candidat (1er jour de formation)", "relative_to": "start", "offset_type": "after", "days": 0},
    {"name": "Le dossier comporte 2 photos d'identité (1 archive, 1 diplôme) pour chaque candidat (1er jour de formation)", "relative_to": "start", "offset_type": "after", "days": 0},
    {"name": "Le dossier comporte le certificat médical conforme à l'Annexe VII de l'arrêté du 2 mai 2005 modifié de chaque candidat (1er jour de formation)", "relative_to": "start", "offset_type": "after", "days": 0},
    {"name": "Le dossier comporte une copie du test de français réalisé par chaque candidat en amont de la formation (1er jour de formation)", "relative_to": "start", "offset_type": "after", "days": 0},
    {"name": "Le dossier comporte le contrat de formation signé par chaque candidat (1er jour de formation)", "relative_to": "start", "offset_type": "after", "days": 0},
    {"name": "Les dossiers de chaque candidat ont été vérifiés avant le démarrage de la session (1er jour de formation)", "relative_to": "start", "offset_type": "after", "days": 0},

    # =======================================
    # 📌 DEMANDE PRÉSIDENCE JURY SDIS (Art 8)
    # =======================================
    {"name": "Le SDIS a été avisé de la date d'organisation des épreuves (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},
    {"name": "La demande comporte le nom, la fonction et la qualification du jury chef de service incendie (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},
    {"name": "La demande comporte l'attestation d'engagement (accord) du jury chef de service incendie (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},
    {"name": "L'engagement écrit, du propriétaire ou de l'exploitant de l'établissement, de mettre à disposition les locaux et d'autoriser la manipulation des installations techniques nécessaires au déroulement de l'épreuve pratique est fournit (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},
    {"name": "Le planning de la session est fournit (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},
    {"name": "Sur le planning le nom, la qualité, la fonction et les qualifications des formateurs sont indiqués (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},
    {"name": "La convention de demande de présidence jury SDIS en fournit en double exemplaire (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},
    {"name": "La demande de présidence de jury SDIS a été envoyé en LRAR (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},

    # =======================================
    # 📌 DOSSIER CANDIDAT (examen)
    # =======================================
    {"name": "Les dossiers examen des candidats sont imprimés pour les membres du jury (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},
    {"name": "Chaque dossier examen comporte la pièce d'identité du candidat (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},
    {"name": "Chaque dossier examen comporte l'attestation de formation au secourisme (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},
    {"name": "Chaque dossier examen comporte le certificat médical conforme à l'Annexe VII de l'arrêté du 2 mai 2005 modifié (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},
    {"name": "Chaque dossier examen comporte le test de français réalisé par le candidat en amont de la formation (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},
    {"name": "Chaque dossier examen comporte le certificat de réalisation de la formation (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},
    {"name": "Chaque dossier examen comporte le PV d'examen individuel pré-rempli (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},
    {"name": "Chaque dossier examen comporte une attestation du directeur certifiant que les candidats ne travaillent pas dans la même entreprise que le jury (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},
    {"name": "Chaque dossier examen comporte une attestation du directeur certifiant que les candidats sont capables d'écrire une main courante (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},

    # =======================================
    # 📌 ORGANISATION DE L’EXAMEN
    # =======================================
    {"name": "Le jury chef de service de sécurité incendie a été nommé (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},
    {"name": "Le lieu d'examen (pratique) a été réservé (65 jours avant début de session)", "relative_to": "start", "offset_type": "before", "days": 65},
    {"name": "Les convocations à l'examen ont été envoyées aux candidats (15 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 15},
    {"name": "Les télécommandes QUIZZBOX ont été vérifiées en vue de l'examen (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},
    {"name": "Le logiciel QUIZZBOX a été paramétré pour l'examen (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},
    {"name": "Le procès verbal collectif a été pré-rempli et imprimé (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},
    {"name": "La salle d'examen théorique a été préparée et vérifiée (2 jours avant l'examen)", "relative_to": "exam", "offset_type": "before", "days": 2},
    {"name": "Les pièces d'identité des candidats ont été vérifié par le jury (jour de l'examen)", "relative_to": "exam", "offset_type": "after", "days": 0},
    {"name": "Le PV de résultats examen théorique (QCM Quizzbox) a été imprimé en double exemplaire : SDIS et archives (jour de l'examen)", "relative_to": "exam", "offset_type": "after", "days": 0},
    {"name": "A l'issue de l'examen les PV d'examen individuels ont été photocopiés en triples exemplaires : SDIS, candidats et archives (jour de l'examen)", "relative_to": "exam", "offset_type": "after", "days": 0},
    {"name": "A l'issue de l'examen le PV d'examen collectif a été photocopié en doubles exemplaires : SDIS et archives (jour de l'examen)", "relative_to": "exam", "offset_type": "after", "days": 0},

    # =======================================
    # 📌 DIPLÔMES — Annexe VIII / Article 11
    # =======================================
    {"name": "Chaque diplôme comporte une photographie couleur dans l'angle droit (2 jours après l'examen)", "relative_to": "exam", "offset_type": "after", "days": 2},
    {"name": "Les numéros de diplômes ont été vérifiés (2 jours après l'examen)", "relative_to": "exam", "offset_type": "after", "days": 2},
    {"name": "La signature du directeur du centre de formation agréé est apposée dans l'angle inférieur gauche (2 jours après l'examen)", "relative_to": "exam", "offset_type": "after", "days": 2},
    {"name": "Les diplômes ont été imprimé sur du papier rigide 180g (2 jours après l'examen)", "relative_to": "exam", "offset_type": "after", "days": 2},
    {"name": "Les diplômes ont été envoyés au SDIS en LRAR (2 jours après l'examen)", "relative_to": "exam", "offset_type": "after", "days": 2},
    {"name": "Les diplômes ont été validé par le SDIS (30 jours après l'examen)", "relative_to": "exam", "offset_type": "after", "days": 30},
    {"name": "Les diplômes ont été distribués aux candidats (35 jours après l'examen)", "relative_to": "exam", "offset_type": "after", "days": 35},
    {"name": "Les candidats ont signé le récépissé de délivrance, preuve de remise du diplôme (35 jours après l'examen)", "relative_to": "exam", "offset_type": "after", "days": 35},
    {"name": "Les diplômes sont référencés dans un tableau Excel pour assurer la traçabilité (2 jours après l'examen)", "relative_to": "exam", "offset_type": "after", "days": 2},

    # =======================================
    # 📌 CLÔTURE DE SESSION
    # =======================================
    {"name": "Le rapport de traçabilité et de conformité a été généré (40 jours après l'examen)", "relative_to": "exam", "offset_type": "after", "days": 40},
    {"name": "Le rapport de traçabilité et de conformité a été envoyé par mail à la préfecture (40 jours après l'examen)", "relative_to": "exam", "offset_type": "after", "days": 40},
    {"name": "Le rapport de traçabilité et de conformité a été imprimé et archivé (40 jours après l'examen)", "relative_to": "exam", "offset_type": "after", "days": 40},
]


GENERAL_STEPS = [
    {"name": "Vérification des extincteurs", "fixed_date": "2026-10-15"},
    {"name": "Contrôle des installations électriques", "fixed_date": "2026-09-10"},
    {"name": "Vérification SSI", "fixed_date": "2026-08-15"},
    {"name": "Contrôle climatisation", "fixed_date": "2026-09-10"},
    {"name": "Inscriptions examen BTS", "fixed_date": "2026-09-10"},
    {"name": "Renouvellement agrément CNAPS", "fixed_date": "2026-09-01"},
]


FORMATION_COLORS = {
    "AFC_APS_SSIAP": "#0f766e",
    "APS": "#1b9aaa",
    "A3P": "#2a9134",
    "SSIAP": "#c0392b",
    "DIRIGEANT": "#8e44ad",
    "GENERAL": "#d4ac0d",
}

FORMATION_LABELS = {
    "AFC_APS_SSIAP": "AFC France Travail APS + SSIAP",
    "APS": "Agent de Prévention et de Sécurité",
    "A3P": "Agent de Protection Rapprochée (A3P)",
    "SSIAP": "Service de Sécurité Incendie et d’Assistance à Personnes (SSIAP)",
    "DIRIGEANT": "Dirigeant",
    "GENERAL": "Général",
}

DIRIGEANT_LOCATIONS = {
    "PARIS": "Paris",
    "PUGET": "Puget",
}

def formation_label(value):
    return FORMATION_LABELS.get(value, value)
app.jinja_env.filters['formation_label'] = formation_label

def default_steps_for(formation):
    steps = steps_rules_for_formation(formation)
    return [{"name": s["name"], "done": False, "done_at": None} for s in steps]


# -----------------------
# Statuts / échéances
# -----------------------
def _rule_for(formation, step_index):
    rules = steps_rules_for_formation(formation)

    # ✅ Protection anti IndexError
    if step_index < 0 or step_index >= len(rules):
        return None

    return rules[step_index]



def parse_date(date_str):
    """Accepte les formats AAAA-MM-JJ ou JJ/MM/AAAA"""
    if not date_str:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None

def deadline_for(step_index, session):
    rule = _rule_for(session["formation"], step_index)

    # ✅ Si cette étape a une date personnalisée enregistrée dans la session → priorité
    custom_date = session["steps"][step_index].get("custom_date")
    if custom_date:
        return parse_date(custom_date)

    if not rule:
        return None

    # ✅ Si l’étape a une date fixe → on la renvoie directement
    if "fixed_date" in rule and rule["fixed_date"]:
        return parse_date(rule["fixed_date"])

    # Sinon on garde le comportement classique (start/exam + offset)
    base_date = parse_date(session.get("date_exam")) if rule["relative_to"] == "exam" else parse_date(session.get("date_debut"))
    if not base_date:
        return None
    return (base_date - timedelta(days=rule["days"])) if rule["offset_type"] == "before" else (base_date + timedelta(days=rule["days"]))



def status_for_step(step_index, session, now=None):
    if now is None:
        now = datetime.now()
    dl = deadline_for(step_index, session)
    if dl is None:
        return ("n/a", None)
    step = session["steps"][step_index]
    if step["done"]:
        return ("done", dl)

    # --- 🔧 Correction : tolérance réelle sur les 24 h ---
    diff_days = (dl.date() - now.date()).days
    if diff_days < 0:
        return ("late", dl)
    elif diff_days == 0:
        return ("on_time", dl)
    elif diff_days == 1:
        return ("upcoming", dl)  # échéance demain → "à venir"
    else:
        return ("on_time", dl)



# ✅ Fonction spéciale pour le template Jinja
def status_for_step_jinja(i, s):
    return status_for_step(i, s, now=datetime.now())

def snapshot_overdue(session):
    overdue = []
    for i, step in enumerate(session["steps"]):
        st, dl = status_for_step(i, session)
        if st == "late":
            overdue.append((step["name"], dl))
    overdue.sort(key=lambda x: (x[1] or datetime.max))
    return overdue

# -----------------------
# Archivage automatique
# -----------------------
def auto_archive_if_all_done(session):
    session["archived"] = all(step["done"] for step in session["steps"])

# -----------------------
# Mails & résumé
# -----------------------
def _late_phrase(dl: datetime) -> str:
    if not dl:
        return "Retard (date N/A)"
    days = (datetime.now().date() - dl.date()).days
    days = max(days, 0)
    return f"Retard de {days} jour{'s' if days>1 else ''} ({dl.strftime('%d-%m-%Y')})"

def normalize_phone_number(value: str):
    if not value:
        return None
    cleaned = value.replace(" ", "").replace(".", "").replace("-", "").replace("(", "").replace(")", "")
    if cleaned.startswith("00"):
        cleaned = "+" + cleaned[2:]
    elif cleaned.startswith("0"):
        cleaned = "+33" + cleaned[1:]
    if not cleaned.startswith("+"):
        return None
    return cleaned

def generate_daily_overdue_email(sessions):
    now_txt = datetime.now().strftime("%d-%m-%Y %H:%M")
    logo_path = os.path.join("static", "img", "logo-integrale.png")
    logo_base64 = ""
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            logo_base64 = base64.b64encode(f.read()).decode("utf-8")

    html = f"""
    <body style="font-family:Arial,Helvetica,sans-serif;background:#f7f7f7;margin:0;padding:0;">
      <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%" style="background:#f7f7f7;">
        <tr>
          <td align="center" style="padding:20px 10px;">
            <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%" style="max-width:600px;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 8px 30px rgba(0,0,0,.1);">
              <tr>
                <td style="background:#121212;color:#fff;padding:20px;text-align:center;">
                  {('<img src="data:image/png;base64,'+logo_base64+'" alt="Intégrale Academy" style="width:100%;max-width:250px;height:auto;margin-bottom:10px;border-radius:12px;">') if logo_base64 else ''}
                  <h1 style="margin:10px 0;font-size:20px;">⚠️ Récapitulatif des retards — Intégrale Academy</h1>
                  <div style="font-size:13px;opacity:.9;">{now_txt}</div>
                </td>
              </tr>

              <tr>
                <td style="padding:20px 18px;">
    """

    found_any = False
    for s in sessions:
        # On ignore les sessions archivées dans le mail quotidien
        # pour éviter d'afficher d'anciennes formations en doublon.
        if s.get("archived"):
            continue
        overdue = snapshot_overdue(s)
        if not overdue:
            continue
        found_any = True
        color = FORMATION_COLORS.get(s["formation"], "#999")
        html += f"""
          <div style="border:1px solid #eee;border-radius:12px;padding:18px 20px;margin-bottom:18px;">
            <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;">
              <div style="background:{color};color:#fff;font-weight:700;border-radius:30px;padding:6px 14px;font-size:14px;letter-spacing:.5px;">{s["formation"]}</div>
              <div style="font-size:14px;color:#444;margin-top:8px;">
                <b>Début :</b> {format_date(s.get("date_debut","—"))} &nbsp;&nbsp;
                <b>Fin :</b> {format_date(s.get("date_fin","—"))} &nbsp;&nbsp;
                <b>Examen :</b> {format_date(s.get("date_exam","—"))}
              </div>
            </div>
            <ul style="margin:12px 0 0 18px;padding:0;color:#333;font-size:15px;line-height:1.6;">
        """
        for name, dl in overdue:
            html += f"<li style='margin-bottom:4px;list-style:none;'>🔸 {name} — {_late_phrase(dl)}</li>"
        html += "</ul></div>"

    if not found_any:
        html += "<p style='text-align:center;font-size:15px;color:#444;margin:20px 0;'>✅ Aucun retard à signaler aujourd’hui.</p>"

    html += """
                </td>
              </tr>
              <tr>
                <td style="background:#fafafa;text-align:center;padding:14px;font-size:13px;color:#666;">
                  Vous recevez ce mail automatiquement chaque matin à 8h.
                </td>
              </tr>
            </table>
          </td>
        </tr>
      </table>
    </body>
    """
    return html

def get_smtp_config():
    if BREVO_SMTP_KEY:
        brevo_login = BREVO_SMTP_LOGIN or "apikey"
        return {
            "server": BREVO_SMTP_SERVER,
            "port": BREVO_SMTP_PORT,
            "login": brevo_login,
            "password": BREVO_SMTP_KEY,
            "from_email": BREVO_FROM_EMAIL or FROM_EMAIL or BREVO_SMTP_LOGIN,
        }
    return {
        "server": SMTP_SERVER,
        "port": SMTP_PORT,
        "login": FROM_EMAIL,
        "password": EMAIL_PASSWORD,
        "from_email": FROM_EMAIL,
    }

def send_daily_overdue_summary():
    smtp_config = get_smtp_config()
    if not smtp_config["login"] or not smtp_config["password"]:
        print("⚠️ EMAIL non configuré")
        return
    data = load_sessions()
    sessions = data["sessions"]
    html = generate_daily_overdue_email(sessions)
    msg = MIMEText(html, "html", _charset="utf-8")
    msg["Subject"] = "⚠️ Récapitulatif des retards — Intégrale Academy"
    msg["From"] = smtp_config["from_email"]
    msg["To"] = "clement@integraleacademy.com"
    try:
        with smtplib.SMTP(smtp_config["server"], smtp_config["port"]) as server:
            server.starttls()
            server.login(smtp_config["login"], smtp_config["password"])
            server.sendmail(smtp_config["from_email"], ["clement@integraleacademy.com"], msg.as_string())
        print("✅ Mail quotidien envoyé avec succès")
    except Exception as e:
        print("❌ Erreur envoi mail quotidien :", e)


def _list_formateur_expired_documents(formateurs):
    today = datetime.now().date()
    expired_docs = []

    for formateur in formateurs:
        nom = (formateur.get("nom") or "").upper().strip()
        prenom = (formateur.get("prenom") or "").strip()
        full_name = f"{prenom} {nom}".strip()

        for doc in formateur.get("documents", []):
            exp_str = (doc.get("expiration") or "").strip()
            if not exp_str:
                continue

            exp_dt = parse_date(exp_str)
            if not exp_dt or exp_dt.date() > today:
                continue

            # Une seule alerte par document et par date d'expiration
            if doc.get("expiration_alert_sent_for") == exp_str:
                continue

            expired_docs.append({
                "formateur_id": formateur.get("id"),
                "formateur_nom": full_name,
                "label": doc.get("label", "Document"),
                "expiration": exp_str,
                "doc": doc,
            })

    return expired_docs


def send_formateur_expiration_alerts():
    smtp_config = get_smtp_config()
    if not smtp_config["login"] or not smtp_config["password"]:
        print("⚠️ SMTP non configuré pour les alertes formateurs")
        return 0

    formateurs = load_formateurs()
    expired_docs = _list_formateur_expired_documents(formateurs)
    if not expired_docs:
        return 0

    html_items = ""
    for item in expired_docs:
        html_items += (
            f"<li><b>{item['formateur_nom'] or 'Formateur sans nom'}</b> — "
            f"{item['label']} (expiration : {format_date(item['expiration'])})</li>"
        )

    now_txt = datetime.now().strftime("%d-%m-%Y %H:%M")
    html = f"""
    <div style=\"font-family:Arial,Helvetica,sans-serif;color:#222;line-height:1.5;\">
      <h2 style=\"margin-bottom:8px;\">⚠️ Documents formateurs expirés</h2>
      <p style=\"margin-top:0;\">Détection automatique du {now_txt}.</p>
      <p>Les documents suivants sont arrivés à expiration :</p>
      <ul>{html_items}</ul>
    </div>
    """

    msg = MIMEText(html, "html", _charset="utf-8")
    msg["Subject"] = "⚠️ Alerte expiration documents formateurs"
    msg["From"] = smtp_config["from_email"]
    msg["To"] = "clement@integraleacademy.com"

    try:
        with smtplib.SMTP(smtp_config["server"], smtp_config["port"]) as server:
            server.starttls()
            server.login(smtp_config["login"], smtp_config["password"])
            server.sendmail(
                smtp_config["from_email"],
                ["clement@integraleacademy.com"],
                msg.as_string(),
            )

        for item in expired_docs:
            item["doc"]["expiration_alert_sent_for"] = item["expiration"]
            item["doc"]["expiration_alert_sent_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_formateurs(formateurs)

        print(f"✅ Alertes expiration formateurs envoyées ({len(expired_docs)} document(s))")
        return len(expired_docs)
    except Exception as e:
        print("❌ Erreur envoi alertes expiration formateurs :", e)
        return 0


def build_jury_invitation_html(session, jury, yes_url, no_url):
    formation = formation_label(session.get("formation", "—"))
    date_exam = format_date(session.get("date_exam", "—"))
    full_name = f"{jury.get('prenom','').strip()} {jury.get('nom','').strip()}".strip()
    return f"""
    <div style="font-family:Arial,Helvetica,sans-serif;color:#222;line-height:1.6;">
      <p>Bonjour{',' if full_name else ''} {full_name}</p>
      <p>
        Nous vous proposons d'intervenir en tant que membre de jury de notre session
        <strong>{formation}</strong>, le <strong>{date_exam}</strong>.
      </p>
      <p>Pourriez-vous svp me confirmer votre présence pour cet examen ?</p>
      <div style="margin:20px 0;">
        <a href="{yes_url}" style="display:inline-block;background:#2a9134;color:#fff;text-decoration:none;padding:10px 16px;border-radius:6px;font-weight:600;margin-right:10px;">
          JE CONFIRME MA PRESENCE
        </a>
        <a href="{no_url}" style="display:inline-block;background:#c0392b;color:#fff;text-decoration:none;padding:10px 16px;border-radius:6px;font-weight:600;">
          JE NE SERAI PAS DISPONIBLE A CETTE DATE
        </a>
      </div>
      <p>Merci par avance,</p>
      <p style="margin-top:10px;">
        Clément VAILLANT<br>
        Intégrale Academy
      </p>
    </div>
    """


def send_jury_invitation_email(session, jury, yes_url, no_url):
    to_email = jury.get("email", "").strip()
    if not to_email:
        print("[jury email] Email jury manquant")
        return False, "Email jury manquant"
    html = build_jury_invitation_html(session, jury, yes_url, no_url)
    if BREVO_API_KEY and (BREVO_SENDER_EMAIL or BREVO_FROM_EMAIL or FROM_EMAIL):
        print("[jury email] Envoi via Brevo API")
        sender_email = BREVO_SENDER_EMAIL or BREVO_FROM_EMAIL or FROM_EMAIL
        sender_name = BREVO_SENDER_NAME or "Intégrale Academy"
        payload = json.dumps({
            "sender": {"email": sender_email, "name": sender_name},
            "to": [{"email": to_email}],
            "subject": f"Invitation jury — Session {session.get('formation', 'Formation')}",
            "htmlContent": html,
        }).encode("utf-8")
        request_obj = urllib.request.Request("https://api.brevo.com/v3/smtp/email")
        request_obj.add_header("Content-Type", "application/json")
        request_obj.add_header("api-key", BREVO_API_KEY)
        try:
            with urllib.request.urlopen(request_obj, data=payload, timeout=10) as response:
                if 200 <= response.status < 300:
                    print("[jury email] Brevo API OK", response.status)
                    return True, "Email envoyé"
                body = response.read().decode("utf-8")
                print("[jury email] Brevo API erreur", response.status, body)
                return False, f"Erreur email: {response.status} {body}"
        except Exception as e:
            print("[jury email] Brevo API exception", e)
            return False, f"Erreur email: {e}"
    if BREVO_API_KEY and not (BREVO_SENDER_EMAIL or BREVO_FROM_EMAIL or FROM_EMAIL):
        print("[jury email] Brevo API configurée mais expéditeur manquant")

    smtp_config = get_smtp_config()
    if not smtp_config["login"] or not smtp_config["password"]:
        print("[jury email] SMTP non configuré", {
            "server": smtp_config["server"],
            "login_set": bool(smtp_config["login"]),
            "password_set": bool(smtp_config["password"]),
            "from_set": bool(smtp_config["from_email"]),
        })
        return False, "EMAIL non configuré"
    msg = MIMEText(html, "html", _charset="utf-8")
    msg["Subject"] = f"Invitation jury — Session {session.get('formation', 'Formation')}"
    msg["From"] = smtp_config["from_email"]
    msg["To"] = to_email
    try:
        print("[jury email] Envoi via SMTP", {
            "server": smtp_config["server"],
            "port": smtp_config["port"],
            "from": smtp_config["from_email"],
            "to": to_email,
        })
        with smtplib.SMTP(smtp_config["server"], smtp_config["port"]) as server:
            server.starttls()
            server.login(smtp_config["login"], smtp_config["password"])
            server.sendmail(smtp_config["from_email"], [to_email], msg.as_string())
        print("[jury email] SMTP OK")
        return True, "Email envoyé"
    except Exception as e:
        print("[jury email] SMTP exception", e)
        return False, f"Erreur email: {e}"


def send_jury_sms(session, jury, yes_url, no_url):
    to_number = jury.get("telephone", "").strip()
    if not to_number:
        print("[jury sms] Téléphone jury manquant")
        return False, "Téléphone jury manquant"
    normalized_number = normalize_phone_number(to_number)
    if not normalized_number:
        print("[jury sms] Téléphone jury invalide", to_number)
        return False, "Téléphone jury au format international requis (ex: +336...)"
    formation = formation_label(session.get("formation", "—"))
    date_exam = format_date(session.get("date_exam", "—"))
    message = (
        "Bonjour,\n\n"
        f"Nous vous proposons d'intervenir en tant que membre de jury de notre session {formation}, le {date_exam}.\n\n"
        "Pourriez-vous svp me confirmer votre présence pour cet examen ?\n"
        f"JE CONFIRME MA PRESENCE: {yes_url}\n"
        f"JE NE SERAI PAS DISPONIBLE A CETTE DATE: {no_url}\n\n"
        "Merci par avance,\n"
        "Clément VAILLANT\n"
        "Intégrale Academy"
    )

    if BREVO_API_KEY and BREVO_SMS_SENDER:
        print("[jury sms] Envoi via Brevo API")
        payload = json.dumps({
            "sender": BREVO_SMS_SENDER,
            "recipient": normalized_number,
            "content": message,
            "type": "transactional",
        }).encode("utf-8")
        request_obj = urllib.request.Request("https://api.brevo.com/v3/transactionalSMS/sms")
        request_obj.add_header("Content-Type", "application/json")
        request_obj.add_header("api-key", BREVO_API_KEY)
        try:
            with urllib.request.urlopen(request_obj, data=payload, timeout=10) as response:
                if 200 <= response.status < 300:
                    print("[jury sms] Brevo API OK", response.status)
                    return True, "SMS envoyé"
                body = response.read().decode("utf-8")
                print("[jury sms] Brevo API erreur", response.status, body)
                return False, f"Erreur SMS: {response.status} {body}"
        except Exception as e:
            print("[jury sms] Brevo API exception", e)
            return False, f"Erreur SMS: {e}"
    elif BREVO_API_KEY and not BREVO_SMS_SENDER:
        print("[jury sms] Brevo API configurée mais sender manquant")
        return False, "SMS non configuré: BREVO_SMS_SENDER manquant"

    account_sid = os.environ.get("TWILIO_ACCOUNT_SID")
    auth_token = os.environ.get("TWILIO_AUTH_TOKEN")
    from_number = os.environ.get("TWILIO_FROM_NUMBER")
    if not account_sid or not auth_token or not from_number:
        print("[jury sms] Twilio non configuré", {
            "account_sid_set": bool(account_sid),
            "auth_token_set": bool(auth_token),
            "from_number_set": bool(from_number),
        })
        return False, "SMS non configuré"
    payload = urllib.parse.urlencode({
        "From": from_number,
        "To": normalized_number,
        "Body": message
    }).encode("utf-8")
    url = f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Messages.json"
    request_obj = urllib.request.Request(url, data=payload, method="POST")
    auth_header = base64.b64encode(f"{account_sid}:{auth_token}".encode("utf-8")).decode("utf-8")
    request_obj.add_header("Authorization", f"Basic {auth_header}")
    try:
        print("[jury sms] Envoi via Twilio", {"from": from_number, "to": normalized_number})
        with urllib.request.urlopen(request_obj, timeout=10) as response:
            if 200 <= response.status < 300:
                print("[jury sms] Twilio OK", response.status)
                return True, "SMS envoyé"
            print("[jury sms] Twilio erreur", response.status)
            return False, f"Erreur SMS: {response.status}"
    except Exception as e:
        print("[jury sms] Twilio exception", e)
        return False, f"Erreur SMS: {e}"

def build_jury_reminder_html(session, jury, yes_url, no_url):
    formation = formation_label(session.get("formation", "Formation"))
    date_exam = format_date(session.get("date_exam", ""))
    full_name = f"{jury.get('prenom','').strip()} {jury.get('nom','').strip()}".strip()
    return f"""
    <div style="font-family:Arial,Helvetica,sans-serif;line-height:1.6;color:#222;">
      <h2>Rappel : jury d'examen</h2>
      <p>Bonjour {full_name or "membre du jury"},</p>
      <p>
        Petit rappel concernant votre participation au jury de la session
        <strong>{formation}</strong> prévue le <strong>{date_exam}</strong>.
      </p>
      <p>Merci de confirmer votre présence :</p>
      <p>
        <a href="{yes_url}" style="background:#2a9134;color:#fff;padding:10px 16px;border-radius:6px;text-decoration:none;margin-right:8px;">✅ Présent</a>
        <a href="{no_url}" style="background:#c0392b;color:#fff;padding:10px 16px;border-radius:6px;text-decoration:none;">❌ Absent</a>
      </p>
      <p>Merci pour votre retour.</p>
    </div>
    """

def send_jury_reminder_email(session, jury, yes_url, no_url):
    to_email = jury.get("email", "").strip()
    if not to_email:
        print("[jury reminder email] Email jury manquant")
        return False, "Email jury manquant"
    html = build_jury_reminder_html(session, jury, yes_url, no_url)
    if BREVO_API_KEY and (BREVO_SENDER_EMAIL or BREVO_FROM_EMAIL or FROM_EMAIL):
        print("[jury reminder email] Envoi via Brevo API")
        sender_email = BREVO_SENDER_EMAIL or BREVO_FROM_EMAIL or FROM_EMAIL
        sender_name = BREVO_SENDER_NAME or "Intégrale Academy"
        payload = json.dumps({
            "sender": {"email": sender_email, "name": sender_name},
            "to": [{"email": to_email}],
            "subject": f"Rappel jury — Session {session.get('formation', 'Formation')}",
            "htmlContent": html,
        }).encode("utf-8")
        request_obj = urllib.request.Request("https://api.brevo.com/v3/smtp/email")
        request_obj.add_header("Content-Type", "application/json")
        request_obj.add_header("api-key", BREVO_API_KEY)
        try:
            with urllib.request.urlopen(request_obj, data=payload, timeout=10) as response:
                if 200 <= response.status < 300:
                    print("[jury reminder email] Brevo API OK", response.status)
                    return True, "Email rappel envoyé"
                body = response.read().decode("utf-8")
                print("[jury reminder email] Brevo API erreur", response.status, body)
                return False, f"Erreur email: {response.status} {body}"
        except Exception as e:
            print("[jury reminder email] Brevo API exception", e)
            return False, f"Erreur email: {e}"
    if BREVO_API_KEY and not (BREVO_SENDER_EMAIL or BREVO_FROM_EMAIL or FROM_EMAIL):
        print("[jury reminder email] Brevo API configurée mais expéditeur manquant")

    smtp_config = get_smtp_config()
    if not smtp_config["login"] or not smtp_config["password"]:
        print("[jury reminder email] SMTP non configuré", {
            "server": smtp_config["server"],
            "login_set": bool(smtp_config["login"]),
            "password_set": bool(smtp_config["password"]),
            "from_set": bool(smtp_config["from_email"]),
        })
        return False, "EMAIL non configuré"
    msg = MIMEText(html, "html", _charset="utf-8")
    msg["Subject"] = f"Rappel jury — Session {session.get('formation', 'Formation')}"
    msg["From"] = smtp_config["from_email"]
    msg["To"] = to_email
    try:
        print("[jury reminder email] Envoi via SMTP", {
            "server": smtp_config["server"],
            "port": smtp_config["port"],
            "from": smtp_config["from_email"],
            "to": to_email,
        })
        with smtplib.SMTP(smtp_config["server"], smtp_config["port"]) as server:
            server.starttls()
            server.login(smtp_config["login"], smtp_config["password"])
            server.sendmail(smtp_config["from_email"], [to_email], msg.as_string())
        print("[jury reminder email] SMTP OK")
        return True, "Email rappel envoyé"
    except Exception as e:
        print("[jury reminder email] SMTP exception", e)
        return False, f"Erreur email: {e}"

def send_jury_reminder_sms(session, jury, yes_url, no_url):
    to_number = jury.get("telephone", "").strip()
    if not to_number:
        print("[jury reminder sms] Téléphone jury manquant")
        return False, "Téléphone jury manquant"
    normalized_number = normalize_phone_number(to_number)
    if not normalized_number:
        print("[jury reminder sms] Téléphone jury invalide", to_number)
        return False, "Téléphone jury au format international requis (ex: +336...)"
    formation = formation_label(session.get("formation", "—"))
    date_exam = format_date(session.get("date_exam", "—"))
    message = (
        f"Rappel jury {formation} du {date_exam}. "
        f"Présent: {yes_url} / Absent: {no_url}"
    )

    if BREVO_API_KEY and BREVO_SMS_SENDER:
        print("[jury reminder sms] Envoi via Brevo API")
        payload = json.dumps({
            "sender": BREVO_SMS_SENDER,
            "recipient": normalized_number,
            "content": message,
            "type": "transactional",
        }).encode("utf-8")
        request_obj = urllib.request.Request("https://api.brevo.com/v3/transactionalSMS/sms")
        request_obj.add_header("Content-Type", "application/json")
        request_obj.add_header("api-key", BREVO_API_KEY)
        try:
            with urllib.request.urlopen(request_obj, data=payload, timeout=10) as response:
                if 200 <= response.status < 300:
                    print("[jury reminder sms] Brevo API OK", response.status)
                    return True, "SMS rappel envoyé"
                body = response.read().decode("utf-8")
                print("[jury reminder sms] Brevo API erreur", response.status, body)
                return False, f"Erreur SMS: {response.status} {body}"
        except Exception as e:
            print("[jury reminder sms] Brevo API exception", e)
            return False, f"Erreur SMS: {e}"
    elif BREVO_API_KEY and not BREVO_SMS_SENDER:
        print("[jury reminder sms] Brevo API configurée mais sender manquant")
        return False, "SMS non configuré: BREVO_SMS_SENDER manquant"

    account_sid = os.environ.get("TWILIO_ACCOUNT_SID")
    auth_token = os.environ.get("TWILIO_AUTH_TOKEN")
    from_number = os.environ.get("TWILIO_FROM_NUMBER")
    if not account_sid or not auth_token or not from_number:
        print("[jury reminder sms] Twilio non configuré", {
            "account_sid_set": bool(account_sid),
            "auth_token_set": bool(auth_token),
            "from_number_set": bool(from_number),
        })
        return False, "SMS non configuré"
    payload = urllib.parse.urlencode({
        "From": from_number,
        "To": normalized_number,
        "Body": message
    }).encode("utf-8")
    url = f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Messages.json"
    request_obj = urllib.request.Request(url, data=payload, method="POST")
    auth_header = base64.b64encode(f"{account_sid}:{auth_token}".encode("utf-8")).decode("utf-8")
    request_obj.add_header("Authorization", f"Basic {auth_header}")
    try:
        print("[jury reminder sms] Envoi via Twilio", {"from": from_number, "to": normalized_number})
        with urllib.request.urlopen(request_obj, timeout=10) as response:
            if 200 <= response.status < 300:
                print("[jury reminder sms] Twilio OK", response.status)
                return True, "SMS rappel envoyé"
            print("[jury reminder sms] Twilio erreur", response.status)
            return False, f"Erreur SMS: {response.status}"
    except Exception as e:
        print("[jury reminder sms] Twilio exception", e)
        return False, f"Erreur SMS: {e}"

def send_jury_reminders(data, base_url):
    today = datetime.now().date()
    reminded = []
    for session in data.get("sessions", []):
        if session.get("archived"):
            continue
        ensure_jury_defaults(session)
        date_exam = parse_date(session.get("date_exam"))
        if not date_exam:
            continue
        if (date_exam.date() - today).days != 5:
            continue
        for jury in session.get("jurys", []):
            if jury.get("status") in ("present", "absent"):
                continue
            if jury.get("reminded_at"):
                continue
            token = jury.get("token") or str(uuid.uuid4())
            jury["token"] = token
            yes_url = f"{base_url}{url_for('jury_response', sid=session['id'], jid=jury['id'], response='present')}?token={token}"
            no_url = f"{base_url}{url_for('jury_response', sid=session['id'], jid=jury['id'], response='absent')}?token={token}"
            email_ok, _ = send_jury_reminder_email(session, jury, yes_url, no_url)
            sms_ok, _ = send_jury_reminder_sms(session, jury, yes_url, no_url)
            if email_ok or sms_ok:
                jury["reminded_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                reminded.append(f"{jury.get('prenom','')} {jury.get('nom','')}")
    return reminded

# ------------------------------------------------------------
# 🔐 Authentification simple pour la préfecture (HTTP Basic)
# ------------------------------------------------------------
from functools import wraps
from flask import request, Response

def pref_auth_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not (auth.username == "prefecture" and auth.password == "pref2025"):
            return Response(
                "Accès réservé à la préfecture.\n",
                401,
                {"WWW-Authenticate": 'Basic realm="Accès Préfecture"'}
            )
        return f(*args, **kwargs)
    return decorated

# ------------------------------------------------------------
# 📋 Résumé conformité globale formateurs (pour l'index)
# ------------------------------------------------------------
def get_formateurs_global_non_conformites():
    formateurs = load_formateurs()
    total_non_conformes = 0

    for f in formateurs:
        for doc in f.get("documents", []):
            auto_update_document_status(doc)
            if doc.get("status") in ("non_conforme", "a_controler"):
                total_non_conformes += 1


    return total_non_conformes




# -----------------------
# Routes principales
# -----------------------
@app.route("/")
def index():
    nb_non_conformes = get_formateurs_global_non_conformites()

    return render_template(
        "index.html",
        title="Plateforme de gestion Intégrale Academy",
        formateurs_non_conformes=nb_non_conformes,
        shortcuts=load_shortcuts()
    )


@app.route("/shortcuts", methods=["GET"])
def shortcuts_data():
    return jsonify(load_shortcuts())


@app.route("/stagiaires/docs-to-control.json")
def stagiaires_docs_to_control():
    now = time.monotonic()
    with _stagiaires_docs_cache_lock:
        cached_payload = _stagiaires_docs_cache["payload"]
        retry_after = _stagiaires_docs_cache["retry_after"]

    if now < retry_after:
        if cached_payload is not None:
            return jsonify(stagiaires_docs_response(cached_payload, stale=True))
        return jsonify({
            "ok": False,
            "pending_count": None,
            "items": [],
            "error": "Données dossiers stagiaires temporairement indisponibles",
        })

    try:
        payload = fetch_json_url(
            STAGIAIRES_DOCS_TO_CONTROL_URL,
            headers=stagiaires_docs_request_headers(),
        )
        if not isinstance(payload, dict) or payload.get("ok") is False:
            raise ValueError("Réponse dossiers stagiaires invalide")
    except (OSError, ValueError, TimeoutError, json.JSONDecodeError) as exc:
        logger.warning("Impossible de récupérer les dossiers stagiaires: %s", exc)
        with _stagiaires_docs_cache_lock:
            _stagiaires_docs_cache["retry_after"] = now + STAGIAIRES_DOCS_RETRY_SECONDS
            cached_payload = _stagiaires_docs_cache["payload"]
        if cached_payload is not None:
            return jsonify(stagiaires_docs_response(cached_payload, stale=True))
        return jsonify({
            "ok": False,
            "pending_count": None,
            "items": [],
            "error": "Données dossiers stagiaires indisponibles",
        })

    with _stagiaires_docs_cache_lock:
        _stagiaires_docs_cache["payload"] = payload
        _stagiaires_docs_cache["retry_after"] = 0.0
    return jsonify(stagiaires_docs_response(payload))


@app.route("/shortcut-images/<path:filename>")
def shortcut_image(filename):
    return send_from_directory(SHORTCUT_UPLOAD_DIR, filename)


@app.route("/shortcuts", methods=["POST"])
def create_shortcut():
    name = (request.form.get("name") or "").strip()
    url = (request.form.get("url") or "").strip()
    image = request.files.get("image")

    if not name or not url:
        return jsonify({"ok": False, "error": "Le nom et le lien sont obligatoires."}), 400

    if not (url.startswith("http://") or url.startswith("https://")):
        return jsonify({"ok": False, "error": "Le lien doit commencer par http:// ou https://"}), 400

    if image is None or not image.filename:
        return jsonify({"ok": False, "error": "Veuillez importer une image."}), 400

    if not allowed_shortcut_image(image.filename):
        return jsonify({"ok": False, "error": "Format d'image non pris en charge."}), 400

    ensure_shortcuts_storage()
    original_name = secure_filename(image.filename)
    extension = original_name.rsplit(".", 1)[1].lower()
    filename = f"{uuid.uuid4().hex}.{extension}"
    image.save(os.path.join(SHORTCUT_UPLOAD_DIR, filename))

    shortcuts = load_shortcuts()
    shortcut = {
        "id": uuid.uuid4().hex,
        "name": name,
        "url": url,
        "image": shortcut_image_url(filename)
    }
    shortcuts.append(shortcut)
    save_shortcuts(shortcuts)
    return jsonify({"ok": True, "shortcut": shortcut}), 201


@app.route("/shortcuts/<shortcut_id>", methods=["DELETE"])
def delete_shortcut(shortcut_id):
    shortcuts = load_shortcuts()
    shortcut_to_delete = next((shortcut for shortcut in shortcuts if shortcut.get("id") == shortcut_id), None)
    remaining = [shortcut for shortcut in shortcuts if shortcut.get("id") != shortcut_id]

    if len(remaining) == len(shortcuts):
        return jsonify({"ok": False, "error": "Raccourci introuvable."}), 404

    if shortcut_to_delete:
        image_path = shortcut_to_delete.get("image") or ""
        filename = os.path.basename(urllib.parse.urlparse(image_path).path)
        if filename:
            stored_image_path = os.path.join(SHORTCUT_UPLOAD_DIR, filename)
            if os.path.exists(stored_image_path):
                os.remove(stored_image_path)

    save_shortcuts(remaining)
    return jsonify({"ok": True})


@app.route("/general-tools")
def general_tools():
    return render_template("general_tools.html", title="Outils généraux")


@app.route("/price-adaptator")
def price_adaptator():
    return render_template("price_adaptator.html", title="Price adaptator")


@app.route("/price-adaptator/data")
def price_adaptator_data():
    data = load_price_adaptator_data()
    return {"prospects": data.get("prospects", []), "dates": data.get("dates", {})}


@app.route("/price-adaptator/prospects", methods=["POST"])
def price_adaptator_add_prospect():
    payload = request.get_json(silent=True) or {}
    nom = normalize_price_adaptator_nom(payload.get("nom"))
    prenom = normalize_price_adaptator_prenom(payload.get("prenom"))
    cpf = payload.get("cpf")
    email = (payload.get("email", "") or "").strip()
    telephone = (payload.get("telephone", "") or "").strip()
    formation = (payload.get("formation", "") or "").strip()

    if not (nom and prenom and formation):
        return {"ok": False, "error": "Données prospect incomplètes"}, 400

    try:
        cpf_value = float(cpf)
        if cpf_value < 0:
            raise ValueError
    except (TypeError, ValueError):
        return {"ok": False, "error": "Montant CPF invalide"}, 400

    data = load_price_adaptator_data()
    prospect = {
        "id": str(uuid.uuid4()),
        "nom": nom,
        "prenom": prenom,
        "cpf": cpf_value,
        "email": email,
        "telephone": telephone,
        "formation": formation,
        "sent": False,
        "sentAt": None,
        "proposed_price": None,
        "last_error": None,
        "last_attempt_at": None,
        "created_at": datetime.now().isoformat(),
    }
    data["prospects"].insert(0, prospect)
    save_price_adaptator_data(data)
    return {"ok": True, "prospects": data["prospects"]}


@app.route("/price-adaptator/prospects/<prospect_id>", methods=["DELETE"])
def price_adaptator_delete_prospect(prospect_id):
    data = load_price_adaptator_data()
    prospects = data.get("prospects", [])
    updated = [prospect for prospect in prospects if prospect.get("id") != prospect_id]
    if len(updated) == len(prospects):
        return {"ok": False, "error": "Prospect introuvable"}, 404
    data["prospects"] = updated
    save_price_adaptator_data(data)
    return {"ok": True, "prospects": data["prospects"]}


@app.route("/price-adaptator/prospects", methods=["DELETE"])
def price_adaptator_clear_prospects():
    data = load_price_adaptator_data()
    data["prospects"] = []
    save_price_adaptator_data(data)
    return {"ok": True, "prospects": data["prospects"]}


@app.route("/price-adaptator/import", methods=["POST"])
def price_adaptator_import():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return {"ok": False, "error": "Fichier Excel manquant"}, 400

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"ok": False, "error": "La bibliothèque openpyxl est manquante"}, 500

    try:
        workbook = load_workbook(filename=BytesIO(upload.read()), data_only=True)
    except Exception:
        return {"ok": False, "error": "Impossible de lire le fichier Excel"}, 400

    sheet = workbook.active
    data = load_price_adaptator_data()
    prospects = data.get("prospects", [])

    existing_emails = {p.get("email", "").strip().lower() for p in prospects if p.get("email")}
    existing_phones = set()
    existing_names = set()
    for prospect in prospects:
        normalized_phone = normalize_phone_number((prospect.get("telephone") or "").strip())
        if normalized_phone:
            existing_phones.add(normalized_phone)
        nom = normalize_price_adaptator_nom(prospect.get("nom")).lower()
        prenom = normalize_price_adaptator_prenom(prospect.get("prenom")).lower()
        formation = (prospect.get("formation") or "").strip()
        if nom and prenom and formation:
            existing_names.add((nom, prenom, formation))

    added_prospects = []
    seen_emails = set()
    seen_phones = set()
    seen_names = set()
    skipped = 0
    errors = []

    def normalize_import_phone(raw_value):
        if raw_value is None:
            return ""
        value = str(raw_value).strip()
        if not value:
            return ""
        if value.startswith("+") or value.startswith("00") or value.startswith("0"):
            return value
        return f"0{value}"

    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = list(row[:6]) if row else []
        values += [None] * (6 - len(values))
        formation_raw, nom, prenom, cpf, email, telephone = values
        if not any([formation_raw, nom, prenom, cpf, email, telephone]):
            continue

        formation = normalize_price_adaptator_formation(formation_raw)
        if not formation:
            errors.append(f"Ligne {idx}: formation invalide")
            continue

        nom_value = normalize_price_adaptator_nom(nom)
        prenom_value = normalize_price_adaptator_prenom(prenom)
        if not nom_value or not prenom_value:
            errors.append(f"Ligne {idx}: nom/prénom manquants")
            continue

        if cpf is None or (isinstance(cpf, str) and not cpf.strip()):
            cpf_value = 0.0
        else:
            try:
                cpf_value = float(cpf)
                if cpf_value < 0:
                    raise ValueError
            except (TypeError, ValueError):
                errors.append(f"Ligne {idx}: montant CPF invalide")
                continue
        formation_price = PRICE_ADAPTATOR_FORMATION_PRICES.get(formation)
        if formation_price is not None and cpf_value > formation_price:
            skipped += 1
            errors.append(
                f"Ligne {idx}: montant CPF supérieur au montant de la formation"
            )
            continue

        email_value = str(email).strip() if email is not None else ""
        email_key = email_value.lower() if email_value else ""
        telephone_value = normalize_import_phone(telephone)
        phone_key = normalize_phone_number(telephone_value) or ""

        name_key = (nom_value.lower(), prenom_value.lower(), formation)

        if (
            (email_key and (email_key in existing_emails or email_key in seen_emails))
            or (phone_key and (phone_key in existing_phones or phone_key in seen_phones))
            or (name_key in existing_names or name_key in seen_names)
        ):
            skipped += 1
            errors.append(f"Ligne {idx}: prospect déjà existant")
            continue

        prospect = {
            "id": str(uuid.uuid4()),
            "nom": nom_value,
            "prenom": prenom_value,
            "cpf": cpf_value,
            "email": email_value,
            "telephone": telephone_value,
            "formation": formation,
            "sent": False,
            "sentAt": None,
            "proposed_price": None,
            "last_error": None,
            "last_attempt_at": None,
            "created_at": datetime.now().isoformat(),
        }
        added_prospects.append(prospect)
        if email_key:
            seen_emails.add(email_key)
        if phone_key:
            seen_phones.add(phone_key)
        seen_names.add(name_key)

    if added_prospects:
        data["prospects"] = added_prospects[::-1] + prospects
        save_price_adaptator_data(data)

    return {
        "ok": True,
        "added": len(added_prospects),
        "skipped": skipped,
        "errors": errors,
        "prospects": data.get("prospects", prospects),
    }


@app.route("/price-adaptator/dates", methods=["POST"])
def price_adaptator_save_dates():
    payload = request.get_json(silent=True) or {}
    dates = payload.get("dates", {})
    data = load_price_adaptator_data()
    cleaned = {}
    for formation, range_data in (dates or {}).items():
        if not isinstance(range_data, dict):
            continue
        cleaned[formation] = {
            "start": range_data.get("start"),
            "end": range_data.get("end"),
            "discount": normalize_price_adaptator_discount(range_data.get("discount")),
        }
    data["dates"] = cleaned
    save_price_adaptator_data(data)
    return {"ok": True, "dates": data["dates"]}


@app.route("/price-adaptator/prospects/<prospect_id>/proposal", methods=["POST"])
def price_adaptator_save_proposal(prospect_id):
    payload = request.get_json(silent=True) or {}
    price = payload.get("price")

    if price is None:
        return {"ok": False, "error": "Prix manquant"}, 400

    try:
        price_value = float(price)
    except (TypeError, ValueError):
        return {"ok": False, "error": "Prix invalide"}, 400

    data = load_price_adaptator_data()
    prospect = next((item for item in data.get("prospects", []) if item.get("id") == prospect_id), None)
    if not prospect:
        return {"ok": False, "error": "Prospect introuvable"}, 404

    price_value = normalize_price_adaptator_proposed_price(price_value)
    prospect["proposed_price"] = price_value
    save_price_adaptator_data(data)

    return {"ok": True, "prospects": data.get("prospects", [])}


@app.route("/price-adaptator/send", methods=["POST"])
def price_adaptator_send():
    payload = request.get_json(silent=True) or {}
    price = payload.get("price")
    prospect_id = payload.get("prospect_id")

    if price is None:
        return {"ok": False, "error": "Prix manquant"}, 400

    try:
        price_value = float(price)
    except (TypeError, ValueError):
        return {"ok": False, "error": "Prix invalide"}, 400

    data = load_price_adaptator_data()
    prospect = next((item for item in data.get("prospects", []) if item.get("id") == prospect_id), None)
    if not prospect:
        return {"ok": False, "error": "Prospect introuvable"}, 404

    price_value = normalize_price_adaptator_proposed_price(price_value)
    result = attempt_price_adaptator_send(prospect, data.get("dates"), price_override=price_value)
    prospect["last_attempt_at"] = datetime.now().isoformat()
    prospect["last_error"] = result["email_error"] or result["sms_error"]
    prospect["proposed_price"] = result["price"]
    prospect["manual_sent"] = True
    prospect["manualSentAt"] = datetime.now().isoformat()
    if result["email_sent"] or result["sms_sent"]:
        prospect["sent"] = True
        prospect["sentAt"] = datetime.now().isoformat()
        prospect["last_sent_price"] = result["price"]
    save_price_adaptator_data(data)

    return {
        "ok": True,
        "email_sent": result["email_sent"],
        "sms_sent": result["sms_sent"],
        "email_error": result["email_error"],
        "sms_error": result["sms_error"],
        "prospects": data.get("prospects", []),
    }


@app.route("/price-adaptator/prospects/<prospect_id>/preview")
def price_adaptator_preview(prospect_id):
    data = load_price_adaptator_data()
    prospect = next((item for item in data.get("prospects", []) if item.get("id") == prospect_id), None)
    if not prospect:
        return {"ok": False, "error": "Prospect introuvable"}, 404

    price_override = prospect.get("last_sent_price")
    if price_override is None:
        price_override = prospect.get("proposed_price")
    if price_override is not None:
        price_override = normalize_price_adaptator_proposed_price(price_override)

    message = build_price_adaptator_message(prospect, data.get("dates"), price_override=price_override)
    return {
        "ok": True,
        "subject": message["subject"],
        "html": message["html"],
        "sent_at": prospect.get("sentAt"),
    }


@app.route("/sessions")
def sessions_home():
    data = load_sessions()
    # 🔄 Synchronise automatiquement les étapes manquantes pour chaque session
    for s in data["sessions"]:
        sync_steps(s)
    save_sessions(data)

    today = datetime.now().date()
    active = []
    archived = []

    for s in data["sessions"]:
        end_date = parse_date(s.get("date_fin"))
        is_finished = bool(end_date and end_date.date() < today)
        s["is_finished"] = is_finished

        if s.get("archived") or is_finished:
            archived.append(s)
        else:
            active.append(s)

    for s in data["sessions"]:
        s["color"] = FORMATION_COLORS.get(s["formation"], "#555")

    # --- DEBUG existant ---
    print("\n=== DEBUG SESSIONS ===")
    for s in data["sessions"]:
        print(f"\nSession: {s['formation']} ({s['date_debut']} → {s['date_exam']})")
        for i, step in enumerate(s["steps"]):
            st, dl = status_for_step(i, s)
            if dl:
                print(f" - {step['name']}: {st} / deadline={dl.strftime('%Y-%m-%d')}")
            else:
                print(f" - {step['name']}: {st} / deadline=N/A")

    # --------- 🧠 On calcule le récap en PYTHON ---------
    recap_map = {}   # { formation: {"late_steps":[(text,days)], "today_steps":[text]} }
    total_late = 0

    # On ne prend que les sessions actives (comme avant)
    for s in active:
        formation = s.get("formation", "—")
        rec = recap_map.setdefault(formation, {"late_steps": [], "today_steps": []})

        for i, step in enumerate(s.get("steps", [])):
            st, dl = status_for_step(i, s)
            # late
            if st == "late" and dl:
                days = max((today - dl.date()).days, 0)
                text = f"[{format_date(s.get('date_debut','—'))}] {step['name']}"
                rec["late_steps"].append((text, days))
                total_late += 1
            # due today
            elif st == "on_time" and dl and dl.date() == today:
                text = f"[{format_date(s.get('date_debut','—'))}] {step['name']}"
                rec["today_steps"].append(text)

    # On transforme en liste triée par nom de formation pour le template
    recap_data = []
    for formation, payload in sorted(recap_map.items(), key=lambda x: x[0]):
        # trier les retards par nb de jours décroissant (les pires d'abord)
        payload["late_steps"].sort(key=lambda t: t[1], reverse=True)
        recap_data.append((formation, payload["late_steps"], payload["today_steps"]))

    return render_template(
        "sessions.html",
        title="Gestion des sessions",
        active_sessions=active,
        archived_sessions=archived,
        status_for_step=status_for_step_jinja,  # garde pour le détail
        now=datetime.now,
        # 👇 nouveaux paramètres pour le récap déjà prêt
        recap_data=recap_data,
        total_late=total_late,
        formations=[f for f, *_ in recap_data],
    )



@app.route("/sessions/create", methods=["POST"])
def create_session():
    formation = request.form.get("formation","").upper().strip()
    date_debut = request.form.get("date_debut","").strip()
    date_fin = request.form.get("date_fin","").strip()
    date_exam = request.form.get("date_exam","").strip()
    contractual_end_date = request.form.get("contractual_end_date", "").strip()
    interruptions = request.form.get("interruptions", "").strip()
    dirigeant_location = request.form.get("dirigeant_location", "").upper().strip()
    if formation not in FORMATION_COLORS:
        flash("Formation invalide.","error")
        return redirect(url_for("sessions_home"))
    if formation == "DIRIGEANT" and dirigeant_location not in DIRIGEANT_LOCATIONS:
        flash("Choisissez le site de la session DIRIGEANT : Paris ou Puget.", "error")
        return redirect(url_for("sessions_home"))
    sid = str(uuid.uuid4())[:8]
    session = {
        "id": sid,
        "formation": formation,
        "date_debut": date_debut,
        "date_fin": date_fin,
        "date_exam": date_exam,
        "color": FORMATION_COLORS.get(formation,"#555"),
        "steps": default_steps_for(formation),
        "archived": False,
        "jurys": [],
        "jury_notification_status": "to_notify",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    if formation == "AFC_APS_SSIAP":
        if not date_debut:
            flash("La date de début est obligatoire pour l'AFC APS + SSIAP.", "error")
            return redirect(url_for("sessions_home"))
        if not contractual_end_date:
            flash("Date de fin obligatoire pour l'AFC APS + SSIAP.", "error")
            return redirect(url_for("sessions_home"))
        session["display_name"] = "AFC France Travail APS + SSIAP"
        session["interruptions"] = interruptions
        session["training_code"] = "AFC_APS_SSIAP"
        session["salle"] = "Intégrale Academy – 54 chemin du Carreou – 83480 PUGET-SUR-ARGENS"
        end_dt = parse_date(contractual_end_date)
        if not end_dt:
            flash("La date de fin souhaitée AFC est invalide.", "error")
            return redirect(url_for("sessions_home"))
        session["contractual_end_date"] = end_dt.date().isoformat()
        planning_data = build_afc_aps_ssiap_planning_data(parse_date(date_debut).date(), "", session["salle"], parse_interruption_ranges(interruptions), contractual_end_date=end_dt.date())
        session["date_fin"] = session["contractual_end_date"]
        session["date_exam"] = planning_data[-1]["date"]
        session["apsPlanningData"] = planning_data
        session["apsPlanningSummary"] = afc_aps_ssiap_summary_from_data(planning_data, parse_interruption_ranges(interruptions), contractual_end_date=end_dt.date())
        session["apsPlanningMode"] = "full_presentiel"
        filename = f"planning_afc_aps_ssiap_session_{sid}.pdf"
        generate_aps_planning_pdf(session, "", os.path.join(PLANNING_DIR, filename), planning_data=planning_data, planning_mode="full_presentiel", document_profile={"validate":"afc_aps_ssiap", "summary": session["apsPlanningSummary"], "planning_title":"PLANNING AFC FRANCE TRAVAIL APS + SSIAP", "short_label":"AFC APS + SSIAP"})
        session["planning_pdf"] = filename
        session["planning_generated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if formation == "DIRIGEANT":
        session["dirigeant_location"] = dirigeant_location
        session["dirigeant_location_label"] = DIRIGEANT_LOCATIONS[dirigeant_location]
    data = load_sessions()
    data["sessions"].append(session)
    save_sessions(data)
    return redirect(url_for("session_detail", sid=sid))

@app.route("/sessions/<sid>")
def session_detail(sid):
    # --- 🔐 Vérification accès préfecture si ?key= est présent ---
    public_key = request.args.get("key")

    PREF_EMAIL = os.getenv("PREF_EMAIL")
    PREF_PASSWORD = os.getenv("PREF_PASSWORD")

    if public_key:
        expected = f"{PREF_EMAIL}:{PREF_PASSWORD}"
        encoded = base64.b64encode(expected.encode()).decode()

        if public_key != encoded:
            abort(403)  # accès refusé

    # --- 🔧 Chargement session ---
    data = load_sessions()
    session = find_session(data, sid)
    if not session:
        abort(404)

    ensure_jury_defaults(session)
    ensure_global_jury_defaults(data)
    sync_global_jurys(data)
    sync_steps(session)
    save_sessions(data)

    statuses = []
    for i in range(len(session["steps"])):
        st, dl = status_for_step(i, session)
        statuses.append({"status": st, "deadline": (dl.strftime("%Y-%m-%d") if dl else None)})

    order = sorted(
        range(len(session["steps"])),
        key=lambda i: deadline_for(i, session) or datetime.max
    )

    auto_archive_if_all_done(session)
    save_sessions(data)

    return render_template(
        "session_detail.html",
        title=f"{session['formation']} — Détail",
        s=session,
        global_jurys=data.get("jurys", []),
        session_jurys_by_id={j.get("id"): j for j in session.get("jurys", [])},
        statuses=statuses,
        order=order,
        now=datetime.now,
        planning_pdf=session.get("planning_pdf"),
        afc_dsf_summary=afc_dsf_summary(session) if is_afc_aps_ssiap_session(session) else None,
        afc_dsf_modules=AFC_DSF_MODULES,
    )



def generate_afc_dsf_pdf(session_data, dsf, output_path):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet(); body=[]
    title = ParagraphStyle('dsfTitle', parent=styles['Title'], textColor=colors.HexColor('#111827'), fontSize=18, leading=22)
    normal = styles['BodyText']
    logo = aps_pdf_logo_path()
    header = []
    if logo: header.append(Image(logo, width=34*mm, height=16*mm, kind='proportional'))
    header.append(Paragraph(f"<b>Demande de service fait</b><br/>{dsf['label']}<br/>{session_data.get('display_name') or session_data.get('formation') or AFC_APS_SSIAP_LABEL}<br/>Session {session_data.get('id')} — période du {format_date(dsf['periodStart'])} au {format_date(dsf['periodEnd'])}", title))
    body.append(Table([header], colWidths=[45*mm, 220*mm] if logo else [265*mm])); body.append(Spacer(1, 8))
    module_labels = [AFC_DSF_MODULES[m]['label'] for m in dsf['modules']]
    body.append(Paragraph(f"Modules sélectionnés : <b>{', '.join(module_labels)}</b><br/>Date de génération : {format_date(dsf.get('createdAt','')[:10])}<br/>{'<b>ANNULÉE</b>' if dsf.get('status') == AFC_DSF_STATUS_CANCELLED else ''}", normal)); body.append(Spacer(1, 8))
    total_per_student = sum(dsf['hoursPerStudent'].values())
    recap = [["Nombre de stagiaires", str(dsf['studentCount']), "Total par stagiaire", f"{total_per_student:g} h", "Total heures-stagiaires", f"{dsf['totalHours']:g} h"]]
    body.append(Table(recap, colWidths=[42*mm,25*mm,42*mm,25*mm,45*mm,25*mm], style=[('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#f8fafc')),('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#cbd5e1')),('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'),('PADDING',(0,0),(-1,-1),6)])); body.append(Spacer(1, 10))
    header_row = ["Nom et prénom"] + module_labels + ["Total"]
    rows = [header_row]
    for st in dsf['students']:
        rows.append([st['displayName']] + [f"{st['modules'].get(m,0):g} h" for m in dsf['modules']] + [f"{st['totalHours']:g} h"])
    rows.append(["Totaux heures-stagiaires"] + [f"{dsf['moduleTotals'].get(m,0):g} h" for m in dsf['modules']] + [f"{dsf['totalHours']:g} h"])
    tbl=Table(rows, repeatRows=1, colWidths=[90*mm]+[55*mm]*len(dsf['modules'])+[35*mm])
    style=[('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#cbd5e1')),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#111827')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#fef3c7')),('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),('ALIGN',(1,1),(-1,-1),'RIGHT'),('PADDING',(0,0),(-1,-1),5)]
    tbl.setStyle(TableStyle(style)); body.append(tbl); body.append(Spacer(1, 18))
    body.append(Paragraph("Fait à : ____________________ &nbsp;&nbsp;&nbsp; Date : ____ / ____ / ______<br/><br/>Nom du responsable : ____________________ &nbsp;&nbsp;&nbsp; Signature et tampon : ____________________<br/><br/>Intégrale Academy — document généré depuis le planning central AFC. Page <seq id='page'/>", normal))
    doc.build(body)

@app.route('/api/sessions/<sid>/afc-dsf/preview', methods=['POST'])
def api_afc_dsf_preview(sid):
    data=load_sessions(); sess=find_session(data,sid)
    if not sess: return jsonify({'ok':False,'error':'Session introuvable'}),404
    payload=request.get_json(silent=True) or {}
    try:
        result=afc_dsf_compute(sess, payload.get('periodStart') or '', payload.get('periodEnd') or '', payload.get('modules') or [])
        return jsonify({'ok':True,'preview':result,'moduleLabels':{k:v['label'] for k,v in AFC_DSF_MODULES.items()}})
    except Exception as exc:
        return jsonify({'ok':False,'error':str(exc)}),400

@app.route('/api/sessions/<sid>/afc-dsf/generate', methods=['POST'])
def api_afc_dsf_generate(sid):
    data=load_sessions(); sess=find_session(data,sid)
    if not sess: return jsonify({'ok':False,'error':'Session introuvable'}),404
    payload=request.get_json(silent=True) or {}
    try:
        result=afc_dsf_compute(sess, payload.get('periodStart') or '', payload.get('periodEnd') or '', payload.get('modules') or [])
        number=afc_dsf_next_number(sess); dsf_id=uuid.uuid4().hex; filename=f"dsf_afc_{sid}_DSF_{number}.pdf"; path=os.path.join(DSF_DIR, filename)
        dsf={"id":dsf_id,"number":number,"label":f"DSF {number}","sessionId":sid,"sessionName":sess.get('display_name') or sess.get('formation'),"createdAt":datetime.now().strftime('%Y-%m-%d %H:%M:%S'),"createdBy":session.get('admin_email') or 'admin',"status":AFC_DSF_STATUS_FINALIZED,"pdfFilename":filename,**result}
        generate_afc_dsf_pdf(sess, dsf, path)
        if not os.path.exists(path): raise RuntimeError('Génération PDF échouée.')
        sess.setdefault('afcDsfs',[]).append(dsf); save_sessions(data)
        return jsonify({'ok':True,'dsf':dsf})
    except Exception as exc:
        app.logger.exception('Erreur génération DSF AFC')
        return jsonify({'ok':False,'error':str(exc)}),400

@app.route('/sessions/<sid>/afc-dsf/<dsf_id>/pdf')
def view_afc_dsf_pdf(sid, dsf_id):
    sess=find_session(load_sessions(),sid); dsf=next((d for d in (sess or {}).get('afcDsfs',[]) if d.get('id')==dsf_id),None)
    if not dsf: abort(404)
    return send_file(os.path.join(DSF_DIR, os.path.basename(dsf.get('pdfFilename',''))), mimetype='application/pdf', as_attachment=False)

@app.route('/sessions/<sid>/afc-dsf/<dsf_id>/download')
def download_afc_dsf_pdf(sid, dsf_id):
    sess=find_session(load_sessions(),sid); dsf=next((d for d in (sess or {}).get('afcDsfs',[]) if d.get('id')==dsf_id),None)
    if not dsf: abort(404)
    return send_file(os.path.join(DSF_DIR, os.path.basename(dsf.get('pdfFilename',''))), mimetype='application/pdf', as_attachment=True, download_name=dsf.get('pdfFilename'))

@app.route('/api/sessions/<sid>/afc-dsf/<dsf_id>/cancel', methods=['POST'])
def api_afc_dsf_cancel(sid, dsf_id):
    data=load_sessions(); sess=find_session(data,sid)
    if not sess: return jsonify({'ok':False,'error':'Session introuvable'}),404
    dsfs=sess.get('afcDsfs') or []
    dsf=next((d for d in dsfs if d.get('id')==dsf_id),None)
    if not dsf: return jsonify({'ok':False,'error':'DSF introuvable'}),404
    if dsf.get('status') != AFC_DSF_STATUS_FINALIZED: return jsonify({'ok':False,'error':'DSF déjà annulée'}),400
    pdf_filename=os.path.basename(dsf.get('pdfFilename',''))
    sess['afcDsfs']=[d for d in dsfs if d.get('id')!=dsf_id]
    save_sessions(data)
    if pdf_filename:
        pdf_path=os.path.join(DSF_DIR, pdf_filename)
        try:
            if os.path.exists(pdf_path): os.remove(pdf_path)
        except OSError:
            app.logger.exception('Erreur suppression PDF DSF AFC %s', pdf_filename)
    best_response = request.accept_mimetypes.best_match(['application/json', 'text/html'])
    wants_json = request.headers.get('X-Requested-With') == 'fetch' or best_response != 'text/html'
    if not wants_json:
        flash('DSF supprimée définitivement.', 'success')
        return redirect(url_for('session_detail', sid=sid))
    return jsonify({'ok':True,'deleted':True})

@app.route("/sessions/<sid>/jury/add", methods=["POST"])
def add_jury(sid):
    data = load_sessions()
    session = find_session(data, sid)
    if not session:
        abort(404)
    ensure_jury_defaults(session)
    ensure_global_jury_defaults(data)
    sync_global_jurys(data)
    nom = request.form.get("nom", "").strip()
    prenom = request.form.get("prenom", "").strip()
    email = request.form.get("email", "").strip()
    telephone = request.form.get("telephone", "").strip()
    if not nom or not prenom:
        flash("Nom et prénom du jury requis.", "error")
        return redirect(url_for("session_detail", sid=sid))
    existing_global = find_global_jury_by_email(data, email)
    if existing_global:
        existing_global.update({
            "nom": nom or existing_global.get("nom", ""),
            "prenom": prenom or existing_global.get("prenom", ""),
            "email": email or existing_global.get("email", ""),
            "telephone": telephone or existing_global.get("telephone", ""),
        })
        jury_id = existing_global["id"]
    else:
        jury_id = str(uuid.uuid4())[:8]
        data["jurys"].append({
            "id": jury_id,
            "nom": nom,
            "prenom": prenom,
            "email": email,
            "telephone": telephone,
        })
    if any(j.get("id") == jury_id for j in session["jurys"]):
        flash("Ce jury est déjà associé à la session.", "info")
        save_sessions(data)
        return redirect(url_for("session_detail", sid=sid))
    session["jurys"].append({
        "id": jury_id,
        "nom": nom,
        "prenom": prenom,
        "email": email,
        "telephone": telephone,
        "status": "pending",
        "token": str(uuid.uuid4()),
        "notified_at": None,
        "reminded_at": None,
    })
    save_sessions(data)
    flash("Jury ajouté.", "success")
    return redirect(url_for("session_detail", sid=sid))


@app.route("/sessions/<sid>/jury/notify", methods=["POST"])
def notify_jury(sid):
    print("🔥 HIT notify_jury", sid, dict(request.form.lists()))
    data = load_sessions()
    session = find_session(data, sid)
    if not session:
        abort(404)
    ensure_jury_defaults(session)
    ensure_global_jury_defaults(data)
    sync_global_jurys(data)
    selected_ids = request.form.getlist("jury_ids")
    logger.info("[jury notify] Déclenchement", extra={"sid": sid, "selected_ids": selected_ids})
    if not selected_ids:
        flash("Sélectionnez au moins un jury à notifier.", "error")
        return redirect(url_for("session_detail", sid=sid))
    base_url = request.url_root.rstrip("/")
    logger.info("[jury notify] base_url=%s", base_url)
    results = []
    any_sent = False
    now_txt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    session_jurys_by_id = {j.get("id"): j for j in session.get("jurys", [])}
    for jury_id in selected_ids:
        jury = session_jurys_by_id.get(jury_id)
        if not jury:
            global_jury = find_global_jury_by_id(data, jury_id)
            if not global_jury:
                results.append(f"Jury introuvable ({jury_id}).")
                continue
            jury = {
                "id": global_jury.get("id"),
                "nom": global_jury.get("nom", ""),
                "prenom": global_jury.get("prenom", ""),
                "email": global_jury.get("email", ""),
                "telephone": global_jury.get("telephone", ""),
                "status": "pending",
                "token": str(uuid.uuid4()),
                "notified_at": None,
                "reminded_at": None,
            }
            session["jurys"].append(jury)
            session_jurys_by_id[jury_id] = jury
        if jury.get("status") in ("present", "absent"):
            results.append(f"{jury.get('prenom','')} {jury.get('nom','')}: déjà répondu")
            continue
        logger.info(
            "[jury notify] Tentative",
            extra={
                "jid": jury.get("id"),
                "email_set": bool(jury.get("email")),
                "phone_set": bool(jury.get("telephone")),
            },
        )
        token = jury.get("token") or str(uuid.uuid4())
        jury["token"] = token
        yes_url = f"{base_url}{url_for('jury_response', sid=sid, jid=jury['id'], response='present')}?token={token}"
        no_url = f"{base_url}{url_for('jury_response', sid=sid, jid=jury['id'], response='absent')}?token={token}"
        email_ok, email_msg = send_jury_invitation_email(session, jury, yes_url, no_url)
        sms_ok, sms_msg = send_jury_sms(session, jury, yes_url, no_url)
        results.append(f"{jury.get('prenom','')} {jury.get('nom','')}: {email_msg} / {sms_msg}")
        logger.info(
            "[jury notify] Résultat email=%s sms=%s",
            email_msg,
            sms_msg,
        )
        if email_ok or sms_ok:
            any_sent = True
        jury["status"] = "pending"
        jury["notified_at"] = now_txt if email_ok or sms_ok else jury.get("notified_at")
    if any_sent:
        session["jury_notification_status"] = "notified"
    save_sessions(data)
    if results:
        flash_message = " | ".join(results)
        if any_sent:
            flash("Notifications envoyées. " + flash_message, "success")
        else:
            flash("Aucune notification envoyée. " + flash_message, "error")
    return redirect(url_for("session_detail", sid=sid))


@app.route("/sessions/<sid>/jury/<jid>/delete", methods=["POST"])
def delete_jury(sid, jid):
    data = load_sessions()
    session = find_session(data, sid)
    if not session:
        abort(404)
    ensure_jury_defaults(session)
    before = len(session["jurys"])
    session["jurys"] = [j for j in session["jurys"] if j.get("id") != jid]
    after = len(session["jurys"])
    save_sessions(data)
    if before == after:
        flash("Jury introuvable.", "error")
    else:
        flash("Jury supprimé.", "success")
    return redirect(url_for("session_detail", sid=sid))


@app.route("/jury-response/<sid>/<jid>/<response>")
def jury_response(sid, jid, response):
    token = request.args.get("token", "")
    if response not in ("present", "absent"):
        abort(400)
    data = load_sessions()
    session = find_session(data, sid)
    if not session:
        abort(404)
    ensure_jury_defaults(session)
    jury = next((j for j in session["jurys"] if j.get("id") == jid), None)
    if not jury or jury.get("token") != token:
        abort(403)
    already_responded = jury.get("status") in ("present", "absent")
    previous_status = jury.get("status")
    if not already_responded:
        jury["status"] = response
        save_sessions(data)
    return render_template(
        "jury_response.html",
        title="Réponse jury",
        response=previous_status if already_responded else response,
        already_responded=already_responded,
        jury=jury,
        session=session
    )


# ------------------------------------------------------------
# 🔐 Route spéciale préfecture : accès en lecture seule
# ------------------------------------------------------------
@app.route("/prefecture/session/<sid>")
@pref_auth_required
def prefecture_session(sid):
    data = load_sessions()
    session = find_session(data, sid)
    if not session:
        abort(404)

    # recalcul des statuts
    statuses = []
    for i in range(len(session["steps"])):
        st, dl = status_for_step(i, session)
        statuses.append({
            "status": st,
            "deadline": (dl.strftime("%Y-%m-%d") if dl else None)
        })

    order = sorted(
        range(len(session["steps"])),
        key=lambda i: deadline_for(i, session) or datetime.max
    )

    # page dédiée "prefecture_session.html"
    return render_template(
        "prefecture_session.html",
        title=f"Dossier session — Préfecture",
        s=session,
        statuses=statuses,
        order=order,
        now=datetime.now
    )

@app.route("/formateurs/<fid>/edit", methods=["GET", "POST"])
def edit_formateur(fid):
    formateurs = load_formateurs()
    profils_docs_config = load_formateur_profils_docs_config()

    formateur = next((f for f in formateurs if f["id"] == fid), None)
    if not formateur:
        abort(404)

    if request.method == "POST":
        formateur["nom"] = request.form.get("nom", "").strip()
        formateur["prenom"] = request.form.get("prenom", "").strip()
        try:
            formateur["nub"] = normalize_formateur_nub(request.form.get("nub", ""))
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template(
                "edit_formateur.html",
                formateur=formateur,
                formateur_profile_options=FORMATEUR_PROFILE_OPTIONS
            )
        formateur["email"] = request.form.get("email", "").strip()
        formateur["telephone"] = request.form.get("telephone", "").strip()
        formateur["siret"] = request.form.get("siret", "").strip()
        formateur["adresse_postale"] = request.form.get("adresse_postale", "").strip()
        formateur["nda"] = request.form.get("nda", "").strip()
        formateur["tarif_journalier_ht"] = request.form.get("tarif_journalier_ht", "").strip()
        formateur["profils"] = normalize_formateur_profils(
            request.form.getlist("profils")
        )
        apply_profile_document_requirements(formateur, profils_docs_config)

        save_formateurs(formateurs)
        return redirect(url_for("formateurs_home"))

    return render_template(
        "edit_formateur.html",
        formateur=formateur,
        formateur_profile_options=FORMATEUR_PROFILE_OPTIONS
    )



@app.route("/sessions/<sid>/edit", methods=["GET","POST"])
def edit_session(sid):
    data = load_sessions()
    session = find_session(data, sid)
    if not session:
        abort(404)
    if request.method == "POST":
        session["date_debut"] = request.form.get("date_debut","").strip()
        session["date_fin"] = request.form.get("date_fin","").strip()
        session["date_exam"] = request.form.get("date_exam","").strip()
        save_sessions(data)
        flash("Session mise à jour.","ok")
        return redirect(url_for("session_detail", sid=sid))
    return render_template("session_edit.html", s=session)

@app.route("/sessions/<sid>/toggle_step", methods=["POST"])
def toggle_step(sid):
    idx = int(request.form.get("index","-1"))
    data = load_sessions()
    session = find_session(data, sid)
    if not session or idx<0 or idx>=len(session["steps"]):
        abort(400)
    step = session["steps"][idx]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    step["done"] = not step["done"]
    step["done_at"] = now if step["done"] else None
    auto_archive_if_all_done(session)
    save_sessions(data)
    return redirect(url_for("session_detail", sid=sid) + f"#step{idx}")

@app.route("/sessions/<sid>/update_date", methods=["POST"])
def update_step_date(sid):
    """Permet de modifier la date fixe d'une étape dans la session GENERAL et la sauvegarder."""
    idx = int(request.form.get("index", "-1"))
    new_date = request.form.get("new_date", "").strip()
    data = load_sessions()
    session = find_session(data, sid)
    if not session or idx < 0 or idx >= len(session["steps"]):
        abort(400)

    if session.get("formation") != "GENERAL":
        flash("❌ Modification de date réservée à la session GENERAL.", "error")
        return redirect(url_for("session_detail", sid=sid))

    try:
        # ✅ On crée un champ 'custom_date' pour cette étape
        session["steps"][idx]["custom_date"] = new_date
        save_sessions(data)
        flash(f"✅ Date mise à jour pour « {session['steps'][idx]['name']} »", "ok")
    except Exception as e:
        flash(f"❌ Erreur modification date : {e}", "error")

    return redirect(url_for("session_detail", sid=sid))


@app.route("/sessions/<sid>/rename", methods=["POST"])
def rename_session(sid):
    data = load_sessions()
    session = find_session(data, sid)
    if not session:
        return {"ok": False, "error": "Session introuvable"}, 404

    payload = request.get_json(silent=True) or {}
    raw_name = payload.get("name", request.form.get("name", ""))
    name = (raw_name or "").strip()

    if not name:
        return {"ok": False, "error": "Le nom ne peut pas être vide."}, 400
    if len(name) > 80:
        return {"ok": False, "error": "Le nom est trop long (80 caractères max)."}, 400

    session["display_name"] = name
    save_sessions(data)
    return {"ok": True, "name": name}


@app.route("/sessions/<sid>/delete", methods=["POST"])
def delete_session(sid):
    data = load_sessions()
    data["sessions"] = [s for s in data["sessions"] if s["id"]!=sid]
    save_sessions(data)
    flash("Session supprimée.","ok")
    return redirect(url_for("sessions_home"))

@app.post("/sessions/<sid>/planning/upload")
def upload_planning_pdf(sid):
    f = request.files.get("planning_pdf")
    if not f or f.filename == "":
        flash("❌ Aucun fichier reçu.", "error")
        return redirect(url_for("session_detail", sid=sid))

    # sécurité : on force PDF
    if not f.filename.lower().endswith(".pdf"):
        flash("❌ Le fichier doit être un PDF.", "error")
        return redirect(url_for("session_detail", sid=sid))

    saved_name = f"planning_session_{sid}.pdf"
    path = os.path.join(PLANNING_DIR, saved_name)
    f.save(path)

    set_planning_for_session(sid, saved_name)
    flash("✅ Planning PDF enregistré.", "ok")
    return redirect(url_for("session_detail", sid=sid))

@app.post("/api/sessions/<sid>/generate-aps-planning")
def generate_aps_planning_route(sid):
    data = load_sessions()
    session_data = find_session(data, sid)
    if not session_data:
        return jsonify({"ok": False, "error": "Session introuvable."}), 404
    try:
        training_code = normalize_training_code(session_data)
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    is_desp = training_code == "DESP"
    is_ssiap1 = training_code == "SSIAP1"
    is_afc = training_code == "AFC_APS_SSIAP"
    if training_code not in {"APS", "SSIAP1", "DESP", "AFC_APS_SSIAP"}:
        return jsonify({"ok": False, "error": "Le planning automatique est réservé aux sessions APS, SSIAP 1, DESP ou AFC APS + SSIAP."}), 400

    payload = request.get_json(silent=True) or {}
    planning_mode = "desp" if is_desp else ("ssiap1" if is_ssiap1 else ("full_presentiel" if is_afc else (payload.get("planningMode") or "").strip()))
    formateur = (payload.get("trainer") or payload.get("formateur") or "").strip()
    room = (payload.get("room") or "Intégrale Academy – 54 chemin du Carreou – 83480 PUGET-SUR-ARGENS").strip() or "Intégrale Academy – 54 chemin du Carreou – 83480 PUGET-SUR-ARGENS"
    if not is_desp and not is_ssiap1 and not is_afc and planning_mode not in {"full_presentiel", "elearning_presentiel"}:
        return jsonify({"ok": False, "error": "Le type de planning APS est obligatoire."}), 400
    if not formateur and not is_desp and not is_afc:
        return jsonify({"ok": False, "error": "Le nom et prénom du formateur sont obligatoires."}), 400
    if not is_afc and not parse_date(session_data.get("date_exam")):
        return jsonify({"ok": False, "error": "La date d'examen est obligatoire pour générer le planning."}), 400

    filename = f"planning_{'afc_aps_ssiap' if is_afc else ('desp' if is_desp else ('ssiap1' if is_ssiap1 else 'aps'))}_session_{sid}.pdf"
    output_path = os.path.join(PLANNING_DIR, filename)
    temp_path = f"{output_path}.tmp"
    try:
        session_data["salle"] = room
        if is_afc:
            interruption_payload = payload.get("interruptions") if "interruptions" in payload else session_data.get("interruptions")
            interruptions = parse_interruption_ranges(interruption_payload)
            start_dt = parse_date(session_data.get("date_debut"))
            if not start_dt:
                raise ValueError("La date de début AFC est invalide.")
            contractual_raw = payload.get("contractual_end_date") or payload.get("date_fin_contractuelle") or payload.get("date_fin") or session_data.get("contractual_end_date") or session_data.get("date_fin_contractuelle") or session_data.get("date_fin")
            contractual_dt = parse_date(contractual_raw) if contractual_raw else None
            auto_end_date = afc_nth_working_day(start_dt.date(), interruptions, 57)
            if not auto_end_date:
                raise ValueError("Impossible de déterminer la 57e date admissible AFC.")
            if not contractual_dt:
                contractual_date = auto_end_date
            else:
                contractual_date = contractual_dt.date()
                eligible_count = 0
                day = start_dt.date()
                while day <= contractual_date:
                    if is_afc_working_day(day, interruptions):
                        eligible_count += 1
                    day += timedelta(days=1)
                if eligible_count != 57 or not is_afc_working_day(contractual_date, interruptions):
                    contractual_date = auto_end_date
            session_data["contractual_end_date"] = contractual_date.isoformat()
            planning_data = build_afc_aps_ssiap_planning_data(start_dt.date(), formateur, room, interruptions, contractual_end_date=contractual_date)
            summary = afc_aps_ssiap_summary_from_data(planning_data, interruptions, contractual_end_date=contractual_date)
            session_data["interruptions"] = "\n".join(f"{start.isoformat()} au {end.isoformat()}" for start, end in interruptions)
            session_data["date_fin"] = planning_data[-1]["date"]
            session_data["date_exam"] = planning_data[-1]["date"]
            result = generate_aps_planning_pdf(session_data, formateur, temp_path, planning_data=planning_data, planning_mode="full_presentiel", document_profile={"validate":"afc_aps_ssiap", "summary": summary, "planning_title":"PLANNING AFC FRANCE TRAVAIL APS + SSIAP", "short_label":"AFC APS + SSIAP"})
        elif is_ssiap1:
            exam = ssiap1_exam_payload(session_data, payload)
            session_data.update({"date_exam": exam["date"], "exam_date": exam["date"], "exam_start_time": exam["start"], "exam_end_time": exam["end"], "exam_room": exam["room"], "ssiapExamStartTime": exam["start"], "ssiapExamEndTime": exam["end"], "ssiapExamRoom": exam["room"], "ssiapSstTrainer": exam.get("sstTrainer") or formateur, "ssiapTrainer": exam.get("ssiapTrainer") or formateur, "ssiapRevisionTrainer": exam.get("revisionTrainer") or formateur, "ssiapExamTrainer": exam.get("examTrainer") or formateur})
            end_dt = parse_date(session_data.get("date_fin"))
            planning_data, totals, total_hours = build_ssiap1_planning_data(parse_date(session_data.get("date_debut")).date(), formateur, room, end_date=end_dt.date() if end_dt else None, exam_iso=exam["date"], exam_payload=exam, excluded_dates=ssiap1_excluded_dates_from_payload(session_data, payload))
            summary = ssiap1_summary_from_data(planning_data)
            result = generate_aps_planning_pdf(session_data, formateur, temp_path, planning_data=planning_data, planning_mode="ssiap1", document_profile={"validate": "ssiap1", "summary": summary, "planning_title": "PLANNING DE FORMATION SSIAP 1", "short_label": "SSIAP 1"})
        elif is_desp:
            elearning_start = parse_date(payload.get("despElearningStart") or session_data.get("despElearningStart") or session_data.get("date_debut"))
            elearning_end = parse_date(payload.get("despElearningEnd") or session_data.get("despElearningEnd") or session_data.get("date_distanciel_fin") or session_data.get("date_elearning_fin"))
            presentiel_start = parse_date(payload.get("despPresentielStart") or session_data.get("despPresentielStart") or session_data.get("date_presentiel_debut"))
            presentiel_end = parse_date(payload.get("despPresentielEnd") or session_data.get("despPresentielEnd") or session_data.get("date_fin"))
            if not all([elearning_start, elearning_end, presentiel_start, presentiel_end]):
                return jsonify({"ok": False, "error": "Les dates de début/fin distanciel et début/fin présentiel DESP sont obligatoires."}), 400
            allow_saturday = bool(payload.get("allowSaturday") or session_data.get("despAllowSaturday"))
            planning_data = generate_desp_planning(elearning_start.date(), elearning_end.date(), presentiel_start.date(), presentiel_end.date(), formateur, room, exam_iso=aps_local_date_iso(session_data.get("date_exam")), allow_saturday=allow_saturday)
            session_data["despAllowSaturday"] = allow_saturday
            summary = desp_summary_from_planning(planning_data)
            result = generate_aps_planning_pdf(session_data, formateur, temp_path, planning_data=planning_data, planning_mode="desp", document_profile={"validate": "desp", "summary": summary, "planning_title": "PLANNING DE FORMATION DESP", "short_label": "DESP"})
        else:
            result = generate_aps_planning_pdf(session_data, formateur, temp_path, planning_mode=planning_mode)
        expected_total = AFC_APS_SSIAP_TOTAL_HOURS if is_afc else (SSIAP1_TOTAL_HOURS if is_ssiap1 else (DESP_TOTAL_HOURS if is_desp else APS_TOTAL_HOURS))
        if round(result["total_hours"], 2) != expected_total :
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return jsonify({"ok": False, "error": f"Le total généré n'est pas exactement de {expected_total}h."}), 500
        exam_iso = aps_local_date_iso(session_data.get("date_exam"))
        if not is_ssiap1 and not is_afc and any(day.get("date") == exam_iso for day in result.get("planning_data", [])):
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return jsonify({"ok": False, "error": f"Sécurité planning APS: la date d'examen ({format_date(exam_iso)}) est réservée à l’examen et ne peut contenir aucun créneau de formation."}), 400
        os.replace(temp_path, output_path)
        session_data["planning_pdf"] = filename
        session_data["apsPlanningData"] = result["planning_data"]
        session_data["apsPlanningSummary"] = result["summary"]
        session_data["apsPlanningMode"] = planning_mode
        session_data["planning_generated_at"] = append_planning_history(session_data, "planning généré")
        save_sessions(data)
        app.logger.info(
            "Planning APS généré session=%s date_debut=%s date_fin=%s date_exam=%s jours=%s total=%sh uv_totals=%s",
            sid,
            session_data.get("date_debut"),
            session_data.get("date_fin"),
            session_data.get("date_exam"),
            len(result["planning_data"]),
            result["total_hours"],
            result["totals"],
        )
        return jsonify({
            "ok": True,
            "url": url_for("view_planning_pdf", sid=sid),
            "filename": filename,
            "generated_at": session_data["planning_generated_at"],
            "warnings": (result.get("summary") or {}).get("warnings") or [],
        })
    except ValueError as exc:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        app.logger.warning("Génération planning APS impossible session=%s erreur=%s", sid, exc)
        return jsonify({"ok": False, "error": str(exc)}), 400
    except Exception as exc:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        app.logger.exception("Erreur génération planning APS session=%s", sid)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.get("/api/sessions/<sid>/aps-trainer-contracts/preview")
def preview_aps_trainer_contracts(sid):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    if (session_data.get("formation") or "").upper() not in {"APS", "DESP", "DIRIGEANT", "AFC_APS_SSIAP"} and not is_ssiap1_session(session_data): return jsonify({"ok": False, "error": "La session n'est pas APS/DESP."}), 400
    planning_data = session_data.get("apsPlanningData") or []
    if not session_data.get("planning_pdf") or not planning_data: return jsonify({"ok": False, "error": "Veuillez générer le planning APS avant de générer un contrat formateur."}), 400
    trainers = []
    for name in aps_detect_trainers(planning_data):
        calc = aps_trainer_interventions(planning_data, name)
        formateur = find_formateur_by_identity(name=name)
        trainers.append({"name": name, **calc, "defaults": formateur_contract_defaults(formateur)})
    return jsonify({"ok": True, "trainers": trainers})


def ensure_aps_trainer_contract_pdf(session_data, contract):
    """Return the local APS trainer contract PDF path, regenerating it if needed.

    Render's filesystem can be reset between deployments/restarts while the
    session metadata remains in persistent storage. In that case, existing
    contract links should keep working instead of returning a bare 404.
    """
    filename = os.path.basename(contract.get("pdfFilename") or "")
    if not filename:
        filename = f"contrat_formateur_aps_{session_data.get('id') or 'session'}_{contract.get('id') or uuid.uuid4()}.pdf"
        contract["pdfFilename"] = filename
    path = os.path.join(APS_CONTRACT_DIR, filename)
    if not os.path.exists(path):
        os.makedirs(APS_CONTRACT_DIR, exist_ok=True)
        generate_aps_trainer_contract_pdf(session_data, contract, path)
    return path


@app.get("/sessions/<sid>/aps-trainer-contracts/<contract_id>/view")
def view_aps_trainer_contract(sid, contract_id):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: abort(404)
    contract = next((c for c in session_data.get("apsTrainerContracts", []) if c.get("id") == contract_id), None)
    if not contract: abort(404)
    try:
        path = ensure_aps_trainer_contract_pdf(session_data, contract)
    except Exception:
        app.logger.exception("Régénération contrat APS impossible session=%s contrat=%s", sid, contract_id)
        abort(404)
    return send_file(path, mimetype="application/pdf", as_attachment=False)


@app.post("/api/sessions/<sid>/aps-trainer-contracts/generate")
def generate_aps_trainer_contracts(sid):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    if (session_data.get("formation") or "").upper() not in {"APS", "DESP", "DIRIGEANT", "AFC_APS_SSIAP"} and not is_ssiap1_session(session_data): return jsonify({"ok": False, "error": "La session n'est pas APS/DESP."}), 400
    planning_data = session_data.get("apsPlanningData") or []
    if not session_data.get("planning_pdf") or not planning_data: return jsonify({"ok": False, "error": "Veuillez générer le planning APS avant de générer un contrat formateur."}), 400
    payload = request.get_json(silent=True) or {}; trainers = payload.get("trainers") or []
    if not trainers: return jsonify({"ok": False, "error": "Aucun formateur sélectionné."}), 400
    saved = []
    for trainer in trainers:
        name = (trainer.get("name") or "").strip(); planning_name = (trainer.get("planningName") or name).strip(); daily_rate = float(trainer.get("dailyRate") or 0)
        if not name or daily_rate <= 0: return jsonify({"ok": False, "error": "Le nom et un tarif journalier HT supérieur à 0 sont obligatoires."}), 400
        calc = aps_trainer_interventions(planning_data, planning_name)
        if not calc["interventions"]: return jsonify({"ok": False, "error": f"Aucun créneau trouvé pour {planning_name}."}), 400
        billed_days = float(trainer.get("billedDays") or calc["calculatedDays"] or 0)
        trainer_attends_exam = bool(trainer.get("trainerAttendsExam") or trainer.get("trainer_attends_exam"))
        exam_trainer_hours = float(trainer.get("examTrainerHours") or trainer.get("exam_trainer_hours") or 0) if trainer_attends_exam else 0.0
        exam_trainer_rate = float(trainer.get("examTrainerRate") or trainer.get("exam_trainer_rate") or 0) if trainer_attends_exam else 0.0
        exam_trainer_amount = round(float(trainer.get("examTrainerAmount") or trainer.get("exam_trainer_amount") or (exam_trainer_hours * exam_trainer_rate) or 0), 2) if trainer_attends_exam else 0.0
        vat_enabled = bool(trainer.get("vatEnabled")); vat_rate = float(trainer.get("vatRate") or 20)
        total_ht = round((billed_days * daily_rate) + exam_trainer_amount, 2); vat_amount = round(total_ht * vat_rate / 100, 2) if vat_enabled else 0; total_ttc = round(total_ht + vat_amount, 2)
        contract_id = str(uuid.uuid4()); filename = f"contrat_formateur_aps_{sid}_{contract_id}.pdf"; path = os.path.join(APS_CONTRACT_DIR, filename)
        trainer = merge_formateur_contract_defaults(trainer, find_formateur_by_identity(name=name, email=trainer.get("email")))
        contract = {"id": contract_id, "trainerName": name, "trainerEmail": (trainer.get("email") or "").strip(), "trainerPhone": (trainer.get("phone") or "").strip(), "dailyRate": daily_rate, "calculatedHours": calc["totalHours"], "calendarDays": calc["calendarDays"], "calculatedDays": calc["calculatedDays"], "billedDays": billed_days, "trainerAttendsExam": trainer_attends_exam, "examTrainerHours": exam_trainer_hours, "examTrainerRate": exam_trainer_rate, "examTrainerAmount": exam_trainer_amount, "totalHT": total_ht, "vatEnabled": vat_enabled, "vatRate": vat_rate, "vatAmount": vat_amount, "totalTTC": total_ttc, "address": (trainer.get("address") or "").strip(), "siret": (trainer.get("siret") or "").strip(), "status": (trainer.get("status") or "").strip(), "commercialName": (trainer.get("commercialName") or "").strip(), "activityDeclaration": (trainer.get("activityDeclaration") or "").strip(), "vatNumber": (trainer.get("vatNumber") or "").strip(), "vatMention": (trainer.get("vatMention") or "").strip(), "rcPro": (trainer.get("rcPro") or "").strip(), "urssafVigilance": (trainer.get("urssafVigilance") or "").strip(), "rneKbis": (trainer.get("rneKbis") or "").strip(), "rib": (trainer.get("rib") or "").strip(), "diplomas": (trainer.get("diplomas") or "").strip(), "cv": (trainer.get("cv") or "").strip(), "interventions": calc["interventions"], "pdfFilename": filename, "pdfUrl": url_for("view_aps_trainer_contract", sid=sid, contract_id=contract_id), "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "sentAt": None}
        generate_aps_trainer_contract_pdf(session_data, contract, path)
        existing_contracts = session_data.setdefault("apsTrainerContracts", [])
        kept_contracts = []
        for existing in existing_contracts:
            existing_planning_name = (existing.get("planningName") or existing.get("trainerName") or "").strip()
            if existing_planning_name == planning_name:
                old_path = os.path.join(APS_CONTRACT_DIR, os.path.basename(existing.get("pdfFilename") or ""))
                if old_path != path and os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except OSError:
                        app.logger.warning("Suppression ancien contrat APS impossible: %s", old_path)
            else:
                kept_contracts.append(existing)
        kept_contracts.append(contract)
        session_data["apsTrainerContracts"] = kept_contracts
        saved.append(contract)
    save_sessions(data)
    return jsonify({"ok": True, "contracts": saved})


@app.post("/api/sessions/<sid>/aps-trainer-contracts/<contract_id>/send")
def send_aps_trainer_contract(sid, contract_id):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    contract = next((c for c in session_data.get("apsTrainerContracts", []) if c.get("id") == contract_id), None)
    if not contract: return jsonify({"ok": False, "error": "Contrat introuvable."}), 404
    if not contract.get("trainerEmail"): return jsonify({"ok": False, "error": "Email formateur manquant."}), 400
    try:
        contract_path = ensure_aps_trainer_contract_pdf(session_data, contract)
    except Exception as exc:
        app.logger.exception("Régénération contrat APS impossible avant envoi mail session=%s contrat=%s", sid, contract_id)
        return jsonify({"ok": False, "error": f"PDF contrat introuvable et régénération impossible: {exc}"}), 400
    planning_name = session_data.get("planning_pdf"); planning_path = os.path.join(PLANNING_DIR, os.path.basename(planning_name or ""))
    if not planning_name or not os.path.exists(planning_path): return jsonify({"ok": False, "error": "PDF planning APS complet introuvable."}), 400
    payload = request.get_json(silent=True) or {}; subject = payload.get("emailSubject") or "Contrat d’intervention formateur — Session APS"; body = payload.get("emailBody") or ""
    ok, message = send_email_with_attachments(contract["trainerEmail"], subject, body, [(contract_path, os.path.basename(contract_path)), (planning_path, os.path.basename(planning_path))])
    if not ok: return jsonify({"ok": False, "error": message}), 500
    contract["sentAt"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S"); save_sessions(data)
    return jsonify({"ok": True, "sentAt": contract["sentAt"]})


def inspect_yousign_pdf_before_upload(pdf_path):
    info = {"path": pdf_path, "size": os.path.getsize(pdf_path) if os.path.exists(pdf_path) else 0, "page_count": None, "signature_tag_present": None}
    try:
        if importlib.util.find_spec("pypdf") is None:
            app.logger.warning("Inspection PDF Yousign ignorée: module pypdf absent")
            return info
        import pypdf

        reader = pypdf.PdfReader(pdf_path)
        info["page_count"] = len(reader.pages)
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        info["signature_tag_present"] = YOUSIGN_TRAINER_SIGNATURE_TAG in text
    except Exception as exc:
        app.logger.warning("Inspection PDF Yousign impossible, envoi poursuivi: %s", exc)
    return info


YOUSIGN_APS_TRAINER_SIGNATURE_FIELD = {"x": 332, "y": 216, "width": 160, "height": 60}
YOUSIGN_APS_TRAINER_NO_FIELD_ERROR = "Le signataire Yousign a été créé mais aucun champ de signature n’a été ajouté au document."


def yousign_trainer_signature_page(pdf_info):
    page_count = int(pdf_info.get("page_count") or 0)
    if page_count < 1:
        raise YousignError("Nombre de pages du PDF APS introuvable: activation Yousign refusée.")
    return page_count


def validate_aps_trainer_signature_field(field_payload, page_count):
    page = int(field_payload.get("page") or 0)
    x = int(field_payload.get("x") or 0)
    y = int(field_payload.get("y") or 0)
    width = int(field_payload.get("width") or 0)
    height = int(field_payload.get("height") or 0)
    logger.info("Yousign APS trainer signature field control page=%s x=%s y=%s width=%s height=%s", page, x, y, width, height)
    if page != int(page_count):
        raise YousignError("Champ signature formateur hors dernière page: activation Yousign refusée.")
    if not (305 <= x <= 518 and 193 <= y <= 282):
        raise YousignError("Champ signature formateur hors zone attendue: activation Yousign refusée.")
    return True


@app.post("/api/sessions/<sid>/aps-trainer-contracts/<contract_id>/yousign/send")
def send_aps_trainer_contract_yousign(sid, contract_id):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    contract = next((c for c in session_data.get("apsTrainerContracts", []) if c.get("id") == contract_id), None)
    if not contract: return jsonify({"ok": False, "error": "Contrat introuvable."}), 404
    email = (contract.get("trainerEmail") or "").strip()
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email): return jsonify({"ok": False, "error": "Email formateur invalide ou manquant."}), 400
    if not is_yousign_configured(): return jsonify({"ok": False, "error": "Yousign n'est pas configuré: renseignez YOUSIGN_API_KEY côté serveur."}), 400

    state = normalize_yousign_state(contract.get("yousign"))
    if state.get("signatureRequestId") and state.get("status") in {"draft", "approval", "ongoing"} and not request.args.get("force"):
        return jsonify({"ok": False, "error": "Une demande Yousign active existe déjà pour ce contrat."}), 409

    try:
        contract_path = ensure_aps_trainer_contract_pdf(session_data, contract)
    except Exception as exc:
        app.logger.exception("Régénération contrat APS impossible avant envoi Yousign session=%s contrat=%s", sid, contract_id)
        return jsonify({"ok": False, "error": f"PDF contrat introuvable et régénération impossible: {exc}"}), 400

    client = YousignClient()
    now = datetime.now().isoformat(timespec="seconds")
    try:
        trainer_name = contract.get("trainerName") or email
        external_id = sanitize_yousign_external_id(f"aps-trainer-contract-{contract_id}")
        app.logger.info("Yousign APS trainer contract external_id=%s", external_id)
        signature_request = client.create_signature_request(f"Contrat formateur APS - {trainer_name}", external_id=external_id)
        signature_request_id = signature_request.get("id")
        try:
            pdf_info = inspect_yousign_pdf_before_upload(contract_path)
            app.logger.info(
                "Yousign APS trainer PDF before upload path=%s size=%s signature_tag_present=%s page_count=%s",
                pdf_info["path"], pdf_info["size"], pdf_info["signature_tag_present"], pdf_info["page_count"]
            )
        except Exception as exc:
            app.logger.warning("Inspection PDF Yousign impossible, envoi poursuivi: %s", exc)
            pdf_info = {"path": contract_path, "size": os.path.getsize(contract_path) if os.path.exists(contract_path) else 0, "page_count": None, "signature_tag_present": None}
        with open(contract_path, "rb") as pdf_file:
            document = client.upload_file(signature_request_id, pdf_file.read(), os.path.basename(contract_path), parse_anchors=False)
        document_id = document.get("id")
        app.logger.info("Yousign APS trainer document uploaded document_id=%s", document_id)
        name_parts = str(trainer_name).split()
        first_name = name_parts[0] if len(name_parts) > 1 else ""
        last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else trainer_name
        normalized_phone = normalizeFrenchPhoneNumber(contract.get("yousignPhoneNumber") or contract.get("trainerPhone") or contract.get("phone") or contract.get("telephone") or "")
        app.logger.info("Yousign APS trainer signer authentication_mode=otp_sms phone_number=%s", mask_phone_number(normalized_phone))
        signer = client.add_signer(signature_request_id, first_name, last_name or trainer_name, email, document_id=document_id, use_text_tags=True, phone_number=normalized_phone, force_sms_otp=True)
        signer_id = signer.get("id") or ""
        app.logger.info("Yousign APS trainer signer created signer_id=%s document_id=%s", signer_id, document_id)
        page_count = yousign_trainer_signature_page(pdf_info)
        field_payload = {**YOUSIGN_APS_TRAINER_SIGNATURE_FIELD, "page": page_count}
        validate_aps_trainer_signature_field(field_payload, page_count)
        field = client.add_signature_field(
            signature_request_id,
            document_id,
            signer_id,
            page=field_payload["page"],
            x=field_payload["x"],
            y=field_payload["y"],
            width=field_payload["width"],
            height=field_payload["height"],
        )
        field_id = field.get("id") if isinstance(field, dict) else ""
        app.logger.info(
            "Yousign APS trainer signature field created field_id=%s signer_id=%s document_id=%s page=%s x=%s y=%s width=%s height=%s",
            field_id, signer_id, document_id, field_payload["page"], field_payload["x"], field_payload["y"], field_payload["width"], field_payload["height"]
        )
        if not field_id:
            raise YousignError(YOUSIGN_APS_TRAINER_NO_FIELD_ERROR, payload={"message": YOUSIGN_APS_TRAINER_NO_FIELD_ERROR})
        activated = client.activate_signature_request(signature_request_id)
        status = extract_yousign_status(activated) or "ongoing"
        signature_url = signer.get("signature_link") or signer.get("signature_url") or activated.get("signature_link") or ""
        contract["yousign"] = normalize_yousign_state({
            "signatureRequestId": signature_request_id,
            "externalId": external_id,
            "documentId": document_id or "",
            "signerId": signer_id,
            "fieldId": field_id,
            "status": status,
            "signatureUrl": signature_url,
            "sentAt": now,
            "lastSyncedAt": now,
            "lastEvent": "signature_request.activated",
            "lastEventAt": now,
            "recipientEmail": email,
            "error": None,
        })
        mirror_yousign_state_on_contract(contract)
        save_sessions(data)
        return jsonify({"ok": True, "status": status, "sentAt": now, "signatureUrl": signature_url})
    except YousignError as exc:
        logger.error("Réponse exacte Yousign APS contract 400/erreur status=%s payload=%r", exc.status_code, exc.payload)
        user_error = yousign_service_access_message(exc.status_code, exc.payload)
        contract["yousign"] = normalize_yousign_state({**state, "status": "error", "lastSyncedAt": now, "lastEvent": "api.error", "lastEventAt": now, "error": user_error, "errorPayload": exc.payload})
        mirror_yousign_state_on_contract(contract)
        save_sessions(data)
        return jsonify({"ok": False, "error": user_error, "errorPayload": exc.payload, "yousignStatus": exc.status_code}), 502


@app.post("/api/sessions/<sid>/aps-trainer-contracts/<contract_id>/yousign/sync")
def sync_aps_trainer_contract_yousign(sid, contract_id):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    contract = next((c for c in session_data.get("apsTrainerContracts", []) if c.get("id") == contract_id), None)
    if not contract: return jsonify({"ok": False, "error": "Contrat introuvable."}), 404
    state = normalize_yousign_state(contract.get("yousign"))
    signature_request_id = state.get("signatureRequestId")
    if not signature_request_id: return jsonify({"ok": False, "error": "Aucune demande Yousign à actualiser."}), 400
    now = datetime.now().isoformat(timespec="seconds")
    try:
        updates = sync_yousign_signature_request_from_api(signature_request_id, now)
        status = updates.get("status") or state.get("status")
        updates.update({"lastEvent": "manual.sync" if updates.get("lastSyncedAt") else "manual.sync.error", "lastEventAt": now})
        contract["yousign"] = normalize_yousign_state({**state, **updates})
        mirror_yousign_state_on_contract(contract)
        save_sessions(data)
        if not updates.get("lastSyncedAt"):
            return jsonify({"ok": False, "error": updates.get("apiError") or "Erreur API Yousign", "status": status, "statusLabel": yousign_status_label(status)}), 502
        return jsonify({"ok": True, "status": status, "statusLabel": yousign_status_label(status)})
    except YousignError as exc:
        contract["yousign"] = normalize_yousign_state({**state, "lastEvent": "manual.sync.error", "lastEventAt": now, "apiError": str(exc), "apiHttpStatus": str(exc.status_code or "network"), "apiStatus": f"erreur {exc.status_code or 'network'}", "error": str(exc)})
        mirror_yousign_state_on_contract(contract)
        save_sessions(data)
        return jsonify({"ok": False, "error": f"Erreur de synchronisation Yousign: {exc}"}), 502


def save_yousign_signed_document(content, target_dir, base_filename):
    os.makedirs(target_dir, exist_ok=True)
    pdf_bytes = None
    if content.startswith(b"%PDF"):
        pdf_bytes = content
    else:
        try:
            with zipfile.ZipFile(BytesIO(content)) as archive:
                pdf_names = [name for name in archive.namelist() if name.lower().endswith(".pdf") and not name.endswith("/")]
                if pdf_names:
                    pdf_names.sort(key=lambda name: ("signed" not in name.lower() and "signe" not in name.lower(), name.lower()))
                    pdf_bytes = archive.read(pdf_names[0])
        except zipfile.BadZipFile:
            pdf_bytes = None
    if pdf_bytes:
        filename = f"{base_filename}.pdf"
        with open(os.path.join(target_dir, filename), "wb") as fh:
            fh.write(pdf_bytes)
        return filename
    filename = f"{base_filename}.zip"
    with open(os.path.join(target_dir, filename), "wb") as fh:
        fh.write(content)
    return filename


@app.route("/api/sessions/<sid>/aps-trainer-contracts/<contract_id>/yousign/download", methods=["GET", "POST"])
def download_aps_trainer_signed_yousign(sid, contract_id):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    contract = next((c for c in session_data.get("apsTrainerContracts", []) if c.get("id") == contract_id), None)
    if not contract: return jsonify({"ok": False, "error": "Contrat introuvable."}), 404
    state = normalize_yousign_state(contract.get("yousign"))
    if not state.get("signatureRequestId"): return jsonify({"ok": False, "error": "Aucune demande Yousign disponible."}), 400
    try:
        content = YousignClient().download_signed_documents(state["signatureRequestId"])
        filename = save_yousign_signed_document(content, APS_CONTRACT_SIGNED_DIR, f"contrat_aps_signe_yousign_{state['signatureRequestId']}")
        contract["yousign"] = normalize_yousign_state({**state, "signedDocumentFilename": filename, "signedDocumentUrl": url_for("download_aps_trainer_signed_yousign_file", filename=filename), "lastSyncedAt": datetime.now().isoformat(timespec="seconds"), "error": None})
        mirror_yousign_state_on_contract(contract)
        save_sessions(data)
        return send_from_directory(APS_CONTRACT_SIGNED_DIR, filename, as_attachment=request.args.get("inline") != "1")
    except Exception as exc:
        return jsonify({"ok": False, "error": f"Téléchargement Yousign impossible: {exc}"}), 502


@app.get("/aps-trainer-contracts/yousign/signed/<path:filename>")
def download_aps_trainer_signed_yousign_file(filename):
    return send_from_directory(APS_CONTRACT_SIGNED_DIR, os.path.basename(filename), as_attachment=True)


@app.patch("/api/sessions/<sid>/aps-trainer-contracts/<contract_id>/email")
def update_aps_trainer_contract_email(sid, contract_id):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    contract = next((c for c in session_data.get("apsTrainerContracts", []) if c.get("id") == contract_id), None)
    if not contract: return jsonify({"ok": False, "error": "Contrat introuvable."}), 404
    payload = request.get_json(silent=True) or {}
    email = (payload.get("email") or "").strip()
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        return jsonify({"ok": False, "error": "Adresse e-mail invalide."}), 400
    contract["trainerEmail"] = email
    state = normalize_yousign_state(contract.get("yousign"))
    if not state.get("sentAt"):
        state["recipientEmail"] = email
        contract["yousign"] = state
        mirror_yousign_state_on_contract(contract)
    save_sessions(data)
    return jsonify({"ok": True, "email": email})


@app.patch("/api/sessions/<sid>/aps-trainer-contracts/<contract_id>/phone")
def update_aps_trainer_contract_phone(sid, contract_id):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    contract = next((c for c in session_data.get("apsTrainerContracts", []) if c.get("id") == contract_id), None)
    if not contract: return jsonify({"ok": False, "error": "Contrat introuvable."}), 404
    payload = request.get_json(silent=True) or {}
    phone = (payload.get("phone") or "").strip()
    try:
        normalized_phone = normalizeFrenchPhoneNumber(phone)
    except YousignError:
        return jsonify({"ok": False, "error": "Téléphone portable invalide pour le code SMS Yousign."}), 400
    contract["trainerPhone"] = phone
    contract["yousignPhoneNumber"] = normalized_phone
    save_sessions(data)
    return jsonify({"ok": True, "phone": phone, "normalizedPhone": normalized_phone})


@app.delete("/api/sessions/<sid>/aps-trainer-contracts/<contract_id>")
def delete_aps_trainer_contract(sid, contract_id):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    contracts = session_data.get("apsTrainerContracts", []); contract = next((c for c in contracts if c.get("id") == contract_id), None)
    if not contract: return jsonify({"ok": False, "error": "Contrat introuvable."}), 404
    if contract.get("pdfFilename"):
        try: os.remove(os.path.join(APS_CONTRACT_DIR, os.path.basename(contract["pdfFilename"])))
        except FileNotFoundError: pass
    session_data["apsTrainerContracts"] = [c for c in contracts if c.get("id") != contract_id]
    save_sessions(data)
    return jsonify({"ok": True})


@app.post("/api/sessions/<sid>/aps-attendance/import-students")
def import_aps_attendance_students(sid):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    formation = (session_data.get("formation") or "").upper()
    if formation not in {"APS", "A3P", "DESP", "DIRIGEANT", "AFC_APS_SSIAP"} and not is_ssiap1_session(session_data): return jsonify({"ok": False, "error": "Cette action est réservée aux sessions APS, A3P, SSIAP 1 et DESP."}), 400
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        return jsonify({"ok": False, "error": "Veuillez importer un fichier PDF."}), 400
    if not uploaded.filename.lower().endswith(".pdf"):
        return jsonify({"ok": False, "error": "Le fichier doit être un PDF."}), 400
    try:
        students, has_text = aps_extract_students_from_pdf(uploaded)
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500
    message = None
    if not has_text or not students:
        message = "Impossible d’extraire automatiquement les noms depuis ce PDF. Merci de saisir ou corriger la liste manuellement."
    if formation == "AFC_APS_SSIAP":
        default_start = session_data.get("date_debut") or ""
        for student in students:
            student["startDate"] = student.get("startDate") or default_start
    return jsonify({"ok": True, "students": students, "message": message})


@app.put("/api/sessions/<sid>/aps-attendance/students")
def save_aps_attendance_students(sid):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    formation = (session_data.get("formation") or "").upper()
    if formation not in {"APS", "A3P", "DESP", "DIRIGEANT", "AFC_APS_SSIAP"} and not is_ssiap1_session(session_data): return jsonify({"ok": False, "error": "Cette action est réservée aux sessions APS, A3P, SSIAP 1 et DESP."}), 400
    payload = request.get_json(silent=True) or {}
    students = []
    for item in payload.get("students") or []:
        last = (item.get("lastName") or "").strip().upper()
        first = (item.get("firstName") or "").strip()
        if last and first:
            student = {"lastName": last, "firstName": first}
            if formation == "AFC_APS_SSIAP":
                start_date = (item.get("startDate") or session_data.get("date_debut") or "").strip()
                if start_date and not re.match(r"^\d{4}-\d{2}-\d{2}$", start_date):
                    return jsonify({"ok": False, "error": "La date d’entrée en formation doit être au format AAAA-MM-JJ."}), 400
                student["startDate"] = start_date
            students.append(student)
    if not students:
        return jsonify({"ok": False, "error": "Veuillez enregistrer au moins un stagiaire."}), 400
    student_key = "a3pAttendanceStudents" if formation == "A3P" else "apsAttendanceStudents"
    updated_key = "a3pAttendanceSheetsUpdatedAt" if formation == "A3P" else "apsAttendanceSheetsUpdatedAt"
    session_data[student_key] = students
    session_data[updated_key] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_sessions(data)
    return jsonify({"ok": True, "students": students})


@app.post("/api/sessions/<sid>/aps-attendance/generate")
def generate_aps_attendance_sheets(sid):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    try:
        training_code = normalize_training_code(session_data)
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    formation = training_code
    planning_key = "a3pPlanningData" if training_code == "A3P" else "apsPlanningData"
    student_key = "a3pAttendanceStudents" if training_code == "A3P" else "apsAttendanceStudents"
    if not session_data.get(planning_key):
        message = "Le planning SSIAP 1 doit être généré avant les feuilles de présence." if is_ssiap1_session(session_data) else f"Veuillez générer le planning {formation} avant de générer les feuilles de présence."
        return jsonify({"ok": False, "error": message}), 400
    if not session_data.get(student_key):
        return jsonify({"ok": False, "error": "Aucune liste de stagiaires n’est enregistrée."}), 400
    shared_session = session_data
    if training_code == "A3P":
        shared_session, converted = _a3p_session_for_shared_docs(session_data)
        if not _aps_presentiel_days(converted, "full_presentiel"):
            return jsonify({"ok": False, "error": "Aucun jour présentiel n’est trouvé."}), 400
    elif not _aps_presentiel_days(session_data.get("apsPlanningData"), "desp" if training_code == "DESP" else ("ssiap1" if training_code == "SSIAP1" else (session_data.get("apsPlanningMode") or "full_presentiel"))):
        return jsonify({"ok": False, "error": "Aucun jour présentiel n’est trouvé."}), 400
    filename_formation = training_code.lower()
    filename = f"feuilles_presence_{filename_formation}_{sid}.pdf"
    output_dir = A3P_DOC_DIR if training_code == "A3P" else APS_ATTENDANCE_DIR
    output_path = os.path.join(output_dir, filename)
    temp_path = f"{output_path}.tmp"
    try:
        generate_a3p_attendance_pdf(session_data, temp_path) if training_code == "A3P" else generate_aps_attendance_pdf(session_data, temp_path)
        if not os.path.exists(temp_path) or os.path.getsize(temp_path) <= 0:
            raise ValueError("Le PDF généré est vide.")
        os.replace(temp_path, output_path)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if training_code == "A3P":
            session_data["a3pAttendanceSheetsPdfUrl"] = url_for("view_a3p_document", sid=sid, kind="attendance")
            session_data["a3pAttendanceSheetsFilename"] = filename
            session_data["a3pAttendanceSheetsGeneratedAt"] = session_data.get("a3pAttendanceSheetsGeneratedAt") or now
            session_data["a3pAttendanceSheetsUpdatedAt"] = now
            docs = session_data.setdefault("a3p_documents", {})
            docs["attendance"] = {"path": output_path, "generated_at": now}
        else:
            session_data["apsAttendanceSheetsPdfUrl"] = url_for("view_aps_attendance_sheets", sid=sid)
            session_data["apsAttendanceSheetsFilename"] = filename
            session_data["apsAttendanceSheetsGeneratedAt"] = session_data.get("apsAttendanceSheetsGeneratedAt") or now
            session_data["apsAttendanceSheetsUpdatedAt"] = now
        save_sessions(data)
        return jsonify({"ok": True, "pdfUrl": session_data["a3pAttendanceSheetsPdfUrl"] if formation == "A3P" else session_data["apsAttendanceSheetsPdfUrl"], "generatedAt": now})
    except Exception as exc:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        app.logger.exception("Erreur génération feuilles présence %s session=%s", formation, sid)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.delete("/api/sessions/<sid>/aps-attendance")
def reset_aps_attendance_sheets(sid):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    formation = (session_data.get("formation") or "").upper()
    if formation not in {"APS", "A3P", "DESP", "DIRIGEANT", "AFC_APS_SSIAP"} and not is_ssiap1_session(session_data): return jsonify({"ok": False, "error": "Cette action est réservée aux sessions APS, A3P, SSIAP 1 et DESP."}), 400
    payload = request.get_json(silent=True) or {}
    filename = session_data.get("a3pAttendanceSheetsFilename" if formation == "A3P" else "apsAttendanceSheetsFilename")
    if filename:
        try: os.remove(os.path.join(A3P_DOC_DIR if formation == "A3P" else APS_ATTENDANCE_DIR, os.path.basename(filename)))
        except FileNotFoundError: pass
    keys = ("a3pAttendanceSheetsPdfUrl", "a3pAttendanceSheetsFilename", "a3pAttendanceSheetsGeneratedAt", "a3pAttendanceSheetsUpdatedAt") if formation == "A3P" else ("apsAttendanceSheetsPdfUrl", "apsAttendanceSheetsFilename", "apsAttendanceSheetsGeneratedAt", "apsAttendanceSheetsUpdatedAt")
    for key in keys:
        session_data.pop(key, None)
    if formation == "A3P":
        (session_data.get("a3p_documents") or {}).pop("attendance", None)
    if payload.get("deleteStudents"):
        session_data.pop("a3pAttendanceStudents" if formation == "A3P" else "apsAttendanceStudents", None)
    save_sessions(data)
    return jsonify({"ok": True})


@app.get("/sessions/<sid>/aps-attendance/view")
def view_aps_attendance_sheets(sid):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: abort(404)
    filename = session_data.get("apsAttendanceSheetsFilename")
    if not filename: abort(404)
    path = os.path.join(APS_ATTENDANCE_DIR, os.path.basename(filename))
    if not os.path.exists(path): abort(404)
    return send_file(path, mimetype="application/pdf", as_attachment=False)



@app.get("/api/admin/sessions/<sid>/a3p-planning-builder")
def get_a3p_planning_builder(sid):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    if (session_data.get("formation") or "").upper() != "A3P": return jsonify({"ok": False, "error": "La session n'est pas A3P."}), 400
    state = session_data.get("a3pPlanningBuilder") or session_data.get("a3pPlanningDraftJson") or {}
    return jsonify({"ok": True, "state": state, "savedAt": session_data.get("a3pPlanningBuilderSavedAt") or session_data.get("a3pPlanningDraftSavedAt")})

@app.put("/api/admin/sessions/<sid>/a3p-planning-builder")
def put_a3p_planning_builder(sid):
    data = load_sessions(); session_data = find_session(data, sid)
    if not session_data: return jsonify({"ok": False, "error": "Session introuvable."}), 404
    if (session_data.get("formation") or "").upper() != "A3P": return jsonify({"ok": False, "error": "La session n'est pas A3P."}), 400
    payload = request.get_json(silent=True) or {}
    state = payload.get("state") if isinstance(payload.get("state"), dict) else payload
    previous_state = session_data.get("a3pPlanningBuilder") or session_data.get("a3pPlanningDraftJson") or {}

    def _has_builder_content(value):
        if not isinstance(value, dict) or not value:
            return False
        if value.get("days") or (value.get("scheduleConfig") or {}).get("days"):
            return True
        locked = (value.get("scheduleConfig") or value).get("lockedModules") or {}
        return any(locked.get(code) for code in locked)

    if _has_builder_content(previous_state) and not _has_builder_content(state):
        return jsonify({"ok": False, "error": "Sauvegarde A3P refusée : état vide ou incomplet, conservation du dernier état valide."}), 400
    session_data["a3pPlanningBuilder"] = state
    modules_data = _a3p_manual_modules_from_state(state)
    if are_a3p_manual_modules_complete(state):
        mark_a3p_manual_modules_admin_validated(session_data, modules_data)
    session_data["a3pPlanningBuilderSavedAt"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_sessions(data)
    return jsonify({"ok": True, "savedAt": session_data["a3pPlanningBuilderSavedAt"]})

@app.post("/api/sessions/<sid>/a3p-documents/draft")
def save_a3p_documents_draft(sid):
    data=load_sessions(); session_data=find_session(data,sid)
    if not session_data: return jsonify({"ok":False,"error":"Session introuvable."}),404
    if (session_data.get("formation") or "").upper() != "A3P": return jsonify({"ok":False,"error":"La session n'est pas A3P."}),400
    payload=request.get_json(silent=True) or {}
    cfg=payload.get("scheduleConfig") or payload
    session_data["a3pPlanningDraftJson"] = cfg
    if are_a3p_manual_modules_complete({"scheduleConfig": cfg, "planning": payload.get("planning") or []}):
        mark_a3p_manual_modules_admin_validated(session_data, cfg.get("lockedModules") or {})
    session_data["a3pTrainerEmail"] = cfg.get("trainerEmail") or session_data.get("a3pTrainerEmail")
    session_data["a3pTrainerName"] = ((cfg.get("trainerFirstName") or "") + " " + (cfg.get("trainerLastName") or "")).strip() or session_data.get("a3pTrainerName")
    session_data["a3pPlanningDraftSavedAt"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_sessions(data)
    return jsonify({"ok":True,"savedAt":session_data["a3pPlanningDraftSavedAt"]})

@app.post("/api/sessions/<sid>/a3p-documents/preview")
def preview_a3p_documents(sid):
    data=load_sessions(); session_data=find_session(data,sid)
    if not session_data: return jsonify({"ok":False,"error":"Session introuvable."}),404
    if (session_data.get("formation") or "").upper() != "A3P": return jsonify({"ok":False,"error":"La session n'est pas A3P."}),400
    payload=request.get_json(silent=True) or {}
    try:
        cfg=payload.get("scheduleConfig") or payload
        if session_data.get("a3pTrainerModulesStatus") in {"completed","validated"}:
            cfg=dict(cfg); cfg["lockedModules"]=session_data.get("a3pTrainerManualModulesData") or cfg.get("lockedModules") or {}
        result=generateA3pSchedule(cfg)
        return jsonify({"ok":True,"planning":result["planning"],"summary":result["summary"]})
    except ValueError as exc:
        return jsonify({"ok":False,"error":str(exc)}),400

@app.post("/api/sessions/<sid>/a3p-documents/generate")
def generate_a3p_documents(sid):
    data=load_sessions(); session_data=find_session(data,sid)
    if not session_data: return jsonify({"ok":False,"error":"Session introuvable."}),404
    if (session_data.get("formation") or "").upper() != "A3P": return jsonify({"ok":False,"error":"La session n'est pas A3P."}),400
    payload=request.get_json(silent=True) or {}
    app.logger.info("Début génération documents A3P session_id=%s", sid)
    try:
        cfg=payload.get("scheduleConfig") or payload
        admin_ready = can_generate_a3p_documents_state({"scheduleConfig": cfg, "planning": payload.get("planning") or []})
        if session_data.get("a3pTrainerModulesStatus") != "validated" and not admin_ready:
            app.logger.warning("Génération documents A3P refusée session_id=%s statut_modules=%s", sid, session_data.get("a3pTrainerModulesStatus"))
            return jsonify({"ok":False,"error":"Les modules formateur A3P doivent être complétés puis validés par l’admin avant génération."}),400
        cfg=dict(cfg)
        if admin_ready:
            mark_a3p_manual_modules_admin_validated(session_data, cfg.get("lockedModules") or {})
        cfg["lockedModules"]=session_data.get("a3pTrainerManualModulesData") or cfg.get("lockedModules") or {}
        supplied_planning = payload.get("planning") if isinstance(payload.get("planning"), list) else None
        if supplied_planning:
            planning = supplied_planning
            errors, summary = validate_a3p_planning(planning, cfg.get("examDate") or session_data.get("date_exam"))
            if errors: raise ValueError(" ".join(errors))
        else:
            result=generateA3pSchedule(cfg); planning=result["planning"]; summary=result["summary"]
        trainer=((cfg.get("trainerFirstName") or "")+" "+(cfg.get("trainerLastName") or "")).strip() or cfg.get("trainerName") or session_data.get("a3pTrainerName") or ""
        if not trainer: return jsonify({"ok":False,"error":"Nom et prénom du formateur obligatoires pour le contrat formateur."}),400
        session_data.update({"a3pPlanningData":planning,"a3pPlanningSummary":summary,"a3pTrainerName":trainer,"a3pRoom":cfg.get("room") or session_data.get("a3pRoom") or "Intégrale Academy – 54 chemin du Carreou – 83480 PUGET-SUR-ARGENS","date_debut":cfg.get("startDate") or session_data.get("date_debut"),"date_fin":cfg.get("endDate") or session_data.get("date_fin"),"date_exam":cfg.get("examDate") or session_data.get("date_exam")})
        app.logger.info("Génération documents A3P session_id=%s lignes_planning=%s", sid, sum(len(d.get("slots", [])) for d in planning))
        now=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        docs=[]
        a3p_documents = {}
        for kind, key, fname in (("planning","a3pPlanningPdfUrl",f"planning_a3p_session_{sid}.pdf"),("attendance","a3pAttendanceSheetsPdfUrl",f"feuilles_presence_a3p_{sid}.pdf")):
            path=os.path.join(A3P_DOC_DIR,fname); generate_a3p_planning_pdf(session_data, path) if kind == "planning" else generate_a3p_attendance_pdf(session_data, path)
            session_data[key]=url_for("view_a3p_document", sid=sid, kind=kind); session_data[key.replace("Url","Filename")]=fname
            a3p_documents[kind] = {"path": path, "generated_at": now}
            docs.append({"kind":kind,"path":path})
        contract_payload=payload.get("contract") or cfg
        cid=str(uuid.uuid4()); cf=f"contrat_formateur_a3p_{sid}_{cid}.pdf"; cp=os.path.join(A3P_DOC_DIR,cf)
        generate_a3p_trainer_contract_pdf(session_data, contract_payload, cp)
        session_data["a3pTrainerContract"]={"id":cid,"pdfFilename":cf,"pdfUrl":url_for("view_a3p_document",sid=sid,kind="contract"),"generatedAt":now,"dailyRate":contract_payload.get("dailyRate"),"vatEnabled":bool(contract_payload.get("vatEnabled"))}
        a3p_documents["contract"] = {"path": cp, "generated_at": now}
        docs.append({"kind":"contract","path":cp})
        session_data["a3p_documents"] = a3p_documents
        session_data["a3pDocumentsGeneratedAt"]=now; save_sessions(data)
        app.logger.info("Documents A3P créés session_id=%s documents=%s chemins_sauvegardés=%s", sid, [d["kind"] for d in docs], docs)
        return jsonify({"ok":True,"generatedAt":now,"planningUrl":session_data["a3pPlanningPdfUrl"],"attendanceUrl":session_data["a3pAttendanceSheetsPdfUrl"],"contractUrl":session_data["a3pTrainerContract"]["pdfUrl"]})
    except ValueError as exc:
        app.logger.warning("Erreur validation génération documents A3P session_id=%s erreur=%s", sid, exc)
        return jsonify({"ok":False,"error":str(exc)}),400
    except Exception as exc:
        app.logger.exception("Erreur précise génération documents A3P session_id=%s", sid); return jsonify({"ok":False,"error":str(exc)}),500



def a3p_document_path(session_data, kind):
    if kind not in {"planning", "attendance", "contract"}:
        return None
    doc = (session_data.get("a3p_documents") or {}).get(kind) or {}
    path = doc.get("path")
    if path:
        return path
    # Compatibilité avec les sessions générées avant la persistance de a3p_documents.
    if kind == "planning" and session_data.get("a3pPlanningFilename"):
        return os.path.join(A3P_DOC_DIR, os.path.basename(session_data["a3pPlanningFilename"]))
    if kind == "attendance" and session_data.get("a3pAttendanceSheetsFilename"):
        return os.path.join(A3P_DOC_DIR, os.path.basename(session_data["a3pAttendanceSheetsFilename"]))
    if kind == "contract" and (session_data.get("a3pTrainerContract") or {}).get("pdfFilename"):
        return os.path.join(A3P_DOC_DIR, os.path.basename(session_data["a3pTrainerContract"]["pdfFilename"]))
    return None

def a3p_document_exists(session_data, kind):
    path = a3p_document_path(session_data, kind)
    return bool(path and os.path.exists(path))

@app.context_processor
def inject_a3p_trainer_helpers():
    return {"a3p_trainer_status": a3p_trainer_status, "a3p_document_exists": a3p_document_exists, "aps_detect_trainers": aps_detect_trainers, "is_ssiap1_session": is_ssiap1_session}

@app.post("/api/admin/sessions/<sid>/a3p/trainer-link")
def generate_a3p_trainer_link(sid):
    data=load_sessions(); session_data=find_session(data,sid)
    if not session_data: return jsonify({"ok":False,"error":"Session introuvable."}),404
    if (session_data.get("formation") or "").upper() != "A3P": return jsonify({"ok":False,"error":"La session n'est pas A3P."}),400
    session_data["a3pTrainerPublicToken"] = secrets.token_urlsafe(48)
    session_data["a3pTrainerPublicLinkCreatedAt"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    session_data["a3pTrainerModulesStatus"] = "waiting"
    session_data.pop("a3pTrainerPublicLinkDisabledAt", None)
    save_sessions(data)
    status=a3p_trainer_status(session_data)
    return jsonify({"ok":True,"url":status["url"],"status":status,"createdAt":session_data["a3pTrainerPublicLinkCreatedAt"]})

@app.post("/api/admin/sessions/<sid>/a3p/trainer-link/send")
def send_a3p_trainer_link(sid):
    data=load_sessions(); session_data=find_session(data,sid)
    if not session_data: return jsonify({"ok":False,"error":"Session introuvable."}),404
    cfg=session_data.get("a3pPlanningDraftJson") or {}
    email=(cfg.get("trainerEmail") or session_data.get("a3pTrainerEmail") or "").strip()
    if not email: return jsonify({"ok":False,"error":"Email formateur non renseigné."}),400
    if not session_data.get("a3pTrainerPublicToken"):
        session_data["a3pTrainerPublicToken"] = secrets.token_urlsafe(48)
        session_data["a3pTrainerPublicLinkCreatedAt"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    url=a3p_trainer_public_url(session_data["a3pTrainerPublicToken"])
    first=(cfg.get("trainerFirstName") or session_data.get("a3pTrainerName") or "formateur").split()[0]
    body=f"Bonjour {first},\n\nDans le cadre de la préparation de la session A3P, merci de compléter les dates des modules imposés dont vous avez la charge.\n\nVous pouvez accéder au formulaire via le lien sécurisé ci-dessous :\n{url}\n\nDates de formation : du {format_date(cfg.get('startDate') or session_data.get('date_debut'))} au {format_date(cfg.get('endDate') or session_data.get('date_fin'))}.\nDate d’examen : {format_date(cfg.get('examDate') or session_data.get('date_exam'))}.\n\nMerci de compléter les 4 modules puis de cliquer sur “J’ai terminé” afin que nous puissions finaliser le planning.\n\nBien cordialement,\nIntégrale Academy"
    smtp_config=get_smtp_config()
    if not smtp_config.get("login") or not smtp_config.get("password"):
        return jsonify({"ok":False,"error":"Email non configuré.","url":url}),400
    msg=MIMEText(body,"plain",_charset="utf-8"); msg["Subject"]="Modules imposés A3P à compléter"; msg["From"]=smtp_config["from_email"]; msg["To"]=email
    with smtplib.SMTP(smtp_config["server"], smtp_config["port"]) as server:
        server.starttls(); server.login(smtp_config["login"], smtp_config["password"]); server.sendmail(smtp_config["from_email"],[email],msg.as_string())
    session_data["a3pTrainerModulesStatus"]="sent"; session_data["a3pTrainerPublicLinkSentAt"]=datetime.now().strftime("%Y-%m-%d %H:%M:%S"); save_sessions(data)
    return jsonify({"ok":True,"url":url,"status":a3p_trainer_status(session_data)})

@app.post("/api/admin/sessions/<sid>/a3p/trainer-modules/validate")
def validate_a3p_trainer_modules_admin(sid):
    data=load_sessions(); session_data=find_session(data,sid)
    if not session_data: return jsonify({"ok":False,"error":"Session introuvable."}),404
    errors=validate_a3p_trainer_manual_data(session_data, session_data.get("a3pTrainerManualModulesData") or {})
    if errors: return jsonify({"ok":False,"errors":errors,"error":"Modules formateur incomplets."}),400
    session_data["a3pTrainerModulesStatus"]="validated"; session_data["a3pTrainerModulesValidatedAt"]=datetime.now().isoformat(); session_data["manual_modules_validated"] = True; session_data["manual_modules_validated_at"] = session_data["a3pTrainerModulesValidatedAt"]
    draft=session_data.setdefault("a3pPlanningDraftJson", {})
    draft["lockedModules"] = session_data.get("a3pTrainerManualModulesData") or {}
    save_sessions(data)
    return jsonify({"ok":True,"status":a3p_trainer_status(session_data)})

@app.get("/public/a3p-planning/<token>")
def public_a3p_planning_page(token):
    return render_template("public_a3p_planning.html", token=token)

@app.get("/api/public/a3p-planning/<token>")
def get_public_a3p_planning(token):
    data=load_sessions(); session_data=find_a3p_public_session(data, token)
    if not session_data: return jsonify({"ok":False,"error":"Lien invalide ou désactivé."}),404
    session_data["a3pTrainerPublicLinkLastAccessAt"]=datetime.now().strftime("%Y-%m-%d %H:%M:%S"); save_sessions(data)
    return jsonify({"ok":True,"data":a3p_public_payload(session_data)})

@app.put("/api/public/a3p-planning/<token>")
def save_public_a3p_planning(token):
    data=load_sessions(); session_data=find_a3p_public_session(data, token)
    if not session_data: return jsonify({"ok":False,"error":"Lien invalide ou désactivé."}),404
    payload=request.get_json(silent=True) or {}; modules=payload.get("modulesData") or {}
    session_data["a3pTrainerManualModulesData"] = {c: modules.get(c, []) for c in A3P_TRAINER_MANUAL_CODES}
    session_data["manual_modules_source"] = "trainer"
    session_data["manual_modules_completed"] = False
    session_data["manual_modules_validated"] = False
    if session_data.get("a3pTrainerModulesStatus") not in {"completed","validated"}: session_data["a3pTrainerModulesStatus"]="in_progress"
    save_sessions(data)
    return jsonify({"ok":True,"status":a3p_trainer_status(session_data),"errors":validate_a3p_trainer_manual_data(session_data, session_data["a3pTrainerManualModulesData"])})

@app.post("/api/public/a3p-planning/<token>/complete")
def complete_public_a3p_planning(token):
    data=load_sessions(); session_data=find_a3p_public_session(data, token)
    if not session_data: return jsonify({"ok":False,"error":"Lien invalide ou désactivé."}),404
    modules=(request.get_json(silent=True) or {}).get("modulesData") or session_data.get("a3pTrainerManualModulesData") or {}
    session_data["a3pTrainerManualModulesData"]={c: modules.get(c, []) for c in A3P_TRAINER_MANUAL_CODES}
    errors=validate_a3p_trainer_manual_data(session_data, session_data["a3pTrainerManualModulesData"])
    session_data["a3pTrainerModulesStatus"]="incomplete" if errors else "completed"
    session_data["manual_modules_source"] = "trainer"
    session_data["manual_modules_completed"] = not bool(errors)
    session_data["manual_modules_validated"] = False
    if not errors: session_data["a3pTrainerModulesCompletedAt"]=datetime.now().isoformat()
    save_sessions(data)
    return (jsonify({"ok":not bool(errors),"status":a3p_trainer_status(session_data),"errors":errors}), 400 if errors else 200)

@app.get("/sessions/<sid>/a3p-documents/<kind>/view")
def view_a3p_document(sid, kind):
    data=load_sessions(); session_data=find_session(data,sid)
    if not session_data:
        return jsonify({"ok": False, "error": "Session introuvable."}), 404
    if (session_data.get("formation") or "").upper() != "A3P":
        return jsonify({"ok": False, "error": "La session n'est pas A3P."}), 400
    if kind not in {"planning", "attendance", "contract"}:
        return jsonify({"ok": False, "error": "Type de document A3P invalide."}), 404

    path = a3p_document_path(session_data, kind)
    exists = bool(path and os.path.exists(path))
    app.logger.info("Téléchargement document A3P session_id=%s kind=%s path=%s exists=%s", sid, kind, path, exists)
    if not exists:
        return jsonify({"ok": False, "error": "Document introuvable, veuillez régénérer les documents."}), 404

    download_names = {
        "planning": f"planning_a3p_session_{sid}.pdf",
        "attendance": f"feuilles_presence_a3p_{sid}.pdf",
        "contract": f"contrat_formateur_a3p_{sid}.pdf",
    }
    return send_file(path, mimetype="application/pdf", as_attachment=False, download_name=download_names[kind])

@app.get("/sessions/<sid>/aps-planning/edit")
def edit_aps_planning_page(sid):
    data = load_sessions()
    session_data = find_session(data, sid)
    if not session_data:
        abort(404)
    if (session_data.get("formation") or "").upper() not in {"APS", "DESP", "DIRIGEANT", "AFC_APS_SSIAP"} and not is_ssiap1_session(session_data):
        abort(404)
    return render_template("aps_planning_editor.html", title="Modifier le planning", s=session_data)

@app.get("/api/sessions/<sid>/aps-planning")
def get_aps_planning_api(sid):
    data = load_sessions()
    session_data = find_session(data, sid)
    if not session_data:
        return jsonify({"ok": False, "error": "Session introuvable."}), 404
    formation = (session_data.get("formation") or "").upper()
    is_desp = formation in {"DESP", "DIRIGEANT"}
    is_ssiap1 = is_ssiap1_session(session_data)
    is_afc = normalize_training_code(session_data) == "AFC_APS_SSIAP" if session_data else False
    if formation != "APS" and not is_desp and not is_ssiap1 and not is_afc:
        return jsonify({"ok": False, "error": "La session n'est pas APS/DESP."}), 400
    planning_data = session_data.get("apsPlanningData") or []
    summary = ssiap1_summary_from_data(planning_data) if is_ssiap1_session(session_data) and planning_data else (desp_summary_from_planning(planning_data) if is_desp and planning_data else (aps_summary_from_data(planning_data) if planning_data else None))
    return jsonify({
        "ok": True,
        "session": session_data,
        "apsPlanningData": planning_data,
        "summary": summary,
        "pdfUrl": url_for("view_planning_pdf", sid=sid) if session_data.get("planning_pdf") else None,
        "needsRegeneration": bool(session_data.get("planning_pdf") and not planning_data),
    })

@app.put("/api/sessions/<sid>/aps-planning")
def update_aps_planning_api(sid):
    data = load_sessions()
    session_data = find_session(data, sid)
    if not session_data:
        return jsonify({"ok": False, "error": "Session introuvable."}), 404
    formation = (session_data.get("formation") or "").upper()
    is_desp = formation in {"DESP", "DIRIGEANT"}
    is_ssiap1 = is_ssiap1_session(session_data)
    if formation != "APS" and not is_desp and not is_ssiap1:
        return jsonify({"ok": False, "error": "La session n'est pas APS/DESP."}), 400
    if not session_data.get("apsPlanningData") and not session_data.get("planning_pdf"):
        return jsonify({"ok": False, "error": "Aucun planning APS n'existe encore."}), 400
    payload = request.get_json(silent=True) or {}
    planning_data = payload.get("planningData")
    if not isinstance(planning_data, list) or not planning_data:
        return jsonify({"ok": False, "error": "planningData est obligatoire."}), 400
    planning_mode = session_data.get("apsPlanningMode") or ("elearning_presentiel" if any(slot.get("modality") == "elearning" for day in planning_data for slot in day.get("slots", [])) else "full_presentiel")
    exam_iso = aps_local_date_iso(session_data.get("date_exam"))
    if exam_iso and not is_ssiap1_session(session_data) and any(day.get("date") == exam_iso for day in planning_data):
        return jsonify({"ok": False, "error": f"Sécurité planning APS: la date d'examen ({format_date(exam_iso)}) est réservée à l’examen et ne peut contenir aucun créneau de formation."}), 400
    if is_ssiap1_session(session_data):
        summary = ssiap1_summary_from_data(planning_data)
        errors = list(summary.get("errors") or [])
    elif is_afc:
        summary = afc_aps_ssiap_summary_from_data(planning_data)
        errors = list(summary.get("errors") or [])
    elif is_desp:
        summary = desp_summary_from_planning(planning_data)
        errors = list(summary.get("errors") or [])
    else:
        errors, summary = validate_aps_planning_data(planning_data, planning_mode)
    if errors:
        return jsonify({"ok": False, "error": "Validation impossible.", "errors": errors, "summary": summary}), 400
    session_data["apsPlanningData"] = planning_data
    session_data["apsPlanningSummary"] = summary
    session_data["planning_modified_at"] = append_planning_history(session_data, "planning modifié")
    pdf_url = url_for("view_planning_pdf", sid=sid) if session_data.get("planning_pdf") else None
    if payload.get("regeneratePdf"):
        filename = f"planning_{'afc_aps_ssiap' if is_afc else ('desp' if is_desp else ('ssiap1' if is_ssiap1 else 'aps'))}_session_{sid}.pdf"
        output_path = os.path.join(PLANNING_DIR, filename)
        temp_path = f"{output_path}.tmp"
        profile = ({"validate": "afc_aps_ssiap", "summary": summary, "planning_title":"PLANNING AFC FRANCE TRAVAIL APS + SSIAP", "short_label":"AFC APS + SSIAP"} if is_afc else ({"validate": "ssiap1", "summary": summary, "planning_title": "PLANNING DE FORMATION SSIAP 1", "short_label": "SSIAP 1"} if is_ssiap1_session(session_data) else ({"validate": "desp", "summary": summary, "planning_title": "PLANNING DE FORMATION DESP", "short_label": "DESP"} if is_desp else None)))
        result = generate_aps_planning_pdf(session_data, "", temp_path, planning_data=planning_data, planning_mode=planning_mode, document_profile=profile)
        if os.path.exists(output_path):
            archive = os.path.join(PLANNING_DIR, f"planning_{'ssiap1' if is_ssiap1_session(session_data) else ('desp' if is_desp else 'aps')}_session_{sid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf")
            try:
                os.replace(output_path, archive)
            except OSError:
                pass
        os.replace(temp_path, output_path)
        session_data["planning_pdf"] = filename
        session_data["apsPlanningSummary"] = result["summary"]
        session_data["apsPlanningMode"] = planning_mode
        session_data["planning_pdf_regenerated_at"] = append_planning_history(session_data, "PDF régénéré")
        pdf_url = url_for("view_planning_pdf", sid=sid)
    save_sessions(data)
    return jsonify({"ok": True, "pdfUrl": pdf_url, "summary": session_data.get("apsPlanningSummary"), "modifiedAt": session_data.get("planning_modified_at")})


@app.delete("/api/sessions/<sid>/aps-planning")
def reset_aps_planning_api(sid):
    data = load_sessions()
    session_data = find_session(data, sid)
    if not session_data:
        return jsonify({"success": False, "message": "Session introuvable."}), 404
    if (session_data.get("formation") or "").upper() not in {"APS", "DESP", "DIRIGEANT", "AFC_APS_SSIAP"} and not is_ssiap1_session(session_data):
        return jsonify({"success": False, "message": "Cette action est disponible uniquement pour les sessions APS/DESP."}), 400

    planning_keys = (
        "apsPlanningData",
        "apsPlanningMode",
        "apsPlanningPdfUrl",
        "apsPlanningGeneratedAt",
        "apsPlanningUpdatedAt",
        "apsPlanningModifiedAt",
        "apsPlanningSummary",
        "apsPlanningHistory",
        "planning_pdf",
        "planning_generated_at",
        "planning_modified_at",
        "planning_pdf_regenerated_at",
        "planning_updated_at",
        "planning_history",
    )
    existing_values = {key: session_data.get(key) for key in planning_keys if session_data.get(key) not in (None, "", [], {})}
    if not existing_values:
        return jsonify({"success": True, "message": "Aucun planning APS à réinitialiser."})

    old_pdf = session_data.get("planning_pdf") or session_data.get("apsPlanningPdfUrl")
    deleted_pdf = None
    pdf_delete_error = None
    if old_pdf:
        pdf_name = os.path.basename(str(old_pdf))
        pdf_path = os.path.join(PLANNING_DIR, pdf_name)
        if os.path.exists(pdf_path):
            try:
                os.remove(pdf_path)
                deleted_pdf = pdf_name
            except OSError as exc:
                pdf_delete_error = str(exc)
                app.logger.warning("Réinitialisation planning APS: suppression PDF impossible session=%s pdf=%s erreur=%s", sid, pdf_name, exc)
        else:
            app.logger.info("Réinitialisation planning APS: PDF déjà absent session=%s pdf=%s", sid, pdf_name)

    for key in planning_keys:
        session_data.pop(key, None)

    reset_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    admin_user = session.get("admin_email") or session.get("admin_user") or (ADMIN_USER if session.get("admin_logged") else "")
    save_sessions(data)
    app.logger.info(
        "Planning APS réinitialisé session=%s reset_at=%s admin=%s ancien_pdf=%s pdf_supprime=%s pdf_erreur=%s champs=%s",
        sid,
        reset_at,
        admin_user or "inconnu",
        old_pdf or "aucun",
        deleted_pdf or "non",
        pdf_delete_error or "aucune",
        sorted(existing_values.keys()),
    )
    return jsonify({"success": True, "message": "Planning APS réinitialisé avec succès"})


@app.get("/sessions/<sid>/planning/view")
def view_planning_pdf(sid):
    data = load_sessions()
    session_data = find_session(data, sid)
    if not session_data:
        abort(404)
    try:
        name = refresh_aps_planning_pdf_file(session_data, sid)
        save_sessions(data)
    except Exception as exc:
        app.logger.exception("Impossible de rafraîchir le planning APS avant affichage session=%s", sid)
        abort(500, description=str(exc))
    if not name:
        abort(404)

    path = os.path.join(PLANNING_DIR, os.path.basename(str(name)))
    if not os.path.exists(path):
        abort(404)

    return send_planning_pdf_file(path, as_attachment=False)


@app.get("/sessions/<sid>/planning/download")
def download_planning_pdf(sid):
    data = load_sessions()
    session_data = find_session(data, sid)
    if not session_data:
        abort(404)
    try:
        name = refresh_aps_planning_pdf_file(session_data, sid)
        save_sessions(data)
    except Exception as exc:
        app.logger.exception("Impossible de rafraîchir le planning APS avant téléchargement session=%s", sid)
        abort(500, description=str(exc))
    if not name:
        abort(404)

    name = os.path.basename(str(name))
    path = os.path.join(PLANNING_DIR, name)
    if not os.path.exists(path):
        abort(404)

    return send_planning_pdf_file(path, as_attachment=True, download_name=name)



@app.post("/api/sessions/<sid>/aps-convocation")
def generate_aps_convocation_route(sid):
    data = load_sessions()
    session_data = find_session(data, sid)
    if not session_data:
        return jsonify({"ok": False, "error": "Session introuvable."}), 404
    if (session_data.get("formation") or "").upper() != "APS":
        return jsonify({"ok": False, "error": "La convocation modèle Word est réservée aux sessions APS."}), 400
    payload = request.get_json(silent=True) or {}
    trainee = payload.get("trainee") if isinstance(payload.get("trainee"), dict) else payload
    try:
        result = generateApsConvocationFromDocxTemplate(trainee, session_data)
        return jsonify({"ok": True, **result})
    except Exception as exc:
        app.logger.exception("Erreur génération convocation APS session=%s", sid)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.get("/convocations/aps/<path:filename>")
def view_aps_convocation_pdf(filename):
    safe_name = secure_filename(filename)
    if not safe_name.lower().endswith(".pdf"):
        abort(404)
    path = os.path.join(CONVOCATION_DIR, safe_name)
    if not os.path.exists(path):
        abort(404)
    return send_file(path, mimetype="application/pdf", as_attachment=False)


@app.get("/convocations/aps/docx/<path:filename>")
def download_aps_convocation_docx(filename):
    safe_name = secure_filename(filename)
    if not safe_name.lower().endswith(".docx"):
        abort(404)
    path = os.path.join(CONVOCATION_DIR, safe_name)
    if not os.path.exists(path):
        abort(404)
    return send_file(path, mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document", as_attachment=True, download_name=safe_name)

@app.route("/healthz")
def healthz():
    return "ok"

@app.route("/cron-check")
def cron_check():
    data = load_sessions()
    for session in data["sessions"]:
        auto_archive_if_all_done(session)
    reminded = send_jury_reminders(data, request.url_root.rstrip("/"))
    expired_alerts = send_formateur_expiration_alerts()
    save_sessions(data)
    message = "Cron check terminé"
    if reminded:
        message = f"{message} | Rappels jury envoyés: {', '.join(reminded)}"
    if expired_alerts:
        message = f"{message} | Alertes expiration formateurs: {expired_alerts}"
    return message, 200

@app.route("/cron-daily-summary")
def cron_daily_summary():
    send_daily_overdue_summary()
    return "Mail récapitulatif envoyé", 200

# ------------------------------------------------------------
# ✅ Route publique pour le suivi auto sur la plateforme principale
#    -> renvoie le nombre total d'étapes en retard (toutes sessions actives)
# ------------------------------------------------------------
@app.route("/data.json")
def data_sessions_json():
    try:
        data = load_sessions()
        sessions = data.get("sessions", [])

        today = datetime.now().date()
        total_retards_steps = 0
        total_sessions_en_retard = 0
        details = []  # utile si tu veux diagnostiquer

        for s in sessions:
            end_date = parse_date(s.get("date_fin"))
            is_finished = bool(end_date and end_date.date() < today)
            if s.get("archived") or is_finished:
                continue  # on ignore les sessions archivées ou terminées (comme /sessions)

            late_steps = []
            for i, step in enumerate(s.get("steps", [])):
                st, dl = status_for_step(i, s)
                if st == "late":
                    total_retards_steps += 1
                    late_steps.append({
                        "name": step.get("name"),
                        "deadline": (dl.strftime("%Y-%m-%d") if dl else None)
                    })

            if late_steps:
                total_sessions_en_retard += 1

            details.append({
                "id": s.get("id"),
                "formation": s.get("formation"),
                "date_debut": s.get("date_debut"),
                "date_exam": s.get("date_exam"),
                "retards": len(late_steps),
                "late_steps": late_steps
            })

        payload = {
            "retards": total_sessions_en_retard,  # 👉 clé utilisée par l'index (compte des sessions ayant au moins 1 retard)
            "retards_steps": total_retards_steps,  # détail: nombre total d'étapes en retard
            "sessions": details
        }

        headers = {
            "Content-Type": "application/json",
            "Access-Control-Allow-Origin": "*"
        }
        return json.dumps(payload, ensure_ascii=False), 200, headers

    except Exception as e:
        print("Erreur /data.json (sessions):", e)
        return json.dumps({"retards": -1, "error": str(e)}), 500, {
            "Access-Control-Allow-Origin": "*"
        }

@app.route("/tz-test")
def tz_test():
    from datetime import datetime
    import time
    return f"Serveur : {datetime.now()}<br>Heure système : {time.tzname}"

# ------------------------------------------------------------
# 📦 GESTION DES DOTATIONS
# ------------------------------------------------------------

DOTATIONS_FILE = os.path.join(DATA_DIR, "dotations.json")

def load_dotations():
    if os.path.exists(DOTATIONS_FILE):
        try:
            with open(DOTATIONS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return []

def save_dotations(data):
    with open(DOTATIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ✉️ Fonction d’envoi d’email (réutilise la conf SMTP)
def send_email(to, subject, body):
    if not FROM_EMAIL or not EMAIL_PASSWORD:
        print("⚠️ Email non configuré")
        return
    msg = MIMEText(body, "html", "utf-8")
    msg["From"] = FROM_EMAIL
    msg["To"] = to
    msg["Subject"] = subject
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as s:
            s.starttls()
            s.login(FROM_EMAIL, EMAIL_PASSWORD)
            s.sendmail(FROM_EMAIL, [to], msg.as_string())
        print(f"✅ Mail envoyé à {to}")
    except Exception as e:
        print("❌ Erreur envoi mail dotation :", e)


def send_price_adaptator_email(to, subject, html):
    smtp_config = get_smtp_config()
    if not smtp_config["login"] or not smtp_config["password"]:
        return False, "SMTP non configuré"
    msg = MIMEText(html, "html", "utf-8")
    msg["From"] = smtp_config["from_email"]
    msg["To"] = to
    msg["Cc"] = "clement@integraleacademy.com"
    msg["Subject"] = subject
    try:
        with smtplib.SMTP(smtp_config["server"], smtp_config["port"]) as server:
            server.starttls()
            server.login(smtp_config["login"], smtp_config["password"])
            server.sendmail(
                smtp_config["from_email"],
                [to, "clement@integraleacademy.com"],
                msg.as_string(),
            )
        return True, None
    except Exception as e:
        print("❌ Erreur envoi mail price adaptator :", e)
        return False, str(e)


def send_price_adaptator_sms(phone, message):
    normalized_phone = normalize_phone_number(phone)
    if not normalized_phone:
        return False, "Téléphone au format international requis (ex: +336...)"

    # ✅ Utilise Brevo comme pour les SMS jury
    if BREVO_API_KEY and BREVO_SMS_SENDER:
        print("[price sms] Envoi via Brevo API", {"to": normalized_phone, "sender": BREVO_SMS_SENDER})

        payload = json.dumps({
            "sender": BREVO_SMS_SENDER,
            "recipient": normalized_phone,
            "content": message,
            "type": "transactional",
        }).encode("utf-8")

        req = urllib.request.Request("https://api.brevo.com/v3/transactionalSMS/sms")
        req.add_header("Content-Type", "application/json")
        req.add_header("api-key", BREVO_API_KEY)

        try:
            with urllib.request.urlopen(req, data=payload, timeout=10) as resp:
                body = resp.read().decode("utf-8", errors="replace")
                print("[price sms] Brevo response", resp.status, body)
                if 200 <= resp.status < 300:
                    return True, None
                return False, f"Brevo SMS erreur {resp.status}: {body}"

        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")
            print("[price sms] Brevo HTTPError", e.code, body)
            return False, f"Brevo HTTP {e.code}: {body}"

        except Exception as e:
            print("[price sms] Brevo exception", repr(e))
            return False, str(e)

    return False, "SMS non configuré (BREVO_API_KEY / BREVO_SMS_SENDER manquants)"


@app.route("/dotations")
def dotations_home():
    data = load_dotations()
    return render_template("dotations.html", title="Gestion des dotations", dotations=data)


@app.route("/dotations/add", methods=["POST"])
def add_dotation():
    data = load_dotations()
    item = {
        "id": str(uuid.uuid4())[:8],
        "nom": request.form.get("nom", "").strip(),
        "prenom": request.form.get("prenom", "").strip(),
        "email": request.form.get("email", "").strip(),
        "ipad": request.form.get("ipad", "").strip(),
        "badge": request.form.get("badge", "").strip(),
        "date_remise": request.form.get("date_remise", "").strip(),
        "statut": "Dotation à distribuer",
        "commentaire": request.form.get("commentaire", "").strip(),
    }
    data.append(item)
    save_dotations(data)
    flash("Dotation ajoutée avec succès.", "ok")
    return redirect(url_for("dotations_home"))


@app.route("/dotations/<id>/delete", methods=["POST"])
def delete_dotation(id):
    data = load_dotations()
    data = [d for d in data if d["id"] != id]
    save_dotations(data)
    flash("Dotation supprimée.", "ok")
    return redirect(url_for("dotations_home"))


@app.route("/dotations/<id>/edit", methods=["POST"])
def edit_dotation(id):
    data = load_dotations()
    for d in data:
        if d["id"] == id:
            d["nom"] = request.form.get("nom", d["nom"])
            d["prenom"] = request.form.get("prenom", d["prenom"])
            d["email"] = request.form.get("email", d["email"])
            d["ipad"] = request.form.get("ipad", d["ipad"])
            d["badge"] = request.form.get("badge", d["badge"])
            d["date_remise"] = request.form.get("date_remise", d["date_remise"])
            d["commentaire"] = request.form.get("commentaire", d["commentaire"])
            break
    save_dotations(data)
    flash("Dotation modifiée.", "ok")
    return redirect(url_for("dotations_home"))

@app.route("/dotations/<id>/update_date", methods=["POST"])
def update_date_remise(id):
    data = load_dotations()
    for d in data:
        if d["id"] == id:
            d["date_remise"] = request.form.get("date_remise", d["date_remise"])
            break
    save_dotations(data)
    flash("Date de remise mise à jour.", "ok")
    return redirect(url_for("dotations_home"))

@app.route("/dotations/<id>/rupture", methods=["POST"])
def rupture_contrat(id):
    data = load_dotations()
    for d in data:
        if d["id"] == id:
            d["statut"] = "Dotation non restituée"
            save_dotations(data)
            body = f"""
            Bonjour {d['prenom']},<br><br>

            Suite à la rupture de votre contrat d’apprentissage, nous vous rappelons que vous devez restituer l’ensemble du matériel mis à disposition (iPad, chargeur et badge distributeur) dans un délai de 5 jours, conformément à la convention signée.<br><br>

            Le matériel peut être déposé directement au centre Intégrale Academy (54 chemin du Carreou, 83480 Puget-sur-Argens) ou envoyé par courrier suivi à la même adresse.<br><br>

            L’iPad doit être restitué en parfait état de fonctionnement et sans dégradation.<br>
            En cas de non-restitution ou de matériel dégradé, des pénalités financières pourront être appliquées :<br>
            – 400 € pour l’iPad<br>
            – 20 € pour le chargeur<br>
            – 20 € pour le badge distributeur<br><br>

            Bien cordialement,<br>
            <b>Clément VAILLANT</b><br>
            Directeur général – Intégrale Academy
            """
            send_email(d["email"], "Restitution du matériel – Intégrale Academy", body)
            break
    flash("📩 Mail de rupture envoyé et statut mis à jour.", "ok")
    return redirect(url_for("dotations_home"))


@app.route("/dotations/<id>/badge_fin", methods=["POST"])
def badge_fin(id):
    data = load_dotations()
    for d in data:
        if d["id"] == id:
            d["statut"] = "Dotation non restituée"  # ✅ au lieu de "Dotation restituée"
            save_dotations(data)
            body = f"""
            Bonjour {d['prenom']},<br><br>
            Votre BTS touche à sa fin, nous vous rappelons que vous devez nous restituer le badge distributeur de boissons et snack avant de quitter l'école, conformément à la convention signée.<br><br>
            Vous pouvez le déposer directement au centre Intégrale Academy (54 chemin du Carreou, 83480 Puget-sur-Argens) ou l’envoyer par courrier suivi à la même adresse.<br><br>
            Nous vous remercions par avance pour votre réactivité.<br><br>
            Bien cordialement,<br>
            <b>L’équipe Intégrale Academy</b>
            """
            send_email(d["email"], "Restitution du badge distributeur – Intégrale Academy", body)
            break
    flash("📩 Mail de fin d’études envoyé et statut mis à jour.", "ok")
    return redirect(url_for("dotations_home"))





@app.route("/dotations/<id>/changer_statut", methods=["POST"])
def changer_statut(id):
    nouveau_statut = request.form.get("statut")
    data = load_dotations()
    for d in data:
        if d["id"] == id:
            d["statut"] = nouveau_statut
            break
    save_dotations(data)
    return redirect(url_for("dotations_home"))

# ------------------------------------------------------------
# 👨‍🏫 GESTION DES FORMATEURS (Contrôle formateurs)
# ------------------------------------------------------------

FORMATEURS_FILE = os.path.join(DATA_DIR, "formateurs.json")
FORMATEURS_LOCK = FORMATEURS_FILE + ".lock"
FORMATEUR_FILES_DIR = os.path.join(DATA_DIR, "formateurs_files")
FORMATEUR_PROFILS_DOCS_FILE = os.path.join(DATA_DIR, "formateur_profils_docs.json")
os.makedirs(FORMATEUR_FILES_DIR, exist_ok=True)

DEFAULT_DOC_LABELS = [
    "Badge formateur indépendant",
    "Pièce d’identité",
    "Carte pro formateur",
    "Carte pro APS",
    "Carte pro A3P",
    "Diplôme APS",
    "Diplôme A3P",
    "Numéro NDA",
    "Extrait SIRENE moins de 3 mois",
    "Attestation d’assurance RC PRO",
    "Extrait KBIS moins de 3 mois",
    "DRACAR moins de 3 mois",
    "Diplôme SSIAP 1 à jour",
    "Diplôme SSIAP 2 à jour",
    "Diplôme SSIAP 3 à jour",
    "Carte formateur SST",
    "Attestation prévention des risques terroristes",
    "Attestation événementiel spécifique",
    "Attestation palpation de sécurité",
    "Attestation gestion des conflits",
    "Attestation gestion des conflits dégradés",
    "Diplôme formateur pédagogie",
    "Attestation sur l’honneur CNAPS",
    "Attestation de vigilance URSSAF de moins de 3 mois",
    "Charte qualité du formateur",
    "Attestation vacataire APS Adef",
    "Attestation vacataire A3P Adef",
    "Agrément dirigeant CNAPS (AGD)",
    "Autorisation d’exercice CNAPS",
    "CV à jour",
    "Photo d'identité",
]

FORMATEUR_PROFILE_OPTIONS = [
    {"key": "APS", "label": "APS", "color": "#1f6feb"},
    {"key": "A3P", "label": "A3P", "color": "#2da44e"},
    {"key": "BTS", "label": "BTS", "color": "#0ea5e9"},
    {"key": "DIRIGEANT", "label": "DIRIGEANT", "color": "#e67e22"},
    {"key": "SALARIE", "label": "Salarié", "color": "#6b7280"},
    {"key": "PRESTATAIRE", "label": "Prestataire", "color": "#111827"},
    {"key": "SSIAP", "label": "SSIAP", "color": "#dc2626"},
    {"key": "SST", "label": "SST", "color": "#2da44e"},
]
FORMATEUR_PROFILE_KEYS = {option["key"] for option in FORMATEUR_PROFILE_OPTIONS}

def load_formateurs():
    def _read_json(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    # 1) Lecture normale
    if os.path.exists(FORMATEURS_FILE):
        try:
            data = _read_json(FORMATEURS_FILE)
            if isinstance(data, list) and len(data) > 0:
                return data
        except Exception:
            pass  # on tentera le backup

    # 2) Tentative restore depuis .bak
    bak_path = FORMATEURS_FILE + ".bak"
    if os.path.exists(bak_path):
        try:
            data = _read_json(bak_path)
            if isinstance(data, list):
                # on restaure le fichier principal
                save_formateurs(data)
                return data
        except Exception:
            pass

    # 3) Dernier recours
    return []




def normalize_lookup_text(value):
    return " ".join(str(value or "").strip().lower().split())

def formateur_full_name(formateur):
    return " ".join(part for part in [(formateur.get("prenom") or "").strip(), (formateur.get("nom") or "").strip()] if part).strip()

def find_formateur_by_identity(name="", email=""):
    wanted_email = normalize_lookup_text(email)
    wanted_name = normalize_lookup_text(name)
    if not wanted_email and not wanted_name:
        return None
    for formateur in load_formateurs():
        if wanted_email and normalize_lookup_text(formateur.get("email")) == wanted_email:
            return formateur
        full_name = normalize_lookup_text(formateur_full_name(formateur))
        reverse_name = normalize_lookup_text(f"{formateur.get('nom', '')} {formateur.get('prenom', '')}")
        if wanted_name and wanted_name in {full_name, reverse_name}:
            return formateur
    return None

def formateur_contract_defaults(formateur):
    if not formateur:
        return {}
    return {
        "email": (formateur.get("email") or "").strip(),
        "phone": (formateur.get("telephone") or "").strip(),
        "address": (formateur.get("adresse_postale") or formateur.get("adresse") or "").strip(),
        "siret": (formateur.get("siret") or "").strip(),
        "activityDeclaration": (formateur.get("nda") or formateur.get("activityDeclaration") or "").strip(),
        "dailyRate": (formateur.get("tarif_journalier_ht") or formateur.get("dailyRate") or "").strip(),
    }

def merge_formateur_contract_defaults(contract, formateur):
    merged = dict(contract or {})
    for key, value in formateur_contract_defaults(formateur).items():
        current = merged.get(key)
        is_empty = not current.strip() if isinstance(current, str) else not current
        if value and is_empty:
            merged[key] = value
    return merged

def save_formateurs(data):
    # verrou simple (anti écritures concurrentes)
    start = time.time()
    while os.path.exists(FORMATEURS_LOCK):
        # évite de bloquer à l’infini si un lock “fantôme” reste
        if time.time() - start > 5:
            try:
                os.remove(FORMATEURS_LOCK)
            except Exception:
                break
        time.sleep(0.05)

    # créer le lock
    with open(FORMATEURS_LOCK, "w") as f:
        f.write(str(os.getpid()))

    try:
        # écriture atomique: tmp -> replace
        tmp_path = FORMATEURS_FILE + ".tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            f.flush()
            os.fsync(f.fileno())

        os.replace(tmp_path, FORMATEURS_FILE)

        # backup
        bak_path = FORMATEURS_FILE + ".bak"
        try:
            with open(bak_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    finally:
        try:
            os.remove(FORMATEURS_LOCK)
        except Exception:
            pass




def find_formateur(formateurs, fid):
    for f in formateurs:
        if f.get("id") == fid:
            return f
    return None


def normalize_formateur_nub(value):
    value = (value or "").strip()
    if not value:
        return ""
    if not value.isdigit() or len(value) != 7:
        raise ValueError("Le NUB doit contenir exactement 7 chiffres.")
    return value


def normalize_formateur_profils(values):
    profils = []
    for value in values or []:
        key = (value or "").strip().upper()
        if key in FORMATEUR_PROFILE_KEYS and key not in profils:
            profils.append(key)
    return profils


def sanitize_doc_labels(values):
    labels = []
    for value in values or []:
        label = (value or "").strip()
        if label and label not in labels:
            labels.append(label)
    return labels


def load_formateur_profils_docs_config():
    default_map = {option["key"]: [] for option in FORMATEUR_PROFILE_OPTIONS}
    if not os.path.exists(FORMATEUR_PROFILS_DOCS_FILE):
        return default_map

    try:
        with open(FORMATEUR_PROFILS_DOCS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return default_map

    if not isinstance(data, dict):
        return default_map

    cleaned = {}
    for option in FORMATEUR_PROFILE_OPTIONS:
        key = option["key"]
        values = data.get(key, [])
        cleaned[key] = sanitize_doc_labels(values if isinstance(values, list) else [])
    return cleaned


def save_formateur_profils_docs_config(config):
    cleaned = {}
    for option in FORMATEUR_PROFILE_OPTIONS:
        key = option["key"]
        values = config.get(key, [])
        cleaned[key] = sanitize_doc_labels(values if isinstance(values, list) else [])

    with open(FORMATEUR_PROFILS_DOCS_FILE, "w", encoding="utf-8") as f:
        json.dump(cleaned, f, ensure_ascii=False, indent=2)


def apply_profile_document_requirements(formateur, profils_docs_config):
    profils = normalize_formateur_profils(formateur.get("profils", []))
    formateur["profils"] = profils
    if not profils:
        return

    required_labels = set()
    for profil in profils:
        required_labels.update(profils_docs_config.get(profil, []))

    for doc in formateur.get("documents", []):
        label = doc.get("label", "")
        if label in required_labels:
            if doc.get("status") == "non_concerne":
                doc["status"] = "non_conforme"
        else:
            doc["status"] = "non_concerne"



def find_formateur_document(formateur, doc_id):
    return next((d for d in formateur.get("documents", []) if d.get("id") == doc_id), None)


def latest_formateur_pdf_attachment(formateur, preferred_doc_id=""):
    docs = formateur.get("documents", [])
    ordered_docs = []
    if preferred_doc_id:
        preferred = find_formateur_document(formateur, preferred_doc_id)
        if preferred:
            ordered_docs.append(preferred)
    contract_docs = [d for d in docs if d not in ordered_docs and "contrat" in (d.get("label") or "").lower()]
    other_docs = [d for d in docs if d not in ordered_docs and d not in contract_docs]
    for doc in ordered_docs + contract_docs + other_docs:
        for attachment in reversed(doc.get("attachments", [])):
            original = attachment.get("original_name") or attachment.get("filename") or ""
            filename = attachment.get("filename") or ""
            if original.lower().endswith(".pdf") or filename.lower().endswith(".pdf"):
                path = os.path.join(FORMATEUR_FILES_DIR, formateur.get("id", ""), doc.get("id", ""), os.path.basename(filename))
                if os.path.exists(path):
                    return doc, attachment, path
    return None, None, None


YOUSIGN_STATUS_LABELS = {
    "draft": "Brouillon", "approval": "En préparation", "ongoing": "En attente de signature",
    "done": "Signé", "signed": "Signé", "partially_signed": "Partiellement signé", "declined": "Refusé", "expired": "Expiré", "canceled": "Annulé",
    "rejected": "Refusé", "error": "Erreur d’envoi",
}
YOUSIGN_EVENT_STATUS = {
    "signature_request.activated": "ongoing",
    "signature_request.done": "done", "signer.done": "done",
    "signature_request.declined": "declined", "signer.declined": "declined",
    "signature_request.expired": "expired", "signature_request.canceled": "canceled",
    "signer.notification_delivery_failed": "error", "signer.error": "error",
}
YOUSIGN_HANDLED_WEBHOOK_EVENTS = {
    "signature_request.activated", "signer.done", "signature_request.done",
    "signer.declined", "signature_request.expired", "signature_request.canceled",
}

def yousign_status_label(status):
    return YOUSIGN_STATUS_LABELS.get((status or "").strip(), YOUSIGN_STATUS_LABELS.get(str(status or "").split(".")[-1], "Statut inconnu"))

def is_yousign_sandbox():
    return "api-sandbox" in (get_yousign_config().base_url or "")

def normalize_yousign_state(state=None):
    defaults = {
        "signatureRequestId": "", "documentId": "", "signerId": "", "fieldId": "", "externalId": "", "status": "draft", "statusLabel": "Brouillon",
        "signatureUrl": "", "sentAt": "", "signedAt": "", "declinedAt": "", "expiredAt": "", "canceledAt": "",
        "lastEvent": "", "lastEventAt": "", "lastSyncedAt": "", "lastWebhookAt": "",
        "apiStatus": "", "apiSignerStatus": "", "apiHttpStatus": "", "apiError": "", "apiRawResponse": "",
        "recipientEmail": "", "signedDocumentFilename": "", "signedDocumentUrl": "", "error": None, "errorPayload": None,
    }
    legacy = {
        "yousign_signature_request_id": "signatureRequestId", "yousign_signer_id": "signerId", "yousign_document_id": "documentId",
        "yousign_external_id": "externalId", "yousign_api_status": "apiStatus", "yousign_api_signer_status": "apiSignerStatus", "yousign_api_http_status": "apiHttpStatus", "yousign_api_error": "apiError", "last_yousign_raw_response": "apiRawResponse",
        "yousign_status": "status", "yousign_status_label": "statusLabel", "yousign_sent_at": "sentAt",
        "yousign_signed_at": "signedAt", "yousign_declined_at": "declinedAt", "yousign_expired_at": "expiredAt",
        "yousign_canceled_at": "canceledAt", "yousign_last_event": "lastEvent", "yousign_last_event_at": "lastEventAt",
        "last_yousign_sync_at": "lastSyncedAt", "yousign_last_sync_at": "lastSyncedAt",
        "yousign_last_error": "error", "yousign_recipient_email": "recipientEmail", "yousign_signed_document_url": "signedDocumentUrl",
    }
    if isinstance(state, dict):
        defaults.update({k: v for k, v in state.items() if k in defaults})
        for old_key, new_key in legacy.items():
            if state.get(old_key) and not defaults.get(new_key):
                defaults[new_key] = state.get(old_key)
    defaults["statusLabel"] = yousign_status_label(defaults.get("status"))
    return defaults

def mirror_yousign_state_on_contract(contract):
    state = normalize_yousign_state(contract.get("yousign"))
    contract["yousign"] = state
    mapping = {
        "yousign_signature_request_id": "signatureRequestId", "yousign_signer_id": "signerId", "yousign_document_id": "documentId",
        "yousign_external_id": "externalId", "yousign_api_status": "apiStatus", "yousign_api_signer_status": "apiSignerStatus", "yousign_api_http_status": "apiHttpStatus", "yousign_api_error": "apiError", "last_yousign_raw_response": "apiRawResponse",
        "yousign_status": "status", "yousign_status_label": "statusLabel", "yousign_sent_at": "sentAt",
        "yousign_signed_at": "signedAt", "yousign_declined_at": "declinedAt", "yousign_expired_at": "expiredAt",
        "yousign_canceled_at": "canceledAt", "yousign_last_event": "lastEvent", "yousign_last_event_at": "lastEventAt",
        "last_yousign_sync_at": "lastSyncedAt", "yousign_last_sync_at": "lastSyncedAt",
        "yousign_last_error": "error", "yousign_recipient_email": "recipientEmail", "yousign_signed_document_url": "signedDocumentUrl",
    }
    for flat_key, state_key in mapping.items():
        contract[flat_key] = state.get(state_key)
    return state

def extract_yousign_status(payload):
    if not isinstance(payload, dict):
        return "error"
    event = payload.get("event_name") or payload.get("event") or payload.get("type") or ""
    if event in YOUSIGN_EVENT_STATUS:
        return YOUSIGN_EVENT_STATUS[event]

    status = payload.get("status") or payload.get("event_name", "").split(".")[-1]
    data_payload = payload.get("data") if isinstance(payload.get("data"), dict) else {}
    signer_payload = payload.get("signer") if isinstance(payload.get("signer"), dict) else data_payload.get("signer", {})
    signer_status = (signer_payload or {}).get("status")
    if signer_status in {"signed", "done"}:
        return "done"
    if signer_status == "declined":
        return "declined"

    signers = payload.get("signers")
    if isinstance(signers, list):
        signer_statuses = [s.get("status") for s in signers if isinstance(s, dict)]
        if signer_statuses and all(s in {"signed", "done"} for s in signer_statuses):
            return "done"
        if any(s == "declined" for s in signer_statuses):
            return "declined"

    return status if status else "ongoing"



def extract_yousign_external_id(payload):
    if not isinstance(payload, dict):
        return ""
    data_payload = payload.get("data") if isinstance(payload.get("data"), dict) else {}
    signature_request = payload.get("signature_request") if isinstance(payload.get("signature_request"), dict) else data_payload.get("signature_request", {})
    return payload.get("external_id") or data_payload.get("external_id") or (signature_request or {}).get("external_id") or ""

def extract_yousign_signers_list(signers_payload):
    if isinstance(signers_payload, list):
        return signers_payload
    if isinstance(signers_payload, dict):
        if isinstance(signers_payload.get("data"), list):
            return signers_payload["data"]
        if isinstance(signers_payload.get("signers"), list):
            return signers_payload["signers"]
    return []

def build_yousign_api_sync_updates(signature_request_payload, signers_payload=None, now=None, http_status=None, raw_response=None):
    now = now or datetime.now().isoformat(timespec="seconds")
    signers = extract_yousign_signers_list(signers_payload)
    signer_statuses = [str(s.get("status") or "") for s in signers if isinstance(s, dict)]
    api_status = signature_request_payload.get("status") if isinstance(signature_request_payload, dict) else ""
    internal_status = extract_yousign_status({**(signature_request_payload or {}), "signers": signers})
    signed_signers = [s for s in signer_statuses if s in {"signed", "done"}]
    if api_status == "done":
        internal_status = "signed"
    elif signed_signers and len(signed_signers) == len(signer_statuses or signed_signers):
        internal_status = "signed"
    elif signed_signers:
        internal_status = "partially_signed"
    elif api_status == "ongoing":
        internal_status = "ongoing"
    updates = {
        "status": internal_status,
        "apiStatus": api_status or internal_status,
        "apiSignerStatus": ",".join([s for s in signer_statuses if s]),
        "apiHttpStatus": str(http_status or ""),
        "apiRawResponse": json.dumps(raw_response if raw_response is not None else signature_request_payload, ensure_ascii=False, default=str)[:4000],
        "apiError": "",
        "lastSyncedAt": now,
        "error": None,
    }
    external_id = extract_yousign_external_id(signature_request_payload or {})
    if external_id:
        updates["externalId"] = external_id
    if internal_status in {"signed", "done"}: updates["signedAt"] = now
    if internal_status == "declined": updates["declinedAt"] = now
    if internal_status == "expired": updates["expiredAt"] = now
    if internal_status == "canceled": updates["canceledAt"] = now
    return updates

def sync_yousign_signature_request_from_api(signature_request_id, now=None):
    now = now or datetime.now().isoformat(timespec="seconds")
    client = YousignClient()
    env = detect_yousign_environment(client.config.base_url)
    logger.info("YOUSIGN ENV=%s base_url=%s", env, client.config.base_url)
    logger.info("YOUSIGN SYNC START signature_request_id=%s", signature_request_id)
    try:
        payload, http_status, url = client.get_signature_request_with_http_status(signature_request_id)
        logger.info("YOUSIGN SYNC REQUEST url=%s signature_request_id=%s", url, signature_request_id)
        logger.info("YOUSIGN SYNC RESPONSE http_status=%s body=%s", http_status, json.dumps(payload, ensure_ascii=False, default=str))
        signers_payload, signers_http_status, signers_url = client.get_signature_request_signers_with_http_status(signature_request_id)
        logger.info("YOUSIGN SYNC SIGNERS REQUEST url=%s signature_request_id=%s", signers_url, signature_request_id)
        logger.info("YOUSIGN SYNC SIGNERS RESPONSE http_status=%s body=%s", signers_http_status, json.dumps(signers_payload, ensure_ascii=False, default=str))
        for signer in extract_yousign_signers_list(signers_payload):
            if not isinstance(signer, dict):
                continue
            info = signer.get("info") if isinstance(signer.get("info"), dict) else {}
            logger.info(
                "YOUSIGN SYNC SIGNER id=%s email=%s status=%s signed_at=%s done_at=%s",
                signer.get("id"), signer.get("email") or info.get("email"), signer.get("status"), signer.get("signed_at"), signer.get("done_at"),
            )
        return build_yousign_api_sync_updates(payload, signers_payload, now, http_status=http_status, raw_response=payload)
    except YousignError as exc:
        status = exc.status_code or "network"
        body = exc.payload if exc.payload is not None else str(exc)
        logger.error("YOUSIGN SYNC ERROR signature_request_id=%s http_status=%s body=%s error=%s", signature_request_id, status, body, exc)
        return {
            "apiStatus": f"erreur {status}",
            "apiHttpStatus": str(status),
            "apiError": str(exc),
            "apiRawResponse": json.dumps(body, ensure_ascii=False, default=str)[:4000],
            "error": str(exc),
        }

def update_formateur_yousign_state(formateur, updates):
    state = normalize_yousign_state(formateur.get("yousign"))
    state.update(updates)
    state["statusLabel"] = yousign_status_label(state.get("status"))
    formateur["yousign"] = state
    return state

app.jinja_env.globals["yousign_status_label"] = yousign_status_label
app.jinja_env.globals["is_yousign_sandbox"] = is_yousign_sandbox

def build_doc_entry(label):
    return {
        "id": str(uuid.uuid4())[:8],
        "label": label,
        "expiration": "",
        "status": "non_conforme",
        "commentaire": "",
        "attachments": []
    }


def get_all_formateur_document_labels(formateurs, profils_docs_config):
    labels = list(DEFAULT_DOC_LABELS)

    for formateur in formateurs:
        for doc in formateur.get("documents", []):
            label = (doc.get("label") or "").strip()
            if label and label not in labels:
                labels.append(label)

    for docs in profils_docs_config.values():
        for label in docs:
            if label and label not in labels:
                labels.append(label)

    return labels


def build_default_documents():
    docs = []
    for label in DEFAULT_DOC_LABELS:
        docs.append(build_doc_entry(label))
    return docs


def auto_update_document_status(doc):
    """
    Si une date d'expiration est renseignée et dépassée,
    on force le statut à 'non_conforme' (sauf si 'non_concerne').
    """
    if doc.get("status") == "non_concerne":
        return

    exp_str = doc.get("expiration", "").strip()
    if not exp_str:
        return

    dt = parse_date(exp_str)
    if not dt:
        return

    if dt.date() < datetime.now().date():
        doc["status"] = "non_conforme"


def replace_formateur_attachment(fid, doc, uploaded_file):
    """Remplace les anciennes pièces jointes d'un contrôle par le dernier fichier reçu."""
    doc_id = doc.get("id")
    if not doc_id or not uploaded_file or not uploaded_file.filename:
        return None

    subdir = os.path.join(FORMATEUR_FILES_DIR, fid, doc_id)
    os.makedirs(subdir, exist_ok=True)

    for attachment in doc.get("attachments", []):
        filename = attachment.get("filename")
        if not filename:
            continue

        old_path = os.path.join(subdir, os.path.basename(filename))
        if os.path.exists(old_path):
            os.remove(old_path)

    original_name = uploaded_file.filename
    safe_name = secure_filename(original_name) or "document"
    stored_name = f"{int(time.time())}_{uuid.uuid4().hex[:8]}_{safe_name}"
    uploaded_file.save(os.path.join(subdir, stored_name))

    attachment = {
        "filename": stored_name,
        "original_name": original_name
    }
    doc["attachments"] = [attachment]
    return attachment


# ------------------------------------------------------------
# 🔑🟦 ÉTAT COMPLET DES CLÉS & BADGES
# ------------------------------------------------------------
def get_etat_cles_badges(formateurs, total_cles=15, total_badges=15):

    # --- Clés ---
    etat_cles = {
        i: {
            "type": TYPES_CLES.get(i, "Inconnu"),
            "attribue_a": "Libre"
        }
        for i in range(1, total_cles + 1)
    }

    # --- Badges ---
    etat_badges = {
        i: {
            "type": "Badge portail",
            "attribue_a": "Libre"
        }
        for i in range(1, total_badges + 1)
    }

    for f in formateurs:
        nom_prenom = f"{f.get('prenom','')} {f.get('nom','').upper()}".strip()

        # ---- CLÉ ----
        cle = f.get("cle", {})
        num_c = str(cle.get("numero", "")).strip()

        # 🔥 Normalisation : True / "true" / "1" / "on"
        attrib_c = str(cle.get("attribuee", "")).lower() in ("true", "1", "yes", "on")

        if attrib_c and num_c.isdigit():
            num = int(num_c)
            if num in etat_cles:
                nom_custom = cle.get("custom_nom", "").strip()
                nom_formateur = nom_prenom
                etat_cles[num]["attribue_a"] = nom_custom if nom_custom else nom_formateur

        # ---- BADGE ----
        badge = f.get("badge", {})
        num_b = str(badge.get("numero", "")).strip()

        attrib_b = str(badge.get("attribue", "")).lower() in ("true", "1", "yes", "on")

        if attrib_b and num_b.isdigit():
            num = int(num_b)
            if num in etat_badges:
                etat_badges[num]["attribue_a"] = nom_prenom

    return etat_cles, etat_badges



# --- CONFIG TYPES DE CLES ---
TYPES_CLES = {
    1: "PASS GENERAL",
    2: "PASS GENERAL",
    3: "PASS GENERAL",

    4: "PASS PARTIEL",
    5: "PASS PARTIEL",
    6: "PASS PARTIEL",      # 🔥 changement demandé

    7: "APPARTEMENT",       # 🔥 renommage

    8: "VIOLET",
    9: "VIOLET",
    10: "VIOLET",
    11: "VIOLET",
    12: "VIOLET",
    13: "VIOLET",
    14: "VIOLET",
    15: "VIOLET",
    16: "VIOLET"            # 🔥 ajout de la 16e clé
}



@app.route("/formateurs")
def formateurs_home():
    filtre_docs = request.args.get("filtre") == "docs_a_controler"
    formateurs = load_formateurs()
    profils_docs_config = load_formateur_profils_docs_config()
    available_doc_labels = get_all_formateur_document_labels(formateurs, profils_docs_config)

    for f in formateurs:
        f["profils"] = normalize_formateur_profils(f.get("profils", []))
        apply_profile_document_requirements(f, profils_docs_config)
        if "cle" not in f:
            f["cle"] = {
                "attribuee": False,
                "numero": "",
                "statut": "non_attribuee"
            }

        if "badge" not in f:
            f["badge"] = {
                "attribue": False,
                "numero": "",
                "statut": "non_attribue"
            }

        # ✅ conformité + simple indicateur "docs à contrôler"
        total = 0
        conformes = 0
        a_controler = False

        for doc in f.get("documents", []):
            auto_update_document_status(doc)

            status = doc.get("status", "non_conforme")
            if status != "non_concerne":
                total += 1
                if status == "conforme":
                    conformes += 1
                if status == "a_controler":
                    a_controler = True

        f["conformite"] = {"conformes": conformes, "total": total}
        f["a_controler"] = a_controler

    if filtre_docs:
        formateurs = [f for f in formateurs if f.get("a_controler")]

    # ===== EXTRACTION DES CLÉS & BADGES =====
    liste_cles = []
    liste_badges = []

    for f in formateurs:
        # --- Clés ---
        cle = f.get("cle", {})
        num = cle.get("numero", "").strip()
        if cle.get("attribuee") and num.isdigit():
            liste_cles.append(int(num))

        # --- Badges ---
        badge = f.get("badge", {})
        num_b = badge.get("numero", "").strip()
        if badge.get("attribue") and num_b.isdigit():
            liste_badges.append(int(num_b))

    liste_cles = sorted(liste_cles)
    liste_badges = sorted(liste_badges)

    # ===== PLAGES TOTALES =====
    total_cles = list(range(1, 17))  # Clés 1 → 16
    total_badges = list(range(1, 16))     # Badges 1 → 15

    # ===== NUMÉROS DISPONIBLES =====
    cles_dispos = [n for n in total_cles if n not in liste_cles]
    badges_dispos = [n for n in total_badges if n not in liste_badges]

    # ===== ÉTAT COMPLET CLÉS & BADGES =====
    etat_cles, etat_badges = get_etat_cles_badges(formateurs, 16, 15)


    return render_template(
        "formateurs.html",
        title="Contrôle formateurs",
        formateurs=formateurs,
        formateur_profile_options=FORMATEUR_PROFILE_OPTIONS,
        profils_docs_config=profils_docs_config,
        available_doc_labels=available_doc_labels,
        liste_cles=liste_cles,
        liste_badges=liste_badges,
        cles_dispos=cles_dispos,
        badges_dispos=badges_dispos,
        filtre_docs=filtre_docs,
        etat_cles=etat_cles,       # 👈 ajouté
        etat_badges=etat_badges   # 👈 ajouté
    )





@app.route("/formateurs/add", methods=["POST"])
def add_formateur():
    formateurs = load_formateurs()
    profils_docs_config = load_formateur_profils_docs_config()
    fid = str(uuid.uuid4())[:8]

    try:
        nub = normalize_formateur_nub(request.form.get("nub", ""))
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("formateurs_home"))

    formateur = {
        "id": fid,
        "nom": request.form.get("nom", "").strip(),
        "prenom": request.form.get("prenom", "").strip(),
        "nub": nub,
        "email": request.form.get("email", "").strip(),
        "telephone": request.form.get("telephone", "").strip(),
        "siret": request.form.get("siret", "").strip(),
        "adresse_postale": request.form.get("adresse_postale", "").strip(),
        "nda": request.form.get("nda", "").strip(),
        "tarif_journalier_ht": request.form.get("tarif_journalier_ht", "").strip(),
        "profils": normalize_formateur_profils(request.form.getlist("profils")),

        "cle": {
            "attribuee": False,
            "numero": "",
            "statut": "non_attribuee"
        },

        "badge": {
            "attribue": False,
            "numero": "",
            "statut": "non_attribue"
        },

        "documents": build_default_documents()
    }

    formateurs.append(formateur)
    apply_profile_document_requirements(formateur, profils_docs_config)
    save_formateurs(formateurs)
    return redirect(url_for("formateur_detail", fid=fid))


@app.route("/formateurs/<fid>/profils/update", methods=["POST"])
def update_formateur_profils(fid):
    formateurs = load_formateurs()
    profils_docs_config = load_formateur_profils_docs_config()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        abort(404)

    formateur["profils"] = normalize_formateur_profils(request.form.getlist("profils"))
    apply_profile_document_requirements(formateur, profils_docs_config)
    save_formateurs(formateurs)
    flash("Profils formateur mis à jour.", "ok")
    return redirect(url_for("formateur_detail", fid=fid))


@app.route("/formateurs/profils-docs/update", methods=["POST"])
def update_formateur_profils_docs():
    config = {}
    for option in FORMATEUR_PROFILE_OPTIONS:
        profile_key = option["key"]
        config[profile_key] = sanitize_doc_labels(request.form.getlist(f"docs_{profile_key}"))

    save_formateur_profils_docs_config(config)

    formateurs = load_formateurs()
    for formateur in formateurs:
        apply_profile_document_requirements(formateur, config)
    save_formateurs(formateurs)

    flash("Règles documents par profil enregistrées.", "ok")
    return redirect(url_for("formateurs_home"))


@app.route("/formateurs/doc-labels/add", methods=["POST"])
def add_formateur_doc_label():
    label = (request.form.get("new_doc_label") or "").strip()
    if not label:
        flash("Merci de renseigner le nom du document.", "error")
        return redirect(url_for("formateurs_home"))

    selected_profiles = normalize_formateur_profils(request.form.getlist("new_doc_profiles"))

    formateurs = load_formateurs()
    profils_docs_config = load_formateur_profils_docs_config()

    for profile_key in selected_profiles:
        docs = profils_docs_config.setdefault(profile_key, [])
        if label not in docs:
            docs.append(label)

    for formateur in formateurs:
        docs = formateur.setdefault("documents", [])
        existing_labels = {(d.get("label") or "").strip() for d in docs}
        if label not in existing_labels:
            doc = build_doc_entry(label)
            if selected_profiles:
                doc["status"] = "non_concerne"
            docs.append(doc)
        apply_profile_document_requirements(formateur, profils_docs_config)

    save_formateur_profils_docs_config(profils_docs_config)
    save_formateurs(formateurs)

    flash("Nouveau document ajouté à la grille.", "ok")
    return redirect(url_for("formateurs_home"))


@app.route("/formateurs/<fid>/identity/update", methods=["POST"])
def update_formateur_identity(fid):
    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        return {"ok": False, "error": "Formateur introuvable"}, 404

    try:
        nub = normalize_formateur_nub(request.form.get("nub", ""))
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}, 400

    formateur["nub"] = nub
    formateur["siret"] = (request.form.get("siret") or "").strip()
    formateur["adresse_postale"] = (request.form.get("adresse_postale") or "").strip()
    formateur["nda"] = (request.form.get("nda") or "").strip()
    formateur["tarif_journalier_ht"] = (request.form.get("tarif_journalier_ht") or "").strip()
    save_formateurs(formateurs)
    return {"ok": True, "nub": nub, "siret": formateur["siret"], "adresse_postale": formateur["adresse_postale"], "nda": formateur["nda"], "tarif_journalier_ht": formateur["tarif_journalier_ht"]}


@app.route("/formateurs/<fid>/cle/update", methods=["POST"])
def update_formateur_cle(fid):
    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        return {"ok": False}, 404

    cle = formateur.setdefault("cle", {})

    cle["attribuee"] = request.form.get("attribuee") == "true"
    cle["numero"] = request.form.get("numero", "").strip()
    cle["statut"] = request.form.get("statut", "non_attribuee")

    # 🆕 AJOUT — nom libre si la clé est donnée à quelqu’un qui n’est pas formateur
    cle["custom_nom"] = request.form.get("custom_nom", "").strip()

    save_formateurs(formateurs)
    return {"ok": True}


@app.route("/formateurs/<fid>/badge/update", methods=["POST"])
def update_formateur_badge(fid):
    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        return {"ok": False}, 404

    badge = formateur.setdefault("badge", {})

    badge["attribue"] = request.form.get("attribue") == "true"
    badge["numero"] = request.form.get("numero", "").strip()
    badge["statut"] = request.form.get("statut", "non_attribue")

    save_formateurs(formateurs)
    return {"ok": True}



@app.route("/formateurs/<fid>")
def formateur_detail(fid):
    formateurs = load_formateurs()
    profils_docs_config = load_formateur_profils_docs_config()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        abort(404)
    formateur["profils"] = normalize_formateur_profils(formateur.get("profils", []))
    apply_profile_document_requirements(formateur, profils_docs_config)

    # 🔑🟦 RÉCUPÉRER TOUTES LES CLÉS / BADGES EXISTANTS
    etat_cles, etat_badges = get_etat_cles_badges(formateurs, 16, 15)
    last_relance_display = ""
    raw_relance = (formateur.get("last_relance") or "").strip()
    if raw_relance:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(raw_relance, fmt)
                last_relance_display = dt.strftime("%d/%m/%Y")
                break
            except ValueError:
                continue

    return render_template(
        "formateur_detail.html",
        title=f"Contrôle formateur — {formateur.get('prenom', '')} {formateur.get('nom', '').upper()}",
        formateur=formateur,
        last_relance_display=last_relance_display,
        formateur_profile_options=FORMATEUR_PROFILE_OPTIONS,
        etat_cles=etat_cles,       # 👈 indispensable
        etat_badges=etat_badges    # 👈 indispensable
    )



@app.route("/formateurs/<fid>/yousign/send", methods=["POST"])
def send_formateur_contract_yousign(fid):
    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        abort(404)

    state = normalize_yousign_state(formateur.get("yousign"))
    if state.get("signatureRequestId") and state.get("status") in {"draft", "approval", "ongoing"} and not request.form.get("force"):
        flash("Une demande Yousign active existe déjà pour ce formateur. Synchronisez le statut ou forcez un remplacement.", "error")
        return redirect(url_for("formateur_detail", fid=fid))

    email = (formateur.get("email") or "").strip()
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        flash("Email formateur invalide ou manquant.", "error")
        return redirect(url_for("formateur_detail", fid=fid))
    if not is_yousign_configured():
        flash("Yousign n'est pas configuré: renseignez YOUSIGN_API_KEY côté serveur.", "error")
        return redirect(url_for("formateur_detail", fid=fid))

    doc_id = (request.form.get("doc_id") or "").strip()
    doc, attachment, pdf_path = latest_formateur_pdf_attachment(formateur, doc_id)
    if not pdf_path:
        flash("Aucun contrat PDF n'a été trouvé dans les pièces jointes du formateur.", "error")
        return redirect(url_for("formateur_detail", fid=fid))

    client = YousignClient()
    now = datetime.now().isoformat(timespec="seconds")
    try:
        trainer_name = formateur_full_name(formateur) or email
        external_id = sanitize_yousign_external_id(f"formateur-{fid}", fallback="formateur-contract")
        app.logger.info("Yousign trainer contract external_id=%s", external_id)
        signature_request = client.create_signature_request(f"Contrat formateur - {trainer_name}", external_id=external_id)
        signature_request_id = signature_request.get("id")
        with open(pdf_path, "rb") as pdf_file:
            document = client.upload_file(signature_request_id, pdf_file.read(), attachment.get("original_name") or "contrat.pdf")
        document_id = document.get("id")
        signer = client.add_signer(signature_request_id, formateur.get("prenom") or "", formateur.get("nom") or trainer_name, email, document_id=document_id)
        activated = client.activate_signature_request(signature_request_id)
        status = extract_yousign_status(activated) or "ongoing"
        signature_url = signer.get("signature_link") or signer.get("signature_url") or activated.get("signature_link") or ""
        update_formateur_yousign_state(formateur, {
            "signatureRequestId": signature_request_id,
            "externalId": external_id,
            "documentId": document_id or "",
            "signerId": signer.get("id") or "",
            "status": status,
            "signatureUrl": signature_url,
            "sentAt": now,
            "lastSyncedAt": now,
            "error": None,
        })
        save_formateurs(formateurs)
        flash("Contrat envoyé à Yousign pour signature.", "ok")
    except YousignError as exc:
        update_formateur_yousign_state(formateur, {"status": "error", "lastSyncedAt": now, "error": str(exc)})
        save_formateurs(formateurs)
        flash(f"Erreur Yousign: {exc}", "error")
    return redirect(url_for("formateur_detail", fid=fid))


@app.route("/formateurs/<fid>/yousign/sync", methods=["POST"])
def sync_formateur_contract_yousign(fid):
    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        abort(404)
    state = normalize_yousign_state(formateur.get("yousign"))
    signature_request_id = state.get("signatureRequestId")
    if not signature_request_id:
        flash("Aucune demande Yousign à synchroniser.", "error")
        return redirect(url_for("formateur_detail", fid=fid))
    try:
        now = datetime.now().isoformat(timespec="seconds")
        updates = sync_yousign_signature_request_from_api(signature_request_id, now)
        status = updates.get("status") or state.get("status")
        update_formateur_yousign_state(formateur, updates)
        save_formateurs(formateurs)
        if updates.get("lastSyncedAt"):
            flash("Statut Yousign synchronisé.", "ok")
        else:
            flash(f"Erreur de synchronisation Yousign: {updates.get('apiError') or 'réponse API inexploitable'}", "error")
    except YousignError as exc:
        update_formateur_yousign_state(formateur, {"apiError": str(exc), "apiHttpStatus": str(exc.status_code or "network"), "apiStatus": f"erreur {exc.status_code or 'network'}", "error": str(exc)})
        save_formateurs(formateurs)
        flash(f"Erreur de synchronisation Yousign: {exc}", "error")
    return redirect(url_for("formateur_detail", fid=fid))


@app.route("/formateurs/<fid>/yousign/download", methods=["POST"])
def download_formateur_signed_yousign(fid):
    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        abort(404)
    state = normalize_yousign_state(formateur.get("yousign"))
    if not state.get("signatureRequestId"):
        flash("Aucune demande Yousign disponible.", "error")
        return redirect(url_for("formateur_detail", fid=fid))
    try:
        content = YousignClient().download_signed_documents(state["signatureRequestId"])
        signed_dir = os.path.join(FORMATEUR_FILES_DIR, fid, "_yousign")
        os.makedirs(signed_dir, exist_ok=True)
        filename = f"contrat_signe_yousign_{state['signatureRequestId']}.zip"
        with open(os.path.join(signed_dir, filename), "wb") as f:
            f.write(content)
        update_formateur_yousign_state(formateur, {"signedDocumentFilename": filename, "lastSyncedAt": datetime.now().isoformat(timespec="seconds"), "error": None})
        save_formateurs(formateurs)
        return send_from_directory(signed_dir, filename, as_attachment=True)
    except Exception as exc:
        flash(f"Téléchargement Yousign impossible: {exc}", "error")
        return redirect(url_for("formateur_detail", fid=fid))


@app.route("/webhooks/yousign", methods=["POST"])
def yousign_webhook():
    raw_body = request.get_data()
    yousign_signature_header_names = [
        "X-Yousign-Signature-256",
        "X-Yousign-Signature",
        "Yousign-Signature",
        "X-Hub-Signature-256",
    ]
    important_headers = {
        key: request.headers.get(key)
        for key in ["Content-Type", "User-Agent", *yousign_signature_header_names]
        if request.headers.get(key)
    }
    payload = request.get_json(silent=True) or {}
    event_name = payload.get("event_name") or payload.get("event") or payload.get("type") or "unknown"
    data_payload = payload.get("data") if isinstance(payload.get("data"), dict) else {}
    signature_request = data_payload.get("signature_request") if isinstance(data_payload.get("signature_request"), dict) else {}
    signer = data_payload.get("signer") if isinstance(data_payload.get("signer"), dict) else {}
    signature_request_id = (signature_request or {}).get("id") or data_payload.get("signature_request_id") or payload.get("signature_request_id") or (signer.get("signature_request") or {}).get("id")
    signer_id = signer.get("id") or data_payload.get("signer_id") or payload.get("signer_id")
    external_id = extract_yousign_external_id(payload)
    logger.info(
        "YOUSIGN WEBHOOK RECEIVED event=%s signature_request_id=%s external_id=%s",
        event_name, signature_request_id or "missing", external_id or "missing",
    )
    logger.info(
        "Yousign webhook diagnostic method=%s url=%s headers=%s body=%s signer_id=%s",
        request.method, request.url, important_headers, raw_body.decode("utf-8", errors="replace"), signer_id or "missing",
    )

    webhook_secret = get_yousign_config().webhook_secret
    signature_header = next((request.headers.get(key) for key in yousign_signature_header_names if request.headers.get(key)), None)
    if webhook_secret:
        if not signature_header:
            logger.warning("YOUSIGN WEBHOOK IGNORED event=%s reason=missing_signature", event_name)
            return {"ok": True, "ignored": True}
        expected = hmac.new(webhook_secret.encode("utf-8"), raw_body, hashlib.sha256).hexdigest()
        provided = signature_header.split("=", 1)[-1].strip()
        if not hmac.compare_digest(expected, provided):
            logger.warning("YOUSIGN WEBHOOK IGNORED event=%s reason=invalid_signature", event_name)
            return {"ok": True, "ignored": True}

    if event_name not in YOUSIGN_HANDLED_WEBHOOK_EVENTS:
        logger.info("YOUSIGN WEBHOOK IGNORED event=%s", event_name)
        return {"ok": True, "ignored": True}

    status = YOUSIGN_EVENT_STATUS.get(event_name) or extract_yousign_status(signature_request or payload)
    now = datetime.now().isoformat(timespec="seconds")
    updates = {"status": status, "externalId": external_id, "lastWebhookAt": now, "lastEvent": event_name, "lastEventAt": now, "error": None}
    if signer_id:
        updates["signerId"] = signer_id
    if status == "done": updates["signedAt"] = now
    elif status == "declined": updates["declinedAt"] = now
    elif status == "expired": updates["expiredAt"] = now
    elif status == "canceled": updates["canceledAt"] = now

    try:
        if signature_request_id:
            updates.update(sync_yousign_signature_request_from_api(signature_request_id, now))
            updates.update({"lastWebhookAt": now, "lastEvent": event_name, "lastEventAt": now})
    except YousignError as exc:
        logger.warning("Yousign webhook API sync failed signature_request_id=%s error=%s", signature_request_id, exc)
        updates["error"] = str(exc)

    def matches_yousign_state(state):
        normalized = normalize_yousign_state(state)
        return any([
            signature_request_id and normalized.get("signatureRequestId") == signature_request_id,
            external_id and normalized.get("externalId") == external_id,
            signer_id and normalized.get("signerId") == signer_id,
        ])

    formateurs = load_formateurs()
    formateur = next((f for f in formateurs if matches_yousign_state(f.get("yousign"))), None)
    if formateur:
        update_formateur_yousign_state(formateur, updates)
        save_formateurs(formateurs)
        logger.info("Webhook Yousign appliqué au formateur id=%s status=%s", formateur.get("id"), updates.get("status"))
        return {"ok": True, "target": "formateur"}

    sessions_data = load_sessions()
    for session_data in sessions_data.get("sessions", []):
        for contract in session_data.get("apsTrainerContracts", []):
            if matches_yousign_state(contract.get("yousign")) or matches_yousign_state(contract):
                contract["yousign"] = normalize_yousign_state({**contract.get("yousign", {}), **updates})
                mirror_yousign_state_on_contract(contract)
                save_sessions(sessions_data)
                logger.info("Webhook Yousign appliqué au contrat APS session=%s contract=%s status=%s", session_data.get("id"), contract.get("id"), updates.get("status"))
                return {"ok": True, "target": "aps_trainer_contract"}

    logger.error("YOUSIGN WEBHOOK CONTRACT NOT FOUND event=%s signature_request_id=%s signer_id=%s external_id=%s", event_name, signature_request_id, signer_id, external_id)
    return {"ok": True, "ignored": True}


@app.route("/formateurs/<fid>/delete", methods=["POST"])
def delete_formateur(fid):
    formateurs = load_formateurs()
    formateurs = [f for f in formateurs if f.get("id") != fid]
    save_formateurs(formateurs)
    flash("Formateur supprimé.", "ok")
    return redirect(url_for("formateurs_home"))


@app.route("/formateurs/<fid>/documents/add", methods=["POST"])
def add_formateur_document(fid):
    label = request.form.get("label", "").strip()
    if not label:
        return redirect(url_for("formateur_detail", fid=fid))

    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        abort(404)

    doc = build_doc_entry(label)
    formateur.setdefault("documents", []).append(doc)
    save_formateurs(formateurs)
    return redirect(url_for("formateur_detail", fid=fid))


@app.route("/formateurs/<fid>/documents/<doc_id>/update", methods=["POST"])
def update_formateur_document(fid, doc_id):
    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        return {"ok": False, "error": "Formateur introuvable"}, 404

    docs = formateur.get("documents", [])
    doc = next((d for d in docs if d.get("id") == doc_id), None)
    if not doc:
        return {"ok": False, "error": "Document introuvable"}, 404

    # champs texte
    if "expiration" in request.form:
        doc["expiration"] = request.form.get("expiration", "").strip()

    if "status" in request.form:
        st = request.form.get("status")
        if st in ("non_concerne", "a_controler", "conforme", "non_conforme"):
            doc["status"] = st


    if "commentaire" in request.form:
        doc["commentaire"] = request.form.get("commentaire", "").strip()

    # pièces jointes : on conserve uniquement le dernier fichier déposé
    files = [f for f in request.files.getlist("piece_jointe") if f.filename]
    if files:
        replace_formateur_attachment(fid, doc, files[-1])

    if "status" not in request.form:
        auto_update_document_status(doc)
    save_formateurs(formateurs)

    # ⛔️ PLUS AUCUN REDIRECT
    return {"ok": True}


@app.route("/formateurs/<fid>/media/upload", methods=["POST"])
def upload_formateur_media(fid):
    media_type = request.form.get("media_type", "").strip()
    if media_type not in ("photo", "badge_photo"):
        return {"ok": False, "error": "Type de média invalide"}, 400

    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        return {"ok": False, "error": "Formateur introuvable"}, 404

    media_file = request.files.get("file")
    if not media_file or not media_file.filename:
        return {"ok": False, "error": "Aucun fichier reçu"}, 400

    ext = os.path.splitext(media_file.filename)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".webp"):
        return {"ok": False, "error": "Format non supporté"}, 400

    storage_dir = os.path.join(FORMATEUR_FILES_DIR, fid, "_media")
    os.makedirs(storage_dir, exist_ok=True)
    unique_name = f"{media_type}_{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(storage_dir, unique_name)
    media_file.save(file_path)

    old_filename = formateur.get(media_type)
    if old_filename:
        old_path = os.path.join(storage_dir, old_filename)
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except OSError:
                pass

    formateur[media_type] = unique_name
    save_formateurs(formateurs)
    return {
        "ok": True,
        "url": url_for("download_formateur_media", fid=fid, filename=unique_name),
    }


@app.route("/formateurs/<fid>/media/<filename>")
def download_formateur_media(fid, filename):
    subdir = os.path.join(FORMATEUR_FILES_DIR, fid, "_media")
    return send_from_directory(subdir, filename, as_attachment=False)



@app.route("/formateurs/<fid>/documents/<doc_id>/attachments/<filename>")
def download_formateur_attachment(fid, doc_id, filename):
    subdir = os.path.join(FORMATEUR_FILES_DIR, fid, doc_id)
    return send_from_directory(subdir, filename, as_attachment=False)


@app.route(
    "/formateurs/<fid>/documents/<doc_id>/attachments/<filename>/delete",
    methods=["POST"]
)
def delete_formateur_attachment(fid, doc_id, filename):

    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        return {"ok": False}, 404

    doc = next(
        (d for d in formateur.get("documents", [])
         if d.get("id") == doc_id),
        None
    )
    if not doc:
        return {"ok": False}, 404

    # 📁 Suppression fichier physique
    file_path = os.path.join(
        FORMATEUR_FILES_DIR,
        fid,
        doc_id,
        filename
    )
    if os.path.exists(file_path):
        os.remove(file_path)

    # 🧹 Suppression dans le JSON
    doc["attachments"] = [
        a for a in doc.get("attachments", [])
        if a.get("filename") != filename
    ]

    # 🔁 Si plus de PJ → non conforme
    if not doc["attachments"]:
        doc["status"] = "non_conforme"

    save_formateurs(formateurs)

    return {"ok": True}




# ------------------------------------------------------------
# 📊 Route JSON pour les dotations (affichage sur index)
# ------------------------------------------------------------
@app.route("/dotations_data.json")
def dotations_data():
    try:
        data = load_dotations()
        a_distribuer = len([d for d in data if d.get("statut") == "Dotation à distribuer"])
        distribuees = len([d for d in data if d.get("statut") == "Dotation distribuée"])
        non_restituees = len([d for d in data if d.get("statut") == "Dotation non restituée"])
        restituees = len([d for d in data if d.get("statut") == "Dotation restituée"])

        payload = {
            "a_distribuer": a_distribuer,
            "distribuees": distribuees,
            "non_restituees": non_restituees,
            "restituees": restituees
        }

        headers = {"Content-Type": "application/json", "Access-Control-Allow-Origin": "*"}
        return json.dumps(payload, ensure_ascii=False), 200, headers
    except Exception as e:
        print("Erreur /dotations_data.json :", e)
        return json.dumps({"error": str(e)}), 500, {"Access-Control-Allow-Origin": "*"}

# ------------------------------------------------------------
# 📊 Route JSON Formateurs (pour tuile dashboard)
# ------------------------------------------------------------
@app.route("/formateurs_data.json")
def formateurs_data():
    try:
        formateurs = load_formateurs()

        total_non_conformes = 0
        total_a_controler = 0

        liste_non_conformes = set()
        liste_a_controler = set()

        for f in formateurs:
            nom_complet = f"{f.get('prenom','')} {f.get('nom','')}".strip()

            has_non_conforme = False
            has_a_controler = False

            for doc in f.get("documents", []):
                auto_update_document_status(doc)
                st = doc.get("status")

                if st == "non_conforme":
                    total_non_conformes += 1
                    has_non_conforme = True

                if st == "a_controler":
                    total_a_controler += 1
                    has_a_controler = True

            if has_non_conforme:
                liste_non_conformes.add(nom_complet)

            if has_a_controler:
                liste_a_controler.add(nom_complet)

        payload = {
            "non_conformes": total_non_conformes,
            "a_controler": total_a_controler,
            "liste_non_conformes": sorted(list(liste_non_conformes)),
            "liste_a_controler": sorted(list(liste_a_controler)),
        }

        headers = {"Content-Type": "application/json", "Access-Control-Allow-Origin": "*"}
        return json.dumps(payload, ensure_ascii=False), 200, headers

    except Exception as e:
        print("Erreur /formateurs_data.json :", e)
        return json.dumps({"error": str(e)}), 500, {"Access-Control-Allow-Origin": "*"}


import zipfile
from flask import send_file
from io import BytesIO

@app.route("/formateurs/<fid>/export")
def export_formateur_dossier(fid):
    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        abort(404)

    nom = (formateur.get("nom") or "").upper()
    prenom = (formateur.get("prenom") or "").strip()
    dossier_name = f"{nom} {prenom}".strip()

    # ZIP en mémoire (pas écrit sur le disque)
    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for doc in formateur.get("documents", []):
            label = doc.get("label", "Document").strip()

            for att in doc.get("attachments", []):
                filename = att.get("filename")
                original = att.get("original_name")

                if not filename or not original:
                    continue

                file_path = os.path.join(
                    FORMATEUR_FILES_DIR,
                    fid,
                    doc["id"],
                    filename
                )

                if not os.path.exists(file_path):
                    continue

                ext = os.path.splitext(original)[1]
                clean_name = f"{label} {prenom} {nom}{ext}"

                arcname = os.path.join(
                    dossier_name,
                    clean_name
                )

                zipf.write(file_path, arcname)

    zip_buffer.seek(0)

    zip_filename = f"Dossier formateur {prenom} {nom}.zip"

    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=zip_filename,
        mimetype="application/zip"
    )

@app.route("/formateurs/<fid>/print")
def print_formateur_dossier(fid):
    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        abort(404)

    # Liste des docs non conformes / manquants
    non_conformes = [
        d for d in formateur.get("documents", [])
        if d.get("status") == "non_conforme"
    ]

    return render_template(
        "formateur_print.html",
        title="État du dossier formateur",
        formateur=formateur,
        non_conformes=non_conformes,
        today=datetime.now().strftime("%d/%m/%Y")
    )

import hashlib
import hmac
import time

def generate_upload_token(fid):
    secret = app.secret_key
    raw = f"{fid}:{secret}"
    return hashlib.sha256(raw.encode()).hexdigest()

def verify_upload_token(fid, token):
    return token == generate_upload_token(fid)

@app.route("/formateurs/<fid>/upload", methods=["GET", "POST"])
def upload_formateur_documents(fid):
    token = request.args.get("token", "")
    if not verify_upload_token(fid, token):
        abort(403)

    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        abort(404)

    # Documents à régulariser
    docs_ko = [
        d for d in formateur.get("documents", [])
        if d.get("status") == "non_conforme"
    ]

    if request.method == "POST":
        doc_id = request.form.get("doc_id")
        files = request.files.getlist("files")

        doc = next((d for d in docs_ko if d["id"] == doc_id), None)
        if not doc:
            abort(400)

        uploaded_files = [f for f in files if f.filename]
        if not uploaded_files:
            flash("❌ Aucun fichier n'a été sélectionné.", "error")
            return redirect(request.url)

        allowed_ext = ["pdf", "png", "jpg", "jpeg"]

        for f in uploaded_files:
            # Vérification extension
            ext = f.filename.lower().rsplit(".", 1)[-1]
            if ext not in allowed_ext:
                flash("❌ Seuls les fichiers PDF, PNG, JPG et JPEG sont acceptés.", "error")
                return redirect(request.url)

        # On remplace l'ancien dépôt et on conserve uniquement le dernier fichier reçu.
        replace_formateur_attachment(fid, doc, uploaded_files[-1])

        # Après upload → à contrôler
        doc["status"] = "a_controler"
        save_formateurs(formateurs)
        flash("Document transmis avec succès.", "ok")
        return redirect(request.url)

    return render_template(
        "formateur_upload.html",
        formateur=formateur,
        docs_ko=docs_ko
    )


@app.route("/formateurs/<fid>/send_mail", methods=["POST"])
def send_formateur_relance(fid):
    formateurs = load_formateurs()
    formateur = find_formateur(formateurs, fid)
    if not formateur:
        abort(404)

    # 📌 Documents non conformes avec commentaire
    docs_ko = [
        {
            "label": d["label"],
            "commentaire": d.get("commentaire", "").strip()
        }
        for d in formateur.get("documents", [])
        if d.get("status") == "non_conforme"
    ]

    if not docs_ko:
        flash("Aucun document à relancer.", "ok")
        return redirect(url_for("formateur_detail", fid=fid))

    # 🔗 Génération lien sécurisé pour upload
    token = generate_upload_token(fid)
    link = url_for(
        "upload_formateur_documents",
        fid=fid,
        token=token,
        _external=True
    )

    # ✉️ Contenu du mail avec bouton visible
    body = f"""
Bonjour {formateur.get('prenom')},<br><br>

Votre dossier formateur nécessite quelques mises à jour. Merci de transmettre vos documents via le bouton ci-dessous. 
<b style='color:#d00000;'>Les envois par mail ne sont plus acceptés.</b><br><br>

<div style="text-align:center;margin:25px 0;">
  <a href="{link}" style="
      display:inline-block;
      padding:14px 28px;
      background:#0f62fe;
      color:#ffffff !important;
      font-size:18px;
      font-weight:700;
      border-radius:8px;
      text-decoration:none;
      box-shadow:0 4px 12px rgba(0,0,0,0.18);
  ">
      📁 Déposer mes documents
  </a>
</div>

Voici les éléments à régulariser :<br><br>

<ul style="font-size:15px;line-height:1.5;">
  {''.join(
    f"<li><b>{d['label']}</b>"
    + (f"<br><span style='color:red;font-weight:600;'>⚠️ {d['commentaire']}</span>" if d['commentaire'] else "")
    + "</li><br>"
    for d in docs_ko
  )}
</ul>

Cordialement,<br>
<b>Intégrale Academy</b>
"""


    # 📩 Envoi
    send_email(
        formateur.get("email"),
        "Documents manquants — Dossier formateur",
        body
    )

    # 🕒 Trace de la relance
    formateur["last_relance"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_formateurs(formateurs)

    flash("📧 Mail envoyé au formateur.", "ok")
    return redirect(url_for("formateur_detail", fid=fid))


@app.route("/cle/assign", methods=["POST"])
def assign_cle():
    payload = request.get_json()
    numero = str(payload.get("numero"))
    fid = payload.get("fid")

    formateurs = load_formateurs()

    # 🔄 Retirer cette clé à tous les formateurs
    for f in formateurs:
        cle = f.setdefault("cle", {})
        if cle.get("numero") == numero:
            cle["attribuee"] = False
            cle["numero"] = ""
            cle["statut"] = "non_attribuee"
            cle["custom_nom"] = ""

    # 🚫 Si Libre → fini
    if not fid:
        save_formateurs(formateurs)
        return {"ok": True}

    # ✅ Sinon attribuer la clé
    formateur = next((f for f in formateurs if f["id"] == fid), None)
    if not formateur:
        return {"ok": False, "error": "Formateur introuvable"}

    formateur["cle"]["attribuee"] = True
    formateur["cle"]["numero"] = numero
    formateur["cle"]["statut"] = "attribuee"
    formateur["cle"]["custom_nom"] = ""

    save_formateurs(formateurs)
    return {"ok": True}


@app.route("/badge/assign", methods=["POST"])
def assign_badge():
    payload = request.get_json()
    numero = str(payload.get("numero"))
    fid = payload.get("fid")

    formateurs = load_formateurs()

    # 🔄 Retirer ce badge à tous les formateurs
    for f in formateurs:
        badge = f.setdefault("badge", {})
        if badge.get("numero") == numero:
            badge["attribue"] = False
            badge["numero"] = ""
            badge["statut"] = "non_attribue"

    # 🚫 Si Libre
    if not fid:
        save_formateurs(formateurs)
        return {"ok": True}

    # ✅ Sinon attribuer le badge
    formateur = next((f for f in formateurs if f["id"] == fid), None)
    if not formateur:
        return {"ok": False, "error": "Formateur introuvable"}

    formateur["badge"]["attribue"] = True
    formateur["badge"]["numero"] = numero
    formateur["badge"]["statut"] = "attribue"

    save_formateurs(formateurs)
    return {"ok": True}

# ------------------------------------------------------------
# 🟩 GESTION DU DISTRIBUTEUR — PERSISTENCE JSON
# ------------------------------------------------------------

DISTRIBUTEUR_FILE = os.path.join(DATA_DIR, "distributeur.json")

def load_distributeur():
    """Charge le distributeur depuis le fichier JSON, ou crée une structure par défaut."""
    default_data = {
        "lignes": [
            {"id": 1, "produits": []},
            {"id": 2, "produits": []},
            {"id": 3, "produits": []},
            {"id": 4, "produits": []},
            {"id": 5, "produits": []}
        ]
    }

    if os.path.exists(DISTRIBUTEUR_FILE):
        try:
            with open(DISTRIBUTEUR_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict) and isinstance(data.get("lignes"), list):
                    return data
        except Exception:
            app.logger.exception("Impossible de lire %s", DISTRIBUTEUR_FILE)

            # Tentative de récupération: certains fichiers ont du texte parasite
            # autour d'un JSON valide (copier/coller, merge, etc.).
            try:
                with open(DISTRIBUTEUR_FILE, "r", encoding="utf-8") as f:
                    content = f.read()
                start = content.find("{")
                end = content.rfind("}")
                if start != -1 and end != -1 and end > start:
                    recovered = json.loads(content[start:end + 1])
                    if isinstance(recovered, dict) and isinstance(recovered.get("lignes"), list):
                        app.logger.warning("Récupération partielle de %s après corruption JSON", DISTRIBUTEUR_FILE)
                        return recovered
            except Exception:
                app.logger.exception("Récupération impossible pour %s", DISTRIBUTEUR_FILE)

    return default_data

def save_distributeur(data):
    """Sauvegarde complète du distributeur."""
    with open(DISTRIBUTEUR_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)



@app.route("/distributeur")
def distributeur_home():
    data = load_distributeur()
    return render_template("distributeur.html", data=data)

@app.route("/distributeur/add/<int:ligne_id>", methods=["POST"])
def distributeur_add(ligne_id):
    data = load_distributeur()

    # Trouver la ligne
    ligne = next((l for l in data["lignes"] if l["id"] == ligne_id), None)
    if not ligne:
        abort(404)

    # Ajouter un produit vide
    new_product = {
        "id": str(uuid.uuid4())[:8],
        "nom": "",
        "qte_cible": 0,
        "qte_actuelle": 0,
        "prix_achat": 0.0,
        "prix_vente": 0.0
    }

    ligne["produits"].append(new_product)
    save_distributeur(data)

    return redirect(url_for("distributeur_home"))

def to_int(x):
    try:
        return int(x)
    except:
        return 0

def to_float(x):
    try:
        return float(x)
    except:
        return 0.0


@app.route("/distributeur/update/<int:ligne_id>/<pid>", methods=["POST"])
def distributeur_update(ligne_id, pid):
    data = load_distributeur()

    for ligne in data["lignes"]:
        if ligne["id"] == ligne_id:
            for produit in ligne["produits"]:
                if produit["id"] == pid:

                    produit["nom"] = request.form.get("nom", "")

                    produit["qte_cible"] = to_int(request.form.get("qte_cible"))
                    produit["qte_actuelle"] = to_int(request.form.get("qte_actuelle"))

                    produit["prix_achat"] = to_float(request.form.get("prix_achat"))
                    produit["prix_vente"] = to_float(request.form.get("prix_vente"))

                    save_distributeur(data)
                    return "ok"

    return "error", 400



@app.route("/distributeur/delete/<int:ligne_id>/<pid>", methods=["POST"])
def distributeur_delete(ligne_id, pid):
    data = load_distributeur()

    # retrouver la ligne
    ligne = next((l for l in data["lignes"] if l["id"] == ligne_id), None)
    if not ligne:
        abort(404)

    # filtrer les produits en supprimant celui voulu
    ligne["produits"] = [p for p in ligne["produits"] if p.get("id") != pid]

    save_distributeur(data)

    return redirect(url_for("distributeur_home"))

@app.route("/reassort")
def distributeur_reassort():
    data = load_distributeur()

    items = []
    for ligne in data["lignes"]:
        for p in ligne["produits"]:
            q_cible = p.get("qte_cible", 0)
            q_actuelle = p.get("qte_actuelle", 0)

            if q_actuelle < q_cible:
                items.append({
                    "ligne_id": ligne["id"],
                    "produit_id": p["id"],
                    "nom": p.get("nom", "Produit"),
                    "reassort": q_cible - q_actuelle,
                    "q_cible": q_cible,
                    "q_actuelle": q_actuelle,
                })

    # Priorité visuelle: afficher d'abord les étages 4 puis 5 dans la liste de réassort.
    etage_priority = {4: 0, 5: 1}
    items.sort(key=lambda item: (etage_priority.get(item["ligne_id"], 2), item["ligne_id"], item["nom"].lower()))

    return render_template("reassort.html", items=items)


@app.route("/distributeur/approvisionnement")
def distributeur_approvisionnement():
    data = load_distributeur()

    produits = []
    for ligne in data["lignes"]:
        for p in ligne["produits"]:
            nom = (p.get("nom") or "").strip()
            if not nom:
                continue
            produits.append({
                "id": p["id"],
                "ligne_id": ligne["id"],
                "nom": nom,
            })

    produits.sort(key=lambda item: (item["ligne_id"], item["nom"].lower()))

    return render_template("approvisionnement.html", produits=produits)

@app.route("/reassort/valider/<int:ligne_id>/<produit_id>", methods=["POST"])
def distributeur_reassort_valider(ligne_id, produit_id):
    data = load_distributeur()

    for ligne in data["lignes"]:
        if ligne["id"] == ligne_id:
            for p in ligne["produits"]:
                if str(p["id"]) == str(produit_id):
                    # Mise à jour automatique
                    p["qte_actuelle"] = p.get("qte_cible", 0)

                    save_distributeur(data)
                    break

    return redirect(url_for("distributeur_reassort"))

start_price_adaptator_scheduler()

import xml.etree.ElementTree as ET
from flask import Response, request
from datetime import datetime
import sqlite3
import calendar
import csv
from io import StringIO, BytesIO
from openpyxl import Workbook

def _first_text(elem, tag_name):
    if elem is None:
        return None
    for child in list(elem):
        if child.tag.endswith(tag_name):
            return (child.text or "").strip()
    return None

@app.post("/webhooks/salesforce/lead-outbound")
def salesforce_lead_outbound():
    # sécurité optionnelle
    secret_expected = os.environ.get("SF_OUTBOUND_SECRET")
    if secret_expected and request.args.get("key") != secret_expected:
        return Response("Unauthorized", status=401)

    raw = request.data.decode("utf-8", errors="ignore")
    if not raw.strip():
        return Response("Empty body", status=400)

    try:
        root = ET.fromstring(raw)
    except Exception as e:
        return Response(f"Bad XML: {e}", status=400)

    # On récupère TON format de stockage (dict)
    data = load_price_adaptator_data()

    now_iso = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    added = 0

    for notif in root.iter():
        if not notif.tag.endswith("notifications"):
            continue

        sobject = None
        for e in notif.iter():
            if e.tag.endswith("sObject"):
                sobject = e
                break
        if sobject is None:
            continue

        # mapping Outbound -> ton modèle prospect
        prospect = {
            "id": str(uuid.uuid4()),
            "nom": normalize_price_adaptator_nom(_first_text(sobject, "LastName")),
            "prenom": normalize_price_adaptator_prenom(_first_text(sobject, "FirstName")),
            "cpf": float(_first_text(sobject, "Montant_CPF__c") or 0),
            "email": (_first_text(sobject, "Email") or "").strip(),
            "telephone": (_first_text(sobject, "Phone") or "").strip(),
            "formation": normalize_price_adaptator_formation(_first_text(sobject, "Type_de_formation__c")),
            "sent": False,
            "sentAt": None,
            "proposed_price": None,
            "last_error": None,
            "last_attempt_at": None,
            "created_at": datetime.now().isoformat(),
            "salesforce": {
                "lead_id": _first_text(sobject, "Id"),
                "received_at": now_iso
            }
        }

        # on n’ajoute que si formation ok + nom/prénom ok
        if prospect["nom"] and prospect["prenom"] and prospect["formation"]:
            data["prospects"].insert(0, prospect)
            added += 1

    save_price_adaptator_data(data)

    soap_response = """<?xml version="1.0" encoding="UTF-8"?>
<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/">
  <soapenv:Body>
    <notificationsResponse xmlns="http://soap.sforce.com/2005/09/outbound">
      <Ack>true</Ack>
    </notificationsResponse>
  </soapenv:Body>
</soapenv:Envelope>
"""
    return Response(soap_response, status=200, content_type="text/xml; charset=utf-8")

# -----------------------
# 📅 Gestion des salles / planning formations (SQLite)
# -----------------------
PLANNING_SALLES = ["Salle 1", "Salle 2", "Salle 1B", "Salle 2B", "Salle 3B"]
PLANNING_TYPES = ["APS", "A3P", "SSIAP", "DESP", "VTC", "BTS", "Autre"]
PERSIST_DIR = os.environ.get("PERSIST_DIR")
PLANNING_DB_NAME = "formations.db"
DB_DIR = PERSIST_DIR if PERSIST_DIR else DATA_DIR
os.makedirs(DB_DIR, exist_ok=True)
PLANNING_DB = os.path.join(DB_DIR, PLANNING_DB_NAME)
LEGACY_PLANNING_DB = os.path.join(BASE_DIR, PLANNING_DB_NAME)

def ensure_planning_db_location():
    """
    Garantit un chemin stable de base de données et migre l'ancienne DB
    si elle existe dans l'ancien emplacement.
    """
    if PLANNING_DB == LEGACY_PLANNING_DB:
        return
    if os.path.exists(PLANNING_DB):
        return
    if os.path.exists(LEGACY_PLANNING_DB):
        try:
            with open(LEGACY_PLANNING_DB, "rb") as src, open(PLANNING_DB, "wb") as dst:
                dst.write(src.read())
            logger.info("Migration planning DB: %s -> %s", LEGACY_PLANNING_DB, PLANNING_DB)
        except OSError as exc:
            logger.warning("Impossible de migrer la DB planning legacy: %s", exc)

def get_db():
    ensure_planning_db_location()
    conn = sqlite3.connect(PLANNING_DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_planning_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS formations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nom TEXT NOT NULL,
                type TEXT NOT NULL,
                date_debut TEXT NOT NULL,
                date_fin TEXT NOT NULL,
                salle TEXT NOT NULL,
                nombre_stagiaires INTEGER NOT NULL,
                commentaire TEXT,
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS salles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nom TEXT NOT NULL UNIQUE,
                capacite_max INTEGER NOT NULL DEFAULT 20,
                equipements TEXT DEFAULT '',
                indisponibilites TEXT DEFAULT '',
                commentaire TEXT DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS formateurs_planning (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nom TEXT NOT NULL,
                prenom TEXT NOT NULL,
                telephone TEXT DEFAULT '',
                email TEXT DEFAULT '',
                competences TEXT DEFAULT '',
                disponibilites TEXT DEFAULT '',
                indisponibilites TEXT DEFAULT '',
                commentaire TEXT DEFAULT ''
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS planning_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                formation_id INTEGER,
                action TEXT NOT NULL,
                details TEXT DEFAULT '',
                user_email TEXT DEFAULT '',
                created_at TEXT NOT NULL
            )
        """)
        for salle in PLANNING_SALLES:
            conn.execute("INSERT OR IGNORE INTO salles(nom, capacite_max, active) VALUES(?, ?, 1)", (salle, 20))

def add_planning_history(formation_id, action, details=""):
    with get_db() as conn:
        conn.execute(
            "INSERT INTO planning_history(formation_id, action, details, user_email, created_at) VALUES(?,?,?,?,?)",
            (formation_id, action, details, session.get("admin_email", ""), datetime.now().isoformat()),
        )

def dates_overlap(start_a, end_a, start_b, end_b):
    return start_a <= end_b and start_b <= end_a

def salle_disponible(salle, date_debut, date_fin, exclude_id=None):
    query = "SELECT id, date_debut, date_fin FROM formations WHERE salle = ?"
    params = [salle]
    if exclude_id is not None:
        query += " AND id != ?"
        params.append(exclude_id)
    with get_db() as conn:
        rows = conn.execute(query, params).fetchall()
    for row in rows:
        if dates_overlap(date_debut, date_fin, row["date_debut"], row["date_fin"]):
            return False
    return True

def choisir_salle(date_debut, date_fin, salle_souhaitee=None, exclude_id=None):
    if salle_souhaitee:
        if salle_disponible(salle_souhaitee, date_debut, date_fin, exclude_id=exclude_id):
            return salle_souhaitee, None
        return None, f"La salle {salle_souhaitee} n'est pas disponible sur cette période."
    for salle in PLANNING_SALLES:
        if salle_disponible(salle, date_debut, date_fin, exclude_id=exclude_id):
            return salle, None
    return None, "Aucune salle disponible sur cette période"

def format_formation(row):
    f = dict(row)
    f["conflit"] = not salle_disponible(f["salle"], f["date_debut"], f["date_fin"], exclude_id=f["id"])
    return f

@app.route("/planning")
def planning_home():
    init_planning_db()
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM formations ORDER BY date_debut ASC, id DESC").fetchall()
    formations = [format_formation(r) for r in rows]
    today = datetime.now().date()
    salles_occupees = {
        f["salle"] for f in formations
        if f["date_debut"] <= today.isoformat() <= f["date_fin"]
    }
    next_sessions = [
        f for f in formations
        if f["date_debut"] >= today.isoformat()
    ][:5]
    stats = {
        "total_formations": len(formations),
        "salles_occupees": len(salles_occupees),
        "salles_disponibles": len(PLANNING_SALLES) - len(salles_occupees),
        "conflits": sum(1 for f in formations if f["conflit"]),
        "prochaines_sessions": next_sessions,
    }
    q = request.args.get("q", "").strip().lower()
    salle_filter = request.args.get("salle", "").strip()
    type_filter = request.args.get("type", "").strip()
    statut_filter = request.args.get("statut", "").strip()
    if q:
        formations = [f for f in formations if q in f["nom"].lower() or q in (f.get("commentaire") or "").lower()]
    if salle_filter:
        formations = [f for f in formations if f["salle"] == salle_filter]
    if type_filter:
        formations = [f for f in formations if f["type"] == type_filter]
    if statut_filter == "conflit":
        formations = [f for f in formations if f["conflit"]]
    if statut_filter == "ok":
        formations = [f for f in formations if not f["conflit"]]
    return render_template("planning.html", formations=formations, salles=PLANNING_SALLES, stats=stats, types=PLANNING_TYPES)

@app.route("/planning/formations")
def planning_formations():
    init_planning_db()
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM formations ORDER BY date_debut ASC, id DESC").fetchall()
    formations = [format_formation(r) for r in rows]
    return render_template("planning_formations.html", formations=formations)

@app.route("/formation/ajouter", methods=["GET", "POST"])
def formation_ajouter():
    init_planning_db()
    if request.method == "POST":
        nom = request.form.get("nom", "").strip()
        type_formation = request.form.get("type", "").strip()
        date_debut = request.form.get("date_debut", "").strip()
        date_fin = request.form.get("date_fin", "").strip()
        nombre_stagiaires = int(request.form.get("nombre_stagiaires") or 0)
        salle_souhaitee = request.form.get("salle_souhaitee", "").strip() or None
        commentaire = request.form.get("commentaire", "").strip()
        if not nom or not type_formation or not date_debut or not date_fin or date_debut > date_fin:
            flash("Merci de remplir correctement le formulaire.", "error")
            return render_template("formation_form.html", salles=PLANNING_SALLES, types=PLANNING_TYPES, mode="ajouter", formation=request.form)
        salle, error = choisir_salle(date_debut, date_fin, salle_souhaitee=salle_souhaitee)
        if error:
            salle = salle_souhaitee or (PLANNING_SALLES[0] if PLANNING_SALLES else "Sans salle")
        with get_db() as conn:
            conn.execute("""INSERT INTO formations(nom, type, date_debut, date_fin, salle, nombre_stagiaires, commentaire, created_at)
                         VALUES(?,?,?,?,?,?,?,?)""",
                         (nom, type_formation, date_debut, date_fin, salle, nombre_stagiaires, commentaire, datetime.now().isoformat()))
            formation_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        add_planning_history(formation_id, "creation", f"{nom} / {salle} / {date_debut}-{date_fin}")
        if error:
            flash(f"{error} Formation ajoutée avec conflit sur {salle}.", "error")
        else:
            flash("Formation ajoutée avec succès.", "success")
        return redirect(url_for("planning_home"))
    return render_template("formation_form.html", salles=PLANNING_SALLES, types=PLANNING_TYPES, mode="ajouter", formation={})

@app.route("/formation/<int:id>/modifier", methods=["GET", "POST"])
def formation_modifier(id):
    init_planning_db()
    with get_db() as conn:
        current = conn.execute("SELECT * FROM formations WHERE id = ?", (id,)).fetchone()
    if not current:
        abort(404)
    if request.method == "POST":
        nom = request.form.get("nom", "").strip()
        type_formation = request.form.get("type", "").strip()
        date_debut = request.form.get("date_debut", "").strip()
        date_fin = request.form.get("date_fin", "").strip()
        nombre_stagiaires = int(request.form.get("nombre_stagiaires") or 0)
        salle_souhaitee = request.form.get("salle_souhaitee", "").strip() or None
        commentaire = request.form.get("commentaire", "").strip()
        salle, error = choisir_salle(date_debut, date_fin, salle_souhaitee=salle_souhaitee, exclude_id=id)
        if error:
            salle = salle_souhaitee or (current["salle"] if current["salle"] else (PLANNING_SALLES[0] if PLANNING_SALLES else "Sans salle"))
        with get_db() as conn:
            conn.execute("""UPDATE formations SET nom=?, type=?, date_debut=?, date_fin=?, salle=?, nombre_stagiaires=?, commentaire=? WHERE id=?""",
                         (nom, type_formation, date_debut, date_fin, salle, nombre_stagiaires, commentaire, id))
        add_planning_history(id, "modification", f"{nom} / {salle} / {date_debut}-{date_fin}")
        if error:
            flash(f"{error} Formation modifiée avec conflit sur {salle}.", "error")
        else:
            flash("Formation modifiée.", "success")
        return redirect(url_for("planning_home"))
    return render_template("formation_form.html", salles=PLANNING_SALLES, types=PLANNING_TYPES, mode="modifier", formation=dict(current), formation_id=id)

@app.route("/formation/<int:id>/edit", methods=["GET", "POST"])
def formation_edit_alias(id):
    return formation_modifier(id)

@app.post("/formation/<int:id>/supprimer")
def formation_supprimer(id):
    with get_db() as conn:
        conn.execute("DELETE FROM formations WHERE id = ?", (id,))
    add_planning_history(id, "suppression", "formation supprimée")
    flash("Formation supprimée.", "success")
    return redirect(url_for("planning_home"))

@app.route("/planning/export.csv")
def planning_export_csv():
    init_planning_db()
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM formations ORDER BY date_debut ASC").fetchall()
    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(["Nom", "Type", "Date début", "Date fin", "Salle", "Stagiaires", "Commentaire"])
    for r in rows:
        writer.writerow([r["nom"], r["type"], format_date(r["date_debut"]), format_date(r["date_fin"]), r["salle"], r["nombre_stagiaires"], r["commentaire"] or ""])
    resp = Response(out.getvalue(), mimetype="text/csv; charset=utf-8")
    resp.headers["Content-Disposition"] = "attachment; filename=planning_formations.csv"
    return resp

@app.route("/planning/export.xlsx")
def planning_export_xlsx():
    init_planning_db()
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM formations ORDER BY date_debut ASC").fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Planning"
    ws.append(["Nom", "Type", "Date début", "Date fin", "Salle", "Stagiaires", "Commentaire"])
    for r in rows:
        ws.append([r["nom"], r["type"], format_date(r["date_debut"]), format_date(r["date_fin"]), r["salle"], r["nombre_stagiaires"], r["commentaire"] or ""])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="planning_formations.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/planning/impression")
def planning_impression():
    init_planning_db()
    mode = request.args.get("mode", "global")
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM formations ORDER BY date_debut ASC").fetchall()
    formations = [dict(r) for r in rows]
    return render_template("planning_print.html", formations=formations, mode=mode, now=datetime.now())

@app.route("/calendrier")
def calendrier():
    init_planning_db()
    room_colors = {
        "Salle 1": "#2563EB",
        "Salle 1B": "#16A34A",
        "Salle 2": "#7C3AED",
        "Salle 3": "#EA580C",
        "Salle 4": "#EC4899",
        "Salle 5": "#334155",
    }
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM formations ORDER BY date_debut ASC").fetchall()
        room_rows = conn.execute("SELECT nom, capacite_max FROM salles WHERE active = 1 ORDER BY nom ASC").fetchall()
    formations = [format_formation(r) for r in rows]
    room_capacity = {r["nom"]: r["capacite_max"] for r in room_rows}
    today_iso = datetime.now().date().isoformat()
    salle_meta = []
    for salle in PLANNING_SALLES:
        occuped_today = []
        for f in formations:
            if f["salle"] == salle and f["date_debut"] <= today_iso <= f["date_fin"]:
                occuped_today.append(f["nom"])
        salle_meta.append({
            "nom": salle,
            "capacite": room_capacity.get(salle, 20),
            "couleur": room_colors.get(salle, "#059669"),
            "statut": "occupée" if occuped_today else "libre",
            "occupations": occuped_today,
        })

    events = [{
        "id": r["id"],
        "title": r["nom"],
        "start": r["date_debut"],
        "end": (datetime.strptime(r["date_fin"], "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d"),
        "backgroundColor": "#DC2626" if r["conflit"] else room_colors.get(r["salle"], "#059669"),
        "borderColor": "#991B1B" if r["conflit"] else room_colors.get(r["salle"], "#059669"),
        "extendedProps": {
            "type": r["type"],
            "salle": r["salle"],
            "stagiaires": r["nombre_stagiaires"],
            "commentaire": r["commentaire"] or "",
            "conflit": bool(r["conflit"]),
            "formateur": "Non renseigné",
            "horaire": "Journée",
        }
    } for r in formations]
    return render_template("calendrier.html", events=events, salles_meta=salle_meta, salles=PLANNING_SALLES, types=PLANNING_TYPES)

@app.route("/salles", methods=["GET", "POST"])
def salles_page():
    init_planning_db()
    if request.method == "POST":
        with get_db() as conn:
            conn.execute(
                "INSERT INTO salles(nom, capacite_max, equipements, indisponibilites, commentaire, active) VALUES(?,?,?,?,?,?)",
                (
                    request.form.get("nom", "").strip(),
                    int(request.form.get("capacite_max") or 20),
                    request.form.get("equipements", "").strip(),
                    request.form.get("indisponibilites", "").strip(),
                    request.form.get("commentaire", "").strip(),
                    1 if request.form.get("active") == "on" else 0,
                ),
            )
        flash("Salle ajoutée.", "success")
        return redirect(url_for("salles_page"))
    with get_db() as conn:
        salles = conn.execute("SELECT * FROM salles ORDER BY nom ASC").fetchall()
    return render_template("salles.html", salles=salles)

@app.route("/formateurs-planning", methods=["GET", "POST"])
def formateurs_planning_page():
    init_planning_db()
    if request.method == "POST":
        with get_db() as conn:
            conn.execute(
                """INSERT INTO formateurs_planning(nom, prenom, telephone, email, competences, disponibilites, indisponibilites, commentaire)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (
                    request.form.get("nom", "").strip(),
                    request.form.get("prenom", "").strip(),
                    request.form.get("telephone", "").strip(),
                    request.form.get("email", "").strip(),
                    request.form.get("competences", "").strip(),
                    request.form.get("disponibilites", "").strip(),
                    request.form.get("indisponibilites", "").strip(),
                    request.form.get("commentaire", "").strip(),
                ),
            )
        flash("Formateur ajouté.", "success")
        return redirect(url_for("formateurs_planning_page"))
    with get_db() as conn:
        formateurs = conn.execute("SELECT * FROM formateurs_planning ORDER BY nom ASC").fetchall()
    return render_template("formateurs_planning.html", formateurs=formateurs)

@app.route("/planning/historique")
def planning_historique():
    init_planning_db()
    with get_db() as conn:
        logs = conn.execute("SELECT * FROM planning_history ORDER BY id DESC LIMIT 300").fetchall()
    return render_template("planning_history.html", logs=logs)

@app.post("/planning/disponibilites")
def planning_disponibilites():
    date_debut = request.form.get("date_debut")
    date_fin = request.form.get("date_fin")
    libres, occupees = [], []
    for salle in PLANNING_SALLES:
        if salle_disponible(salle, date_debut, date_fin):
            libres.append(salle)
        else:
            occupees.append(salle)
    return jsonify({"libres": libres, "occupees": occupees})










    
