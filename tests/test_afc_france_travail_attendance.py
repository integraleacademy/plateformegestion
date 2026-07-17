from datetime import date
from pathlib import Path
import hashlib, sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook
from services.afc_france_travail_attendance import (
    is_afc_session, update_afc_france_travail_settings, save_france_travail_ids,
    preview, generate_france_travail_workbook, build_session_weeks, build_week_schedule,
    get_week_trainees, attendance_students
)

ROOT = Path(__file__).resolve().parents[1]

def sha(p): return hashlib.sha256(Path(p).read_bytes()).hexdigest()

def slot(start,end,cat,trainer="Formateur Un"):
    h1,m1=map(int,start.split(':')); h2,m2=map(int,end.split(':'))
    return {"start":start,"end":end,"durationMinutes":(h2*60+m2)-(h1*60+m1),"duration":((h2*60+m2)-(h1*60+m1))/60,"afcCategory":cat,"afcKind":"FT" if cat in {"APS","SSIAP1"} else cat,"uv":cat,"trainer":trainer,"modality":"presentiel"}

def sample_session(student_count=4):
    days=[]
    d=date(2026,11,18)
    while d <= date(2026,12,4):
        if d.weekday()<5:
            days.append({"date":d.isoformat(),"slots":[slot("08:30","12:30","RAN"),slot("13:30","16:30","RAN")]})
        d = date.fromordinal(d.toordinal()+1)
    days[3]["slots"] = []  # journée sans cours explicite dans le planning
    days[4]["slots"] = days[4]["slots"][:1]  # demi-journée sans cours
    days[6]["slots"][1] = slot("13:30","16:30","SP","Formateur Deux")
    students=[]
    for i in range(student_count):
        students.append({"id": f"s{i}", "lastName": f"NOM{i}", "firstName": f"Prenom{i}", "startDate": days[0]["date"], "france_travail_id": f"00A{i}"})
    students[1]["startDate"] = days[2]["date"]
    if len(students) > 2:
        students[2]["france_travail_id"] = ""
    return {"id":"afc-ft","formation":"AFC_APS_SSIAP","training_code":"AFC_APS_SSIAP","display_name":"AFC Test","date_debut":days[0]["date"],"date_fin":days[-1]["date"],"apsPlanningData":days,"apsAttendanceStudents":students,"france_travail":{"marche_afc":"M1","brs":"B1","convention":"C1","bon_commande":"BC1","type_session":"ESF","intitule":"Intitulé FT"}}

def test_card_condition_exact_afc_only():
    assert is_afc_session({"formation":"AFC_APS_SSIAP"})
    for code in ["APS","SSIAP1","A3P","DIRIGEANT","DESP","BTS"]:
        assert not is_afc_session({"formation": code})

def test_settings_are_session_scoped_and_ids_remain_text_with_zero():
    a=sample_session(); b=sample_session(); b["id"]="b"
    update_afc_france_travail_settings(a,{"marche_afc":"000MARCHE","intitule":"A"})
    update_afc_france_travail_settings(b,{"marche_afc":"M2","intitule":"B"})
    assert a["france_travail"]["marche_afc"] == "000MARCHE"
    assert b["france_travail"]["marche_afc"] == "M2"
    save_france_travail_ids(a,{"s0":"00123ABC"})
    assert a["apsAttendanceStudents"][0]["france_travail_id"] == "00123ABC"

def test_preview_warns_missing_ids_and_settings_and_counts_three_weeks_trainers():
    s=sample_session(); s["france_travail"]["brs"]=""
    p=preview(s)
    assert p["weekCount"] == 3
    assert p["missingIdCount"] == 1 and "NOM2 Prenom2" in p["missingIds"]
    assert "brs" in p["missingSettings"]
    assert p["trainerCount"] == 2

def test_workbook_generation_keeps_template_intact_removes_old_sheets_and_opens():
    template = ROOT/"static/upload/tableau.xlsx"; before=sha(template)
    s=sample_session(); bio=generate_france_travail_workbook(s, ROOT)
    assert sha(template) == before
    wb=load_workbook(bio)
    assert wb.sheetnames == ["1611 au 2011", "2311 au 2711", "3011 au 0412"]
    assert "2024 Exemple pour mode emploi" not in wb.sheetnames
    assert all("LAM ALAM" not in str(ws.values) for ws in wb.worksheets)

def test_partial_week_grey_days_blank_signatures_and_header_30_hours():
    wb=load_workbook(generate_france_travail_workbook(sample_session(), ROOT)); ws=wb[wb.sheetnames[0]]
    assert ws["A7"].value == "Durée hebdomadaire : 30 heures"
    assert ws["C14"].fill.fgColor.rgb in ("00D9D9D9","D9D9D9")
    assert ws["E14"].fill.fgColor.rgb in ("00D9D9D9","D9D9D9")
    assert ws["G14"].value is None and ws["G14"].fill.fill_type is None
    assert ws["B14"].value == "00A0"
    assert ws["B14"].number_format == "@"

def test_different_student_entry_dates_and_weekly_totals_differ():
    wb=load_workbook(generate_france_travail_workbook(sample_session(), ROOT)); ws=wb[wb.sheetnames[0]]
    assert ws["A14"].value.startswith("1 NOM0")
    assert ws["A16"].value.startswith("2 NOM1")
    assert ws["M14"].value != ws["M16"].value
    assert ws["C16"].fill.fgColor.rgb in ("00D9D9D9","D9D9D9")

def test_empty_day_half_day_support_row_and_totals_are_computed():
    wb=load_workbook(generate_france_travail_workbook(sample_session(), ROOT)); ws=wb[wb.sheetnames[1]]
    # First day of this week was emptied: all signature cells for first trainee are grey.
    assert ws["C14"].fill.fgColor.rgb in ("00D9D9D9","D9D9D9") and ws["D14"].fill.fgColor.rgb in ("00D9D9D9","D9D9D9")
    # Support slot appears on the second line only.
    assert ws["J15"].fill.fill_type is None
    assert ws["J14"].fill.fill_type is None
    total_row = next(r for r in range(1, ws.max_row+1) if ws.cell(r,1).value == "Total des heures facturables")
    assert ws.cell(total_row,13).value is not None
    assert ws.cell(total_row+1,13).value is None

def test_less_and_more_than_twelve_students_keep_totals_trainers_notes_after_rows():
    for count in (2, 13):
        wb=load_workbook(generate_france_travail_workbook(sample_session(count), ROOT)); ws=wb[wb.sheetnames[0]]
        last_student_row = 14 + count*2 - 1
        total_row = next(r for r in range(1, ws.max_row+1) if ws.cell(r,1).value == "Total des heures facturables")
        trainer_row = next(r for r in range(1, ws.max_row+1) if str(ws.cell(r,1).value).startswith("Formateur"))
        assert total_row > last_student_row
        assert trainer_row > total_row
        assert ws.print_area

def test_week_trainees_come_from_planning_and_status_filtered():
    s=sample_session(); s["apsAttendanceStudents"].append({"lastName":"BAD","firstName":"Deleted","status":"supprimé"})
    week=build_session_weeks(s)[0]; sched=build_week_schedule(s, week); trainees=get_week_trainees(s, week, sched)
    assert all(t["lastName"] != "BAD" for t in trainees)
    assert len(attendance_students(s)) == 4

def test_generation_route_refuses_non_afc(monkeypatch):
    import app as app_module
    monkeypatch.setattr(app_module, "load_sessions", lambda: {"sessions":[{"id":"aps1","formation":"APS","training_code":"APS"}]})
    client = app_module.app.test_client()
    with client.session_transaction() as sess:
        sess["admin_logged"] = True
        sess["admin_session_version"] = app_module.ADMIN_SESSION_VERSION
    response = client.get("/api/sessions/aps1/afc-france-travail/generate")
    assert response.status_code == 403

def test_france_travail_groups_successive_morning_slots_without_changing_durations():
    s = sample_session(2)
    s["date_debut"] = "2026-11-16"; s["date_fin"] = "2026-11-16"
    for st in s["apsAttendanceStudents"]: st["startDate"] = "2026-11-16"
    s["apsPlanningData"] = [{"date":"2026-11-16","slots":[slot("08:30","10:00","FT"),slot("10:00","10:30","SP"),slot("10:30","12:30","FT")]}]
    wb=load_workbook(generate_france_travail_workbook(s, ROOT)); ws=wb.active
    assert ws["C12"].value == "08h30-12h30"
    assert ws["C13"].value == "FT / S"
    assert ws["M14"].value == 4
    assert ws["C15"].value == 0.5


def test_france_travail_groups_successive_afternoon_slots():
    s = sample_session(2)
    s["date_debut"] = "2026-11-16"; s["date_fin"] = "2026-11-16"
    for st in s["apsAttendanceStudents"]: st["startDate"] = "2026-11-16"
    s["apsPlanningData"] = [{"date":"2026-11-16","slots":[slot("13:30","14:00","FT"),slot("14:00","15:30","PAF"),slot("15:30","16:30","SP")]}]
    wb=load_workbook(generate_france_travail_workbook(s, ROOT)); ws=wb.active
    assert ws["D12"].value == "13h30-16h30"


def test_france_travail_splits_noon_boundary_into_correct_half_days():
    s = sample_session(2)
    s["date_debut"] = "2026-11-16"; s["date_fin"] = "2026-11-16"
    for st in s["apsAttendanceStudents"]: st["startDate"] = "2026-11-16"
    s["apsPlanningData"] = [{"date":"2026-11-16","slots":[slot("08:30","12:00","FT"),slot("12:00","12:30","SP"),slot("13:30","16:30","FT")]}]
    wb=load_workbook(generate_france_travail_workbook(s, ROOT)); ws=wb.active
    assert ws["C12"].value == "08h30-12h30"
    assert ws["D12"].value == "13h30-16h30"
    assert ws["C15"].value == 0.5


def test_france_travail_support_hours_by_student_and_totals():
    s = sample_session(2)
    s["date_debut"] = "2026-11-16"; s["date_fin"] = "2026-11-20"
    s["apsAttendanceStudents"][0]["startDate"] = "2026-11-16"
    s["apsAttendanceStudents"][1]["startDate"] = "2026-11-16"
    s["apsPlanningData"] = [{"date":"2026-11-16","slots":[slot("08:30","12:30","FT")]}]
    for offset, day in enumerate(["2026-11-17", "2026-11-20"]):
        s["apsPlanningData"].append({"date":day,"slots":[]})
    sp1 = slot("13:30","15:30","SP"); sp1["studentIds"]=["s0"]
    sp2 = slot("08:30","10:30","SP"); sp2["studentIds"]=["s0", "s1"]
    sp3 = slot("13:30","16:30","SP"); sp3["studentIds"]=["s0"]
    s["apsPlanningData"][1]["slots"]=[sp1]
    s["apsPlanningData"][2]["slots"]=[sp2, sp3]
    wb=load_workbook(generate_france_travail_workbook(s, ROOT)); ws=wb.active
    assert ws["F15"].value == 2
    assert ws["K15"].value == 2
    assert ws["L15"].value == 3
    assert ws["M15"].value == 7
    assert ws["M17"].value == 2
    assert ws["M14"].value == 11


def test_france_travail_support_absence_keeps_cells_blank_and_zero_total():
    s = sample_session(2)
    s["date_debut"] = "2026-11-16"; s["date_fin"] = "2026-11-16"
    for st in s["apsAttendanceStudents"]: st["startDate"] = "2026-11-16"
    s["apsPlanningData"] = [{"date":"2026-11-16","slots":[slot("08:30","12:30","FT")]}]
    wb=load_workbook(generate_france_travail_workbook(s, ROOT)); ws=wb.active
    assert all(ws.cell(15, c).value is None for c in range(3,13))
    assert ws["M15"].value == 0
