from datetime import date
from decimal import Decimal
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from app import (
    AFC_DSF_STATUS_CANCELLED, AFC_DSF_STATUS_FINALIZED,
    afc_dsf_compute, afc_dsf_next_number, afc_dsf_summary, afc_dsf_session_snapshot,
    build_afc_aps_ssiap_planning_data, generate_afc_dsf_pdf,
    is_afc_aps_ssiap_session,
)


def sample_session():
    planning = build_afc_aps_ssiap_planning_data(date(2026, 11, 16), "Formateur", "Salle", [])
    return {"id":"s1","formation":"AFC_APS_SSIAP","training_code":"AFC_APS_SSIAP","display_name":"AFC France Travail APS + SSIAP","date_debut":planning[0]["date"],"date_fin":planning[-1]["date"],"apsPlanningData":planning,"apsAttendanceStudents":[{"id":"a","lastName":"DUPONT","firstName":"Jean"},{"id":"b","lastName":"MARTIN","firstName":"Sophie"}],"afcDsfs":[]}

def test_tab_visibility_condition_exact_afc_only():
    assert is_afc_aps_ssiap_session({"formation":"AFC_APS_SSIAP"})
    assert not is_afc_aps_ssiap_session({"formation":"APS"})
    assert not is_afc_aps_ssiap_session({"formation":"SSIAP1"})
    assert not is_afc_aps_ssiap_session({"formation":"A3P"})
    assert not is_afc_aps_ssiap_session({"formation":"DIRIGEANT"})

def test_module_selection_one_or_two_and_third_refused():
    s=sample_session(); day=s["apsPlanningData"][0]["date"]
    assert afc_dsf_compute(s, day, day, ["RAN"])["totalHours"] > 0
    assert afc_dsf_compute(s, day, day, ["RAN","FT"])["totalHours"] > 0
    try: afc_dsf_compute(s, day, day, ["RAN","FT","SP"]); assert False
    except ValueError as e: assert "deux modules" in str(e)

def test_inclusive_dates_and_planning_hours_absences_ignored_no_presence_dependency():
    s=sample_session(); d=s["apsPlanningData"][0]["date"]; s["absences"]={"a":[d]}; s["presenceSheets"]="must-not-be-read"
    res=afc_dsf_compute(s,d,d,["RAN"])
    assert res["hoursPerStudent"]["RAN"] == 7
    assert res["totalHours"] == 14


def test_student_entry_date_reduces_only_that_student_hours():
    s = sample_session()
    first_day = s["apsPlanningData"][0]["date"]
    second_day = s["apsPlanningData"][1]["date"]
    s["apsAttendanceStudents"][1]["startDate"] = second_day
    res = afc_dsf_compute(s, first_day, second_day, ["RAN"])
    rows = {row["lastName"]: row for row in res["students"]}
    assert rows["DUPONT"]["modules"]["RAN"] == 14
    assert rows["MARTIN"]["modules"]["RAN"] == 7
    assert res["moduleTotals"]["RAN"] == 21
    summary = afc_dsf_summary(s)
    ran = next(card for card in summary["cards"] if card["code"] == "RAN")
    assert ran["plannedTotal"] == 103
    martin = next(row for row in summary["detail"] if row["student"]["lastName"] == "MARTIN")
    assert martin["modules"]["RAN"]["planned"] == 48

def test_afc_categories_grouped_and_modules_separated():
    s=sample_session(); all_start=s["date_debut"]; all_end=s["date_fin"]
    res=afc_dsf_compute(s,all_start,all_end,["FT","RAN"])
    assert res["hoursPerStudent"]["FT"] == 273
    assert res["hoursPerStudent"]["RAN"] == 55
    assert afc_dsf_compute(s,all_start,all_end,["SP"])["hoursPerStudent"]["SP"] == 45
    assert afc_dsf_compute(s,all_start,all_end,["PAF"])["hoursPerStudent"]["PAF"] == 20

def test_totals_numbering_double_billing_cancel_and_remaining_snapshot():
    s=sample_session(); d=s["apsPlanningData"][0]["date"]
    r1=afc_dsf_compute(s,d,d,["RAN"]); assert r1["totalHours"]==14
    dsf1={"id":"1","number":1,"label":"DSF 1","status":AFC_DSF_STATUS_FINALIZED,**r1}; s["afcDsfs"].append(dsf1)
    assert afc_dsf_next_number(s)==2
    try: afc_dsf_compute(s,d,d,["RAN"]); assert False
    except ValueError as e: assert "Aucune heure restante" in str(e)
    dsf2={"id":"2","number":2,"label":"DSF 2","status":AFC_DSF_STATUS_CANCELLED,**r1}; s["afcDsfs"].append(dsf2)
    assert afc_dsf_next_number(s)==3
    dsf1["status"]=AFC_DSF_STATUS_CANCELLED
    assert afc_dsf_compute(s,d,d,["RAN"])["totalHours"]==14
    summary=afc_dsf_summary(s); assert summary["cards"][0]["remainingTotal"] >= 0
    snap=dsf2["students"][0]["modules"]["RAN"]; s["apsPlanningData"][0]["slots"][0]["durationMinutes"]=60
    assert dsf2["students"][0]["modules"]["RAN"] == snap


def test_summary_does_not_crash_when_finalized_dsf_exceeds_current_planning():
    s = sample_session()
    d = s["apsPlanningData"][0]["date"]
    r = afc_dsf_compute(s, d, d, ["RAN"])
    r["students"][0]["modules"]["RAN"] = 999
    r["moduleTotals"]["RAN"] = 999
    r["totalHours"] = 999
    s["afcDsfs"].append({"id": "over", "number": 1, "label": "DSF 1", "status": AFC_DSF_STATUS_FINALIZED, **r})

    summary = afc_dsf_summary(s)

    ran = next(card for card in summary["cards"] if card["code"] == "RAN")
    assert ran["remainingTotal"] == 0
    assert ran["overbilledTotal"] > 0


def test_period_without_hours_refused():
    s=sample_session()
    try: afc_dsf_compute(s,"2026-11-21","2026-11-22",["PAF"]); assert False
    except ValueError as e: assert "Aucune heure" in str(e)


def test_editable_hourly_rate_recalculates_total_ca_and_coherence():
    s = sample_session()
    summary = afc_dsf_summary(s, hourly_rate="12.50")

    assert summary["rate"] == "12.50"
    assert summary["total"]["amountTotal"] == summary["total"]["amountBilled"] + summary["total"]["amountToInvoice"] + summary["total"]["amountRemaining"]
    assert summary["total"]["amountTotal"] == summary["total"]["planned"] * Decimal("12.50")


def test_compute_uses_custom_hourly_rate_for_dsf_preview():
    s = sample_session()
    d = s["apsPlanningData"][0]["date"]
    result = afc_dsf_compute(s, d, d, ["RAN"], hourly_rate="13.00")

    assert result["totalHours"] == 14
    assert result["amountTotal"] == "182.00"

def test_pdf_contains_students_labels_logo_and_no_nulls(tmp_path):
    from pypdf import PdfReader
    s=sample_session(); d=s["apsPlanningData"][0]["date"]; r=afc_dsf_compute(s,d,d,["RAN"])
    dsf={"id":"1","number":1,"label":"DSF 1","status":AFC_DSF_STATUS_FINALIZED,"createdAt":"2026-11-16 10:00:00",**r}
    out=tmp_path/"dsf.pdf"; generate_afc_dsf_pdf(s,dsf,str(out))
    text="\n".join(p.extract_text() or "" for p in PdfReader(str(out)).pages)
    assert "DUPONT Jean" in text and "MARTIN Sophie" in text
    assert "Remise à niveau (RAN)" in text and "Intégrale Academy" in text
    assert all(x not in text for x in ["None","null","undefined"])


def test_delete_dsf_route_deletes_session_entry_and_pdf(monkeypatch, tmp_path):
    import app as application

    application.app.config.update(TESTING=True, SECRET_KEY="test")
    s = sample_session()
    d = s["apsPlanningData"][0]["date"]
    result = afc_dsf_compute(s, d, d, ["RAN"])
    pdf_path = tmp_path / "dsf.pdf"
    pdf_path.write_bytes(b"pdf")
    s["afcDsfs"].append({
        "id": "dsf-to-delete",
        "number": 1,
        "label": "DSF 1",
        "status": AFC_DSF_STATUS_FINALIZED,
        "pdfFilename": pdf_path.name,
        **result,
    })
    saved = {"sessions": [s], "jurys": []}
    monkeypatch.setattr(application, "DSF_DIR", str(tmp_path))
    monkeypatch.setattr(application, "load_sessions", lambda: saved)
    monkeypatch.setattr(application, "save_sessions", lambda data: saved.update(data))

    with application.app.test_client() as client:
        with client.session_transaction() as flask_session:
            flask_session["admin_logged"] = True
            flask_session["admin_session_version"] = application.ADMIN_SESSION_VERSION
        response = client.post("/api/sessions/s1/afc-dsf/dsf-to-delete/delete")

    assert response.status_code == 200
    assert response.get_json()["deleted"] is True
    assert saved["sessions"][0]["afcDsfs"] == []
    assert not pdf_path.exists()


def render_afc_dsf_page(monkeypatch, session_data):
    import app as application

    application.app.config.update(TESTING=True, SECRET_KEY="test")
    saved = {"sessions": [session_data], "jurys": []}
    monkeypatch.setattr(application, "load_sessions", lambda: saved)
    monkeypatch.setattr(application, "save_sessions", lambda data: saved.update(data))
    with application.app.test_client() as client:
        with client.session_transaction() as flask_session:
            flask_session["admin_logged"] = True
            flask_session["admin_session_version"] = application.ADMIN_SESSION_VERSION
        response = client.get(f"/sessions/{session_data['id']}")
    assert response.status_code == 200
    return response.get_data(as_text=True)


def test_afc_dsf_dashboard_contains_five_cards_generate_button_actions_and_dynamic_data(monkeypatch):
    s = sample_session()
    result = afc_dsf_compute(s, s["date_debut"], s["date_debut"], ["RAN"])
    s["afcDsfs"].append({
        "id": "dsf-actions",
        "number": 1,
        "label": "DSF 1",
        "status": AFC_DSF_STATUS_FINALIZED,
        "createdAt": "2026-11-16 10:00:00",
        **result,
    })

    html = render_afc_dsf_page(monkeypatch, s)

    assert html.count('data-testid="afc-dsf-card-') == 5
    assert 'id="openAfcDsfModal"' in html and "Générer une DSF" in html
    assert "273 h" in html and "Remise à niveau (RAN)" in html
    assert "Voir le PDF" in html
    assert "Télécharger" in html
    assert "Supprimer la DSF" in html
    assert "/sessions/s1/afc-dsf/dsf-actions/pdf" in html
    assert "/sessions/s1/afc-dsf/dsf-actions/download" in html
    assert "/api/sessions/s1/afc-dsf/dsf-actions/delete" in html
    assert "window.SaasDialog?await window.SaasDialog.confirm" in html


def test_afc_dsf_overbilling_alert_is_conditional(monkeypatch):
    normal = sample_session()
    normal_html = render_afc_dsf_page(monkeypatch, normal)
    assert 'data-testid="afc-dsf-overbilling-alert"' not in normal_html

    overbilled = sample_session()
    result = afc_dsf_compute(overbilled, overbilled["date_debut"], overbilled["date_debut"], ["RAN"])
    result["students"][0]["modules"]["RAN"] = 999
    overbilled["afcDsfs"].append({
        "id": "over",
        "number": 1,
        "label": "DSF 1",
        "status": AFC_DSF_STATUS_FINALIZED,
        "createdAt": "2026-11-16 10:00:00",
        **result,
    })

    overbilled_html = render_afc_dsf_page(monkeypatch, overbilled)
    assert 'data-testid="afc-dsf-overbilling-alert"' in overbilled_html
    assert "Surfacturation détectée" in overbilled_html
    assert "h-stagiaires au-delà du planning actuel" in overbilled_html


def test_afc_dsf_student_detail_uses_grouped_module_headers_and_dynamic_values(monkeypatch):
    s = sample_session()
    html = render_afc_dsf_page(monkeypatch, s)

    assert 'data-testid="afc-dsf-student-detail"' in html
    assert 'afc-dsf-accordion afc-dsf-accordion--detail' in html
    assert 'Lecture par stagiaire, regroupée par module AFC' in html
    assert html.index('scope="colgroup" title="Remise à niveau (RAN)">RAN</th>') < html.index('scope="colgroup" title="Formation technique (FT)">FT</th>')
    assert 'scope="colgroup" title="Formation technique (FT)">FT</th>' in html
    assert 'scope="colgroup" title="Remise à niveau (RAN)">RAN</th>' in html
    assert 'scope="colgroup" title="Soutien personnalisé (SP)">SP</th>' in html
    assert 'scope="colgroup" title="Préparation à l’après-formation (PAF)">PAF</th>' in html
    assert html.count('class="afc-dsf-metric-head afc-dsf-metric-head--ft" scope="col">Prévue</th>') == 1
    assert html.count('scope="col">Facturée</th>') >= 4
    assert html.count('scope="col">Restante</th>') >= 4
    assert '<span>DUPONT Jean</span>' in html
    assert 'afc-dsf-detail-legend' in html
    assert 'Rouge = heures restantes à facturer' in html
    assert 'afc-dsf-module-cell--ft afc-dsf-module-cell--first' in html
    assert 'afc-dsf-module-cell--ft afc-dsf-module-cell--last is-positive' in html

def render_afc_dsf_print(monkeypatch, session_data):
    import app as application

    application.app.config.update(TESTING=True, SECRET_KEY="test")
    saved = {"sessions": [session_data], "jurys": []}
    monkeypatch.setattr(application, "load_sessions", lambda: saved)
    with application.app.test_client() as client:
        with client.session_transaction() as flask_session:
            flask_session["admin_logged"] = True
            flask_session["admin_session_version"] = application.ADMIN_SESSION_VERSION
        response = client.get(f"/sessions/{session_data['id']}/afc-detail/print")
    assert response.status_code == 200
    return response.get_data(as_text=True)


def test_afc_dsf_print_template_is_dedicated_without_navigation_or_sidebar(monkeypatch):
    html = render_afc_dsf_print(monkeypatch, sample_session())

    assert 'data-testid="afc-dsf-print-report"' in html
    assert "Détail des heures par stagiaire" in html
    assert "Suivi des heures prévues, facturées et restant à facturer" in html
    assert '<aside' not in html.lower()
    assert 'class="sidebar' not in html.lower()
    assert "detail-nav" not in html
    assert "Vue imprimable" not in html
    assert "Exporter en PDF" not in html


def test_afc_dsf_print_css_prevents_blank_pages_and_bad_page_breaks(monkeypatch):
    html = render_afc_dsf_print(monkeypatch, sample_session())

    assert "@page { size: A4 portrait; margin: 10mm 12mm; }" in html
    assert "height: auto !important" in html
    assert "min-height: 0 !important" in html
    assert "max-height: none !important" in html
    assert "overflow: visible !important" in html
    assert "page-break-inside: avoid" in html
    assert "break-inside: avoid" in html
    assert "page-break-after: auto" in html


def test_afc_dsf_print_statuses_and_overbilling_are_visible(monkeypatch):
    s = sample_session()
    full = afc_dsf_compute(s, s["date_debut"], s["date_fin"], ["RAN", "FT"])
    over = afc_dsf_compute(s, s["date_debut"], s["date_debut"], ["RAN"])
    over["students"][0]["modules"]["RAN"] = 999
    s["afcDsfs"].append({"id": "all", "number": 1, "label": "DSF 1", "status": AFC_DSF_STATUS_FINALIZED, **full})
    s["afcDsfs"].append({"id": "over", "number": 2, "label": "DSF 2", "status": AFC_DSF_STATUS_FINALIZED, **over})

    html = render_afc_dsf_print(monkeypatch, s)

    assert "green" in html
    assert " h à facturer" in html
    assert "Dépassement de " in html
    assert "orange : dépassement des heures prévues" in html


def test_dsf_france_travail_excel_snapshot_fills_template_and_keeps_original_intact():
    import hashlib
    from openpyxl import load_workbook
    from services.afc_dsf_france_travail_excel import generate_dsf_excel_from_snapshot
    from app import afc_dsf_session_snapshot

    s = sample_session()
    s["france_travail"] = {"convention": "041C", "intitule": "Intitulé AFC"}
    s["apsAttendanceStudents"][0]["france_travail_id"] = "0123A"
    s["apsAttendanceStudents"][1]["france_travail_id"] = ""
    d = s["apsPlanningData"][0]["date"]
    result = afc_dsf_compute(s, d, d, ["RAN"])
    snapshot = afc_dsf_session_snapshot(s, result, "1", "12.10")
    template = Path(__file__).resolve().parents[1] / "static/upload/dsf.xlsx"
    before = hashlib.sha256(template.read_bytes()).hexdigest()

    wb = load_workbook(generate_dsf_excel_from_snapshot(snapshot, Path(__file__).resolve().parents[1]))

    assert hashlib.sha256(template.read_bytes()).hexdigest() == before
    ws = wb["DSF1"]
    assert ws["D4"].value == "041C"
    assert ws["K3"].value == "Intitulé AFC"
    assert ws["E6"].value == "16/11/2026"
    assert ws["E7"].value == "16/11/2026"
    assert ws["I7"].value == "1"
    assert ws["C10"].value == "DUPONT Jean\n0123A"
    assert ws["D10"].value == "MARTIN Sophie"
    assert "nom prénom" not in " ".join(str(ws.cell(10, c).value or "") for c in range(3, 19))
    assert ws["C15"].value == 7
    assert ws["D15"].value == 7
    assert ws["C16"].value == 0 and ws["C17"].value == 0
    assert ws["C18"].value == 7
    assert ws["C27"].value == 7
    assert ws["C28"].value == 0
    assert ws["C29"].value == 7
    assert ws["B15"].value == 14
    assert ws["B29"].value == 14
    assert ws["E10"].value is None and ws["E15"].value is None


def test_dsf_france_travail_excel_hour_formats_are_value_specific_and_numeric():
    from openpyxl import load_workbook
    from services.afc_dsf_france_travail_excel import generate_dsf_excel_from_snapshot

    snapshot = {
        "number": "9",
        "periodStart": "2026-11-16",
        "periodEnd": "2026-11-27",
        "session": {
            "name": "AFC France Travail APS + SSIAP",
            "date_debut": "2026-11-16",
            "date_fin": "2026-11-27",
            "convention": "041C",
        },
        "students": [
            {"lastName": "ENTIER", "firstName": "VingtHuit", "totalHours": 28, "modules": {"FT": 28}},
            {"lastName": "ZERO", "firstName": "Heure", "totalHours": 0.5, "modules": {"FT": 0.5}},
            {"lastName": "QUATORZE", "firstName": "DecimalZero", "totalHours": 14.0, "modules": {"FT": 14.0}},
            {"lastName": "TROIS", "firstName": "Demie", "totalHours": 3.5, "modules": {"FT": 3.5}},
            {"lastName": "DEUX", "firstName": "Quart", "totalHours": 2.25, "modules": {"FT": 2.25}},
            {"lastName": "DIXSEPT", "firstName": "Demie", "totalHours": 17.5, "modules": {"FT": 17.5}},
        ],
    }

    wb = load_workbook(generate_dsf_excel_from_snapshot(snapshot, Path(__file__).resolve().parents[1]))
    ws = wb["DSF9"]

    assert ws["C11"].value == 28 and isinstance(ws["C11"].value, int)
    assert ws["C11"].number_format == "0"
    assert ws["C12"].value == 0 and isinstance(ws["C12"].value, int)
    assert ws["C12"].number_format == "0"
    assert ws["F11"].value == 3.5 and isinstance(ws["F11"].value, float)
    assert ws["F11"].number_format == "0.##"
    assert ws["G11"].value == 2.25 and isinstance(ws["G11"].value, float)
    assert ws["G11"].number_format == "0.##"
    assert ws["E11"].value == 14 and isinstance(ws["E11"].value, int)
    assert ws["E11"].number_format == "0"
    assert ws["I10"].value is None and ws["I11"].value is None and ws["I11"].number_format == "General"
    assert isinstance(ws["B11"].value, float)
    assert ws["B11"].value == 65.75
    assert ws["B11"].number_format == "0.##"
    assert load_workbook(generate_dsf_excel_from_snapshot(snapshot, Path(__file__).resolve().parents[1]))
    for row in range(11, 34):
        for col in range(2, 19):
            cell = ws.cell(row, col)
            if isinstance(cell.value, int):
                assert cell.number_format != "0.##"
    assert 15 not in [ws.cell(7, col).value for col in range(1, 19)]


def test_dsf_france_travail_excel_splits_after_sixteen_trainees():
    from openpyxl import load_workbook
    from services.afc_dsf_france_travail_excel import generate_dsf_excel_from_snapshot
    from app import afc_dsf_session_snapshot

    s = sample_session()
    s["apsAttendanceStudents"] = [{"id": f"s{i}", "lastName": f"NOM{i}", "firstName": "Test", "france_travail_id": f"00{i}"} for i in range(17)]
    d = s["apsPlanningData"][0]["date"]
    result = afc_dsf_compute(s, d, d, ["RAN"])
    snapshot = afc_dsf_session_snapshot(s, result, "2", "12.10")

    wb = load_workbook(generate_dsf_excel_from_snapshot(snapshot, Path(__file__).resolve().parents[1]))

    assert wb.sheetnames == ["DSF2 - 1", "DSF2 - 2"]
    assert wb["DSF2 - 1"]["R10"].value.startswith("NOM15")
    assert wb["DSF2 - 2"]["C10"].value.startswith("NOM16")
    assert wb["DSF2 - 2"]["D10"].value is None


def test_dsf_france_travail_route_refuses_non_afc(monkeypatch):
    import app as application

    application.app.config.update(TESTING=True, SECRET_KEY="test")
    saved = {"sessions": [{"id": "aps", "formation": "APS", "afcDsfs": []}], "jurys": []}
    monkeypatch.setattr(application, "load_sessions", lambda: saved)
    with application.app.test_client() as client:
        with client.session_transaction() as flask_session:
            flask_session["admin_logged"] = True
            flask_session["admin_session_version"] = application.ADMIN_SESSION_VERSION
        response = client.post("/api/sessions/aps/afc-dsf/excel-preview", json={})
    assert response.status_code == 404

def test_afc_invoice_excel_snapshot_and_template_intact():
    import hashlib
    from openpyxl import load_workbook
    from services.afc_france_travail_invoice_excel import build_invoice_snapshot, generate_invoice_excel_from_snapshot, amount_to_french_words
    s = sample_session(); s["france_travail"] = {"convention":"41C32B061177", "engagement_kairos":"41C32B061177", "marche_afc":"51190", "intitule":"Intitulé AFC"}
    d=s["apsPlanningData"][0]["date"]
    r=afc_dsf_compute(s,d,d,["RAN"], hourly_rate="12.10")
    snap=afc_dsf_session_snapshot(s, r, "1", "12.10")
    dsf={"id":"dsf1","number":"1","status":AFC_DSF_STATUS_FINALIZED,"amountTotal":r["amountTotal"],"modules":["RAN"],"franceTravailExcelSnapshot":snap}
    template = Path(__file__).resolve().parents[1] / "static/upload/facture.xlsx"
    before = hashlib.sha256(template.read_bytes()).hexdigest()
    inv=build_invoice_snapshot(s, dsf, {"invoice_number":"00042","invoice_date":"2026-07-17","invoice_place":"PUGET SUR ARGENS","invoice_type":"intermediate","kairos_engagement_reference":"41C32B061177_N1"}, "tester")
    wb=load_workbook(generate_invoice_excel_from_snapshot(inv, Path(__file__).resolve().parents[1]), data_only=False)
    assert hashlib.sha256(template.read_bytes()).hexdigest() == before
    assert wb.sheetnames == ["FACTURE DSF1"]
    ws=wb["FACTURE DSF1"]
    assert ws["H2"].value == "41C32B061177_N1" and ws["H4"].value == "00042"
    assert ws["D22"].value == "16/11/2026" and ws["G22"].value == "16/11/2026"
    assert ws["A33"].value == "RAN" and ws["A34"].value is None
    assert ws["G33"].value == ws["C33"].value and ws["I33"].value == float(inv["amount_total"])
    assert ws["I38"].value == float(inv["amount_total"]) and ws["C42"].number_format.endswith('[$€-fr-FR]')
    assert amount_to_french_words("1.50") == "UN EURO ET CINQUANTE CENTIMES"
    assert not any(isinstance(c.value, str) and '#REF!' in c.value for row in ws.iter_rows() for c in row)


def test_afc_invoice_routes_unique_and_historical_download(monkeypatch):
    import app as application
    application.app.config.update(TESTING=True, SECRET_KEY="test")
    s=sample_session(); s["france_travail"]={"engagement_kairos":"41C32B061177","marche_afc":"51190"}
    d=s["apsPlanningData"][0]["date"]; r=afc_dsf_compute(s,d,d,["RAN"], hourly_rate="12.10")
    dsf={"id":"dsf1","number":"1","label":"DSF 1","status":AFC_DSF_STATUS_FINALIZED,"amountTotal":r["amountTotal"],"modules":["RAN"],"franceTravailExcelSnapshot":afc_dsf_session_snapshot(s,r,"1","12.10"), **r}
    s["afcDsfs"].append(dsf); saved={"sessions":[s],"jurys":[]}
    monkeypatch.setattr(application,"load_sessions",lambda: saved)
    monkeypatch.setattr(application,"save_sessions",lambda data: saved.update(data))
    with application.app.test_client() as client:
        with client.session_transaction() as flask_session:
            flask_session["admin_logged"] = True; flask_session["admin_session_version"] = application.ADMIN_SESSION_VERSION
        assert client.get('/api/sessions/s1/afc-dsf/dsf1/invoice/preview').json['preview']['invoiceType'] == ''
        payload={"invoice_number":"INV001","invoice_date":"2026-07-17","invoice_place":"PUGET SUR ARGENS","invoice_type":"intermediate","kairos_engagement_reference":"41C32B061177_N1"}
        resp=client.post('/api/sessions/s1/afc-dsf/dsf1/invoice', json=payload)
        assert resp.status_code == 200
        assert saved['sessions'][0]['afcDsfs'][0]['invoice']['invoice_number'] == 'INV001'
        assert client.post('/api/sessions/s1/afc-dsf/dsf1/invoice', json=payload).status_code == 400
        assert client.get('/sessions/s1/afc-dsf/dsf1/invoice.xlsx').status_code == 200


def test_afc_invoice_button_visibility(monkeypatch):
    s=sample_session(); d=s["apsPlanningData"][0]["date"]; r=afc_dsf_compute(s,d,d,["RAN"])
    s["afcDsfs"].append({"id":"dsf-actions","number":1,"label":"DSF 1","status":AFC_DSF_STATUS_FINALIZED,"franceTravailExcelSnapshot":afc_dsf_session_snapshot(s,r,"1","12.10"),**r})
    html=render_afc_dsf_page(monkeypatch, s)
    assert "Générer la facture" in html
    s_non=sample_session(); s_non["formation"]="APS"; s_non["afcDsfs"]=[s["afcDsfs"][0]]
    assert "Générer la facture" not in render_afc_dsf_page(monkeypatch, s_non)


def test_afc_invoice_decimal_money_reference_and_type_cases():
    from openpyxl import load_workbook
    from services.afc_france_travail_invoice_excel import (
        MONEY_FORMAT,
        amount_to_french_words,
        build_full_kairos_reference,
        build_invoice_snapshot,
        build_kairos_dsf_reference,
        normalize_kairos_base_reference,
        generate_invoice_excel_from_snapshot,
        normalize_dsf_sequence_number,
    )
    s = sample_session()
    s["france_travail"] = {"engagement_kairos": "41C32B061177_N1", "marche_afc": "51190"}
    dsf_snapshot = {
        "number": "1",
        "periodStart": "2026-07-01",
        "periodEnd": "2026-07-31",
        "hourlyRate": "12.10",
        "session": {"france_travail": s["france_travail"], "name": "AFC", "date_debut": "2026-07-01", "date_fin": "2026-07-31"},
        "students": [],
        "moduleTotals": {"FT": 200, "RAN": 75},
        "studentCount": 1,
        "totalHours": 275,
    }
    dsf = {"id": "dsf1", "number": "1", "status": AFC_DSF_STATUS_FINALIZED, "amountTotal": "3327.50", "modules": ["FT", "RAN"], "franceTravailExcelSnapshot": dsf_snapshot}
    inv = build_invoice_snapshot(s, dsf, {"invoice_number": "INV-DEC", "invoice_date": "2026-07-17", "invoice_place": "PUGET", "invoice_type": "intermediate"}, "tester")
    wb = load_workbook(generate_invoice_excel_from_snapshot(inv, Path(__file__).resolve().parents[1]), data_only=False)
    ws = wb["FACTURE DSF1"]
    assert ws["H33"].value == 12.10 and isinstance(ws["H33"].value, float)
    assert ws["H33"].number_format == MONEY_FORMAT
    assert ws["H34"].value == 12.10 and ws["H34"].number_format == MONEY_FORMAT
    assert ws["I33"].value == 2420.00 and ws["I34"].value == 907.50
    assert ws["I38"].value == 3327.50 and ws["C42"].value == 3327.50
    assert ws["F42"].value == "TROIS MILLE TROIS CENT VINGT-SEPT EUROS ET CINQUANTE CENTIMES"
    assert ws["H2"].value == ws["D24"].value == "41C32B061177_N1"
    assert ws["A22"].value.startswith("☒") and ws["I22"].value.startswith("☐")
    inv_final = {**inv, "invoice_type": "final"}
    ws_final = load_workbook(generate_invoice_excel_from_snapshot(inv_final, Path(__file__).resolve().parents[1]), data_only=False)["FACTURE DSF1"]
    assert ws_final["A22"].value.startswith("☐") and ws_final["I22"].value.startswith("☒")
    assert amount_to_french_words("3327.50") == "TROIS MILLE TROIS CENT VINGT-SEPT EUROS ET CINQUANTE CENTIMES"
    assert build_kairos_dsf_reference("41C32B061177", 1) == "41C32B061177_N1"
    assert build_kairos_dsf_reference("41C32B061177", 2) == "41C32B061177_N2"
    assert [normalize_dsf_sequence_number(v) for v in (1, "1", "DSF1", "N1", "_N1", "41C32B061177_N1")] == [1, 1, 1, 1, 1, 1]
    assert normalize_kairos_base_reference("41C32B061177_N1") == "41C32B061177"
    assert build_kairos_dsf_reference("41C32B061177_N1", 2) == "41C32B061177_N2"
    assert build_full_kairos_reference("41C32B061177", "DSF1") == "41C32B061177_N1"
    assert all(cell.value not in ("N1_N1", "N1 N1") for row in ws.iter_rows() for cell in row)


def test_afc_invoice_type_required_backend_and_frontend(monkeypatch):
    import app as application
    application.app.config.update(TESTING=True, SECRET_KEY="test")
    s = sample_session(); s["france_travail"] = {"engagement_kairos": "41C32B061177", "marche_afc": "51190"}
    dsf_snapshot = {"number": "1", "periodStart": "2026-07-01", "periodEnd": "2026-07-31", "hourlyRate": "12.10", "session": {"france_travail": s["france_travail"]}, "moduleTotals": {"RAN": 75}, "students": [], "studentCount": 1, "totalHours": 75}
    s["afcDsfs"].append({"id": "dsf1", "number": "1", "label": "DSF 1", "status": AFC_DSF_STATUS_FINALIZED, "amountTotal": "907.50", "modules": ["RAN"], "franceTravailExcelSnapshot": dsf_snapshot})
    saved = {"sessions": [s], "jurys": []}
    monkeypatch.setattr(application, "load_sessions", lambda: saved)
    monkeypatch.setattr(application, "save_sessions", lambda data: saved.update(data))
    html = render_afc_dsf_page(monkeypatch, s)
    assert 'name="afcInvoiceType" value="intermediate"' in html
    assert 'name="afcInvoiceType" value="final"' in html
    assert "confirmAfcInvoice').disabled=true" in html
    with application.app.test_client() as client:
        with client.session_transaction() as flask_session:
            flask_session["admin_logged"] = True; flask_session["admin_session_version"] = application.ADMIN_SESSION_VERSION
        payload = {"invoice_number": "INV-NOTYPE", "invoice_date": "2026-07-17", "invoice_place": "PUGET"}
        resp = client.post('/api/sessions/s1/afc-dsf/dsf1/invoice', json=payload)
        assert resp.status_code == 400
        assert resp.json["error"] == "Veuillez sélectionner Facture intermédiaire ou Facture de solde."
