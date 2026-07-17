from __future__ import annotations

import copy
import io
from decimal import Decimal, InvalidOperation
import math
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TRAINEE_COLUMNS = list(range(3, 19))  # C:R
MODULE_ROWS = {
    "FT": (11, 12, 13, 14),
    "RAN": (15, 16, 17, 18),
    "SP": (19, 20, 21, 22),
    "PAF": (23, 24, 25, 26),
}
TOTAL_PLAN_ROW = 27
NON_BILLABLE_ROW = 28
BILLABLE_ROW = 29
DISTANCE_ROW = 30
FESTE_ROWS = (31, 32, 33)
HOUR_ROWS = list(range(11, 34))
HEADER_CELLS = ["D3", "K3", "D4", "K4", "L5", "O5", "E6", "E7", "I7", "B7"]


def load_dsf_template(app_root: str | Path):
    path = Path(app_root) / "static" / "upload" / "dsf.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"Modèle DSF France Travail introuvable : {path}")
    return load_workbook(path)


def _parse_date(value):
    if not value:
        return None
    if hasattr(value, "strftime"):
        return value
    return datetime.strptime(str(value)[:10], "%Y-%m-%d")


def fmt_date(value: Any) -> str:
    d = _parse_date(value)
    return d.strftime("%d/%m/%Y") if d else ""


def split_trainees_into_pages(trainees: list[dict[str, Any]], per_page: int = 16) -> list[list[dict[str, Any]]]:
    return [trainees[i:i + per_page] for i in range(0, len(trainees), per_page)] or [[]]


def clear_dsf_template_data(ws):
    for cell in HEADER_CELLS:
        ws[cell].value = None
    for col in TRAINEE_COLUMNS:
        ws.cell(10, col).value = None
    for row in HOUR_ROWS:
        for col in range(2, max(TRAINEE_COLUMNS) + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.number_format = "General"


def set_hour_value(cell, value):
    if value is None or value == "":
        cell.value = None
        cell.number_format = "General"
        return

    try:
        number = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(f"Valeur horaire invalide : {value!r}") from exc

    if number == number.to_integral_value():
        cell.value = int(number)
        cell.number_format = "0"
    else:
        cell.value = float(number)
        cell.number_format = "0.##"


def populate_dsf_header(ws, snapshot: dict[str, Any]):
    session = snapshot.get("session") or {}
    ws["D3"] = session.get("organisme") or "INTEGRALE SECURITE FORMATIONS"
    ws["K3"] = session.get("intitule") or session.get("name") or ""
    ws["D4"] = str(session.get("convention") or "")
    ws["D4"].number_format = "@"
    ws["K4"] = session.get("lieu") or session.get("organisme_adresse") or ""
    ws["L5"] = fmt_date(session.get("date_debut"))
    ws["O5"] = fmt_date(session.get("date_fin"))
    ws["E6"] = fmt_date(snapshot.get("periodStart"))
    ws["E7"] = fmt_date(snapshot.get("periodEnd"))
    ws["I7"] = str(snapshot.get("number") or "")
    ws["I7"].number_format = "@"


def populate_dsf_trainee_headers(ws, trainees: list[dict[str, Any]]):
    for idx, trainee in enumerate(trainees):
        col = TRAINEE_COLUMNS[idx]
        name = f"{trainee.get('lastName','')} {trainee.get('firstName','')}".strip()
        ft_id = str(trainee.get("france_travail_id") or "")
        ws.cell(10, col).value = f"{name}\n{ft_id}" if ft_id else name
        ws.cell(10, col).number_format = "@"


def _num(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def populate_dsf_module_hours(ws, trainees: list[dict[str, Any]]):
    row_totals = {row: 0.0 for row in HOUR_ROWS}
    for idx, trainee in enumerate(trainees):
        col = TRAINEE_COLUMNS[idx]
        modules = trainee.get("modules") or {}
        total_plan = 0.0
        non_billable = 0.0
        for module, rows in MODULE_ROWS.items():
            main_row, auth_abs_row, unjust_abs_row, total_row = rows
            hours = _num(modules.get(module))
            for row, value in ((main_row, hours), (auth_abs_row, 0), (unjust_abs_row, 0), (total_row, hours)):
                set_hour_value(ws.cell(row, col), value)
                row_totals[row] += value
            total_plan += hours
        distance = _num(trainee.get("distanceHours"))
        feste = _num(trainee.get("festeHours"))
        values = {
            TOTAL_PLAN_ROW: total_plan + feste,
            NON_BILLABLE_ROW: non_billable,
            BILLABLE_ROW: total_plan + feste - non_billable,
            DISTANCE_ROW: distance,
            FESTE_ROWS[0]: feste,
            FESTE_ROWS[1]: 0,
            FESTE_ROWS[2]: 0,
        }
        for row, value in values.items():
            set_hour_value(ws.cell(row, col), value)
            row_totals[row] += value
    return row_totals


def populate_dsf_totals(ws, row_totals: dict[int, float]):
    for row in HOUR_ROWS:
        set_hour_value(ws.cell(row, 2), round(row_totals.get(row, 0), 2))


def configure_dsf_print_settings(ws):
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if not ws.print_area:
        ws.print_area = "A1:R33"


def populate_page(ws, snapshot: dict[str, Any], trainees: list[dict[str, Any]]):
    clear_dsf_template_data(ws)
    populate_dsf_header(ws, snapshot)
    populate_dsf_trainee_headers(ws, trainees)
    row_totals = populate_dsf_module_hours(ws, trainees)
    populate_dsf_totals(ws, row_totals)
    configure_dsf_print_settings(ws)


def generate_dsf_excel_from_snapshot(snapshot: dict[str, Any], app_root: str | Path):
    wb = load_dsf_template(app_root)
    template = wb["DSF1"] if "DSF1" in wb.sheetnames else wb[wb.sheetnames[0]]
    trainees = [t for t in snapshot.get("students", []) if _num(t.get("totalHours")) > 0]
    pages = split_trainees_into_pages(trainees)
    sheets = []
    for i, page in enumerate(pages, start=1):
        ws = template if i == 1 else wb.copy_worksheet(template)
        if len(pages) == 1:
            ws.title = f"DSF{snapshot.get('number') or 1}"
        else:
            ws.title = f"DSF{snapshot.get('number') or 1} - {i}"
        populate_page(ws, snapshot, page)
        sheets.append(ws)
    for ws in list(wb.worksheets):
        if ws not in sheets:
            wb.remove(ws)
    wb.active = 0
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def generate_dsf_excel(snapshot: dict[str, Any], app_root: str | Path):
    return generate_dsf_excel_from_snapshot(snapshot, app_root)


def safe_filename_part(value: str) -> str:
    value = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-").lower()
    return value or "session"


def dsf_excel_filename(snapshot: dict[str, Any]) -> str:
    name = safe_filename_part((snapshot.get("session") or {}).get("name") or snapshot.get("sessionName") or "session")
    start = fmt_date(snapshot.get("periodStart")).replace("/", "-")
    end = fmt_date(snapshot.get("periodEnd")).replace("/", "-")
    return f"dsf_france_travail_{name}_dsf_{snapshot.get('number')}_{start}_{end}.xlsx"


def page_count_for_snapshot(snapshot: dict[str, Any]) -> int:
    return max(1, math.ceil(len([t for t in snapshot.get("students", []) if _num(t.get("totalHours")) > 0]) / 16))
