from __future__ import annotations

import copy, io, logging, re, time, unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

WEEK_RE = re.compile(r"^\d{4} au \d{4}$")
FT_KEYS = ("marche_afc", "brs", "convention", "bon_commande", "type_session", "intitule")
DAY_COLS = [("Lundi", 3, 4), ("Mardi", 5, 6), ("Mercredi", 7, 8), ("Jeudi", 9, 10), ("Vendredi", 11, 12)]
STUDENT_START_ROW = 14
BASE_STUDENT_PAIRS = 12
BASE_TOTAL_ROW = 39
BASE_TRAINER_ROW = 41
GREY_FILL = PatternFill("solid", fgColor="D9D9D9")
WHITE_FILL = PatternFill(fill_type=None)
AFC_CODE_MAP = {"SP": "S", "S": "S", "RAN": "RAN", "PAF": "PAF", "E": "E", "DIS": "DIS"}
FT_CATEGORIES = {"ACCUEIL", "APS", "EXAM_APS", "H0B0", "SSIAP1", "EXAM_SSIAP1", "BILAN", "FT"}

@dataclass
class SlotInfo:
    date: date; col: int; part: str; start: str; end: str; minutes: int; module: str; trainer: str; student_ids: tuple[str, ...] = ()


def is_afc_session(session: dict[str, Any]) -> bool:
    return str(session.get("training_code") or session.get("formation") or "").upper() == "AFC_APS_SSIAP"


def template_path(app_root: str | Path) -> Path:
    return Path(app_root) / "static" / "upload" / "tableau.xlsx"


def load_france_travail_template(app_root: str | Path):
    path = template_path(app_root)
    if not path.exists():
        logging.getLogger(__name__).error("Modèle France Travail introuvable: %s", path)
        raise FileNotFoundError(f"Modèle Excel introuvable : {path}")
    return load_workbook(path)


def get_afc_france_travail_settings(session: dict[str, Any]) -> dict[str, str]:
    ft = dict(session.get("france_travail") or {})
    default_title = session.get("display_name") or session.get("formation") or ""
    return {k: str(ft.get(k) or (default_title if k == "intitule" else "")) for k in FT_KEYS}


def update_afc_france_travail_settings(session: dict[str, Any], payload: dict[str, Any]) -> dict[str, str]:
    settings = {k: str(payload.get(k) or "").strip() for k in FT_KEYS}
    if not settings["intitule"]:
        settings["intitule"] = session.get("display_name") or session.get("formation") or ""
    session["france_travail"] = settings
    return settings


def attendance_students(session):
    out = []
    for i, s in enumerate(session.get("apsAttendanceStudents") or []):
        last = (s.get("lastName") or s.get("nom") or "").strip().upper(); first = (s.get("firstName") or s.get("prenom") or "").strip()
        if not (last and first): continue
        status = str(s.get("status") or s.get("statut") or "").lower()
        if any(x in status for x in ("supprim", "annul")): continue
        out.append({"index": i, "id": s.get("id") or f"student-{i+1}", "lastName": last, "firstName": first, "displayName": f"{last} {first}", "entryDate": s.get("entryDate") or s.get("dateEntree") or s.get("startDate") or session.get("date_debut") or "", "exitDate": s.get("exitDate") or s.get("dateSortie") or s.get("endDate") or s.get("abandonDate") or session.get("date_fin") or "", "france_travail_id": str(s.get("france_travail_id") or "")})
    return out


def save_france_travail_ids(session, ids: dict[str, Any]):
    students = session.setdefault("apsAttendanceStudents", [])
    for i, st in enumerate(students):
        key = str(st.get("id") or f"student-{i+1}")
        if key in ids or str(i) in ids:
            st["france_travail_id"] = str(ids.get(key, ids.get(str(i), "")))
    return attendance_students(session)


def parse_date(v):
    if isinstance(v, date): return v
    if not v: return None
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()

def build_session_weeks(session):
    dates = [parse_date(d.get("date")) for d in session.get("apsPlanningData") or [] if d.get("date")]
    start = parse_date(session.get("date_debut")) or (min(dates) if dates else None); end = parse_date(session.get("date_fin")) or (max(dates) if dates else None)
    if not start or not end or end < start: raise ValueError("Dates de session AFC invalides.")
    monday = start - timedelta(days=start.weekday()); weeks=[]
    while monday <= end:
        friday = monday + timedelta(days=4); weeks.append({"monday": monday, "friday": friday, "name": f"{monday:%d%m} au {friday:%d%m}"}); monday += timedelta(days=7)
    return weeks

def slot_module(slot):
    cat = slot.get("afcKind") or slot.get("afcCategory") or slot.get("category") or slot.get("uv") or slot.get("sequence")
    cat = str(cat or "").upper()
    if cat in FT_CATEGORIES: return "FT"
    return AFC_CODE_MAP.get(cat, cat if cat in {"RAN","PAF","E","DIS","S"} else "FT")

def minutes(slot): return int(slot.get("durationMinutes") or round(float(slot.get("duration") or 0) * 60))
def time_to_minutes(value):
    h, m = map(int, str(value or "00:00")[:5].split(":"))
    return h * 60 + m
def minutes_to_time(value): return f"{value//60:02d}:{value%60:02d}"
def half(start): return "am" if time_to_minutes(start) < 13 * 60 else "pm"
def fmt_slot(s,e): return f"{s.replace(':','h')}-{e.replace(':','h')}"
def slot_student_ids(slot):
    ids = slot.get("studentIds") or slot.get("students") or slot.get("traineeIds") or slot.get("stagiaireIds") or []
    if isinstance(ids, str): ids = [ids]
    return tuple(str(i) for i in ids if str(i))
def slot_applies_to_student(slot_info, student):
    return not slot_info.student_ids or str(student.get("id")) in slot_info.student_ids or str(student.get("index")) in slot_info.student_ids
def fmt_hours(h): return int(h) if abs(h-int(h))<0.001 else str(round(h,2)).replace('.', ',')

def build_week_schedule(session, week):
    day_to_col = {week["monday"]+timedelta(days=i): cols for i, cols in enumerate(DAY_COLS)}; slots=[]
    morning_end = 12 * 60 + 30; afternoon_start = 13 * 60 + 30
    for day in session.get("apsPlanningData") or []:
        d=parse_date(day.get("date"))
        if d not in day_to_col: continue
        for sl in day.get("slots") or []:
            start_min, end_min = time_to_minutes(sl.get("start")), time_to_minutes(sl.get("end"))
            if end_min <= start_min:
                m=minutes(sl)
                if m<=0: raise ValueError("Créneau AFC incohérent (durée négative ou nulle).")
                end_min = start_min + m
            pieces=[]
            if start_min < morning_end and end_min > start_min:
                pieces.append((start_min, min(end_min, morning_end), "am"))
            if end_min > afternoon_start:
                pieces.append((max(start_min, afternoon_start), end_min, "pm"))
            if not pieces and start_min >= morning_end and end_min <= afternoon_start:
                pieces.append((start_min, end_min, half(sl.get("start"))))
            for piece_start, piece_end, part in pieces:
                m = piece_end - piece_start
                if m<=0: continue
                col = day_to_col[d][1 if part=="am" else 2]
                slots.append(SlotInfo(d,col,part,minutes_to_time(piece_start),minutes_to_time(piece_end),m,slot_module(sl),sl.get("trainer") or sl.get("formateur") or "",slot_student_ids(sl)))
    return slots

def applicable(student, d):
    entry=parse_date(student.get("entryDate")); exit=parse_date(student.get("exitDate"))
    return (not entry or d>=entry) and (not exit or d<=exit)

def get_week_trainees(session, week, schedule=None):
    schedule = schedule if schedule is not None else build_week_schedule(session, week)
    result=[]
    for st in attendance_students(session):
        if any(applicable(st, sl.date) for sl in schedule): result.append(st)
    return result

def get_week_trainers(schedule):
    seen={};
    for sl in schedule:
        name=sl.trainer.strip()
        if name and name not in seen: seen[name]=sl
    return [(n, s) for n,s in seen.items()]

def preview(session):
    weeks=build_session_weeks(session); schedules=[build_week_schedule(session,w) for w in weeks]; students=attendance_students(session); settings=get_afc_france_travail_settings(session)
    trainers={sl.trainer.strip() for sch in schedules for sl in sch if sl.trainer.strip()}
    missing_ids=[s["displayName"] for s in students if not s.get("france_travail_id")]
    total=sum(sl.minutes/60 for sch in schedules for sl in sch)*len(students)
    return {"sessionName":session.get("display_name") or session.get("formation"),"dateStart":session.get("date_debut"),"dateEnd":session.get("date_fin"),"weekCount":len(weeks),"sheetNames":[w["name"] for w in weeks],"studentCount":len(students),"missingIdCount":len(missing_ids),"missingIds":missing_ids,"missingSettings":[k for k,v in settings.items() if not v],"trainerCount":len(trainers),"totalPlannedHours":fmt_hours(total),"settings":settings}

# Excel helpers
def copy_row_style(ws, src, dst):
    ws.row_dimensions[dst].height = ws.row_dimensions[src].height
    for c in range(1,14):
        a,b=ws.cell(src,c),ws.cell(dst,c); b._style=copy.copy(a._style); b.font=copy.copy(a.font); b.fill=copy.copy(a.fill); b.border=copy.copy(a.border); b.alignment=copy.copy(a.alignment); b.number_format=a.number_format

def clone_week_template(wb):
    src = wb["0604 au 1004"] if "0604 au 1004" in wb.sheetnames else next((wb[n] for n in wb.sheetnames if WEEK_RE.match(n)), None)
    if src is None: raise ValueError("Aucun onglet hebdomadaire utilisable dans le modèle.")
    return src

def prepare_sheet(ws, student_count, trainer_count):
    need=max(student_count,1); extra=max(0,need-BASE_STUDENT_PAIRS)
    if extra: ws.insert_rows(BASE_TOTAL_ROW, extra*2); [copy_row_style(ws, STUDENT_START_ROW, BASE_TOTAL_ROW+i) for i in range(extra*2)]
    elif need<BASE_STUDENT_PAIRS: ws.delete_rows(STUDENT_START_ROW+need*2, (BASE_STUDENT_PAIRS-need)*2)
    total_row=STUDENT_START_ROW+need*2+1; trainer_row=total_row+2
    if trainer_count>3: ws.insert_rows(trainer_row+3, trainer_count-3)
    return total_row, trainer_row

def populate_week_header(ws, session, week, settings, student_count):
    ws["A1"]="FEUILLE DE PRESENCE - Marchés AFC France Travail"; ws["A3"]=f"N° marché AFC : {settings['marche_afc']}"; ws["E3"]=f"N° BRS : {settings['brs']}"; ws["A4"]=f"N° convention : {settings['convention']}"; ws["E4"]=f"N° Bon de commande : {settings['bon_commande']}"; ws["A5"]=f"Intitulé de la formation : {settings['intitule']}"; ws["A6"]=f"Type de session  : {settings['type_session']}"; ws["A7"]="Durée hebdomadaire : 30 heures"; ws["E7"]="Dates de session : "; ws["E8"]=f"{parse_date(session.get('date_debut')).strftime('%d/%m/%Y')} au {parse_date(session.get('date_fin')).strftime('%d/%m/%Y')}"; ws["A8"]=f"Nombre de stagiaires inscrits : {student_count}"
    ws["H4"]="INTEGRALE SECURITE FORMATIONS"; ws["H5"]="54 chemin du Carreou"; ws["H6"]="83480 PUGET SUR ARGENS"; ws["H7"]="SIRET 840 899 884 00026"
    for i,(name,c1,c2) in enumerate(DAY_COLS): ws.cell(10,c1).value=f"{name} {(week['monday']+timedelta(days=i)):%d/%m}"

def populate_week_slots(ws, week, schedule):
    for _, c1, c2 in DAY_COLS:
        for c in (c1,c2): ws.cell(12,c).value=None; ws.cell(13,c).value=None; ws.cell(12,c).fill=copy.copy(GREY_FILL); ws.cell(13,c).fill=copy.copy(GREY_FILL)
    for col in range(3,13):
        ss=[s for s in schedule if s.col==col]
        if ss:
            ws.cell(12,col).value=fmt_slot(minutes_to_time(min(time_to_minutes(s.start) for s in ss)), minutes_to_time(max(time_to_minutes(s.end) for s in ss))); ws.cell(13,col).value=" / ".join(dict.fromkeys(s.module for s in ss)); ws.cell(12,col).fill=copy.copy(WHITE_FILL); ws.cell(13,col).fill=copy.copy(WHITE_FILL)

def populate_week_trainees(ws, trainees, schedule):
    totals=[0]*10
    for idx, st in enumerate(trainees,1):
        r=STUDENT_START_ROW+(idx-1)*2; ws.cell(r,1).value=f"{idx} {st['displayName']}"; ws.cell(r,2).value=st.get('france_travail_id') or ""; ws.cell(r,2).number_format='@'; ws.cell(r+1,1).value="Soutien personnalisé"; total=0
        for c in range(3,13):
            day_slots=[s for s in schedule if s.col==c and applicable(st,s.date) and slot_applies_to_student(s, st)]; classic=sum(s.minutes for s in day_slots if s.module!='S'); support=sum(s.minutes for s in day_slots if s.module=='S'); total+=(classic+support)/60; totals[c-3]+=(classic+support)/60
            ws.cell(r,c).value=None; ws.cell(r,c).fill=copy.copy(WHITE_FILL if (classic + support) else GREY_FILL)
            ws.cell(r+1,c).value=None; ws.cell(r+1,c).fill=copy.copy(WHITE_FILL if support else GREY_FILL)
            if support:
                ws.cell(r+1,c).value=support/60; ws.cell(r+1,c).number_format='0,## "h"'
        support_total=sum((s.minutes/60) for s in schedule if applicable(st,s.date) and slot_applies_to_student(s, st) and s.module=='S')
        ws.cell(r,13).value=fmt_hours(total) if total else None; ws.cell(r+1,13).value=support_total if support_total else 0; ws.cell(r+1,13).number_format='0,## "h"'
    return totals

def populate_week_totals(ws, total_row, totals):
    ws.cell(total_row,1).value="Total des heures facturables"
    for i,v in enumerate(totals,3): ws.cell(total_row,i).value=fmt_hours(v) if v else None
    ws.cell(total_row,13).value=fmt_hours(sum(totals)) if sum(totals) else None
    ws.cell(total_row+1,1).value="Total des heures non facturables"
    for c in range(3,14): ws.cell(total_row+1,c).value=None

def populate_week_trainers(ws, trainer_row, trainers, schedule):
    rows_to_clear = max(3, len(trainers))
    for i in range(rows_to_clear):
        r = trainer_row + i
        ws.cell(r, 1).value = None
        for c in range(2, 14):
            ws.cell(r, c).value = None
            if 3 <= c <= 12:
                ws.cell(r, c).fill = copy.copy(GREY_FILL)
    for i,(name,_) in enumerate(trainers):
        r=trainer_row+i; ws.cell(r,1).value=f"Formateur {name}"
        for c in range(3,13): ws.cell(r,c).value=None; ws.cell(r,c).fill=copy.copy(WHITE_FILL if any(s.col==c and s.trainer.strip()==name for s in schedule) else GREY_FILL)

def configure_print_settings(ws, last_row):
    ws.sheet_properties.pageSetUpPr.fitToPage=True; ws.page_setup.orientation='landscape'; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0; ws.print_area=f"A1:M{last_row}"; ws.freeze_panes="C14"

def generate_france_travail_workbook(session, app_root):
    start_time=time.time()
    if not is_afc_session(session): raise PermissionError("Génération réservée aux sessions AFC APS + SSIAP.")
    wb=load_france_travail_template(app_root); template=clone_week_template(wb); copied=[]; settings=get_afc_france_travail_settings(session); weeks=build_session_weeks(session); all_students=attendance_students(session)
    for week in weeks:
        ws=wb.copy_worksheet(template); ws.title=week['name']; copied.append(ws)
        schedule=build_week_schedule(session, week); trainees=get_week_trainees(session, week, schedule); trainers=get_week_trainers(schedule); total_row, trainer_row=prepare_sheet(ws, len(trainees), len(trainers)); populate_week_header(ws, session, week, settings, len(all_students)); populate_week_slots(ws, week, schedule); totals=populate_week_trainees(ws, trainees, schedule); populate_week_totals(ws,total_row,totals); populate_week_trainers(ws,trainer_row,trainers,schedule); configure_print_settings(ws, max(ws.max_row, trainer_row+len(trainers)+2))
    for ws in list(wb.worksheets):
        if ws not in copied: wb.remove(ws)
    wb.active=0
    bio=io.BytesIO(); wb.save(bio); bio.seek(0)
    logging.getLogger(__name__).info("France Travail AFC généré session=%s semaines=%s stagiaires=%s onglets=%s durée=%.2fs", session.get('id'), len(weeks), len(all_students), len(copied), time.time()-start_time)
    return bio

def safe_filename(session_name, today=None):
    base=unicodedata.normalize('NFKD', session_name or 'session').encode('ascii','ignore').decode('ascii').lower(); base=re.sub(r'[^a-z0-9]+','_',base).strip('_') or 'session'; d=(today or datetime.now()).strftime('%d-%m-%Y'); return f"feuilles_presence_france_travail_{base}_{d}.xlsx"
