import io
from datetime import date, timedelta

import pytest
from openpyxl import load_workbook

import prospecting
import app as application
from app import app


@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("PERSIST_DIR", str(tmp_path))
    app.config.update(TESTING=True, SECRET_KEY="test")
    with app.test_client() as client:
        with client.session_transaction() as session:
            session["admin_logged"] = True
            session["admin_session_version"] = application.ADMIN_SESSION_VERSION
        yield client


def iso_days_ago(days):
    return (date.today() - timedelta(days=days)).isoformat()


def make_candidate(name, siren, company_days, **extra):
    raw = {
        "nom": name,
        "siren": siren,
        "ville": "Paris",
        "code_postal": "75001",
        "ape": "85.59A",
        "date_creation_entreprise": iso_days_ago(company_days),
        "signal": extra.pop("signal", "Entreprise de formation"),
        **extra,
    }
    return prospecting._candidate(raw, "Test")


def test_recent_scoring_prioritizes_signal_and_ape_is_secondary():
    recent = make_candidate("Nouveau centre SSIAP", "123456789", 20)
    assert recent["est_recent"] == 1
    assert recent["archive"] == 0
    assert recent["type_signal_recent"] == "Création entreprise récente"
    assert recent["score"] >= 65

    old = make_candidate("Ancienne société 8559A", "223456789", 800)
    assert old["est_recent"] == 0
    assert old["archive"] == 1
    assert old["score"] == 0
    assert "APE 8559A (secondaire)" in old["raison_detection"]


def test_recent_web_signal_keeps_old_company_visible():
    prospect = make_candidate(
        "Société historique qui ouvre un centre",
        "323456789",
        800,
        signal="Ouverture d'un nouveau centre de formation SSIAP",
        type_signal_recent="Ouverture centre détectée",
        date_signal_recent=iso_days_ago(12),
    )
    assert prospect["est_recent"] == 1
    assert prospect["archive"] == 0
    assert prospect["anciennete_signal_jours"] == 12
    assert prospect["type_signal_recent"] == "Ouverture centre détectée"


def test_default_view_hides_archives_and_filters_show_them(client):
    with app.app_context():
        prospecting.init_prospect_db()
        prospecting._upsert(make_candidate("Prospect récent", "423456789", 15))
        prospecting._upsert(make_candidate("Prospect ancien", "523456789", 900))

    default_page = client.get("/admin").get_data(as_text=True)
    assert "Prospect récent" in default_page
    assert "Prospect ancien" not in default_page
    assert "Vue par défaut :" in default_page

    archive_page = client.get("/admin?signal_filter=archives").get_data(as_text=True)
    assert "Prospect ancien" in archive_page
    assert "Prospect récent" not in archive_page

    all_page = client.get("/admin?signal_filter=all").get_data(as_text=True)
    assert "Prospect récent" in all_page
    assert "Prospect ancien" in all_page


def test_upsert_deduplicates_siren_and_preserves_recent_signal(client):
    with app.app_context():
        prospecting.init_prospect_db()
        signaled = make_candidate(
            "Centre avec actualité", "623456789", 700,
            type_signal_recent="Recrutement formateur sécurité",
            date_signal_recent=iso_days_ago(5),
            signal="Recrutement formateur APS",
        )
        assert prospecting._upsert(signaled) is True
        stale_ape_scan = make_candidate("Centre renommé", "623456789", 700)
        assert prospecting._upsert(stale_ape_scan) is False
        with prospecting.get_prospect_db() as connection:
            rows = connection.execute("SELECT * FROM prospects WHERE siren='623456789'").fetchall()
        assert len(rows) == 1
        assert rows[0]["type_signal_recent"] == "Recrutement formateur sécurité"
        assert rows[0]["archive"] == 0


def test_admin_actions_and_excel_export(client):
    with app.app_context():
        prospecting.init_prospect_db()
        prospecting._upsert(make_candidate("Sécurité Formation Test", "723456789", 10, email="contact@example.fr"))

    assert client.get("/cron-prospects-scan").status_code == 401
    response = client.get("/admin/prospects/1/mail")
    assert response.status_code == 200
    assert "Intégrale Academy" in response.get_data(as_text=True)
    assert "Contacté" in client.post("/admin/prospects/1/contacted", follow_redirects=True).get_data(as_text=True)

    response = client.get("/admin/export.xlsx")
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.data), read_only=True)
    headers = [cell.value for cell in next(workbook.active.iter_rows())]
    assert headers == [
        "Score", "Nom", "SIREN", "SIRET", "Ville", "Département", "Date création entreprise",
        "Date création établissement", "Type signal récent", "Date signal récent", "Ancienneté signal jours",
        "Est récent", "Archive", "Source", "Raison détection", "Statut commercial", "Commentaire",
    ]


def test_existing_database_is_migrated_and_old_company_archived(tmp_path, monkeypatch):
    import sqlite3

    monkeypatch.setenv("PERSIST_DIR", str(tmp_path))
    database = sqlite3.connect(tmp_path / "prospects.db")
    database.execute("""CREATE TABLE prospects (
        id INTEGER PRIMARY KEY AUTOINCREMENT, fingerprint TEXT UNIQUE, score INTEGER, name TEXT,
        siren TEXT, siret TEXT, city TEXT, department TEXT, manager TEXT, email TEXT, phone TEXT, website TEXT,
        source TEXT, source_url TEXT, signal TEXT, commercial_status TEXT, comment TEXT, detected_at TEXT,
        company_created_at TEXT, ape_code TEXT, qualiopi INTEGER, nda TEXT, ai_analysis TEXT, updated_at TEXT
    )""")
    database.execute(
        """INSERT INTO prospects(fingerprint,score,name,siren,source,detected_at,company_created_at,ape_code,qualiopi,updated_at)
           VALUES('legacy',10,'Ancienne entreprise','823456789','RNE','2025-01-01','2020-01-01','85.59A',0,'2025-01-01')"""
    )
    database.commit()
    database.close()

    with app.app_context():
        prospecting.init_prospect_db()
        with prospecting.get_prospect_db() as connection:
            row = connection.execute("SELECT * FROM prospects WHERE fingerprint='legacy'").fetchone()
            columns = {column["name"] for column in connection.execute("PRAGMA table_info(prospects)")}

    assert set(prospecting.NEW_COLUMNS) <= columns
    assert row["date_creation_entreprise"] == "2020-01-01"
    assert row["archive"] == 1
    assert row["est_recent"] == 0
    assert row["score"] == 0


def test_run_scan_tolerates_invalid_limit_and_unexpected_source_error(client, monkeypatch):
    monkeypatch.setenv("PROSPECT_SCAN_LIMIT", "not-a-number")
    monkeypatch.delenv("SERPER_API_KEY", raising=False)

    def broken_source(_limit):
        raise RuntimeError("unexpected payload")

    monkeypatch.setattr(prospecting, "_rne_rows", broken_source)

    with app.app_context():
        result = prospecting.run_scan()
        with prospecting.get_prospect_db() as connection:
            scan = connection.execute("SELECT * FROM prospect_scans ORDER BY id DESC LIMIT 1").fetchone()

    assert result["found"] == 0
    assert result["errors"] == ["Annuaire des Entreprises: erreur inattendue (RuntimeError)"]
    assert scan["status"] == "partial"
    assert scan["finished_at"]


def test_admin_scan_completes_and_redirects_with_results(client, monkeypatch):
    monkeypatch.delenv("SERPER_API_KEY", raising=False)
    monkeypatch.setattr(
        prospecting,
        "_rne_rows",
        lambda limit: [make_candidate("Nouveau centre sécurité", "923456789", 10)],
    )

    response = client.post("/admin/scan", follow_redirects=True)
    page = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Scan terminé : 1 détecté(s), 1 ajouté(s), 0 actualisé(s)." in page
    assert "Nouveau centre sécurité" in page
    assert "Scan en cours" not in page


def test_admin_scan_reports_source_failure_without_staying_running(client, monkeypatch):
    monkeypatch.delenv("SERPER_API_KEY", raising=False)

    def unavailable(_limit):
        raise OSError("service unavailable")

    monkeypatch.setattr(prospecting, "_rne_rows", unavailable)

    response = client.post("/admin/scan", follow_redirects=True)
    page = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Scan terminé avec une source indisponible" in page
    assert "Scanner maintenant" in page
    with app.app_context():
        with prospecting.get_prospect_db() as connection:
            scan = connection.execute("SELECT * FROM prospect_scans ORDER BY id DESC LIMIT 1").fetchone()
    assert scan["status"] == "partial"
    assert scan["finished_at"]


def test_admin_closes_running_rows_left_by_old_async_version(client):
    with app.app_context():
        prospecting.init_prospect_db()
        with prospecting.get_prospect_db() as connection:
            connection.execute(
                "INSERT INTO prospect_scans(started_at,status,sources) VALUES (?,'running','data.gouv / DGEFP')",
                ("2026-06-10T06:00:00+00:00",),
            )

    response = client.get("/admin")
    page = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Le scan précédent a été interrompu" in page
    assert "Scanner maintenant" in page
    with app.app_context():
        with prospecting.get_prospect_db() as connection:
            scan = connection.execute("SELECT * FROM prospect_scans ORDER BY id DESC LIMIT 1").fetchone()
    assert scan["status"] == "failed"
    assert scan["finished_at"]


def test_rne_scanner_uses_training_organization_filters(monkeypatch):
    captured = {}

    def request_json(url, timeout):
        captured["url"] = url
        captured["timeout"] = timeout
        return {
            "results": [{
                "nom_complet": "Centre API",
                "siren": "123456789",
                "date_creation": date.today().isoformat(),
                "activite_principale": "85.59A",
                "siege": {
                    "siret": "12345678900010",
                    "libelle_commune": "PARIS",
                    "code_postal": "75001",
                    "date_creation": date.today().isoformat(),
                    "liste_id_organisme_formation": ["11750000075"],
                },
                "complements": {"est_qualiopi": True},
                "dirigeants": [],
            }]
        }

    monkeypatch.setattr(prospecting, "_request_json", request_json)

    rows = prospecting._rne_rows(25)

    assert len(rows) == 1
    assert rows[0]["nda"] == "11750000075"
    assert rows[0]["qualiopi"] is True
    assert "est_organisme_formation=true" in captured["url"]
    assert "etat_administratif=A" in captured["url"]
    assert "activite_principale=85.59A" in captured["url"]
    assert captured["timeout"] == 20


def test_expire_stale_scans_symbol_used_by_admin_exists(client):
    assert callable(prospecting._expire_stale_scans)

    with app.app_context():
        prospecting.init_prospect_db()
        with prospecting.get_prospect_db() as connection:
            connection.execute(
                "INSERT INTO prospect_scans(started_at,status) VALUES (?,'running')",
                ("2026-06-10T06:00:00+00:00",),
            )
            prospecting._expire_stale_scans(connection)
            scan = connection.execute(
                "SELECT * FROM prospect_scans ORDER BY id DESC LIMIT 1"
            ).fetchone()

    assert scan["status"] == "failed"
    assert scan["finished_at"]


def test_admin_route_has_no_reference_to_removed_cleanup_symbol(client):
    response = client.get("/admin")

    assert response.status_code == 200
    assert b"Internal Server Error" not in response.data
