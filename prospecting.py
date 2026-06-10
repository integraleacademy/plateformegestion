"""Prospection commerciale centrée sur les signaux récents de formation sécurité."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import os
import re
import sqlite3
import threading
import urllib.error
import urllib.parse
import urllib.request
import unicodedata
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from flask import Blueprint, Response, current_app, flash, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
prospecting_bp = Blueprint("prospecting", __name__)

SECURITY_TERMS = (
    "sécurité", "securite", "sûreté", "surete", "protection", "aps", "ssiap", "cnaps",
    "sécurité privée", "securite privee", "vidéoprotection", "videoprotection",
    "protection rapprochée", "protection rapprochee", "formation sécurité", "formation securite",
)
RECRUITMENT_TERMS = (
    "formateur aps", "formateur ssiap", "formateur sécurité", "formateur securite",
    "responsable pédagogique sécurité", "responsable pedagogique securite",
)
OPENING_TERMS = (
    "ouverture de centre", "nouveau centre de formation", "nouvel organisme de formation",
    "lancement de formation sécurité", "lancement de formation securite",
)
STATUSES = ("Nouveau", "À qualifier", "À contacter", "Contacté", "À relancer", "Converti", "Non pertinent")
SIGNAL_FILTERS = (
    ("recent_7", "Signal récent depuis 7 jours"),
    ("recent_30", "Signal récent depuis 30 jours"),
    ("recent_90", "Signal récent depuis 90 jours"),
    ("company_30", "Création entreprise depuis 30 jours"),
    ("company_90", "Création entreprise depuis 90 jours"),
    ("establishment_30", "Création établissement depuis 30 jours"),
    ("establishment_90", "Création établissement depuis 90 jours"),
    ("new_training_org", "Nouvel organisme de formation"),
    ("recent_qualiopi", "Qualiopi récent"),
    ("recent_recruitment", "Recrutement récent"),
    ("recent_security_page", "Page sécurité récente"),
    ("all", "Tous les prospects"),
    ("archives", "Archives / anciens prospects"),
)
DATA_GOUV_DATASET = "liste-publique-des-organismes-de-formation-l-6351-7-1-du-code-du-travail"
DATA_GOUV_API = f"https://www.data.gouv.fr/api/1/datasets/{DATA_GOUV_DATASET}/"
RNE_SEARCH_API = "https://recherche-entreprises.api.gouv.fr/search"

FIELD_ALIASES = {
    "name": ("nom", "denomination", "raison_sociale", "raison sociale", "nom_organisme", "name"),
    "siren": ("siren", "numero_siren"),
    "siret": ("siret", "numero_siret", "siret_etablissement_declarant"),
    "city": ("ville", "commune", "libelle_commune", "adresse_ville"),
    "postal_code": ("code_postal", "code postal", "cp", "adresse_code_postal"),
    "email": ("email", "courriel", "adresse_electronique"),
    "phone": ("telephone", "téléphone", "tel"),
    "website": ("site_internet", "site internet", "url", "website"),
    "manager": ("dirigeant", "representant_legal", "nom_representant"),
    "ape": ("code_ape", "ape", "activite_principale", "naf"),
    "company_created": ("date_creation_entreprise", "date_creation", "date de creation"),
    "establishment_created": ("date_creation_etablissement", "date_debut_activite", "date_creation_siret"),
    "training_signal_date": ("date_organisme_formation", "date_declaration_activite", "date_enregistrement_nda", "date_derniere_declaration"),
    "qualiopi_date": ("date_qualiopi", "date_certification_qualite", "date_certification"),
    "nda": ("numero_de_declaration_d_activite", "numero declaration activite", "nda"),
    "qualiopi": ("qualiopi", "certifications", "certification_qualite", "certifie_qualite"),
    "signal_date": ("date_signal", "date_publication", "published_at", "date"),
}

NEW_COLUMNS = {
    "date_creation_entreprise": "TEXT DEFAULT ''",
    "date_creation_etablissement": "TEXT DEFAULT ''",
    "date_detection": "TEXT DEFAULT ''",
    "date_signal_recent": "TEXT DEFAULT ''",
    "type_signal_recent": "TEXT DEFAULT 'Aucun signal récent'",
    "anciennete_signal_jours": "INTEGER",
    "est_recent": "INTEGER NOT NULL DEFAULT 0",
    "raison_detection": "TEXT DEFAULT ''",
    "archive": "INTEGER NOT NULL DEFAULT 0",
}


def _db_path() -> Path:
    root = os.environ.get("PERSIST_DIR") or os.environ.get("DATA_DIR") or current_app.root_path
    path = Path(root)
    path.mkdir(parents=True, exist_ok=True)
    return path / "prospects.db"


def get_prospect_db() -> sqlite3.Connection:
    connection = sqlite3.connect(_db_path())
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA journal_mode=WAL")
    connection.execute("PRAGMA busy_timeout=5000")
    return connection


def init_prospect_db() -> None:
    """Crée la base et migre sans perte les bases issues de la première version."""
    with get_prospect_db() as connection:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS prospects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fingerprint TEXT NOT NULL UNIQUE,
                score INTEGER NOT NULL DEFAULT 0,
                name TEXT NOT NULL,
                siren TEXT DEFAULT '', siret TEXT DEFAULT '', city TEXT DEFAULT '', department TEXT DEFAULT '',
                manager TEXT DEFAULT '', email TEXT DEFAULT '', phone TEXT DEFAULT '', website TEXT DEFAULT '',
                source TEXT NOT NULL, source_url TEXT DEFAULT '', signal TEXT DEFAULT '',
                commercial_status TEXT NOT NULL DEFAULT 'Nouveau', comment TEXT DEFAULT '',
                detected_at TEXT NOT NULL, company_created_at TEXT DEFAULT '', ape_code TEXT DEFAULT '',
                qualiopi INTEGER NOT NULL DEFAULT 0, nda TEXT DEFAULT '', ai_analysis TEXT DEFAULT '', updated_at TEXT NOT NULL,
                date_creation_entreprise TEXT DEFAULT '', date_creation_etablissement TEXT DEFAULT '',
                date_detection TEXT DEFAULT '', date_signal_recent TEXT DEFAULT '',
                type_signal_recent TEXT DEFAULT 'Aucun signal récent', anciennete_signal_jours INTEGER,
                est_recent INTEGER NOT NULL DEFAULT 0, raison_detection TEXT DEFAULT '', archive INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS prospect_scans (
                id INTEGER PRIMARY KEY AUTOINCREMENT, started_at TEXT NOT NULL, finished_at TEXT,
                status TEXT NOT NULL, sources TEXT DEFAULT '', found_count INTEGER NOT NULL DEFAULT 0,
                added_count INTEGER NOT NULL DEFAULT 0, updated_count INTEGER NOT NULL DEFAULT 0,
                error_message TEXT DEFAULT ''
            );
            """
        )
        existing = {row["name"] for row in connection.execute("PRAGMA table_info(prospects)")}
        for column, definition in NEW_COLUMNS.items():
            if column not in existing:
                connection.execute(f"ALTER TABLE prospects ADD COLUMN {column} {definition}")
        connection.executescript(
            """
            CREATE INDEX IF NOT EXISTS idx_prospects_score ON prospects(score DESC);
            CREATE INDEX IF NOT EXISTS idx_prospects_status ON prospects(commercial_status);
            CREATE INDEX IF NOT EXISTS idx_prospects_recent ON prospects(archive, est_recent, date_signal_recent DESC);
            CREATE INDEX IF NOT EXISTS idx_prospects_siren ON prospects(siren);
            CREATE INDEX IF NOT EXISTS idx_prospects_siret ON prospects(siret);
            UPDATE prospects SET date_creation_entreprise = company_created_at
             WHERE COALESCE(date_creation_entreprise, '') = '' AND COALESCE(company_created_at, '') != '';
            UPDATE prospects SET date_detection = detected_at
             WHERE COALESCE(date_detection, '') = '' AND COALESCE(detected_at, '') != '';
            """
        )
        rows = connection.execute("SELECT * FROM prospects").fetchall()
        for row in rows:
            refreshed = qualify_prospect(dict(row))
            connection.execute(
                """UPDATE prospects SET score=?, date_signal_recent=?, type_signal_recent=?,
                   anciennete_signal_jours=?, est_recent=?, raison_detection=?, archive=? WHERE id=?""",
                (refreshed["score"], refreshed["date_signal_recent"], refreshed["type_signal_recent"],
                 refreshed["anciennete_signal_jours"], refreshed["est_recent"],
                 refreshed["raison_detection"], refreshed["archive"], row["id"]),
            )


def _now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _clean(value) -> str:
    return "" if value is None else " ".join(str(value).strip().split())


def _normalized_key(value: str) -> str:
    cleaned = unicodedata.normalize("NFKD", _clean(value).lower().replace("’", "'"))
    ascii_value = "".join(character for character in cleaned if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", "_", ascii_value).strip("_")


def _row_value(row: dict, field: str) -> str:
    normalized = {_normalized_key(key): value for key, value in row.items()}
    for alias in FIELD_ALIASES[field]:
        value = normalized.get(_normalized_key(alias))
        if value not in (None, ""):
            return _clean(value)
    return ""


def _department(postal_code: str) -> str:
    digits = re.sub(r"\D", "", postal_code)
    if len(digits) < 2:
        return ""
    return digits[:3] if digits.startswith(("97", "98")) and len(digits) >= 3 else digits[:2]


def _parse_date(value: str) -> date | None:
    value = _clean(value)
    if not value:
        return None
    cleaned = value[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    match = re.search(r"(\d+)\s+(?:day|jour)s?", value.lower())
    return date.today() - timedelta(days=int(match.group(1))) if match else None


def _iso_date(value: str) -> str:
    parsed = _parse_date(value)
    return parsed.isoformat() if parsed else ""


def _age_days(value: str) -> int | None:
    parsed = _parse_date(value)
    return max((date.today() - parsed).days, 0) if parsed else None


def _is_truthy(value: str) -> bool:
    lowered = _clean(value).lower()
    return lowered in {"1", "true", "oui", "yes", "certifié", "certifie", "qualiopi"} or "qualiopi" in lowered


def qualify_prospect(prospect: dict) -> dict:
    """Détermine le signal récent, l'archivage et le score commercial."""
    company_age = _age_days(prospect.get("date_creation_entreprise", ""))
    establishment_age = _age_days(prospect.get("date_creation_etablissement", ""))
    text = " ".join(_clean(prospect.get(field)) for field in ("name", "signal", "raison_detection")).lower()
    explicit_signal_date = _iso_date(prospect.get("date_signal_recent", ""))
    explicit_signal_age = _age_days(explicit_signal_date)
    signal_kind = _clean(prospect.get("type_signal_recent"))

    candidates: list[tuple[int, str, str]] = []
    if company_age is not None and company_age <= 90:
        candidates.append((company_age, "Création entreprise récente", prospect["date_creation_entreprise"]))
    if establishment_age is not None and establishment_age <= 90:
        candidates.append((establishment_age, "Création établissement récente", prospect["date_creation_etablissement"]))
    if explicit_signal_age is not None and explicit_signal_age <= 90 and signal_kind and signal_kind != "Aucun signal récent":
        candidates.append((explicit_signal_age, signal_kind, explicit_signal_date))

    if candidates:
        signal_age, signal_kind, signal_date = min(candidates, key=lambda item: item[0])
        is_recent = True
    else:
        signal_age, signal_kind, signal_date = None, "Aucun signal récent", ""
        is_recent = False

    score = 0
    reasons = []
    if company_age is not None and company_age <= 30:
        score += 35; reasons.append("entreprise créée depuis moins de 30 jours")
    elif company_age is not None and company_age <= 90:
        score += 25; reasons.append("entreprise créée depuis moins de 90 jours")
    if establishment_age is not None and establishment_age <= 30:
        score += 25; reasons.append("établissement créé depuis moins de 30 jours")
    elif establishment_age is not None and establishment_age <= 90:
        score += 20; reasons.append("établissement créé depuis moins de 90 jours")
    if signal_kind == "Nouvel organisme de formation":
        score += 30; reasons.append("nouvel organisme de formation")
    if prospect.get("qualiopi"):
        score += 25; reasons.append("Qualiopi")
    if signal_kind == "Qualiopi récent":
        score += 35; reasons.append("Qualiopi récent")
    if signal_kind == "Recrutement formateur sécurité":
        score += 30; reasons.append("recrutement sécurité récent")
    if signal_kind in {"Nouvelle page formation sécurité", "Ouverture centre détectée"}:
        score += 25; reasons.append("actualité web sécurité récente")
    if any(term in text for term in ("aps", "ssiap", "sécurité privée", "securite privee")):
        score += 20; reasons.append("activité APS / SSIAP / sécurité privée")
    if "cnaps" in text:
        score += 15; reasons.append("mention CNAPS")
    if _clean(prospect.get("ape_code")).replace(".", "").upper() == "8559A":
        score += 10; reasons.append("APE 8559A (secondaire)")

    company_old = company_age is not None and company_age > 365
    establishment_old = establishment_age is None or establishment_age > 365
    archive = company_old and establishment_old and not is_recent
    if archive:
        score -= 50; reasons.append("entreprise ancienne sans signal récent")
    elif not is_recent:
        score -= 25; reasons.append("aucun signal récent vérifié")

    prospect.update({
        "score": max(0, min(score, 100)),
        "date_signal_recent": signal_date,
        "type_signal_recent": signal_kind,
        "anciennete_signal_jours": signal_age,
        "est_recent": int(is_recent),
        "archive": int(archive),
        "raison_detection": " · ".join(reasons) or "Aucun critère commercial récent vérifié",
    })
    return prospect


def score_prospect(prospect: dict) -> tuple[int, str]:
    """Compatibilité publique : retourne le nouveau score et son explication."""
    mapped = dict(prospect)
    mapped.setdefault("date_creation_entreprise", mapped.get("company_created_at", ""))
    mapped.setdefault("date_creation_etablissement", mapped.get("establishment_created_at", ""))
    qualified = qualify_prospect(mapped)
    return qualified["score"], qualified["raison_detection"]


def _candidate(raw: dict, source: str, source_url: str = "") -> dict:
    postal_code = _row_value(raw, "postal_code")
    prospect = {
        "name": _row_value(raw, "name") or "Organisme sans dénomination",
        "siren": re.sub(r"\D", "", _row_value(raw, "siren")),
        "siret": re.sub(r"\D", "", _row_value(raw, "siret")),
        "city": _row_value(raw, "city"), "department": _department(postal_code),
        "manager": _row_value(raw, "manager"), "email": _row_value(raw, "email"),
        "phone": _row_value(raw, "phone"), "website": _row_value(raw, "website"),
        "source": source, "source_url": source_url,
        "signal": _clean(raw.get("signal")) or "Présence dans une source de formation",
        "date_creation_entreprise": _iso_date(_row_value(raw, "company_created")),
        "date_creation_etablissement": _iso_date(_row_value(raw, "establishment_created")),
        "ape_code": _row_value(raw, "ape"), "qualiopi": _is_truthy(_row_value(raw, "qualiopi")),
        "nda": _row_value(raw, "nda"), "date_signal_recent": "", "type_signal_recent": "Aucun signal récent",
        "raison_detection": _clean(raw.get("raison_detection")),
    }
    supplied_type = _clean(raw.get("type_signal_recent"))
    supplied_date = _iso_date(_clean(raw.get("date_signal_recent")) or _row_value(raw, "signal_date"))
    training_date = _iso_date(_row_value(raw, "training_signal_date"))
    qualiopi_date = _iso_date(_row_value(raw, "qualiopi_date"))
    if supplied_type and supplied_date:
        prospect.update(type_signal_recent=supplied_type, date_signal_recent=supplied_date)
    elif qualiopi_date and _age_days(qualiopi_date) <= 90:
        prospect.update(type_signal_recent="Qualiopi récent", date_signal_recent=qualiopi_date)
    elif training_date and _age_days(training_date) <= 90:
        prospect.update(type_signal_recent="Nouvel organisme de formation", date_signal_recent=training_date)
    qualify_prospect(prospect)
    identity = prospect["siret"] or prospect["siren"] or f'{prospect["name"]}|{prospect["city"]}'
    prospect["fingerprint"] = hashlib.sha256(identity.lower().encode()).hexdigest()
    return prospect


def _request_json(url: str, *, timeout: int = 25) -> dict:
    req = urllib.request.Request(url, headers={"Accept": "application/json", "User-Agent": "IntegraleAcademyProspector/2.0"})
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return json.loads(response.read().decode(response.headers.get_content_charset() or "utf-8"))


def _download(url: str, max_bytes: int = 60 * 1024 * 1024) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "IntegraleAcademyProspector/2.0"})
    with urllib.request.urlopen(req, timeout=45) as response:
        content = response.read(max_bytes + 1)
    if len(content) > max_bytes:
        raise ValueError("La ressource dépasse 60 Mo.")
    return content


def _data_gouv_rows(limit: int) -> list[dict]:
    metadata = _request_json(DATA_GOUV_API)
    resources = [r for r in metadata.get("resources", []) if (r.get("format") or "").lower() in {"csv", "xlsx", "xls"}]
    if not resources:
        raise ValueError("Aucune ressource CSV/XLSX data.gouv trouvée.")
    preferred = [r for r in resources if (r.get("format") or "").lower() == "csv"] or resources
    resource = max(preferred, key=lambda r: r.get("last_modified") or r.get("created_at") or "")
    content = _download(resource.get("latest") or resource.get("url"))
    raw_rows = []
    if (resource.get("format") or "").lower() == "csv":
        text = content.decode("utf-8-sig", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=";,\t|")
        except csv.Error:
            dialect = csv.excel; dialect.delimiter = ";"
        raw_rows = csv.DictReader(io.StringIO(text), dialect=dialect)
    else:
        sheet = load_workbook(io.BytesIO(content), read_only=True, data_only=True).active
        iterator = sheet.iter_rows(values_only=True)
        headers = [_clean(value) for value in next(iterator)]
        raw_rows = (dict(zip(headers, values)) for values in iterator)
    prospects = []
    for row in raw_rows:
        candidate = _candidate(row, "data.gouv / DGEFP", resource.get("url", ""))
        if candidate["est_recent"] or candidate["archive"]:
            prospects.append(candidate)
        if len(prospects) >= limit:
            break
    return prospects


def _rne_rows(limit: int) -> list[dict]:
    params = urllib.parse.urlencode({"activite_principale": "85.59A", "per_page": min(limit, 25), "page": 1})
    payload = _request_json(f"{RNE_SEARCH_API}?{params}")
    prospects = []
    for company in payload.get("results", [])[:limit]:
        headquarters, managers = company.get("siege") or {}, company.get("dirigeants") or []
        raw = {
            "denomination": company.get("nom_complet") or company.get("nom_raison_sociale"),
            "siren": company.get("siren"), "siret": headquarters.get("siret"),
            "ville": headquarters.get("libelle_commune"), "code_postal": headquarters.get("code_postal"),
            "activite_principale": company.get("activite_principale"),
            "date_creation_entreprise": company.get("date_creation"),
            "date_creation_etablissement": headquarters.get("date_creation"),
            "dirigeant": " ".join(_clean(managers[0].get(k)) for k in ("prenom", "nom")) if managers else "",
            "signal": "Entreprise APE 8559A repérée dans l’Annuaire des Entreprises / RNE",
        }
        prospects.append(_candidate(raw, "RNE / Annuaire des Entreprises", "https://annuaire-entreprises.data.gouv.fr"))
    return prospects


def _web_rows(limit: int) -> list[dict]:
    api_key = os.environ.get("SERPER_API_KEY")
    if not api_key:
        return []
    searches = (
        ('("nouveau centre de formation" OR "ouverture de centre" OR "lancement formation sécurité") France', "Ouverture centre détectée"),
        ('("formateur APS" OR "formateur SSIAP" OR "responsable pédagogique sécurité") recrutement France', "Recrutement formateur sécurité"),
        ('("formation sécurité" OR APS OR SSIAP OR CNAPS) France', "Nouvelle page formation sécurité"),
    )
    prospects = []
    for query, signal_type in searches:
        body = json.dumps({"q": query, "num": min(limit, 10), "tbs": "qdr:m3"}).encode()
        req = urllib.request.Request("https://google.serper.dev/search", data=body, method="POST", headers={"Content-Type": "application/json", "X-API-KEY": api_key})
        with urllib.request.urlopen(req, timeout=25) as response:
            payload = json.loads(response.read().decode())
        for item in payload.get("organic", []):
            text = f'{item.get("title", "")} {item.get("snippet", "")}'.lower()
            if signal_type == "Recrutement formateur sécurité" and not any(term in text for term in RECRUITMENT_TERMS):
                continue
            if signal_type == "Ouverture centre détectée" and not any(term in text for term in OPENING_TERMS):
                continue
            if signal_type == "Nouvelle page formation sécurité" and not any(term in text for term in SECURITY_TERMS):
                continue
            item_date = _iso_date(item.get("date", "")) or date.today().isoformat()
            prospects.append(_candidate({
                "nom": item.get("title"), "site_internet": item.get("link"),
                "signal": item.get("snippet"), "type_signal_recent": signal_type,
                "date_signal_recent": item_date,
            }, "Recherche web récente", item.get("link", "")))
            if len(prospects) >= limit:
                return prospects
    return prospects


def _find_existing(connection: sqlite3.Connection, prospect: dict):
    if prospect["siret"]:
        row = connection.execute("SELECT * FROM prospects WHERE siret=?", (prospect["siret"],)).fetchone()
        if row:
            return row
    if prospect["siren"]:
        row = connection.execute("SELECT * FROM prospects WHERE siren=?", (prospect["siren"],)).fetchone()
        if row:
            return row
    return connection.execute("SELECT * FROM prospects WHERE fingerprint=?", (prospect["fingerprint"],)).fetchone()


def _upsert(prospect: dict) -> bool:
    now = _now()
    columns = (
        "fingerprint", "score", "name", "siren", "siret", "city", "department", "manager", "email", "phone", "website",
        "source", "source_url", "signal", "date_creation_entreprise", "date_creation_etablissement", "date_detection",
        "date_signal_recent", "type_signal_recent", "anciennete_signal_jours", "est_recent", "raison_detection", "archive",
        "ape_code", "qualiopi", "nda", "updated_at",
    )
    with get_prospect_db() as connection:
        existing = _find_existing(connection, prospect)
        if existing:
            # Un scan APE sans actualité ne doit jamais effacer un signal récent déjà connu.
            existing_signal_age = _age_days(existing["date_signal_recent"])
            if not prospect.get("est_recent") and existing["est_recent"] and existing_signal_age is not None and existing_signal_age <= 90:
                for field in ("date_signal_recent", "type_signal_recent", "signal", "source", "source_url"):
                    prospect[field] = existing[field]
            for field in ("date_creation_entreprise", "date_creation_etablissement", "siren", "siret", "manager", "email", "phone", "website"):
                if not prospect.get(field) and existing[field]:
                    prospect[field] = existing[field]
            prospect["qualiopi"] = int(bool(prospect.get("qualiopi") or existing["qualiopi"]))
        qualify_prospect(prospect)
        values = [prospect.get(column, "") for column in columns]
        values[columns.index("date_detection")] = existing["date_detection"] if existing and existing["date_detection"] else now
        values[columns.index("updated_at")] = now
        if existing:
            assignments = [f"{column}=?" for column in columns if column != "fingerprint"]
            update_values = [value for column, value in zip(columns, values) if column != "fingerprint"] + [existing["id"]]
            connection.execute(f"UPDATE prospects SET {', '.join(assignments)} WHERE id=?", update_values)
            return False
        legacy_values = {
            "detected_at": now, "company_created_at": prospect.get("date_creation_entreprise", ""),
            "commercial_status": "Nouveau", "comment": "", "ai_analysis": "",
        }
        all_columns = list(columns) + list(legacy_values)
        all_values = values + list(legacy_values.values())
        connection.execute(
            f"INSERT INTO prospects ({', '.join(all_columns)}) VALUES ({', '.join('?' for _ in all_columns)})",
            all_values,
        )
        return True


def _scan_limit() -> int:
    raw_limit = os.environ.get("PROSPECT_SCAN_LIMIT", "250")
    try:
        return max(5, min(int(raw_limit), 2000))
    except (TypeError, ValueError):
        logger.warning("PROSPECT_SCAN_LIMIT invalide (%r), utilisation de 250.", raw_limit)
        return 250


def _create_scan_run() -> int | None:
    """Réserve un scan de façon atomique, même avec plusieurs workers Gunicorn."""
    init_prospect_db()
    now = _now()
    stale_before = (datetime.now(timezone.utc) - timedelta(minutes=30)).replace(microsecond=0).isoformat()
    with get_prospect_db() as connection:
        connection.execute("BEGIN IMMEDIATE")
        connection.execute(
            """UPDATE prospect_scans
               SET finished_at=?, status='failed', error_message='Scan interrompu avant sa finalisation'
               WHERE status='running' AND started_at<?""",
            (now, stale_before),
        )
        running = connection.execute(
            "SELECT id FROM prospect_scans WHERE status='running' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        if running:
            return None
        return connection.execute(
            "INSERT INTO prospect_scans(started_at,status) VALUES (?,'running')",
            (now,),
        ).lastrowid


def run_scan(scan_id: int | None = None) -> dict:
    init_prospect_db()
    limit = _scan_limit()
    if scan_id is None:
        with get_prospect_db() as connection:
            scan_id = connection.execute(
                "INSERT INTO prospect_scans(started_at,status) VALUES (?,'running')",
                (_now(),),
            ).lastrowid
    found = added = updated = 0
    sources, errors = [], []
    expected_errors = (OSError, ValueError, KeyError, TypeError, AttributeError, json.JSONDecodeError, urllib.error.URLError)
    for source_name, scanner in (("data.gouv / DGEFP", _data_gouv_rows), ("RNE", _rne_rows), ("Web", _web_rows)):
        try:
            rows = scanner(limit)
            if rows or source_name != "Web":
                sources.append(source_name)
            for prospect in rows:
                found += 1
                if _upsert(prospect):
                    added += 1
                else:
                    updated += 1
        except expected_errors as exc:
            logger.warning("Échec scanner %s: %s", source_name, exc)
            errors.append(f"{source_name}: {exc}")
        except Exception as exc:
            # Une source externe ou une ligne mal formée ne doit pas faire tomber toute la page admin.
            logger.exception("Erreur inattendue pendant le scanner %s", source_name)
            errors.append(f"{source_name}: erreur inattendue ({type(exc).__name__})")
    with get_prospect_db() as connection:
        connection.execute(
            "UPDATE prospect_scans SET finished_at=?,status=?,sources=?,found_count=?,added_count=?,updated_count=?,error_message=? WHERE id=?",
            (_now(), "partial" if errors else "success", ", ".join(sources), found, added, updated, " | ".join(errors), scan_id),
        )
    return {"found": found, "added": added, "updated": updated, "errors": errors}


def _run_scan_in_background(app, scan_id: int) -> None:
    with app.app_context():
        try:
            run_scan(scan_id)
        except Exception as exc:
            logger.exception("Échec du scan de prospection en arrière-plan")
            try:
                with get_prospect_db() as connection:
                    connection.execute(
                        """UPDATE prospect_scans
                           SET finished_at=?, status='failed', error_message=?
                           WHERE id=? AND status='running'""",
                        (_now(), f"{type(exc).__name__}: {exc}", scan_id),
                    )
            except Exception:
                logger.exception("Impossible d'enregistrer l'échec du scan %s", scan_id)


def start_background_scan() -> bool:
    scan_id = _create_scan_run()
    if scan_id is None:
        return False

    app = current_app._get_current_object()
    thread = threading.Thread(
        target=_run_scan_in_background,
        args=(app, scan_id),
        name=f"prospect-scan-{scan_id}",
        daemon=True,
    )
    thread.start()
    return True


def _openai_mail(prospect: sqlite3.Row) -> str:
    fallback = f"Objet : Échange autour de votre actualité formation sécurité\n\nBonjour,\n\nNous avons identifié {prospect['name']} à la suite du signal suivant : {prospect['type_signal_recent']}.\n\nIntégrale Academy accompagne les acteurs de la formation sécurité. Seriez-vous disponible pour un échange de 15 minutes ?\n\nBien cordialement,\nIntégrale Academy"
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key: return fallback
    prompt = f"Rédige un email B2B français de moins de 170 mots, sans invention. Prospect: {prospect['name']}; ville: {prospect['city']}; signal récent: {prospect['type_signal_recent']} le {prospect['date_signal_recent']}; raison: {prospect['raison_detection']}. CTA: échange de 15 minutes."
    body = json.dumps({"model": os.environ.get("OPENAI_MODEL", "gpt-5-mini"), "input": prompt}).encode()
    req = urllib.request.Request("https://api.openai.com/v1/responses", data=body, method="POST", headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=45) as response: payload = json.loads(response.read().decode())
        text = _clean(payload.get("output_text"))
        if not text:
            text = "\n".join(content.get("text", "") for item in payload.get("output", []) for content in item.get("content", []) if content.get("type") == "output_text").strip()
        return text or fallback
    except (OSError, urllib.error.URLError, json.JSONDecodeError):
        return fallback


def _filter_clause(filter_name: str) -> tuple[str, list]:
    today = date.today()
    cutoffs = {days: (today - timedelta(days=days)).isoformat() for days in (7, 30, 90)}
    if filter_name in {"recent_7", "recent_30", "recent_90"}:
        days = int(filter_name.split("_")[1]); return "archive=0 AND est_recent=1 AND date_signal_recent>=?", [cutoffs[days]]
    if filter_name in {"company_30", "company_90"}:
        days = int(filter_name.split("_")[1]); return "archive=0 AND date_creation_entreprise>=?", [cutoffs[days]]
    if filter_name in {"establishment_30", "establishment_90"}:
        days = int(filter_name.split("_")[1]); return "archive=0 AND date_creation_etablissement>=?", [cutoffs[days]]
    type_filters = {"new_training_org": "Nouvel organisme de formation", "recent_qualiopi": "Qualiopi récent", "recent_recruitment": "Recrutement formateur sécurité"}
    if filter_name in type_filters: return "archive=0 AND type_signal_recent=?", [type_filters[filter_name]]
    if filter_name == "recent_security_page": return "archive=0 AND type_signal_recent IN ('Nouvelle page formation sécurité','Ouverture centre détectée')", []
    if filter_name == "archives": return "archive=1", []
    if filter_name == "all": return "1=1", []
    return "archive=0 AND est_recent=1 AND date_signal_recent>=?", [cutoffs[90]]


@prospecting_bp.get("/prospection")
def prospecting_shortcut():
    return redirect(url_for("prospecting.admin_prospects"))


@prospecting_bp.get("/admin")
def admin_prospects():
    init_prospect_db()
    search, status = _clean(request.args.get("q")), _clean(request.args.get("status"))
    signal_filter = _clean(request.args.get("signal_filter")) or "recent_90"
    minimum_score = request.args.get("score", type=int) or 0
    base_clause, parameters = _filter_clause(signal_filter)
    clauses = [base_clause, "score>=?"]; parameters.append(minimum_score)
    if search:
        clauses.append("(name LIKE ? OR siren LIKE ? OR siret LIKE ? OR city LIKE ? OR raison_detection LIKE ?)"); parameters.extend([f"%{search}%"] * 5)
    if status in STATUSES: clauses.append("commercial_status=?"); parameters.append(status)
    with get_prospect_db() as connection:
        prospects = connection.execute(f"SELECT * FROM prospects WHERE {' AND '.join(clauses)} ORDER BY score DESC, date_signal_recent DESC, date_detection DESC LIMIT 1000", parameters).fetchall()
        stats = connection.execute("SELECT COUNT(*) total, SUM(est_recent=1 AND archive=0) new_count, SUM(commercial_status='À relancer' AND archive=0) followup_count, COALESCE(ROUND(AVG(CASE WHEN archive=0 THEN score END)),0) average_score FROM prospects").fetchone()
        last_scan = connection.execute("SELECT * FROM prospect_scans ORDER BY id DESC LIMIT 1").fetchone()
    scan_running = bool(last_scan and last_scan["status"] == "running")
    return render_template("admin_prospects.html", prospects=prospects, stats=stats, last_scan=last_scan, statuses=STATUSES,
        signal_filters=SIGNAL_FILTERS, filters={"q": search, "status": status, "score": minimum_score, "signal_filter": signal_filter},
        scan_running=scan_running, openai_enabled=bool(os.environ.get("OPENAI_API_KEY")),
        web_enabled=bool(os.environ.get("SERPER_API_KEY")))


@prospecting_bp.route("/cron-prospects-scan", methods=["GET", "POST"])
def cron_scan_prospects():
    expected = os.environ.get("CRON_SECRET")
    provided = request.headers.get("Authorization", "").removeprefix("Bearer ") or request.args.get("key")
    if not expected or provided != expected: return jsonify({"ok": False, "error": "unauthorized"}), 401
    return jsonify({"ok": True, **run_scan()})


@prospecting_bp.post("/admin/scan")
def scan_prospects():
    try:
        started = start_background_scan()
    except Exception:
        logger.exception("Impossible de démarrer le scan de prospection")
        flash("Le scan n'a pas pu démarrer. Réessayez dans quelques instants.", "error")
        return redirect(url_for("prospecting.admin_prospects"))

    if started:
        flash("Le scan a démarré en arrière-plan. La page sera actualisée automatiquement.", "success")
    else:
        flash("Un scan est déjà en cours.", "warning")
    return redirect(url_for("prospecting.admin_prospects"))


@prospecting_bp.post("/admin/prospects/<int:prospect_id>/update")
def update_prospect(prospect_id: int):
    status = _clean(request.form.get("commercial_status")); status = status if status in STATUSES else "À qualifier"
    with get_prospect_db() as connection:
        connection.execute("UPDATE prospects SET commercial_status=?,comment=?,updated_at=? WHERE id=?", (status, _clean(request.form.get("comment")), _now(), prospect_id))
    flash("Prospect mis à jour.", "success"); return redirect(url_for("prospecting.admin_prospects"))


@prospecting_bp.post("/admin/prospects/<int:prospect_id>/contacted")
def contact_prospect(prospect_id: int):
    with get_prospect_db() as connection: connection.execute("UPDATE prospects SET commercial_status='Contacté',updated_at=? WHERE id=?", (_now(), prospect_id))
    flash("Prospect marqué comme contacté.", "success"); return redirect(url_for("prospecting.admin_prospects"))


@prospecting_bp.post("/admin/prospects/<int:prospect_id>/follow-up")
def follow_up_prospect(prospect_id: int):
    with get_prospect_db() as connection: connection.execute("UPDATE prospects SET commercial_status='À relancer',updated_at=? WHERE id=?", (_now(), prospect_id))
    flash("Prospect ajouté aux relances.", "success"); return redirect(url_for("prospecting.admin_prospects"))


@prospecting_bp.post("/admin/prospects/<int:prospect_id>/delete")
def delete_prospect(prospect_id: int):
    with get_prospect_db() as connection: connection.execute("DELETE FROM prospects WHERE id=?", (prospect_id,))
    flash("Prospect supprimé.", "success"); return redirect(url_for("prospecting.admin_prospects"))


@prospecting_bp.get("/admin/prospects/<int:prospect_id>/mail")
def prepare_mail(prospect_id: int):
    init_prospect_db()
    with get_prospect_db() as connection: prospect = connection.execute("SELECT * FROM prospects WHERE id=?", (prospect_id,)).fetchone()
    return Response(_openai_mail(prospect), content_type="text/plain; charset=utf-8") if prospect else ("Prospect introuvable", 404)


@prospecting_bp.get("/admin/export.xlsx")
def export_prospects():
    init_prospect_db()
    with get_prospect_db() as connection: prospects = connection.execute("SELECT * FROM prospects ORDER BY archive ASC, score DESC, date_signal_recent DESC, date_detection DESC").fetchall()
    workbook = Workbook(); sheet = workbook.active; sheet.title = "Prospects sécurité"
    headers = ["Score", "Nom", "SIREN", "SIRET", "Ville", "Département", "Date création entreprise", "Date création établissement", "Type signal récent", "Date signal récent", "Ancienneté signal jours", "Est récent", "Archive", "Source", "Raison détection", "Statut commercial", "Commentaire"]
    sheet.append(headers)
    for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="171717")
    for p in prospects:
        sheet.append([p["score"], p["name"], p["siren"], p["siret"], p["city"], p["department"], p["date_creation_entreprise"], p["date_creation_etablissement"], p["type_signal_recent"], p["date_signal_recent"], p["anciennete_signal_jours"], "Oui" if p["est_recent"] else "Non", "Oui" if p["archive"] else "Non", p["source"], p["raison_detection"], p["commercial_status"], p["comment"]])
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate((9,32,15,18,20,14,22,24,34,20,22,12,12,26,55,20,38), 1): sheet.column_dimensions[get_column_letter(index)].width = width
    output = io.BytesIO(); workbook.save(output); output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"prospects-securite-{date.today().isoformat()}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
